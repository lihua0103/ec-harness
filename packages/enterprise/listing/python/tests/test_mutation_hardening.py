"""变异补强：错误路径回执标志、拦截层纯函数性、边界常量、协议循环。"""
import io
import json
import sys

import pandas as pd
import pytest

import worker
from discovery import read_spec_files
from test_worker_dispatch import inspect_and_read
from source_registry import SourceTag


# ---------------------------------------------------------------------------
# 错误路径回执标志（mutant: ok/retryable 标志翻转）
# ---------------------------------------------------------------------------

def test_error_receipts_carry_ok_false(project):
    cases = [
        {"operation": "listing_inspect", "project": str(project / "missing")},
        {"operation": "listing_publish", "project": str(project), "scenario": "manual"},
        {"operation": "listing_nope"},
    ]
    for request in cases:
        result = worker.dispatch(request)
        assert result["ok"] is False, request


def test_duplicate_dataset_representations_are_deduplicated(tmp_path):
    (tmp_path / "one").mkdir(); (tmp_path / "two").mkdir()
    (tmp_path / "one" / "AE.csv").write_text("ID\n1\n")
    (tmp_path / "two" / "AE.csv").write_text("ID\n2\n")
    inspect = worker.dispatch({"operation": "listing_inspect", "project": str(tmp_path)})
    assert inspect["ok"] is True
    assert [item["name"] for item in inspect["inspection"]["datasets"]] == ["AE"]
    run = worker.dispatch({
        "operation": "listing_run_code", "project": str(tmp_path), "code": "outputs = {}"})
    assert run["ok"] is False and run["code"] == "INVALID_OUTPUTS"

    other = tmp_path / "broken"
    other.mkdir()
    (other / "AE.csv").write_text("ID\n1\n")
    (other / "bad.zip").write_bytes(b"not a zip")
    load = worker.dispatch({
        "operation": "listing_run_code", "project": str(other), "code": "outputs = {}"})
    assert load["ok"] is False and load["code"] == "INVALID_OUTPUTS"


def test_run_code_error_flags(project):
    inspect_and_read(project)
    result = worker.dispatch({"operation": "listing_run_code", "project": str(project), "code": "1/0"})
    assert result["ok"] is False and result["retryable"] is True
    bad = worker.dispatch({
        "operation": "listing_run_code", "project": str(project), "code": "outputs = {42: pd.DataFrame()}"})
    assert bad["ok"] is False and bad["code"] == "INVALID_OUTPUTS" and bad["retryable"] is True
    not_dict = worker.dispatch({
        "operation": "listing_run_code", "project": str(project), "code": "outputs = 'oops'"})
    assert not_dict["ok"] is False and not_dict["code"] == "INVALID_OUTPUTS"
    blank_name = worker.dispatch({
        "operation": "listing_run_code", "project": str(project), "code": 'outputs = {" ": datasets["AE"]}'})
    assert blank_name["ok"] is False and blank_name["code"] == "INVALID_OUTPUTS"


def test_publish_error_on_bad_labels(project):
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]="oops"\noutputs={"AE":out}'})
    result = worker.dispatch({
        "operation": "listing_publish", "project": str(project), "scenario": "manual"})
    assert result["ok"] is False and result["code"] == "PUBLISH_ERROR"


def test_publish_track_changes_default_true(project):
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'})
    worker.dispatch({"operation": "listing_publish", "project": str(project), "scenario": "rbqm"})
    log = project / ".clinical-listing" / "output" / "rbqm" / "RBQM_LISTINGS_changes.json"
    assert log.exists()
    body = json.loads(log.read_text(encoding="utf-8"))
    assert body["changes"]["AE"] == {"new": 2, "modified": 0, "old": 0}


def test_republish_same_data_counts_zero(project):
    inspect_and_read(project)
    code = ('out = datasets["AE"].copy()\n'
            'out.attrs["labels"]={"USUBJID":"Subject","AETERM":"Term"}\noutputs={"AE":out}')
    worker.dispatch({"operation": "listing_run_code", "project": str(project), "code": code})
    first = worker.dispatch({"operation": "listing_publish", "project": str(project), "scenario": "manual"})
    second = worker.dispatch({"operation": "listing_publish", "project": str(project), "scenario": "manual"})
    assert first["ok"] and second["ok"]
    assert second["receipt"]["outputCount"] == 1
    from openpyxl import load_workbook
    ws = load_workbook(project / ".clinical-listing" / "output" / "manual" / "MANUAL_LISTINGS.xlsx")["Content"]
    assert [ws.cell(3, c).value for c in range(4, 8)] == [2, 0, 0, 0]   # Total/New/Modified/Old


def test_calculate_changes_counts_zero_for_identical():
    from excel import calculate_changes
    frame = pd.DataFrame({"A": [1, 2]})
    changes = calculate_changes({"KEPT": frame.copy()}, {"KEPT": frame.copy()})
    assert changes == {"KEPT": {"new": 0, "modified": 0, "old": 0}}


def test_report_scenario_reserves_cover_not_content(tmp_path):
    from excel import create_multi_sheet_excel
    frame = pd.DataFrame({"A": [1]})
    frame.attrs["report_metadata"] = {"sponsor": "s"}
    output = tmp_path / "REPORT_LISTINGS.xlsx"
    stats = create_multi_sheet_excel({"Content": frame}, output, "report", track_changes=False)
    assert stats["sheetNames"] == ["Cover Page", "Content"]
    assert stats["reportStructureApplied"] is True
    assert stats["rbqmStructureFlexible"] is False


def test_rbqm_statistics_flag(tmp_path):
    from excel import create_multi_sheet_excel
    frame = pd.DataFrame({"A": [1]})
    frame.attrs["labels"] = {"A": "a"}
    stats = create_multi_sheet_excel({"R": frame}, tmp_path / "RBQM_LISTINGS.xlsx", "rbqm", track_changes=False)
    assert stats["rbqmStructureFlexible"] is True
    assert stats["standardStructureApplied"] is False


def test_invalid_outputs_shapes_rejected(tmp_path):
    from excel import create_multi_sheet_excel
    frame = pd.DataFrame({"A": [1]})
    frame.attrs["labels"] = {"A": "a"}
    for bad in ({}, {" ": frame}, {"x": 42}):
        with pytest.raises((ValueError, TypeError)):
            create_multi_sheet_excel(bad, tmp_path / "x.xlsx", "manual", track_changes=False)


def test_change_log_exact_format(project):
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"受试表":out}'})
    worker.dispatch({"operation": "listing_publish", "project": str(project), "scenario": "rbqm"})
    raw = (project / ".clinical-listing" / "output" / "rbqm" / "RBQM_LISTINGS_changes.json").read_text(encoding="utf-8")
    assert "受试表" in raw                    # ensure_ascii=False：中文 sheet 名原样落盘
    assert '\n  "' in raw                     # indent=2


# ---------------------------------------------------------------------------
# 拦截层纯函数性 / 恒等 / 审计断言见 test_data_guard.py
# ---------------------------------------------------------------------------



# ---------------------------------------------------------------------------
# 边界常量（mutant: 上限 ±1）
# ---------------------------------------------------------------------------

def test_stream_contents_not_returned(project):
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": f'print("Z" * 20000)\nout = datasets["AE"].copy()\nout.attrs["labels"]={{"USUBJID":"S"}}\noutputs={{"AE":out}}'})
    assert result["receipt"]["stdoutOmitted"] is True
    assert result["receipt"]["stderrOmitted"] is True


def test_line_count_exact(tmp_path):
    doc = tmp_path / "doc"
    doc.mkdir()
    (doc / "a.txt").write_text("l1\nl2\nl3", encoding="utf-8")
    documents, _ = read_spec_files(doc)
    assert documents[0]["lineCount"] == 3


def test_spec_excel_has_no_rows_key(tmp_path):
    doc = tmp_path / "doc"
    doc.mkdir()
    pd.DataFrame({"A": [1]}).to_excel(doc / "a.xlsx", index=False)
    documents, _ = read_spec_files(doc)
    assert documents[0]["rows"][0]["rows"] == [[1]]
    assert documents[0]["structure"]["sheets"][0]["headerRows"] == [["A"]]


def test_file_kind_for_zip(tmp_path):
    from discovery import list_files
    (tmp_path / "a.zip").write_bytes(b"PK")
    kinds = {entry["path"]: entry["kind"] for entry in list_files(tmp_path)}
    assert kinds["a.zip"] == "archive"


def test_source_tag_frozen():
    tag = SourceTag.of("model-output")
    with pytest.raises(Exception):
        tag.source = "dataset"


# ---------------------------------------------------------------------------
# 协议主循环（mutant: WORKER_ERROR / ensure_ascii）
# ---------------------------------------------------------------------------

def test_main_loop_protocol(monkeypatch, capsys, project):
    lines = "\n".join([
        json.dumps({"operation": "listing_inspect", "project": str(project)}),
        json.dumps(["not", "a", "dict"]),                       # 触发 WORKER_ERROR
        json.dumps({"operation": "listing_read_document", "project": str(project),
                    "documentId": "doc-000001", "chunkIndex": 0}),
        json.dumps({"operation": "listing_run_code", "project": str(project),
                    "code": 'print("受试者")\nout = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'}),
    ])
    monkeypatch.setattr(sys, "stdin", io.StringIO(lines + "\n"))
    worker._reset_session()
    worker.main()
    out = capsys.readouterr().out
    responses = [json.loads(line) for line in out.splitlines()]
    assert responses[0]["ok"] is True
    assert responses[1]["ok"] is False and responses[1]["code"] == "WORKER_ERROR"
    assert responses[2]["document"]["isFinal"] is True
    assert "需求文档" in responses[2]["document"]["content"]
    assert "受试者" not in out                                  # stdout 数据值不回流
    assert responses[3]["receipt"]["stdoutOmitted"] is True
    assert "需求文档" in out                                     # ensure_ascii=False：doc 分片原样出协议


def test_publish_returns_safe_stage_code(project, monkeypatch):
    from excel import ListingPublishError
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'})
    monkeypatch.setattr(worker, "create_multi_sheet_excel",
                        lambda *args, **kwargs: (_ for _ in ()).throw(ListingPublishError("RENDER_WORKBOOK_FAILED")))
    result = worker.dispatch({"operation": "listing_publish", "project": str(project), "scenario": "manual"})
    assert result == {"ok": False, "code": "PUBLISH_ERROR", "reason": "发布失败",
                      "stage": "RENDER_WORKBOOK_FAILED", "retryable": True}
