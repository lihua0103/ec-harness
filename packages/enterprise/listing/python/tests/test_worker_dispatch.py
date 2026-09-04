"""Worker 端到端：三操作调度 + 数据拦截出口 + 会话语义 + 审计。"""
import json

import worker


def read_requirement_documents(project, inspection):
    documents = []
    for manifest in inspection["requirementDocuments"]:
        chunks = []
        for index in range(manifest["totalChunks"]):
            result = worker.dispatch({
                "operation": "listing_read_document", "project": str(project),
                "documentId": manifest["documentId"], "chunkIndex": index,
            })
            assert result["ok"] is True, result
            chunks.append(result["document"]["content"])
        documents.append(json.loads("".join(chunks)))
    return documents


def inspect_and_read(project, **extra):
    result = worker.dispatch({"operation": "listing_inspect", "project": str(project), **extra})
    assert result["ok"] is True, result
    read_requirement_documents(project, result["inspection"])
    return result

def test_inspect_projects_both_data_value_classes(project):
    from openpyxl import Workbook
    aux = Workbook(); aux.active.title = "ALS"
    aux.active.append(["Dataset Name", "Variable Name", "Label"])
    aux.active.append(["AE", "AETERM", "Term"])
    aux.save(project / "aux.xlsx")

    wb = Workbook(); ws = wb.active; ws.title = "ALS"
    ws.append(["Dataset Name", "Variable Name", "Label"])
    ws.append(["AE", "AETERM", "Term"])
    notes = wb.create_sheet("Notes")
    notes.append(["note"])
    for _ in range(3):
        notes.append(["padding"])
    notes.append(["SECRET-CELL-42"])
    wb.save(project / "doc" / "als.xlsx")

    result = inspect_and_read(project)
    payload = json.dumps(result, ensure_ascii=False)
    assert result["ok"] is True
    assert "SUBJ-777" not in payload and "SUBJ-888" not in payload     # 数据集行值
    assert "SECRET-CELL-42" not in payload                             # doc 外辅助 Excel 单元格值
    assert "DOC-CELL-99" not in payload                                # inspect 只给 doc manifest
    inspection = result["inspection"]
    documents = read_requirement_documents(project, inspection)
    assert "REQUIREMENT-TAIL" in json.dumps(documents, ensure_ascii=False)
    assert "SECRET-CELL-42" in json.dumps(documents, ensure_ascii=False)
    assert inspection["supportedScenarios"] == ["manual", "medical", "rbqm", "report"]
    datasets = inspection["datasets"][0]
    assert datasets["name"] == "AE" and datasets["rowCount"] == 2
    assert datasets["columns"] == ["USUBJID", "AETERM"]
    assert "sample" not in datasets                                    # 构建期节流：根本没建
    als = inspection["auxiliaryDocuments"][0]
    assert "rows" not in als
    assert als["mappings"][0]["datasetName"] == "AE"                   # ALS 语义映射保留


def test_inspect_ignores_model_interception_false_request(project):
    from openpyxl import Workbook
    wb = Workbook(); wb.active.append(["note"]); wb.active.append(["SECRET-CELL-42"])
    wb.save(project / "notes.xlsx")

    result = worker.dispatch(
        {"operation": "listing_inspect", "project": str(project), "dataInterception": False})
    payload = json.dumps(result, ensure_ascii=False)
    assert result["ok"] is True
    assert "SUBJ-777" not in payload and "SUBJ-888" not in payload
    assert "SECRET-CELL-42" not in payload
    assert "REQUIREMENT-TAIL" not in payload
    documents = read_requirement_documents(project, result["inspection"])
    assert "REQUIREMENT-TAIL" in json.dumps(documents)
    assert "sample" not in result["inspection"]["datasets"][0]
    assert all("rows" not in doc for doc in result["inspection"]["auxiliaryDocuments"])


def test_inspect_writes_audit_jsonl(project):
    inspect_and_read(project)
    audit_file = project / ".clinical-listing" / "audit.jsonl"
    assert audit_file.exists()
    record = json.loads(audit_file.read_text(encoding="utf-8").splitlines()[-1])
    assert record["operation"] == "listing_inspect"
    assert {"source": "dataset", "path": "AE.csv"} in record["projections"]
    assert json.dumps(record).count("SUBJ") == 0                       # 审计无数据值
    # 第二次投影不因目录已存在而失败（exist_ok）
    inspect_and_read(project)
    assert len(audit_file.read_text(encoding="utf-8").splitlines()) == 2


def test_audit_writes_non_ascii_path_raw(tmp_path):
    (tmp_path / "受试.csv").write_text("ID\n1\n", encoding="utf-8")
    worker.dispatch({"operation": "listing_inspect", "project": str(tmp_path)})
    raw = (tmp_path / ".clinical-listing" / "audit.jsonl").read_text(encoding="utf-8")
    assert "受试.csv" in raw                                        # ensure_ascii=False：中文原样落审计


def test_inspect_writes_audit_even_for_interception_false_request(project):
    worker.dispatch({"operation": "listing_inspect", "project": str(project), "dataInterception": False})
    assert (project / ".clinical-listing" / "audit.jsonl").exists()


def test_host_disabled_inspection_returns_both_value_classes(project):
    from openpyxl import Workbook
    wb = Workbook(); wb.active.append(["note"]); wb.active.append(["SECRET-CELL-42"])
    wb.save(project / "notes.xlsx")

    result = worker.dispatch({
        "operation": "listing_inspect", "project": str(project),
        "hostDataInterception": False})
    payload = json.dumps(result, ensure_ascii=False)
    assert result["ok"] is True
    assert "SUBJ-777" in payload and "SUBJ-888" in payload
    assert "SECRET-CELL-42" in payload
    assert result["inspection"]["datasets"][0]["rows"][0] == ["SUBJ-777", "Headache"]
    assert result["inspection"]["auxiliaryDocuments"][0]["rows"][0]["rows"] == [["SECRET-CELL-42"]]
    assert not (project / ".clinical-listing" / "audit.jsonl").exists()


def test_inspect_seeds_session_so_run_code_skips_reload(project):
    assert inspect_and_read(project)["ok"]
    # 数据文件在 inspect 后消失：run_code 用会话数据照常工作（免二次读取）
    (project / "AE.csv").unlink()
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'})
    assert result["ok"] is True, result
    assert result["receipt"]["outputCount"] == 1


def test_run_code_requires_all_documents_when_session_cold(project):
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'outputs={"AE": datasets["AE"].copy()}'})
    assert result["ok"] is False
    assert result["code"] == "REQUIREMENTS_NOT_FULLY_READ"

def test_run_code_stdout_value_omitted(project):
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'print(datasets["AE"].iloc[0]["AETERM"])\n'
                'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'})
    payload = json.dumps(result, ensure_ascii=False)
    assert "Headache" not in payload and "SUBJ-777" not in payload
    assert result["receipt"]["stdoutOmitted"] is True
    assert result["receipt"]["stderrOmitted"] is True
    assert "stdout" not in result["receipt"] and "stderr" not in result["receipt"]


def test_run_code_host_disabled_returns_stdout_and_controlled_names(project):
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "hostDataInterception": False,
        "code": 'print(datasets["AE"].iloc[0]["AETERM"])\n'
                'df = datasets["AE"].copy()\n'
                'df = df.rename(columns={"USUBJID": datasets["AE"].iloc[0]["USUBJID"]})\n'
                'outputs={datasets["AE"].iloc[0]["AETERM"]: df}'})
    payload = json.dumps(result, ensure_ascii=False)
    assert result["ok"] is True, result
    assert "Headache" in payload and "SUBJ-777" in payload
    assert result["receipt"]["stdout"] == "Headache\n"
    assert result["receipt"]["outputs"][0]["name"] == "Headache"
    assert result["receipt"]["outputs"][0]["columns"][0]["name"] == "SUBJ-777"


def test_run_code_requires_outputs(project):
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project), "code": "x = 1"})
    assert result["ok"] is False
    assert result["code"] == "OUTPUTS_REQUIRED"
    assert result["reason"] == "代码必须定义 outputs: dict[str, pandas.DataFrame]"
    assert result["stdoutOmitted"] is True and result["stderrOmitted"] is True
    assert "datasets" in result["environmentHint"]                    # 环境自描述（信息供给）


def test_run_code_rejects_invalid_outputs(project):
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project), "code": "outputs = {'bad': 42}"})
    assert result["code"] == "INVALID_OUTPUTS"


def test_run_code_allows_partial_dataset_failure_when_usable_source_remains(tmp_path):
    (tmp_path / "AE.csv").write_text("ID\n1\n", encoding="utf-8")
    (tmp_path / "broken.zip").write_bytes(b"not a zip")
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(tmp_path), "code": "outputs = {}"})
    assert result["ok"] is False
    assert result["code"] == "INVALID_OUTPUTS"

    inspection = worker.dispatch({"operation": "listing_inspect", "project": str(tmp_path)})
    assert inspection["ok"] is True
    assert inspection["inspection"]["failures"][0]["stage"] == "extract-archive"


def test_publish_flow_and_fingerprint(project):
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S","AETERM":"T"}\noutputs={"AE":out}'})
    result = worker.dispatch({
        "operation": "listing_publish", "project": str(project), "scenario": "medical",
        "trackChanges": False})
    assert result["ok"] is True, result
    receipt = result["receipt"]
    assert receipt["format"] == "single-workbook-multi-sheet-xlsx"
    assert receipt["outputFile"] == ".clinical-listing/output/medical/MEDICAL_LISTINGS.xlsx"
    assert (project / receipt["outputFile"]).exists()
    assert receipt["outputCount"] == 1
    assert "statistics" not in receipt and "sheetNames" not in receipt


def test_publish_requires_successful_run(project):
    result = worker.dispatch({"operation": "listing_publish", "project": str(project), "scenario": "manual"})
    assert result["code"] == "NO_SUCCESSFUL_RUN"


def test_publish_rejects_invalid_scenario_before_output_path(project):
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'})
    result = worker.dispatch({
        "operation": "listing_publish", "project": str(project), "scenario": "invalid"})
    assert result["ok"] is False
    assert result["code"] == "INVALID_SCENARIO"
    assert not (project / ".clinical-listing" / "output").exists()


def test_publish_project_mismatch(project, tmp_path_factory):
    other = tmp_path_factory.mktemp("other-project")
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'})
    result = worker.dispatch({
        "operation": "listing_publish", "project": str(other), "scenario": "manual"})
    assert result["ok"] is False
    assert result["code"] == "PROJECT_SESSION_MISMATCH"


def test_unknown_operation():
    assert worker.dispatch({"operation": "listing_nope"})["code"] == "UNKNOWN_OPERATION"


def test_project_not_found(tmp_path):
    result = worker.dispatch({"operation": "listing_inspect", "project": str(tmp_path / "missing")})
    assert result["code"] == "PROJECT_NOT_FOUND"


def test_scenario_is_not_inferred_from_project_name(tmp_path):
    project = tmp_path / "XX-Medical-Study"
    project.mkdir()
    (project / "raw").mkdir(parents=True)
    (project / "raw" / "AE.csv").write_text("ID\n1\n", encoding="utf-8")
    result = worker.dispatch({"operation": "listing_inspect", "project": str(project)})
    assert "inferredScenario" not in result["inspection"]
    assert result["inspection"]["supportedScenarios"] == ["manual", "medical", "rbqm", "report"]


def test_publish_requires_explicit_scenario(project):
    inspect_and_read(project)
    worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": 'out = datasets["AE"].copy()\nout.attrs["labels"]={"USUBJID":"S"}\noutputs={"AE":out}'})
    result = worker.dispatch({"operation": "listing_publish", "project": str(project)})
    assert result["ok"] is False
    assert result["code"] == "SCENARIO_REQUIRED"


def test_error_receipt_hides_dynamic_exception_values(project):
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": "raise ValueError(datasets['AE'].iloc[0]['AETERM'])"})
    payload = json.dumps(result, ensure_ascii=False)
    assert result["ok"] is False
    assert result["code"] == "CODE_EXECUTION_ERROR"
    assert "Headache" not in payload
    assert result["stdoutOmitted"] is True and result["stderrOmitted"] is True


def test_scan_text_keeps_protected_values_in_worker_session(project):
    inspect_and_read(project)
    protected = worker.dispatch({
        "operation": "listing_scan_text", "project": str(project),
        "text": json.dumps({"output": "SUBJ-777"})})
    ordinary = worker.dispatch({
        "operation": "listing_scan_text", "project": str(project),
        "text": "ordinary diagnostic output"})
    assert protected == {"ok": True, "containsProtectedValue": True}
    assert ordinary == {"ok": True, "containsProtectedValue": False}
    assert "SUBJ-777" not in json.dumps(protected)


def test_scan_text_requires_initialized_session(project):
    result = worker.dispatch({
        "operation": "listing_scan_text", "project": str(project), "text": "safe"})
    assert result["ok"] is False
    assert result["code"] == "SCAN_SESSION_NOT_INITIALIZED"


def test_read_metadata_pages_never_return_protected_rows(project):
    inspection = worker.dispatch({"operation": "listing_inspect", "project": str(project)})
    assert inspection["ok"] is True
    result = worker.dispatch({
        "operation": "listing_read_metadata", "project": str(project),
        "pageIndex": 0, "pageSize": 1,
    })
    assert result["ok"] is True
    metadata = result["metadata"]
    assert metadata["totalDatasets"] == 1
    assert metadata["totalPages"] == 1
    assert len(metadata["datasets"]) == 1
    assert "rows" not in metadata["datasets"][0]
    assert all("rows" not in document for document in metadata["auxiliaryDocuments"])
    payload = json.dumps(result, ensure_ascii=False)
    assert "SUBJ-777" not in payload and "Headache" not in payload


def test_read_metadata_rejects_invalid_page(project):
    worker.dispatch({"operation": "listing_inspect", "project": str(project)})
    result = worker.dispatch({
        "operation": "listing_read_metadata", "project": str(project),
        "pageIndex": -1,
    })
    assert result["ok"] is False
    assert result["code"] == "INVALID_METADATA_PAGE"


def test_read_metadata_respects_disabled_interception(project):
    worker.dispatch({"operation": "listing_inspect", "project": str(project),
                     "hostDataInterception": False})
    result = worker.dispatch({"operation": "listing_read_metadata", "project": str(project),
                              "hostDataInterception": False})
    assert result["ok"] is True
    assert "rows" in result["metadata"]["datasets"][0]


def test_run_code_receipt_omits_model_controlled_names(project):
    """输出名/列名可被改成数据值；回执不返回任何名称。"""
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": "df = datasets['AE'].copy()\n"
                "df = df.rename(columns={'USUBJID': datasets['AE'].iloc[0]['USUBJID']})\n"
                "outputs = {datasets['AE'].iloc[0]['AETERM']: df}"})
    assert result["ok"] is True, result
    payload = json.dumps(result, ensure_ascii=False)
    assert "SUBJ-777" not in payload and "Headache" not in payload
    assert result["receipt"]["outputCount"] == 1
    assert all("name" not in column for output in result["receipt"]["outputs"] for column in output["columns"])

def test_model_code_cannot_patch_trusted_receipt(project):
    inspect_and_read(project)
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(project),
        "code": "import worker\nworker._omitted_stream_payload=lambda: {'stdout':'SUBJ-777'}\n"
                "out=datasets['AE'].copy()\noutputs={'AE':out}"})
    payload = json.dumps(result, ensure_ascii=False)
    assert result["ok"] is True, result
    assert "SUBJ-777" not in payload
    assert result["receipt"]["stdoutOmitted"] is True

def test_run_code_blocks_incomplete_inspection_snapshot(tmp_path):
    (tmp_path / "doc").mkdir()
    (tmp_path / "doc" / "broken.xlsx").write_bytes(b"not-an-excel-file")
    (tmp_path / "AE.csv").write_text("ID\n1\n", encoding="utf-8")
    inspection = worker.dispatch({"operation": "listing_inspect", "project": str(tmp_path)})
    assert inspection["ok"] is True
    result = worker.dispatch({
        "operation": "listing_run_code", "project": str(tmp_path),
        "code": "outputs = {'AE': datasets['AE'].copy()}"})
    assert result["ok"] is False
    assert result["code"] == "REQUIREMENTS_LOAD_FAILED"
    assert result["retryable"] is True
