"""Excel 精确契约（pytest 侧锚定）：把 vitest 锁定的几何值在 Python 侧同样钉死。

变异测试发现的缺口：精确样式/几何断言此前只存在于 vitest（worker.test.ts），
pytest 侧未覆盖——渲染层 mutant 因此幸存。本文件补齐。
"""
from openpyxl import load_workbook

from excel import calculate_changes, create_multi_sheet_excel
from excel.templates import build_content_sheet, build_report_cover
import pandas as pd


def _listing_frame():
    frame = pd.DataFrame({"USUBJID": ["S1"], "AETERM": ["Head"]})
    frame.attrs["labels"] = {"USUBJID": "Subject Identifier", "AETERM": "Adverse Event Term"}
    return frame


def _report_frame(columns):
    frame = pd.DataFrame({column: ["v"] for column in columns})
    frame.attrs["report_metadata"] = {"sponsor": "S", "protocol_no": "P",
                                      "project_id": "W", "report_date": "2026-08-27"}
    return frame


def test_degenerate_frames_geometry(tmp_path):
    """退化帧（0 行 / 1 列）的几何下限：筛选、合并、列宽不越界。"""
    empty = _listing_frame().iloc[0:0]
    single = pd.DataFrame({"ONLY": ["v"]})
    single.attrs["labels"] = {"ONLY": "Only Label"}
    single.attrs["_skip_default_template"] = True          # 保持 1 列，测真正的单列分支
    two = pd.DataFrame({"A": [1], "B": [2]})
    two.attrs["labels"] = {"A": "LA", "B": "LB"}
    two.attrs["_skip_default_template"] = True             # 2 列：标题合并恰为 B1:B1
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    create_multi_sheet_excel({"EMPTY": empty, "SINGLE": single, "TWO": two}, output, "manual", track_changes=False)
    wb = load_workbook(output)
    empty_ws, single_ws, two_ws = wb["EMPTY"], wb["SINGLE"], wb["TWO"]
    # 空帧：筛选下限 A2:G2（表头行自身），无数据行
    assert empty_ws.auto_filter.ref == "A2:G2"
    # 单列：不做标题合并（>=2 才合并），B1 不写；1 行数据 → 筛选 A2:A3
    assert single_ws.auto_filter.ref == "A2:A3"
    assert [str(r) for r in single_ws.merged_cells.ranges] == []
    assert single_ws.max_column == 1                     # B1 访问会实例化单元格，先断言列数
    assert single_ws["B1"].value is None
    # 两列：合并恰为 B1（openpyxl 归一化单格区间 B1:B1 → "B1"）
    assert [str(r) for r in two_ws.merged_cells.ranges] == ["B1"]
    assert two_ws["B1"].value == "TWO"

    report_out = tmp_path / "REPORT_LISTINGS.xlsx"
    empty_report = _report_frame(["C1"])
    create_multi_sheet_excel({"EMPTYR": empty_report.iloc[0:0]}, report_out, "report", track_changes=False)
    report_ws = load_workbook(report_out)["EMPTYR"]
    assert report_ws.auto_filter.ref == "A1:A1"          # 报表空帧筛选下限


def test_listing_sheet_exact_geometry(tmp_path):
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    create_multi_sheet_excel({"LISTING_AE_01": _listing_frame()}, output, "manual", track_changes=False)
    ws = load_workbook(output)["LISTING_AE_01"]
    # 2 数据列 + 5 审核列 = 7 列
    assert ws["A1"].value == '=HYPERLINK("#\'Content\'!A1","Go back")'
    assert ws["A1"].font.color.rgb == "FF0000FF" and ws["A1"].font.underline is None
    assert ws["B1"].value == "LISTING_AE_01"
    assert [str(r) for r in ws.merged_cells.ranges] == ["B1:F1"]       # merge_end = min(7,6)
    # 标题带锚点样式：B1 是 SHEET_TITLE 样式（TNR 14 加粗 浅蓝）；
    # 合并区非锚点（C1..F1）回读恒为默认样式（openpyxl 语义），以合并范围为准
    assert (ws["B1"].font.name, ws["B1"].font.sz, ws["B1"].font.bold,
            ws["B1"].fill.fgColor.rgb) == ("Times New Roman", 14, True, "FFEDF2F9")
    assert ws.cell(1, 7).fill.fgColor.rgb == "FFEDF2F9"                # 第 7 列标题带补样式
    assert ws.max_column == 7                                          # 标题带不越界到 H 列
    assert ws.cell(2, 1).value == "Subject Identifier"
    assert [ws.cell(2, c).value for c in range(3, 8)] == [
        "Flag1", "FLAG(New/Modified/Old)", "Update Detail", "Review Comments", "Initial/Date"]
    assert ws.row_dimensions[2].height == 60
    assert [ws.cell(3, c).value for c in (1, 2)] == ["S1", "Head"]      # 数据从第 3 行第 1 列起
    assert ws.auto_filter.ref == "A2:G3"                                # 表头行 + 数据行
    assert ws.freeze_panes == "A3" and ws.sheet_view.showGridLines is False
    assert ws.cell(3, 1).font.name == "Times New Roman" and ws.cell(3, 1).font.sz == 13
    assert ws.cell(3, 1).fill.fgColor.rgb == "FFFFFFFF"


def test_listing_column_width_heuristic(tmp_path):
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    create_multi_sheet_excel({"W": _listing_frame()}, output, "manual", track_changes=False)
    ws = load_workbook(output)["W"]
    # "Subject Identifier" 18 字符 → 18+2=20；"Adverse Event Term" 18 → 20
    assert ws.column_dimensions["A"].width == 20.0
    assert ws.column_dimensions["B"].width == 20.0


def test_content_sheet_exact_geometry():
    from openpyxl import Workbook
    outputs = {"A": _listing_frame()}
    wb = Workbook()
    build_content_sheet(wb, outputs, calculate_changes(None, outputs))
    ws = wb["Content"]
    assert ws["A1"].value == "Comparison Summary"
    assert (ws["A1"].font.name, ws["A1"].font.sz, ws["A1"].font.bold, ws["A1"].fill.fgColor.rgb) \
        == ("Times New Roman", 16, True, "FFEDF2F9")
    assert [ws.cell(1, c).fill.fgColor.rgb for c in range(1, 8)] == ["FFEDF2F9"] * 7
    assert ws.cell(3, 1).value == 1                                     # Listing Seq = row - 2
    assert ws["B3"].font.color.rgb == "FF0000FF" and ws["B3"].font.underline == "single"
    assert ws["B3"].alignment.horizontal == "left"
    assert ws.max_column == 7                                           # 表格只到 G 列


def test_report_cover_exact_geometry():
    from openpyxl import Workbook
    wb = Workbook()
    build_report_cover(wb, {key: f"v-{key}" for key in ("sponsor", "protocol_no", "project_id", "report_date")})
    cover = wb["Cover Page"]
    a1 = cover["A1"]
    assert (a1.font.name, a1.font.sz, a1.font.bold, a1.fill.fgColor.rgb) == ("宋体", 16, True, "FFD9D9D9")
    assert (a1.border.right.style, a1.border.bottom.style) == ("medium", "medium")
    label = cover["A3"]
    assert (label.font.name, label.font.sz, label.font.bold, label.fill.fgColor.rgb) \
        == ("宋体", 14, True, "FFD9D9D9")
    assert (label.border.right.style, label.border.top.style, label.border.bottom.style) == ("thin",) * 3
    assert label.alignment.horizontal == "left" and label.alignment.wrap_text is True
    value = cover["B3"]
    assert (value.font.name, value.font.sz, value.font.bold) == ("微软雅黑", 16, True)
    assert value.alignment.wrap_text is True
    assert [cover.row_dimensions[r].height for r in range(1, 7)] == [75, 12.6, 54, 47.25, 54, 39.75]
    assert cover.column_dimensions["A"].width == 38.140625
    assert cover.max_column == 7                                          # 样式带只到 G 列


def test_report_sheet_exact_geometry(tmp_path):
    output = tmp_path / "REPORT_LISTINGS.xlsx"
    create_multi_sheet_excel(
        {"Missing Page": _report_frame(["Site Name", "Site Number", "Subject Number"])},
        output, "report", track_changes=False)
    ws = load_workbook(output)["Missing Page"]
    header = ws["A1"]
    assert (header.font.name, header.font.sz, header.font.bold, header.fill.fgColor.rgb) \
        == ("Calibri", 12, True, "FFC5D9F1")
    assert header.alignment.horizontal == "center" and header.alignment.wrap_text is True
    assert ws.row_dimensions[1].height == 31.5
    assert [ws.column_dimensions[c].width for c in "ABC"] == [59.7109375, 13.7109375, 17.140625]
    assert [ws.cell(2, c).value for c in range(1, 4)] == ["v", "v", "v"]   # 数据从第 2 行起
    assert ws.auto_filter.ref == "A1:C2"
    assert ws.freeze_panes == "A2"


def test_report_template_width_boundary_and_fallback(tmp_path):
    output = tmp_path / "REPORT_LISTINGS.xlsx"
    create_multi_sheet_excel(
        {"Missing Page": _report_frame(["Only"]), "Dynamic Sheet": _report_frame(["Only"])},
        output, "report", track_changes=False)
    wb = load_workbook(output)
    # 列号 == 模板宽度长度边界：Missing Page 模板有 18 个宽度，第 1 列取模板值
    assert wb["Missing Page"].column_dimensions["A"].width == 59.7109375
    # 非模板 Sheet：回退 max(13.0, len+2)，封顶 100.71
    assert wb["Dynamic Sheet"].column_dimensions["A"].width == 13.0


def test_report_width_exact_boundary_and_fallback_formula(tmp_path):
    """18 列恰触模板边界；回退宽度公式 len+2（非地板值）。"""
    from openpyxl.utils import get_column_letter

    output = tmp_path / "REPORT_LISTINGS.xlsx"
    columns18 = [f"C{index}" for index in range(18)]
    long_header = "VeryLongHeaderX"      # 15 字符 → 15+2 = 17
    create_multi_sheet_excel({
        "Missing Page": _report_frame(columns18),
        "Dynamic Sheet": _report_frame([long_header]),
    }, output, "report", track_changes=False)
    wb = load_workbook(output)
    # 第 18 列恰在模板长度边界上（column <= len → 模板值 12.42578125）
    assert wb["Missing Page"].column_dimensions[get_column_letter(18)].width == 12.42578125
    # 回退公式：len("VeryLongHeaderX")+2 = 17（地板 13 不生效）
    assert wb["Dynamic Sheet"].column_dimensions["A"].width == 17.0


def test_direct_excel_default_track_changes_writes_log(tmp_path):
    """直调 create_multi_sheet_excel 不传 track_changes → 默认 True 写变化日志。"""
    output = tmp_path / "RBQM_LISTINGS.xlsx"
    create_multi_sheet_excel({"R": _listing_frame()}, output, "rbqm")
    assert (tmp_path / "RBQM_LISTINGS_changes.json").exists()


def test_report_row_height_from_template(tmp_path):
    output = tmp_path / "REPORT_LISTINGS.xlsx"
    create_multi_sheet_excel(
        {"Matrix by Site": _report_frame(["A"]), "Dynamic Sheet": _report_frame(["A"])},
        output, "report", track_changes=False)
    wb = load_workbook(output)
    assert wb["Matrix by Site"].row_dimensions[1].height == 63.0
    assert wb["Dynamic Sheet"].row_dimensions[1].height == 31.5
