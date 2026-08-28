"""漏洞扫描 V-5 回归:Excel 公式注入中和(字面量写入 + HYPERLINK 引号转义)。"""
from pathlib import Path

from openpyxl import load_workbook

from excel import create_multi_sheet_excel


def test_literal_cell_neutralizes_formula_strings(tmp_path):
    from excel.templates import literal_cell
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    literal_cell(ws, 1, 1, "=1+1")
    literal_cell(ws, 2, 1, "=WEBSERVICE(\"http://evil\")")
    literal_cell(ws, 3, 1, "普通文本")
    literal_cell(ws, 4, 1, 42)
    assert ws.cell(1, 1).data_type == "s" and ws.cell(1, 1).value == "=1+1"
    assert ws.cell(2, 1).data_type == "s"
    assert ws.cell(3, 1).data_type == "s"
    assert ws.cell(4, 1).data_type == "n"


def _evil_outputs():
    import pandas as pd
    frame = pd.DataFrame({"A": ["=CMD(\"calc\")", "=1+1", "ok"]})
    frame.attrs["labels"] = {"A": "=HYPERLINK(\"http://evil\",\"x\")"}
    return {"=bad\"name": frame}


def test_publish_pipeline_neutralizes_all_model_strings(tmp_path):
    output = tmp_path / "out.xlsx"
    sheet_name = "=bad\"name"
    create_multi_sheet_excel(_evil_outputs(), output, scenario="rbqm", track_changes=False)
    wb = load_workbook(output, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if ws.title == "Content" and cell.coordinate == "B3":
                        continue  # 设计内跳转公式,由下方精确断言
                    if cell.coordinate == "A1":
                        continue  # 设计内固定返回链接,由下方精确断言
                    assert cell.data_type == "s", f"{ws.title}!{cell.coordinate} 未中和"
    content = wb["Content"]
    from excel.templates import hyperlink_formula
    assert content["B3"].value == hyperlink_formula(sheet_name)   # 引号双写后的精确形态
    business = wb[sheet_name]
    assert business["A1"].value == '=HYPERLINK("#\'Content\'!A1","Go back")'  # 固定常量未被模型污染


def test_content_link_formula_escapes_sheet_name_quotes(tmp_path):
    import pandas as pd
    frame = pd.DataFrame({"A": ["x"]})
    frame.attrs["labels"] = {"A": "A"}
    outputs = {'sh"eet': frame}
    output = tmp_path / "out.xlsx"
    create_multi_sheet_excel(outputs, output, scenario="medical", track_changes=False)
    wb = load_workbook(output, data_only=False)
    content = wb["Content"]
    formulas = [c.value for row in content.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=HYPERLINK")]
    assert formulas, "Content 页应存在跳转公式"
    # 引号双写:公式字符串里不应出现未成对的 " 突破(结构性断言:值可被 openpyxl 完整回读)
    assert any('sh""eet' in f for f in formulas), formulas
