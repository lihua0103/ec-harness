"""自定义排版（ADR-0022）：_layout 解析 + 渲染 + 变化计数回退。"""
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel import create_multi_sheet_excel
from excel.layout import read_layout


def _frame():
    frame = pd.DataFrame({"A": ["a1", "a2"], "B": ["b1", "b2"], "C": ["c1", "c2"]})
    frame.attrs["labels"] = {"A": "Alpha", "B": "Beta", "C": "Gamma"}
    return frame


def test_read_layout_none_when_absent():
    assert read_layout(_frame()) is None


def test_parse_defaults():
    frame = _frame()
    frame.attrs["_layout"] = {"header_rows": 2}
    layout = read_layout(frame)
    assert layout.anchor_cell == (3, 1)               # 表头带下一行
    assert layout.freeze_panes is None                 # 渲染时回退 A{anchor}
    assert layout.back_link is not None                # 默认写 A1 返回链接
    assert layout.header_columns is None


def test_parse_explicit():
    frame = _frame()
    spec = {
        "header_rows": 3,
        "header_columns": [["组1", "组1", "组2"], ["组1", "组2", "组2"], ["A", "B", "C"]],
        "anchor_cell": [4, 1],
        "freeze_panes": "A4",
        "back_link": None,
        "column_widths": [20.0, 1.0, 40.0],
    }
    frame.attrs["_layout"] = spec
    layout = read_layout(frame)
    assert layout.header_rows == 3
    assert layout.anchor_cell == (4, 1)
    assert layout.freeze_panes == "A4"
    assert layout.back_link is None
    assert layout.column_widths == [20.0, 1.0, 40.0]      # 宽度 1.0 合法（> 0 即可）
    # Layout 是冻结数据类，且 raw 不进 repr
    with pytest.raises(Exception):
        layout.header_rows = 2
    assert "raw=" not in repr(layout)


def test_anchor_error_message_contains_row():
    frame = _frame()
    frame.attrs["_layout"] = {"header_rows": 3, "anchor_cell": [3, 1]}
    with pytest.raises(ValueError) as excinfo:
        read_layout(frame)
    assert "行 3" in str(excinfo.value)


@pytest.mark.parametrize("spec,match", [
    ({"header_rows": 0}, "header_rows"),
    ({"anchor_cell": [2, 1], "header_rows": 3}, "anchor_cell"),
    ({"anchor_cell": [3, 1], "header_rows": 3}, "anchor_cell"),   # 锚点恰在表头带内也拒绝
    ({"anchor_cell": [0, 0]}, "anchor_cell"),
    ({"anchor_cell": [1]}, "anchor_cell"),
    ({"anchor_cell": [1, 0]}, "anchor_cell"),                     # 列 0 非法（1 基）
    ({"anchor_cell": [True, 1]}, "anchor_cell"),                  # bool 不是合法锚点分量
    ({"header_rows": 1, "anchor_cell": [5, True]}, "anchor_cell"),  # 行合法但分量是 bool 仍拒绝
    ({"freeze_panes": ""}, "freeze_panes"),
    ({"back_link": {"cell": "", "formula": "x"}}, "cell"),
    ({"back_link": {"cell": "A1"}}, "formula"),
    ({"header_columns": [["a"]], "header_rows": 2}, "header_columns"),
    ({"header_columns": "nope"}, "header_columns"),
    ({"header_columns": []}, "header_columns"),                   # 空表头列非法
    ({"header_columns": [["a"], ["b", 42]]}, "header_columns"),
    ({"column_widths": [10, -1]}, "column_widths"),
    ({"column_widths": [0]}, "column_widths"),                    # 非正宽度非法
])
def test_parse_fail_closed(spec, match):
    frame = _frame()
    frame.attrs["_layout"] = spec
    with pytest.raises(ValueError, match=match):
        read_layout(frame)


def test_empty_layout_dict_defaults():
    """空 _layout：单行表头（frame labels）、锚点 (2,1)。"""
    frame = _frame()
    frame.attrs["_skip_default_template"] = True
    frame.attrs["_layout"] = {}
    layout = read_layout(frame)
    assert layout.header_rows == 1
    assert layout.anchor_cell == (2, 1)


def test_custom_sheet_rendering(tmp_path):
    frame = _frame()
    frame.attrs["_skip_default_template"] = True
    frame.attrs["_layout"] = {
        "header_rows": 2,
        "header_columns": [["组1", "组1", "组2"], ["Alpha", "Beta", "Gamma"]],
        "anchor_cell": [3, 1],
        "freeze_panes": "A3",
        "back_link": None,
        "column_widths": [11.0, 22.0, 33.0],
    }
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    stats = create_multi_sheet_excel({"CUSTOM": frame}, output, "manual", track_changes=False)
    assert stats["customLayoutSheets"] == 1

    wb = load_workbook(output)
    assert wb.sheetnames == ["Content", "CUSTOM"]
    ws = wb["CUSTOM"]
    # 两层表头：同值相邻横向合并（组1 A1:B1）——合并区域恰好这一个，无单体合并
    assert [str(r) for r in ws.merged_cells.ranges] == ["A1:B1"]
    assert ws["A1"].value == "组1" and ws["C1"].value == "组2"
    assert [ws.cell(2, c).value for c in range(1, 4)] == ["Alpha", "Beta", "Gamma"]
    # 锚点数据从第 3 行开始
    assert [ws.cell(3, c).value for c in range(1, 4)] == ["a1", "b1", "c1"]
    assert ws.cell(4, 1).value == "a2"
    # 无返回链接 / 冻结 / 筛选覆盖表头带 + 数据区 / 网格线隐藏
    assert ws["A1"].value == "组1"  # back_link None → A1 留给表头
    assert ws.freeze_panes == "A3"
    assert ws.auto_filter.ref == "A1:C4"
    assert ws.sheet_view.showGridLines is False
    assert [ws.column_dimensions[c].width for c in "ABC"] == [11.0, 22.0, 33.0]
    # 默认模板被跳过：无审核列
    assert ws.max_column == 3


def test_custom_layout_defaults_freeze_and_width_heuristic(tmp_path):
    """无 freeze_panes → 回退锚点行；无 column_widths → 按末行表头自适应。"""
    frame = _frame()
    frame.attrs["_skip_default_template"] = True
    frame.attrs["_layout"] = {"header_rows": 1}
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    create_multi_sheet_excel({"PLAIN2": frame}, output, "manual", track_changes=False)
    ws = load_workbook(output)["PLAIN2"]
    assert ws.freeze_panes == "A2"                                  # layout.freeze_panes 回退
    assert ws.sheet_view.showGridLines is False
    # 表头 "Alpha"(5)/"Beta"(4)/"Gamma"(5) → len+2 后低于下限 → MIN_COLUMN_WIDTH
    assert ws.column_dimensions["A"].width == 14.7109375


def test_layout_without_header_columns_uses_frame_labels(tmp_path):
    frame = _frame()
    frame.attrs["_skip_default_template"] = True
    frame.attrs["_layout"] = {"header_rows": 1}
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    create_multi_sheet_excel({"PLAIN": frame}, output, "manual", track_changes=False)
    ws = load_workbook(output)["PLAIN"]
    assert [ws.cell(1, c).value for c in range(1, 4)] == ["Alpha", "Beta", "Gamma"]
    assert ws["A2"].value == "a1"


def test_custom_layout_changes_fall_back_to_all_new(tmp_path):
    frame = _frame()
    frame.attrs["_skip_default_template"] = True
    frame.attrs["_layout"] = {"header_rows": 1}
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    create_multi_sheet_excel({"CUSTOM": frame}, output, "manual", track_changes=True)
    stats = create_multi_sheet_excel({"CUSTOM": frame}, output, "manual", track_changes=True)
    wb = load_workbook(output)
    content = wb["Content"]
    # 自定义 layout 无稳定回读结构：重跑不与上一版比较，全量计为 new
    assert [content.cell(3, c).value for c in range(4, 8)] == [2, 2, 0, 0]
    assert stats["customLayoutSheets"] == 1


def test_back_link_formula_whitelist():
    """V-8:自定义返回链接仅接受 =HYPERLINK 形态;=WEBSERVICE/=CMD 拒绝。"""
    import pytest
    from excel.layout import parse_layout
    ok = parse_layout({"back_link": {"cell": "A1", "formula": '=HYPERLINK("#\'Content\'!A1","返回")'}})
    assert ok.back_link["cell"] == "A1"
    for evil in ('=WEBSERVICE("http://evil")', '=CMD("calc")', 'HYPERLINK("x")'):
        with pytest.raises(ValueError, match="HYPERLINK"):
            parse_layout({"back_link": {"cell": "A1", "formula": evil}})
