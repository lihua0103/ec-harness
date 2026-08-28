"""固定模板（ADR-0016/0020）：Content Sheet / Cover Page / ALS 审核列保留 + 可跳过。"""
import pytest
from openpyxl import Workbook, load_workbook

from excel import (
    COMPARISON_COLUMNS,
    CONTENT_COLUMNS,
    apply_default_template,
    calculate_changes,
    create_multi_sheet_excel,
)
from excel.style_atoms import CONTENT_WIDTHS
from excel.templates import (
    CONTENT_SHEET,
    REPORT_COVER_LABELS,
    REPORT_METADATA_KEYS,
    SKIP_TEMPLATE_ATTR,
    build_content_sheet,
    build_report_cover,
    report_metadata,
)
import pandas as pd


def _frame(labels=True, **extra_columns):
    frame = pd.DataFrame({"USUBJID": ["S1", "S2"], "AETERM": ["Head", "Nausea"], **extra_columns})
    if labels:
        frame.attrs["labels"] = {"USUBJID": "Subject Identifier", "AETERM": "Adverse Event Term"}
    return frame


# ---------------------------------------------------------------------------
# ALS 审核列模板
# ---------------------------------------------------------------------------

def test_als_columns_added_by_default_for_standard_scenarios():
    for scenario in ("manual", "medical"):
        frame = _frame()
        apply_default_template({"AE": frame}, scenario)
        for column in COMPARISON_COLUMNS:
            assert column in frame.columns
            assert frame.attrs["labels"][column]


def test_als_columns_not_added_for_report_or_rbqm():
    for scenario in ("report", "rbqm"):
        frame = _frame()
        apply_default_template({"AE": frame}, scenario)
        assert not any(column in frame.columns for column in COMPARISON_COLUMNS)


def test_skip_default_template_attr_leaves_frame_alone():
    frame = _frame()
    frame.attrs[SKIP_TEMPLATE_ATTR] = True
    apply_default_template({"AE": frame}, "manual")
    assert list(frame.columns) == ["USUBJID", "AETERM"]
    assert "Flag1" not in frame.attrs.get("labels", {})


# ---------------------------------------------------------------------------
# Content Sheet 模板（§6.4）
# ---------------------------------------------------------------------------

def test_content_sheet_template_intact():
    outputs = {"LISTING_AE_01": _frame(), "LISTING_DM_01": _frame()}
    changes = calculate_changes(None, outputs)
    wb = Workbook()
    build_content_sheet(wb, outputs, changes)
    ws = wb[CONTENT_SHEET]
    assert ws["A1"].value == "Comparison Summary"
    assert [cell.value for cell in ws[2]] == CONTENT_COLUMNS
    assert ws.freeze_panes == "A3"
    assert ws.sheet_view.showGridLines is False
    assert [ws.column_dimensions[c].width for c in "ABCDEFG"] == CONTENT_WIDTHS
    # 每个业务表一行统计 + 跳转链接
    assert ws["B3"].value.startswith('=HYPERLINK("#\'LISTING_AE_01\'!A1"')
    assert ws["C4"].value == "Yes" and ws["D4"].value == 2


# ---------------------------------------------------------------------------
# Cover Page 模板（§6.4）
# ---------------------------------------------------------------------------

def test_report_cover_template_intact():
    wb = Workbook()
    build_report_cover(wb, {key: f"v-{key}" for key in REPORT_METADATA_KEYS})
    cover = wb["Cover Page"]
    assert cover["A1"].value == "数据管理状态报告\nDM Status Report"
    assert [cover.cell(row, 1).value for row in range(3, 7)] == REPORT_COVER_LABELS
    assert [cover.cell(row, 2).value for row in range(3, 7)] == [
        "v-sponsor", "v-protocol_no", "v-project_id", "v-report_date"]
    assert sorted(str(item) for item in cover.merged_cells.ranges) == [
        "A1:G1", "B3:G3", "B4:G4", "B5:G5", "B6:G6"]
    assert [cover.row_dimensions[row].height for row in range(1, 7)] == [75, 12.6, 54, 47.25, 54, 39.75]


def test_report_metadata_from_first_frame_attrs():
    first = _frame()
    first.attrs["report_metadata"] = {"sponsor": "S", "protocol_no": "P", "project_id": "W", "report_date": "D"}
    assert report_metadata({"A": first}) == {"sponsor": "S", "protocol_no": "P", "project_id": "W", "report_date": "D"}
    with pytest.raises(ValueError, match="report_metadata"):
        bad = _frame(); bad.attrs["report_metadata"] = "oops"
        report_metadata({"A": bad})


# ---------------------------------------------------------------------------
# 默认发布结构（模板不动，§6.4）
# ---------------------------------------------------------------------------

def test_manual_publish_generates_content_and_listings(tmp_path):
    output = tmp_path / "MANUAL_LISTINGS.xlsx"
    stats = create_multi_sheet_excel(
        {"LISTING_AE_01": _frame()}, output, "manual", track_changes=False)
    wb = load_workbook(output)
    assert wb.sheetnames == ["Content", "LISTING_AE_01"]
    ws = wb["LISTING_AE_01"]
    # 审核列已注入，Label 行含审核列 Label
    header_labels = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert header_labels[:2] == ["Subject Identifier", "Adverse Event Term"]
    assert "FLAG(New/Modified/Old)" in header_labels
    assert stats["standardStructureApplied"] is True
    assert stats["customLayoutSheets"] == 0


def test_report_publish_generates_cover(tmp_path):
    frame = _frame()
    frame.attrs["report_metadata"] = {"sponsor": "S", "protocol_no": "P", "project_id": "W", "report_date": "D"}
    output = tmp_path / "REPORT_LISTINGS.xlsx"
    stats = create_multi_sheet_excel({"Missing Page": frame}, output, "report", track_changes=False)
    wb = load_workbook(output)
    assert wb.sheetnames == ["Cover Page", "Missing Page"]
    assert [cell.value for cell in wb["Missing Page"][1]] == ["USUBJID", "AETERM"]
    assert stats["reportStructureApplied"] is True


def test_unsupported_scenario_rejected(tmp_path):
    with pytest.raises(ValueError, match="不支持的 Listing 场景"):
        create_multi_sheet_excel({"A": _frame()}, tmp_path / "x.xlsx", "unknown", track_changes=False)
