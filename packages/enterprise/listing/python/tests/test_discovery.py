"""全量读取层：doc/ 文本全文 / xlsx 整表 / 源头标注 / 构建期节流 / 密码推导。"""
import base64
import json
import zipfile

import pytest
from openpyxl import Workbook

import discovery
from discovery import (
    list_files,
    load_datasets,
    read_spec_files,
    scan_excel_structures,
)


def test_read_spec_files_full_text_content(project):
    documents, failures = read_spec_files(project / "doc")
    assert failures == []
    doc = documents[0]
    assert doc["_source"] == "spec-document"      # 审计标记；不在投影表 = 全量直通
    assert doc["path"] == "spec.txt"
    assert doc["type"] == "text"
    assert "REQUIREMENT-TAIL" in doc["content"]   # 全文在场，永不被投影
    assert doc["lineCount"] >= 3


def test_read_spec_files_has_no_text_cap(tmp_path):
    (tmp_path / "doc").mkdir()
    (tmp_path / "doc" / "big.txt").write_text("Y" * 250_000, encoding="utf-8")
    (tmp_path / "doc" / "small.txt").write_text("Z" * 1000, encoding="utf-8")
    documents, _ = read_spec_files(tmp_path / "doc")
    by_path = {doc["path"]: doc for doc in documents}
    assert len(by_path["big.txt"]["content"]) == 250_000
    assert "truncated" not in by_path["big.txt"]
    assert len(by_path["small.txt"]["content"]) == 1000


def test_read_spec_files_preserves_decoding_failure_marker(tmp_path):
    """UTF-8 非法字节显式替换，不静默丢字。"""
    (tmp_path / "doc").mkdir()
    (tmp_path / "doc" / "broken.txt").write_bytes(b"ok\xff\xfe\n")
    documents, failures = read_spec_files(tmp_path / "doc")
    assert failures == []
    assert documents[0]["content"] == "ok\ufffd\ufffd\n"


def _write_als(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ALS"
    ws.append(["Dataset Name", "Variable Name", "Label"])
    ws.append(["AE", "AETERM", "Adverse Event Term"])
    ws.append(["AE", "USUBJID", "Subject"])
    notes = wb.create_sheet("Notes")
    notes.append(["note"])
    notes.append(["SECRET-CELL-42"])
    wb.save(path)


def test_read_spec_files_excel_rows_and_semantics(project):
    _write_als(project / "doc" / "als.xlsx")
    documents, failures = read_spec_files(project / "doc")
    assert failures == []
    als = next(doc for doc in documents if doc["type"] == "als")
    assert als["_source"] == "spec-document"
    assert als["path"] == "als.xlsx"
    assert als["mappings"] == [
        {"datasetName": "AE", "sourceColumn": "AETERM", "label": "Adverse Event Term"},
        {"datasetName": "AE", "sourceColumn": "USUBJID", "label": "Subject"},
    ]
    assert als["datasets"] == ["AE"]
    rows = {sheet["sheet"]: sheet["rows"] for sheet in als["rows"]}
    assert rows["Notes"][-1] == ["SECRET-CELL-42"]
    assert "SECRET-CELL-42" in json.dumps(als, ensure_ascii=False)
    sheets = {sheet["name"]: sheet for sheet in als["structure"]["sheets"]}
    assert sheets["ALS"]["rowCount"] == 3
    assert sheets["ALS"]["headerRows"][0] == ["Dataset Name", "Variable Name", "Label"]
    assert len(sheets["ALS"]["headerRows"]) == 1
    assert sheets["Notes"]["rowCount"] == 2


def test_doc_files_without_known_suffix_are_losslessly_included(tmp_path):
    doc = tmp_path / "doc"
    doc.mkdir()
    (doc / "template.bin").write_bytes(b"\x00DSH\xff")
    documents, failures = read_spec_files(doc)
    assert failures == []
    payload = documents[0]
    assert payload["type"] == "binary" and payload["encoding"] == "base64"
    assert base64.b64decode(payload["content"]) == b"\x00DSH\xff"


def test_spreadsheet_ml_2003_xls_is_fully_read(tmp_path):
    doc = tmp_path / "doc"
    doc.mkdir()
    xml = '''<?xml version="1.0"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet">
 <Worksheet xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet" ss:Name="ALS">
  <Table xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
   <Row><Cell><Data>Dataset Name</Data></Cell><Cell><Data>Variable Name</Data></Cell></Row>
   <Row><Cell><Data>AE</Data></Cell><Cell ss:Index="3"><Data>DOC-XML-CELL</Data></Cell></Row>
  </Table>
 </Worksheet>
</Workbook>'''
    (doc / "legacy.xls").write_bytes(b"\xef\xbb\xbf" + xml.encode("utf-8"))
    documents, failures = read_spec_files(doc)
    assert failures == []
    payload = documents[0]
    assert payload["type"] == "excel"
    assert payload["rows"][0]["rows"] == [["AE", None, "DOC-XML-CELL"]]


def test_load_datasets_tags_dataset_source(project):
    datasets, failures, sources = load_datasets(project)
    assert failures == []
    assert sources == {"AE": "AE.csv"}
    assert datasets["AE"].attrs["_source"] == "dataset"
    assert list(datasets["AE"].columns) == ["USUBJID", "AETERM"]


def test_doc_dataset_like_file_is_requirement_not_rawdata(tmp_path):
    doc = tmp_path / "doc"
    doc.mkdir()
    (doc / "requirement.csv").write_text("Requirement\nDOC-CELL-99\n", encoding="utf-8")
    datasets, failures, sources = load_datasets(tmp_path)
    assert datasets == {} and failures == [] and sources == {}


def test_load_datasets_falls_back_to_gbk_csv(tmp_path):
    """Windows 中文 CSV：UTF-8 解码失败时回退 GBK，不把数据源读失败。"""
    (tmp_path / "DM.csv").write_bytes("ID,名称\n1,受试者\n".encode("gbk"))
    datasets, failures, _ = load_datasets(tmp_path)
    assert failures == []
    assert datasets["DM"].loc[0, "名称"] == "受试者"


def test_dataset_payloads_never_build_samples(project):
    datasets, _, sources = load_datasets(project)
    payload = discovery.dataset_payloads(datasets, sources)[0]
    assert "sample" not in payload
    assert payload["rowCount"] == 2 and payload["columns"] == ["USUBJID", "AETERM"]
    assert payload["dtypes"]["USUBJID"] in {"object", "str"}
    assert payload["nullCount"] == {"USUBJID": 0, "AETERM": 0}
    assert payload["uniqueCount"] == {"USUBJID": 2, "AETERM": 2}


def test_scan_excel_structures_structure_only(project):
    _write_als(project / "doc" / "als.xlsx")
    payload = scan_excel_structures(project / "doc" / "als.xlsx")
    assert payload["_source"] == "aux-excel"
    assert payload["path"] == "als.xlsx"
    names = [sheet["name"] for sheet in payload["structure"]["sheets"]]
    assert names == ["ALS", "Notes"]
    assert "rows" not in payload  # 结构扫描天生只有结构，没有行值通道


def test_scan_excel_structures_rejects_non_excel(project):
    with pytest.raises(ValueError, match="NOT_EXCEL"):
        scan_excel_structures(project / "AE.csv")


def test_aux_excel_scope_covers_root_and_work_but_not_doc_or_output(tmp_path):
    (tmp_path / "doc").mkdir(parents=True)
    (tmp_path / "_work" / "nested").mkdir(parents=True)
    (tmp_path / ".clinical-listing" / "output").mkdir(parents=True)
    _write_als(tmp_path / "root.xlsx")
    _write_als(tmp_path / "_work" / "nested" / "work.xlsx")
    _write_als(tmp_path / "doc" / "doc.xlsx")
    _write_als(tmp_path / ".clinical-listing" / "output" / "published.xlsx")
    documents, _ = discovery.read_aux_excel_files(tmp_path)
    assert [item["path"] for item in documents] == ["_work/nested/work.xlsx", "root.xlsx"]
    assert all("rows" not in item for item in documents)
    assert all("SECRET-CELL-42" not in json.dumps(item) for item in documents)


def test_list_files_kinds_and_ignores(project):
    (project / ".clinical-listing" / "_work").mkdir(parents=True)
    (project / ".clinical-listing" / "_work" / "LEAK.csv").write_text("X\n1\n")
    entries = list_files(project)
    kinds = {entry["path"]: entry["kind"] for entry in entries}
    assert kinds["AE.csv"] == "dataset"
    assert kinds["doc/spec.txt"] == "text"
    assert not any("_work" in path for path in kinds)


def test_list_files_direct_call_has_project_fence(project, tmp_path_factory):
    outside = tmp_path_factory.mktemp("outside")
    (outside / "x.csv").write_text("A\n1\n", encoding="utf-8")
    with pytest.raises(ValueError, match="ESCAPE_PROJECT_ROOT"):
        list_files(project, str(outside))


def test_directory_named_like_dataset_is_skipped(tmp_path):
    """目录名带数据后缀（trap.csv/）不是数据源——不能进加载候选。"""
    (tmp_path / "trap.csv").mkdir()
    datasets, failures, sources = load_datasets(tmp_path)
    assert datasets == {} and failures == [] and sources == {}


def test_archive_non_data_members_ignored(tmp_path):
    """归档里的 .txt 与子目录不进数据集候选；只有数据扩展名文件被加载。"""
    import zipfile as _zip

    archive = tmp_path / "data.zip"
    with _zip.ZipFile(archive, "w") as zf:
        zf.writestr("nested/DM.csv", "ID\n1\n")
        zf.writestr("nested/readme.txt", "not a dataset")
    datasets, failures, sources = load_datasets(tmp_path)
    assert sources == {"DM": "archive/data.zip/nested/DM.csv"}
    assert failures == []


def test_nested_archive_datasets_are_loaded_without_mutating_input(tmp_path):
    inner = tmp_path / "inner.zip"
    with zipfile.ZipFile(inner, "w") as zf:
        zf.writestr("raw/DM.csv", "ID\n1\n")
    outer = tmp_path / "outer.zip"
    with zipfile.ZipFile(outer, "w") as zf:
        zf.writestr("package/inner.zip", inner.read_bytes())
    inner.unlink()

    before = outer.read_bytes()
    datasets, failures, sources = load_datasets(tmp_path)

    assert failures == []
    assert list(datasets) == ["DM"]
    assert sources["DM"] == "archive/outer.zip/package/inner.zip/raw/DM.csv"
    assert outer.read_bytes() == before


def test_repeated_inspect_does_not_register_archive_cache_twice(tmp_path):
    """归档缓存只由归档管线加载，重复 inspect 不产生同名冲突。"""
    import zipfile as _zip

    archive = tmp_path / "data.zip"
    with _zip.ZipFile(archive, "w") as zf:
        zf.writestr("nested/AE.csv", "ID\n1\n")
    first = load_datasets(tmp_path)
    second = load_datasets(tmp_path)
    assert first[2] == second[2] == {"AE": "archive/data.zip/nested/AE.csv"}
    assert first[1] == second[1] == []


def test_repeated_load_reloads_updated_same_named_source(tmp_path):
    """同名数据源更新后，下一次运行必须重新读取而非复用旧 DataFrame。"""
    source = tmp_path / "AE.csv"
    source.write_text("ID\n1\n", encoding="utf-8")
    first, failures, _ = load_datasets(tmp_path)
    assert failures == []
    source.write_text("ID\n2\n", encoding="utf-8")
    second, failures, _ = load_datasets(tmp_path)
    assert failures == []
    assert first["AE"].iloc[0, 0] == 1
    assert second["AE"].iloc[0, 0] == 2


# ---------------------------------------------------------------------------
# 密码推导必须保留——工程现实，密码不在 doc/ AI 无法推断
# ---------------------------------------------------------------------------

def test_password_candidates_contract(tmp_path):
    from archive_passwords import password_candidates

    (tmp_path / "doc").mkdir()
    (tmp_path / "doc" / "hint.txt").write_text("SideCarPW99!", encoding="utf-8")
    (tmp_path / "raw").mkdir()
    (tmp_path / "raw" / "STEM-HINT.txt").write_text("SiblingPW7", encoding="utf-8")
    archive = tmp_path / "raw" / "STEM-HINT.zip"

    decoded = [value.decode() for value in password_candidates(tmp_path, archive) if value is not None]
    assert "STEM-HINT" in decoded        # zip 同名 .txt / zip 自身 stem
    assert "hint" in decoded             # 全树 rglob("*.txt") 的 stem（doc/ 下）
    assert "SiblingPW7" in decoded       # 归档同目录 sidecar 的内容
    explicit = [value.decode() for value in password_candidates(tmp_path, archive, "EXPLICIT-PW") if value is not None]
    assert explicit[0] == "EXPLICIT-PW"  # 显式凭据排第一
    assert password_candidates(tmp_path, archive)[-1] is None  # 无密码兜底在最后


def test_password_candidates_combine_adjacent_tokens(tmp_path):
    """项目名中的相邻标识 token 可组合为常见连接符形态。"""
    from archive_passwords import password_candidates

    project = tmp_path / "ABC-123-X"
    project.mkdir()
    archive = project / "package_DATA_20260903.zip"

    decoded = [
        value.decode()
        for value in password_candidates(project, archive)
        if value is not None
    ]
    assert "ABC-123" in decoded
    assert "ABC_123" in decoded
    assert "ABC123" in decoded
    assert "DATA-20260903" in decoded


def test_extract_with_password_derived_adjacent_token(project):
    """无需外部凭据文件时，相邻 token 组合也能解锁受控派生密码。"""
    import pyzipper

    nested_project = project / "ABC-123-X"
    nested_project.mkdir()
    archive = nested_project / "data.zip"
    with pyzipper.AESZipFile(
        archive, "w", compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES
    ) as zf:
        zf.setpassword(b"ABC-123")
        zf.writestr("nested/DM.csv", "ID\n1\n")

    result = discovery.extract_with_password(
        archive, project / "_extract-derived", nested_project
    )
    assert result.extracted_count == 1
    assert result.password_required_count == 0
    assert (project / "_extract-derived" / "nested" / "DM.csv").read_text() == "ID\n1\n"


def test_extract_with_password_sidecar(project):
    archive = project / "data.zip"
    with zipfile.ZipFile(archive, "w") as zf:
        zf.writestr("nested/DM.csv", "ID\n1\n")
    # zip 同名 sidecar .txt 携带密码（工程现实：密码文件与归档同目录）
    (project / "data.txt").write_text("SideCarPW99!", encoding="utf-8")
    # 无加密归档在任意候选下都能解出（zipfile 对未加密成员忽略 pwd）；
    # 这里验证 sidecar 候选与显式凭据都能走通解压管线
    discovery.extract_with_password(archive, project / "_extract", project, "SideCarPW99!")
    assert (project / "_extract" / "nested" / "DM.csv").read_text() == "ID\n1\n"


def test_extract_with_password_aes_sidecar(project):
    """AES ZIP 不能因标准 zipfile 只识别 method 99 而误报缺密码。"""
    import pyzipper

    archive = project / "encrypted.zip"
    with pyzipper.AESZipFile(
        archive, "w", compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES
    ) as zf:
        zf.setpassword(b"AES-PW-77")
        zf.writestr("nested/AE.csv", "ID\n1\n")
    (project / "encrypted.txt").write_text("AES-PW-77", encoding="utf-8")

    discovery.extract_with_password(archive, project / "_extract-aes", project)
    assert (project / "_extract-aes" / "nested" / "AE.csv").read_text() == "ID\n1\n"


def test_mixed_archive_extracts_available_members(project):
    """混合归档不能因加密成员整体失败；缺失成员必须返回显式计数。"""
    import pyzipper

    archive = project / "mixed.zip"
    with pyzipper.AESZipFile(
        archive, "w", compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES
    ) as zf:
        zf.setpassword(b"Locked-PW-9")
        zf.writestr("locked/AE.csv", "ID\n1\n")
    with zipfile.ZipFile(archive, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("open/DM.csv", "ID\n2\n")

    result = discovery.extract_with_password(archive, project / "_extract-mixed", project)
    assert result.partial
    assert result.extracted_count == 1
    assert result.password_required_count == 1
    assert (project / "_extract-mixed" / "open" / "DM.csv").read_text() == "ID\n2\n"
    assert not (project / "_extract-mixed" / "locked" / "AE.csv").exists()


def test_xlsx_value_error_does_not_fall_back_to_xml(tmp_path, monkeypatch):
    from openpyxl import Workbook

    valid = tmp_path / "valid.xlsx"
    Workbook().save(valid)
    source = tmp_path / "source.xlsx"
    with zipfile.ZipFile(valid, "r") as original, zipfile.ZipFile(
        source, "w", zipfile.ZIP_DEFLATED
    ) as package:
        for info in original.infolist():
            data = original.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                data = (
                    '<?xml version="1.0"?>'
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                    '<sheetData/><autoFilter ref="1:1"/></worksheet>'
                ).encode()
            package.writestr(info, data)

    calls = []

    def fake_read_excel(path, **kwargs):
        calls.append(path)
        if path == source:
            raise ValueError("Value does not match pattern")
        return {"Sheet1": "read-from-normalized-copy"}

    monkeypatch.setattr(discovery.pd, "read_excel", fake_read_excel)
    monkeypatch.setattr(
        discovery, "_read_spreadsheet_ml",
        lambda path: pytest.fail("XLSX parse errors must not enter SpreadsheetML fallback"),
    )

    assert discovery._read_excel_sheets(source) == {"Sheet1": "read-from-normalized-copy"}
    assert calls[0] == source
    assert calls[1].name == source.name

    normalized = tmp_path / "normalized.xlsx"
    assert discovery._normalize_xlsx_row_autofilters(source, normalized)

    with zipfile.ZipFile(normalized, "r") as package:
        assert b"A1:XFD1" in package.read("xl/worksheets/sheet1.xml")


def test_password_value_never_in_receipt(project):
    """密码推导是程序职责，但密码值永不回执；失败清单带 path 键。"""
    import worker

    archive = project / "broken.zip"
    archive.write_bytes(b"not a zip")
    (project / "broken.txt").write_text("SideCarPW99!", encoding="utf-8")
    result = worker.dispatch({"operation": "listing_inspect", "project": str(project)})
    payload = json.dumps(result, ensure_ascii=False)
    assert "SideCarPW99!" not in payload
    failure = result["inspection"]["failures"][0]
    assert failure["stage"] == "extract-archive"
    assert failure["path"] == "broken.zip"


def test_archive_directory_named_like_zip_skipped(tmp_path):
    """目录名带 .zip 后缀不是归档——不能进解压候选。"""
    (tmp_path / "evil.zip").mkdir()
    datasets, failures, sources = load_datasets(tmp_path)
    assert datasets == {} and failures == [] and sources == {}


def test_large_aux_excel_never_builds_cell_rows(tmp_path):
    import pandas as _pd

    with _pd.ExcelWriter(tmp_path / "big.xlsx") as writer:
        for sheet in ("A", "B"):
            _pd.DataFrame({f"C{i}": [f"v{i}"] * 1000 for i in range(21)}).to_excel(
                writer, sheet_name=sheet, index=False)
    documents, _ = discovery.read_aux_excel_files(tmp_path)
    assert "rows" not in documents[0]
    assert [sheet["name"] for sheet in documents[0]["structure"]["sheets"]] == ["A", "B"]
