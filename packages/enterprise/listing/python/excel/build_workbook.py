"""工作簿唯一入口：固定模板（Content/Cover）+ 默认业务页 + 自定义 layout 渲染。

职责（ADR-0016 / ADR-0020 / ADR-0022）：
- 默认路径：manual/medical/rbqm 走 RT01 业务页结构，report 走 DM Status
  单层表头 + Cover Page——与既有测试锁定的样式逐字节一致。
- 自定义路径：``df.attrs["_layout"]`` 存在时业务页由 layout 接管排版，
  样式原子复用。
- 变化计数与原子写出为机械职责，与模板无关。
"""
import json
import os
import tempfile
import uuid
from collections import Counter
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from . import style_atoms as atoms
from .layout import Layout, read_layout
from .templates import (
    CONTENT_SHEET,
    literal_cell,
    REPORT_COVER_SHEET,
    STANDARD_SCENARIOS,
    SUPPORTED_SCENARIOS,
    apply_default_template,
    build_content_sheet,
    build_report_cover,
    frame_labels,
    report_metadata,
)

#: Excel 工作表硬边界；超界数据在写入单元格时本就会抛错，这里把
#: auto_filter.ref 一并钳制，避免产出 Excel 需要修复的越界引用。
EXCEL_MAX_ROW = 1_048_576
EXCEL_MAX_COLUMN = 16_384


def _autofilter_ref(first_column: int, first_row: int, last_column: int, last_row: int) -> str:
    return (
        f"{get_column_letter(first_column)}{min(first_row, EXCEL_MAX_ROW)}:"
        f"{get_column_letter(min(last_column, EXCEL_MAX_COLUMN))}{min(last_row, EXCEL_MAX_ROW)}"
    )


def normalize_sheet_outputs(
    outputs: Dict[str, pd.DataFrame], reserved_names: Optional[List[str]] = None,
) -> Dict[str, pd.DataFrame]:
    """验证模型输出并规范 Excel 工作表名称；冲突时 fail-closed。"""
    if not isinstance(outputs, dict) or not outputs:
        raise ValueError("outputs 必须是非空 dict[str, pandas.DataFrame]")

    normalized: Dict[str, pd.DataFrame] = {}
    used = {name.casefold() for name in (reserved_names or [CONTENT_SHEET])}
    for raw_name, frame in outputs.items():
        if not isinstance(raw_name, str) or not raw_name.strip():
            raise ValueError("每个 outputs 键必须是非空工作表名称")
        if not isinstance(frame, pd.DataFrame):
            raise TypeError(f"outputs[{raw_name!r}] 必须是 pandas DataFrame")
        safe_name = atoms.INVALID_SHEET_CHARS.sub("_", raw_name.strip()).strip("'")[:atoms.SHEET_NAME_MAX].strip()
        if not safe_name:
            raise ValueError(f"工作表名称 {raw_name!r} 规范化后为空")
        key = safe_name.casefold()
        if key in used:
            raise ValueError(f"工作表名称冲突或使用保留名称: {raw_name!r} -> {safe_name!r}")
        used.add(key)
        normalized[safe_name] = _normalized_frame(frame)
    return normalized


# ---------------------------------------------------------------------------
# 版本间变化计数（机械职责）
# ---------------------------------------------------------------------------

def _excel_scalar(value: Any) -> Any:
    """将 pandas/numpy/复杂 Python 值归一化为 openpyxl 可写入的标量。"""
    if value is None:
        return None
    if isinstance(value, (list, tuple, dict, set)):
        return json.dumps(value, ensure_ascii=False, default=str, sort_keys=True)
    try:
        missing = pd.isna(value)
        if isinstance(missing, (bool, np.bool_)) and bool(missing):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.to_pydatetime() if isinstance(value, pd.Timestamp) else value
    if isinstance(value, pd.Timedelta):
        return value.to_pytimedelta()
    if isinstance(value, (str, bool, int, float)):
        return value
    return str(value)


def _normalized_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """发布前统一单元格类型，避免真实数据类型让渲染器随机失败。"""
    normalized = frame.copy()
    for column in normalized.columns:
        normalized[column] = [_excel_scalar(value) for value in normalized[column].tolist()]
    normalized.attrs = dict(frame.attrs)
    return normalized


def _row_counter(frame: pd.DataFrame) -> Counter:
    values = frame.astype(object).where(pd.notna(frame), None)
    values = values.replace("", None)
    return Counter(tuple(_excel_scalar(value) for value in row)
                   for row in values.itertuples(index=False, name=None))


def _align_previous_columns(
    previous: Dict[str, pd.DataFrame], current: Dict[str, pd.DataFrame],
) -> Dict[str, pd.DataFrame]:
    """把上一版按位置重命名为当前 sheets 的列名。

    旧版业务 Sheet 不再单独展示变量名（oid），读取时只能拿到匿名列。
    calculate_changes 做整行多重集比较，列名只要位置对齐即可正确计数。
    """
    aligned: Dict[str, pd.DataFrame] = {}
    for sheet_name, old_frame in previous.items():
        new_frame = current.get(sheet_name)
        if new_frame is None:
            aligned[sheet_name] = old_frame
            continue
        new_columns = [str(column) for column in new_frame.columns]
        renamed = old_frame.copy()
        if len(renamed.columns) >= len(new_columns):
            renamed = renamed.iloc[:, :len(new_columns)]
            renamed.columns = new_columns
        else:
            extra = [f"_extra_{index}" for index in range(len(renamed.columns), len(new_columns))]
            renamed.columns = [*renamed.columns, *extra]
            for column in extra:
                renamed[column] = None
            renamed = renamed[new_columns]
        aligned[sheet_name] = renamed
    return aligned


def load_previous_version(
    output_file: Path, index_sheet: str = CONTENT_SHEET, scenario: str = "manual",
) -> Optional[Dict[str, pd.DataFrame]]:
    """按场景读取上一版业务页；读取失败即不做变化比较。

    medical/manual/rbqm 业务 Sheet 的 Row 1 是返回链接 + 标题合并区（不携带 oid），
    因此无法用 ``pd.read_excel(header=...)`` 直接定位变量名。这里改用 openpyxl
    按单元格逐列读取，把 Row 2+ 视为数据，并按列序重建无列名的 DataFrame。
    calculate_changes 只做整行多重集比较，列名是否对齐不影响 old/new 计数。
    自定义 layout 的业务页没有稳定的回读结构，由 calculate_changes 按
    "全量新增"处理（见 treat_as_new）。
    """
    if not output_file.exists():
        return None
    try:
        previous: Dict[str, pd.DataFrame] = {}
        workbook = load_workbook(output_file, data_only=True, read_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                if sheet_name == index_sheet:
                    continue
                sheet = workbook[sheet_name]
                if scenario == "report":
                    rows_iter = sheet.iter_rows(values_only=True)
                    header = next(rows_iter, None)
                    if not header:
                        continue
                    columns = [str(value) if value is not None else "" for value in header]
                    data = [list(row) for row in rows_iter]
                    previous[sheet_name] = pd.DataFrame(data, columns=columns)
                else:
                    rows_iter = sheet.iter_rows(min_row=3, values_only=True)
                    data = [list(row) for row in rows_iter]
                    if not data:
                        previous[sheet_name] = pd.DataFrame()
                        continue
                    width = max(len(row) for row in data)
                    columns = [str(index) for index in range(width)]
                    previous[sheet_name] = pd.DataFrame(data, columns=columns)
        finally:
            workbook.close()
        return previous
    except Exception:
        return None


def calculate_changes(
    previous: Optional[Dict[str, pd.DataFrame]], current: Dict[str, pd.DataFrame],
    treat_as_new: Optional[set] = None,
) -> Dict[str, Dict[str, int]]:
    """按完整行多重集计算新增/删除；无业务唯一键时不伪造 modified。

    ``treat_as_new`` 中的表不与上一版比较（自定义 layout 无稳定回读结构），
    全量计为 new。
    """
    treat_as_new = treat_as_new or set()
    changes: Dict[str, Dict[str, int]] = {}
    for sheet_name, frame in current.items():
        if sheet_name in treat_as_new:
            changes[sheet_name] = {"new": len(frame), "modified": 0, "old": 0}
            continue
        old = previous.get(sheet_name) if previous else None
        if old is None or list(old.columns) != list(frame.columns):
            changes[sheet_name] = {"new": len(frame), "modified": 0, "old": 0}
            continue
        old_rows = _row_counter(old)
        new_rows = _row_counter(frame)
        changes[sheet_name] = {
            "new": sum((new_rows - old_rows).values()),
            "modified": 0,
            "old": sum((old_rows - new_rows).values()),
        }
    return changes


# ---------------------------------------------------------------------------
# 业务页渲染
# ---------------------------------------------------------------------------

def _style_cell(cell, *, font, fill, alignment) -> None:
    cell.font = copy(font)
    cell.fill = copy(fill)
    cell.border = copy(atoms.GRID_BORDER)
    cell.alignment = copy(alignment)


def build_listing_sheet(wb: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    """默认业务 Sheet：第 1 行返回链接 + 标题，第 2 行 Label，第 3 行起数据。"""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    labels = frame_labels(frame)
    columns = [str(column) for column in frame.columns]
    last_column = max(1, len(columns))

    ws["A1"] = '=HYPERLINK("#\'Content\'!A1","Go back")'
    ws["A1"].font = copy(atoms.BACK_LINK_FONT)
    ws["A1"].border = copy(atoms.GRID_BORDER)
    if last_column >= 2:
        merge_end = min(last_column, 6)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=merge_end)
        literal_cell(ws, 1, 2, sheet_name)
        for column in range(2, merge_end + 1):
            _style_cell(ws.cell(1, column), font=atoms.SHEET_TITLE_FONT, fill=atoms.PALE_BLUE, alignment=atoms.CENTER)
    for column in range(7, last_column + 1):
        _style_cell(ws.cell(1, column), font=atoms.DATA_FONT, fill=atoms.PALE_BLUE, alignment=atoms.CENTER)

    for column, variable in enumerate(columns, 1):
        _style_cell(literal_cell(ws, 2, column, labels[variable]), font=atoms.HEADER_FONT, fill=atoms.PALE_BLUE, alignment=atoms.HEADER_ALIGNMENT)
        width = max(atoms.MIN_COLUMN_WIDTH, min(atoms.MAX_COLUMN_WIDTH, len(labels[variable]) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[2].height = atoms.LABEL_ROW_HEIGHT

    for row_index, values in enumerate(frame.itertuples(index=False, name=None), 3):
        for column, value in enumerate(values, 1):
            cell = literal_cell(ws, row_index, column, _excel_scalar(value))
            _style_cell(cell, font=atoms.DATA_FONT, fill=atoms.WHITE, alignment=Alignment(vertical="center"))
    ws.auto_filter.ref = _autofilter_ref(1, 2, last_column, 2 + len(frame))


def _report_column_width(sheet_name: str, column: int, header: str) -> float:
    template_widths = atoms.REPORT_COLUMN_WIDTHS.get(sheet_name, [])
    if column <= len(template_widths):
        return template_widths[column - 1]
    return min(atoms.REPORT_WIDTH_MAX, max(atoms.REPORT_WIDTH_MIN, len(header) + 2))


def build_report_sheet(wb: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    """report 单层表头业务页；数据始终从第 2 行开始。"""
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    columns = [str(column) for column in frame.columns]
    last_column = max(1, len(columns))
    for column, header in enumerate(columns, 1):
        cell = literal_cell(ws, 1, column, header)
        cell.font = copy(atoms.REPORT_HEADER_FONT)
        cell.fill = copy(atoms.REPORT_HEADER_FILL)
        cell.border = copy(atoms.REPORT_HEADER_BORDER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(column)].width = _report_column_width(sheet_name, column, header)
    ws.row_dimensions[1].height = atoms.REPORT_HEADER_HEIGHTS.get(sheet_name, atoms.REPORT_HEADER_DEFAULT_HEIGHT)

    for row_index, values in enumerate(frame.itertuples(index=False, name=None), 2):
        for column, value in enumerate(values, 1):
            cell = literal_cell(ws, row_index, column, _excel_scalar(value))
            cell.font = copy(atoms.REPORT_DATA_FONT)
            cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = _autofilter_ref(1, 1, last_column, 1 + len(frame))


def _merge_equal_runs(ws, row: int, start_col: int, labels: List[str]) -> None:
    """同一行内相邻且同值的表头横向合并（多层表头的常规语义）。"""
    run_start = 0
    for index in range(1, len(labels) + 1):
        if index == len(labels) or labels[index] != labels[run_start]:
            if index - run_start > 1:
                ws.merge_cells(
                    start_row=row, start_column=start_col + run_start,
                    end_row=row, end_column=start_col + index - 1,
                )
            run_start = index


def build_custom_sheet(wb: Workbook, sheet_name: str, frame: pd.DataFrame, layout: Layout) -> None:
    """自定义 layout 业务页：表头带 + 锚点数据区，样式原子复用。"""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    columns = [str(column) for column in frame.columns]
    anchor_row, anchor_col = layout.anchor_cell
    header_top = anchor_row - layout.header_rows
    last_column = anchor_col - 1 + max(1, len(columns))

    if layout.back_link is not None:
        link_cell = ws[layout.back_link["cell"]]
        link_cell.value = layout.back_link["formula"]
        link_cell.font = copy(atoms.BACK_LINK_FONT)
        link_cell.border = copy(atoms.GRID_BORDER)

    labels = frame_labels(frame)
    band_labels: List[List[str]] = []
    for band_index in range(layout.header_rows):
        row_number = header_top + band_index
        row_labels: List[str]
        if layout.header_columns is not None and band_index < len(layout.header_columns):
            row_labels = list(layout.header_columns[band_index])
            while len(row_labels) < len(columns):
                row_labels.append("")
        elif band_index == layout.header_rows - 1:
            row_labels = [labels[column] for column in columns]
        else:
            row_labels = [""] * len(columns)
        band_labels.append(row_labels)
        for offset, text in enumerate(row_labels):
            cell = literal_cell(ws, row_number, anchor_col + offset, text)
            _style_cell(cell, font=atoms.HEADER_FONT, fill=atoms.PALE_BLUE, alignment=atoms.HEADER_ALIGNMENT)
        _merge_equal_runs(ws, row_number, anchor_col, row_labels)
        ws.row_dimensions[row_number].height = atoms.LABEL_ROW_HEIGHT

    for row_index, values in enumerate(frame.itertuples(index=False, name=None), anchor_row):
        for column, value in enumerate(values, anchor_col):
            cell = literal_cell(ws, row_index, column, _excel_scalar(value))
            _style_cell(cell, font=atoms.DATA_FONT, fill=atoms.WHITE, alignment=Alignment(vertical="center"))

    width_source = layout.column_widths or [
        max(atoms.MIN_COLUMN_WIDTH, min(atoms.MAX_COLUMN_WIDTH, len(text) + 2))
        for text in band_labels[-1]
    ]
    for offset, width in enumerate(width_source[:max(1, len(columns))]):
        ws.column_dimensions[get_column_letter(anchor_col + offset)].width = width

    ws.freeze_panes = layout.freeze_panes or f"{get_column_letter(anchor_col)}{anchor_row}"
    bottom = anchor_row - 1 + max(1, len(frame))
    ws.auto_filter.ref = _autofilter_ref(anchor_col, header_top, last_column, bottom)


def _business_sheet(wb: Workbook, sheet_name: str, frame: pd.DataFrame, scenario: str) -> None:
    layout = read_layout(frame)
    if layout is not None:
        build_custom_sheet(wb, sheet_name, frame, layout)
    elif scenario == "report":
        build_report_sheet(wb, sheet_name, frame)
    else:
        build_listing_sheet(wb, sheet_name, frame)


# ---------------------------------------------------------------------------
# 单一入口
# ---------------------------------------------------------------------------

class ListingPublishError(ValueError):
    """只携带稳定阶段码，调用方绝不把底层异常文本回传给 AI。"""

    def __init__(self, code: str) -> None:
        super().__init__(code)
        self.code = code


def _stage_workbook(wb: Workbook, output_file: Path) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(prefix=f".{output_file.stem}-", suffix=".xlsx", dir=output_file.parent)
    os.close(fd)
    temporary_file = Path(temporary_name)
    try:
        try:
            wb.save(temporary_file)
        finally:
            wb.close()
        return temporary_file
    except Exception as exc:
        temporary_file.unlink(missing_ok=True)
        raise ListingPublishError("WRITE_WORKBOOK_FAILED") from exc


def _stage_text(path: Path, content: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".tmp", dir=path.parent)
    os.close(fd)
    temporary_file = Path(temporary_name)
    try:
        temporary_file.write_text(content, encoding="utf-8")
        return temporary_file
    except Exception as exc:
        temporary_file.unlink(missing_ok=True)
        raise ListingPublishError("WRITE_CHANGE_HISTORY_FAILED") from exc


def _commit_publication(staged: list[tuple[Path, Path]]) -> None:
    """提交工作簿和 sidecar；任一步失败都恢复提交前文件。"""
    backups: list[tuple[Path, Path | None]] = []
    committed: list[Path] = []
    try:
        for target, _ in staged:
            backup = None
            if target.exists():
                fd, name = tempfile.mkstemp(prefix=f".{target.stem}-backup-", suffix=".bak", dir=target.parent)
                os.close(fd)
                backup = Path(name)
                os.replace(target, backup)
            backups.append((target, backup))
        for target, temporary in staged:
            os.replace(temporary, target)
            committed.append(target)
    except Exception as exc:
        for target in committed:
            target.unlink(missing_ok=True)
        for target, backup in reversed(backups):
            if backup is not None and backup.exists():
                os.replace(backup, target)
        raise ListingPublishError("COMMIT_PUBLICATION_FAILED") from exc
    finally:
        for _, backup in backups:
            if backup is not None:
                backup.unlink(missing_ok=True)
        for _, temporary in staged:
            temporary.unlink(missing_ok=True)

def create_multi_sheet_excel(
    outputs: Dict[str, pd.DataFrame],
    output_file: Path,
    scenario: str = "manual",
    unique_key_columns: Optional[Dict[str, List[str]]] = None,
    track_changes: bool = True,
    cover_labels: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """原子生成唯一 Excel，并按场景选择固定工作簿结构。

    - 默认模板（Content/Cover/ALS）来自 templates，AI 可逐表跳过；
    - 带 ``_layout`` 的业务页由 layout 接管排版；
    - 变化计数对自定义 layout 页按全量新增处理；
    - ``cover_labels``：report 场景 Cover 行标签（部署方 patch 配置，
      None 用中性默认）。
    """
    del unique_key_columns
    scenario = scenario.lower()
    if scenario not in SUPPORTED_SCENARIOS:
        raise ValueError(f"不支持的 Listing 场景: {scenario}")

    try:
        reserved = [REPORT_COVER_SHEET] if scenario == "report" else [CONTENT_SHEET]
        prepared = normalize_sheet_outputs(outputs, reserved)
        prepared = apply_default_template(prepared, scenario)
        custom_layout_sheets = {
            sheet_name for sheet_name, frame in prepared.items() if read_layout(frame) is not None
        }

        index_sheet = REPORT_COVER_SHEET if scenario == "report" else CONTENT_SHEET
        previous = load_previous_version(output_file, index_sheet, scenario) if track_changes else None
        if previous is not None:
            previous = _align_previous_columns(previous, prepared)
        changes = calculate_changes(previous, prepared, treat_as_new=custom_layout_sheets)

        wb = Workbook()
        if scenario == "report":
            build_report_cover(wb, report_metadata(prepared), cover_labels)
        else:
            build_content_sheet(wb, prepared, changes)
        for sheet_name, frame in prepared.items():
            _business_sheet(wb, sheet_name, frame, scenario)
    except ListingPublishError:
        raise
    except Exception as exc:
        raise ListingPublishError("RENDER_WORKBOOK_FAILED") from exc

    timestamp = datetime.now().isoformat()
    run_id = uuid.uuid4().hex
    change_body = {
        "timestamp": timestamp,
        "runId": run_id,
        "scenario": scenario,
        "baselineAvailable": previous is not None,
        "changes": changes,
    }
    staged: list[tuple[Path, Path]] = []
    try:
        staged.append((output_file, _stage_workbook(wb, output_file)))
        if track_changes:
            change_log = output_file.parent / f"{output_file.stem}_changes.json"
            history_log = output_file.parent / f"{output_file.stem}_run_history.jsonl"
            previous_history = history_log.read_text(encoding="utf-8") if history_log.exists() else ""
            staged.append((change_log, _stage_text(change_log, json.dumps(change_body, ensure_ascii=False, indent=2))))
            staged.append((history_log, _stage_text(history_log, previous_history + json.dumps(change_body, ensure_ascii=False) + "\n")))
        _commit_publication(staged)
    except ListingPublishError:
        for _, temporary in staged:
            temporary.unlink(missing_ok=True)
        raise
    except Exception as exc:
        for _, temporary in staged:
            temporary.unlink(missing_ok=True)
        raise ListingPublishError("WRITE_CHANGE_HISTORY_FAILED") from exc

    return {
        "outputFile": str(output_file),
        "format": "single-workbook-multi-sheet-xlsx",
        "sheetNames": [index_sheet, *prepared.keys()],
        "listingSheetCount": len(prepared),
        "totalSheets": len(prepared) + 1,
        "totalRows": sum(len(frame) for frame in prepared.values()),
        "scenario": scenario,
        "standardStructureApplied": scenario in STANDARD_SCENARIOS,
        "reportStructureApplied": scenario == "report",
        "rbqmStructureFlexible": scenario == "rbqm",
        "customLayoutSheets": len(custom_layout_sheets),
    }
