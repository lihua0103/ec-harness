"""固定模板（ADR-0016 / ADR-0020）：Content Sheet、Report Cover、ALS 审核列。

Content Sheet / Cover Page / ALS 审核列是**输出标准**，必须保留为模板
（反馈 4：固定模板不能动）。模板只对默认场景生效；AI 可通过
``df.attrs["_skip_default_template"] = True`` 跳过模板注入（ADR-0020），
也可用 ``df.attrs["_layout"]`` 改业务表排版（见 layout.py）。

systemPrompt 中的"标准输出范例"是对这些模板的文字引导（ADR-0014），
与这里的代码模板是两件事：程序不强制 AI 产出什么内容，只为默认路径
固化交付结构。
"""
from copy import copy
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from . import style_atoms as atoms

#: 触发固定结构的标准场景（Content Sheet + ALS 审核列）。
STANDARD_SCENARIOS = {"manual", "medical"}
#: 全部支持场景。
SUPPORTED_SCENARIOS = STANDARD_SCENARIOS | {"report", "rbqm"}

CONTENT_SHEET = "Content"
CONTENT_TITLE = "Comparison Summary"
CONTENT_COLUMNS = [
    "Listing Seq.", "Form Name", "New/Modified ?", "Total",
    "New", "Modified", "Old",
]

#: ALS 审核列（RT01 标准，输出标准的一部分）。
COMPARISON_COLUMNS = [
    "Flag1", "__cmp_FLAG__", "__cmp_UpdateDetail__",
    "__cmp_RCcomment__", "__cmp_Idate__",
]
COMPARISON_LABELS = {
    "Flag1": "Flag1",
    "__cmp_FLAG__": "FLAG(New/Modified/Old)",
    "__cmp_UpdateDetail__": "Update Detail",
    "__cmp_RCcomment__": "Review Comments",
    "__cmp_Idate__": "Initial/Date",
}

REPORT_COVER_SHEET = "Cover Page"
REPORT_TITLE = "数据管理状态报告\nDM Status Report"
#: Cover 行标签默认值（中性）。申办方特定文案（如"康德弘翼/WuXi"）属部署
#: 参数，经 cordis.patch.yml 的 reportCoverLabels 配置下发（CODING_STANDARDS
#: 禁止把部署参数写死在插件代码）。
REPORT_COVER_LABELS = [
    "申办方：\nSponsor:",
    "方案编号：\nProtocol No:",
    "项目编号：\nProject ID:",
    "最新报告生成日期：",
]
REPORT_METADATA_KEYS = ["sponsor", "protocol_no", "project_id", "report_date"]

#: 跳过默认模板的 attrs 开关（ADR-0020）。
SKIP_TEMPLATE_ATTR = "_skip_default_template"


def frame_labels(frame: pd.DataFrame) -> Dict[str, str]:
    """读取 attrs["labels"] 并补齐缺省（列名即 Label），格式错误 fail-closed。"""
    labels = frame.attrs.get("labels", {})
    if not isinstance(labels, dict):
        raise ValueError("DataFrame.attrs['labels'] 必须是 {变量名: 字段Label} 字典")
    return {str(column): str(labels.get(column, column)) for column in frame.columns}


def literal_cell(ws, row: int, column: int, value: Any):
    """字面量单元格写入（漏洞扫描 V-5）：openpyxl 会把 '=' 开头的字符串
    推断为公式——数据值/表名/标签一律经此写入，命中即强制 data_type='s'
    中和公式注入；非字符串与普通值零影响。程序自有公式（HYPERLINK）不走此路。"""
    cell = ws.cell(row, column, value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


def hyperlink_formula(sheet_name: str) -> str:
    """Content 页跳转公式（V-5）：sheet 名内的双引号双写转义，
    防止经 outputs 键名打破公式字符串实现注入。"""
    safe = sheet_name.replace('"', '""')
    return f'=HYPERLINK("#\'{safe}\'!A1","{safe}")'


def apply_default_template(
    outputs: Dict[str, pd.DataFrame], scenario: str,
) -> Dict[str, pd.DataFrame]:
    """默认模板注入（标准输出范例的机械化落地）。

    - manual/medical：补齐五个 ALS 审核列 + Label（AI 逐表可用
      ``_skip_default_template`` 跳过）。
    - report/rbqm：不注入审核列。
    """
    if scenario not in STANDARD_SCENARIOS:
        return outputs
    for frame in outputs.values():
        if frame.attrs.get(SKIP_TEMPLATE_ATTR):
            continue
        labels = frame_labels(frame)
        for column in COMPARISON_COLUMNS:
            if column not in frame.columns:
                frame[column] = ""
            labels[column] = COMPARISON_LABELS[column]
        frame.attrs["labels"] = labels
    return outputs


def report_metadata(outputs: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
    """从首个业务表的 attrs 提取封面字段，不读取业务行。"""
    first = next(iter(outputs.values()))
    metadata = first.attrs.get("report_metadata", {})
    if not isinstance(metadata, dict):
        raise ValueError("DataFrame.attrs['report_metadata'] 必须是字典")
    return {key: metadata.get(key, "") for key in REPORT_METADATA_KEYS}


def _style_cell(cell, *, font, fill, alignment) -> None:
    cell.font = copy(font)
    cell.fill = copy(fill)
    cell.border = copy(atoms.GRID_BORDER)
    cell.alignment = copy(alignment)


def build_content_sheet(
    wb: Workbook, outputs: Dict[str, pd.DataFrame], changes: Dict[str, Dict[str, int]],
) -> None:
    """Content Sheet 固定模板：标题 + 表头 + 每业务表一行变化统计。"""
    ws = wb.active
    ws.title = CONTENT_SHEET
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:G1")
    ws["A1"] = CONTENT_TITLE
    _style_cell(ws["A1"], font=atoms.CONTENT_TITLE_FONT, fill=atoms.PALE_BLUE, alignment=atoms.CENTER)
    for column in range(2, 8):
        _style_cell(ws.cell(1, column), font=atoms.CONTENT_TITLE_FONT, fill=atoms.PALE_BLUE, alignment=atoms.CENTER)

    for column, name in enumerate(CONTENT_COLUMNS, 1):
        cell = ws.cell(2, column, name)
        _style_cell(cell, font=atoms.HEADER_FONT, fill=atoms.PALE_BLUE, alignment=atoms.CENTER)
        ws.column_dimensions[get_column_letter(column)].width = atoms.CONTENT_WIDTHS[column - 1]

    for row, (sheet_name, frame) in enumerate(outputs.items(), 3):
        delta = changes[sheet_name]
        values = [
            row - 2,
            sheet_name,
            "Yes" if delta["new"] or delta["modified"] else "No",
            len(frame),
            delta["new"],
            delta["modified"],
            delta["old"],
        ]
        for column, value in enumerate(values, 1):
            cell = literal_cell(ws, row, column, value)
            _style_cell(cell, font=atoms.DATA_FONT, fill=atoms.WHITE, alignment=atoms.CENTER)
        link = ws.cell(row, 2)
        link.value = hyperlink_formula(sheet_name)
        link.font = copy(atoms.LINK_FONT)
        link.alignment = Alignment(horizontal="left", vertical="center")


def build_report_cover(wb: Workbook, metadata: Dict[str, Any], cover_labels: Optional[List[str]] = None) -> None:
    """DM Status Report 固定 Cover Page 模板。

    ``cover_labels``：部署方经 patch 配置下发的行标签（长度须与
    REPORT_METADATA_KEYS 对齐，取前 N 项）；None 用中性默认。
    """
    labels = list(cover_labels) if cover_labels else list(REPORT_COVER_LABELS)
    ws = wb.active
    ws.title = REPORT_COVER_SHEET
    ws.merge_cells("A1:G1")
    ws["A1"] = REPORT_TITLE
    for column in range(1, 8):
        cell = ws.cell(1, column)
        cell.font = copy(atoms.REPORT_COVER_TITLE_FONT)
        cell.fill = copy(atoms.REPORT_COVER_FILL)
        cell.border = copy(atoms.REPORT_COVER_TITLE_BORDER)
        cell.alignment = copy(atoms.CENTER)
    ws.row_dimensions[1].height = atoms.REPORT_COVER_TITLE_ROW_HEIGHT
    ws.row_dimensions[2].height = atoms.REPORT_COVER_SPACER_ROW_HEIGHT
    for letter, width in atoms.REPORT_COVER_COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    for row, (label, key, height) in enumerate(
        zip(labels, REPORT_METADATA_KEYS, atoms.REPORT_COVER_ROW_HEIGHTS), 3,
    ):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        literal_cell(ws, row, 1, label)
        label_cell = ws.cell(row, 1)
        label_cell.font = copy(atoms.REPORT_COVER_LABEL_FONT)
        label_cell.fill = copy(atoms.REPORT_COVER_FILL)
        label_cell.border = copy(atoms.REPORT_COVER_ROW_BORDER)
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        literal_cell(ws, row, 2, metadata[key])
        for column in range(2, 8):
            value_cell = ws.cell(row, column)
            value_cell.font = copy(atoms.REPORT_COVER_VALUE_FONT)
            value_cell.fill = copy(atoms.WHITE)
            value_cell.border = copy(atoms.REPORT_COVER_ROW_BORDER)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = height
