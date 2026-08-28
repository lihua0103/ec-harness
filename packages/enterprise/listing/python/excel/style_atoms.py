"""样式原子（ADR-0022）：颜色 / 字体 / 边框 / 对齐 / 行高列宽常量。

提炼自 RT01 Manual Listing 与 DM Status Report 标准范例，是输出标准的
一部分（反馈 4：固定样式不能动）。排版**决策**（横/纵/多层表头/锚点）
不在本层——那是 layout.py 按 ``df.attrs["_layout"]`` 的事。
"""
import re

from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

THIN_AUTO = Side(style="thin", color=Color(auto=True))
MEDIUM_AUTO = Side(style="medium", color=Color(auto=True))
GRID_BORDER = Border(left=THIN_AUTO, right=THIN_AUTO, top=THIN_AUTO, bottom=THIN_AUTO)
PALE_BLUE = PatternFill(fill_type="solid", fgColor="FFEDF2F9")
WHITE = PatternFill(fill_type="solid", fgColor="FFFFFFFF")

CONTENT_TITLE_FONT = Font(name="Times New Roman", size=16, bold=True)
SHEET_TITLE_FONT = Font(name="Times New Roman", size=14, bold=True)
HEADER_FONT = Font(name="Times New Roman", size=13, bold=True)
DATA_FONT = Font(name="Times New Roman", size=13)
LINK_FONT = Font(name="Times New Roman", size=13, color="FF0000FF", underline="single")
BACK_LINK_FONT = Font(name="Times New Roman", size=13, color="FF0000FF")

REPORT_HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FF000000")
REPORT_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFC5D9F1")
REPORT_HEADER_BORDER = Border(top=THIN_AUTO, bottom=THIN_AUTO)
REPORT_DATA_FONT = Font(name="Calibri", size=11, color="FF000000")

REPORT_COVER_LABEL_FONT = Font(name="宋体", size=14, bold=True, color="FF000000")
REPORT_COVER_TITLE_FONT = Font(name="宋体", size=16, bold=True, color="FF000000")
REPORT_COVER_VALUE_FONT = Font(name="微软雅黑", size=16, bold=True, color="FF000000")
REPORT_COVER_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
REPORT_COVER_TITLE_BORDER = Border(right=MEDIUM_AUTO, bottom=MEDIUM_AUTO)
REPORT_COVER_ROW_BORDER = Border(right=THIN_AUTO, top=THIN_AUTO, bottom=THIN_AUTO)

CENTER = Alignment(horizontal="center", vertical="center")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="center")
LEFT_CENTER = Alignment(horizontal="left", vertical="center")

#: 工作表名非法字符与长度上限（Excel 硬限制）。
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
SHEET_NAME_MAX = 31

#: Manual/Medical 业务页列宽启发式边界。
MIN_COLUMN_WIDTH = 14.7109375
MAX_COLUMN_WIDTH = 50.7109375
#: Label 行行高。
LABEL_ROW_HEIGHT = 60

#: Content Sheet 固定列宽（RT01 范例提炼）。
CONTENT_WIDTHS = [16.7109375, 50.7109375, 18.7109375, 9.7109375, 8.7109375, 12.7109375, 8.7109375]

#: Report 场景：范例业务页表头行高。
REPORT_HEADER_HEIGHTS = {
    "Matrix by Study": 15.75,
    "Matrix by Site": 63.0,
    "Matrix by Subject": 63.0,
    "Missing Page": 31.5,
    "Missing Lab": 31.5,
    "UnSDV Page": 31.5,
    "Queries Not Resolved": 31.5,
    "All Queries Matrix by page": 15.75,
}
REPORT_HEADER_DEFAULT_HEIGHT = 31.5

#: Report 场景：范例业务页列宽。
REPORT_COLUMN_WIDTHS = {
    "Matrix by Study": [43.7109375, 12.42578125, *([13.0] * 15)],
    "Matrix by Site": [12.7109375, *([13.0] * 40)],
    "Matrix by Subject": [12.7109375, *([13.0] * 43)],
    "Missing Page": [59.7109375, 13.7109375, 17.140625, 32.7109375, 7.7109375,
                     15.7109375, 49.7109375, 46.7109375, 12.42578125, 10.140625,
                     21.85546875, 18.42578125, 10.140625, 13.0, 16.0, 17.140625,
                     14.85546875, 12.42578125],
    "Missing Lab": [59.7109375, 21.7109375, 9.0, 32.7109375, 7.7109375,
                    15.7109375, 29.7109375, 24.7109375, 12.42578125, 11.7109375,
                    21.85546875, 18.42578125, 11.7109375, 10.140625, 16.0,
                    17.140625, 12.42578125, 64.7109375],
    "UnSDV Page": [59.7109375, 13.7109375, 17.140625, 32.7109375, 7.7109375,
                   15.7109375, 49.7109375, 72.7109375, 12.42578125, 14.85546875,
                   9.0, 18.42578125],
    "Queries Not Resolved": [59.7109375, 13.7109375, 17.140625, 32.7109375,
                             7.7109375, 15.7109375, 49.7109375, 12.42578125,
                             63.7109375, 18.42578125, 8.7109375, 10.140625,
                             17.7109375, 14.85546875, 100.7109375, 11.7109375,
                             14.7109375, 11.7109375, 13.7109375, 100.7109375,
                             10.140625, 13.7109375, 18.42578125, 17.140625,
                             16.0, 13.0, 17.140625, 18.42578125],
    "All Queries Matrix by page": [93.7109375, 16.0, 20.7109375, 14.85546875,
                                   19.5703125, 18.42578125, 6.5703125],
}
#: Report 场景：列宽回退边界。
REPORT_WIDTH_MIN = 13.0
REPORT_WIDTH_MAX = 100.7109375

#: Cover Page 版式。
REPORT_COVER_TITLE_ROW_HEIGHT = 75.0
REPORT_COVER_SPACER_ROW_HEIGHT = 12.6
REPORT_COVER_ROW_HEIGHTS = [54, 47.25, 54, 39.75]
REPORT_COVER_COLUMN_WIDTHS = {"A": 38.140625, "B": 38.42578125, "C": 7.140625, "D": 9.140625}
