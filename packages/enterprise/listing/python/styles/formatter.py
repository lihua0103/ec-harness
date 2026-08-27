"""
Excel 输出格式化器 - 应用样式规范

负责：
1. 应用样式到 Excel workbook
2. 设置列宽、行高
3. 冻结窗格
4. 自动过滤器
"""
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import pandas as pd

from .style_spec import (
    get_style_spec,
    is_system_field,
    CONTENTS_SPEC,
    CellStyle,
    SheetStyleSpec
)


def apply_cell_style(cell, style: CellStyle):
    """应用单元格样式"""
    cell.font = style.font
    cell.fill = style.fill
    cell.alignment = style.alignment
    cell.border = style.border


def format_contents_sheet(ws: Worksheet):
    """格式化 Contents 页"""
    spec = CONTENTS_SPEC
    
    # 应用表头样式
    for col_idx, cell in enumerate(ws[1], 1):
        apply_cell_style(cell, spec.header_style)
    
    # 应用数据样式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            apply_cell_style(cell, spec.data_style)
    
    # 设置列宽
    for col_idx, cell in enumerate(ws[1], 1):
        col_letter = get_column_letter(col_idx)
        col_name = cell.value
        width = spec.column_widths.get(col_name, 15)
        ws.column_dimensions[col_letter].width = width
    
    # 设置行高
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = spec.row_height
    
    # 冻结窗格
    ws.freeze_panes = spec.freeze_panes
    
    # 自动过滤器
    ws.auto_filter.ref = ws.dimensions


def format_data_sheet(ws: Worksheet, scenario: str, columns: List[str]):
    """格式化数据页
    
    Args:
        ws: worksheet 对象
        scenario: 场景类型
        columns: 列名列表
    """
    spec = get_style_spec(scenario)
    
    if ws.max_row < 1:
        return
    
    # 应用表头样式
    for col_idx, cell in enumerate(ws[1], 1):
        apply_cell_style(cell, spec.header_style)
    
    # 应用数据样式（区分系统字段列和业务数据列）
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if is_system_field(col_name, scenario):
                apply_cell_style(cell, spec.system_column_style)
            else:
                apply_cell_style(cell, spec.data_style)
    
    # 设置列宽
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        width = spec.column_widths.get(col_name, spec.column_widths.get("default", 15))
        ws.column_dimensions[col_letter].width = width
    
    # 设置行高
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = spec.row_height
    
    # 表头行高
    ws.row_dimensions[1].height = 25
    
    # 冻结窗格
    ws.freeze_panes = spec.freeze_panes
    
    # 自动过滤器
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def apply_conditional_formatting(ws: Worksheet, scenario: str):
    """应用条件格式（可选，用于特殊标记）
    
    例如：
    - manual: Flag 列的值高亮
    - medical: Review Comments 非空高亮
    - rbqm: Risk Level 根据级别着色
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    
    if scenario == "manual":
        # Flag 列非空时黄色高亮
        flag_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "Flag":
                flag_col = get_column_letter(col_idx)
                break
        
        if flag_col and ws.max_row > 1:
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws.conditional_formatting.add(
                f'{flag_col}2:{flag_col}{ws.max_row}',
                CellIsRule(operator='notEqual', formula=['""'], fill=yellow_fill)
            )
    
    elif scenario == "rbqm":
        # Risk Level 着色：High=红, Medium=黄, Low=绿
        risk_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "Risk Level":
                risk_col = get_column_letter(col_idx)
                break
        
        if risk_col and ws.max_row > 1:
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            
            ws.conditional_formatting.add(
                f'{risk_col}2:{risk_col}{ws.max_row}',
                CellIsRule(operator='equal', formula=['"High"'], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f'{risk_col}2:{risk_col}{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Medium"'], fill=yellow_fill)
            )
            ws.conditional_formatting.add(
                f'{risk_col}2:{risk_col}{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Low"'], fill=green_fill)
            )


def format_workbook(wb: Workbook, scenario: str, sheet_columns: Dict[str, List[str]]):
    """格式化整个 workbook
    
    Args:
        wb: Workbook 对象
        scenario: 场景类型
        sheet_columns: 每个 sheet 的列名，格式 {sheet_name: [col1, col2, ...]}
    """
    # 格式化 Contents 页
    if "Contents" in wb.sheetnames:
        format_contents_sheet(wb["Contents"])
    
    # 格式化数据页
    for sheet_name in wb.sheetnames:
        if sheet_name == "Contents":
            continue
        
        ws = wb[sheet_name]
        columns = sheet_columns.get(sheet_name, [])
        
        if columns:
            format_data_sheet(ws, scenario, columns)
            # 可选：应用条件格式
            # apply_conditional_formatting(ws, scenario)
