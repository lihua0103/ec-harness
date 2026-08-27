"""
Excel 样式规范定义 - 临床 Listing 输出标准

定义三个场景的统一样式：manual、medical、rbqm
每个场景包含：Contents 页样式 + 数据页样式
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from dataclasses import dataclass


@dataclass
class CellStyle:
    """单元格样式定义"""
    font: Font
    fill: PatternFill
    alignment: Alignment
    border: Border


@dataclass
class SheetStyleSpec:
    """Sheet 样式规范"""
    header_style: CellStyle
    data_style: CellStyle
    system_column_style: CellStyle  # 系统字段列样式（如 Flag, Update Details）
    freeze_panes: str  # 冻结窗格位置，如 "A2"
    column_widths: Dict[str, float]  # 列宽配置
    row_height: float  # 数据行高度


# ============================================================================
# 通用样式元素
# ============================================================================

# 边框样式
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# 对齐方式
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=False)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ============================================================================
# Contents 页样式（所有场景通用）
# ============================================================================

CONTENTS_HEADER_STYLE = CellStyle(
    font=Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    fill=PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

CONTENTS_DATA_STYLE = CellStyle(
    font=Font(name='Calibri', size=11, color='000000'),
    fill=PatternFill(fill_type=None),
    alignment=ALIGN_LEFT,
    border=THIN_BORDER
)

CONTENTS_SPEC = SheetStyleSpec(
    header_style=CONTENTS_HEADER_STYLE,
    data_style=CONTENTS_DATA_STYLE,
    system_column_style=CONTENTS_DATA_STYLE,  # Contents 页无系统字段区别
    freeze_panes="A2",
    column_widths={
        "No.": 8,
        "Listing": 30,
        "Description": 50,
        "Rows": 12,
        "Columns": 12,
        "Status": 15,
        "Last Updated": 20
    },
    row_height=20
)


# ============================================================================
# Manual 场景样式
# ============================================================================

MANUAL_HEADER_STYLE = CellStyle(
    font=Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    fill=PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

MANUAL_DATA_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(fill_type=None),
    alignment=ALIGN_LEFT,
    border=THIN_BORDER
)

MANUAL_SYSTEM_COLUMN_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

MANUAL_SPEC = SheetStyleSpec(
    header_style=MANUAL_HEADER_STYLE,
    data_style=MANUAL_DATA_STYLE,
    system_column_style=MANUAL_SYSTEM_COLUMN_STYLE,
    freeze_panes="A2",
    column_widths={
        "default": 15,
        "Flag": 8,
        "Update Details": 25,
        "Review Comments": 30,
        "Initial_Date": 15
    },
    row_height=18
)


# ============================================================================
# Medical 场景样式
# ============================================================================

MEDICAL_HEADER_STYLE = CellStyle(
    font=Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    fill=PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

MEDICAL_DATA_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(fill_type=None),
    alignment=ALIGN_LEFT,
    border=THIN_BORDER
)

MEDICAL_SYSTEM_COLUMN_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

MEDICAL_SPEC = SheetStyleSpec(
    header_style=MEDICAL_HEADER_STYLE,
    data_style=MEDICAL_DATA_STYLE,
    system_column_style=MEDICAL_SYSTEM_COLUMN_STYLE,
    freeze_panes="A2",
    column_widths={
        "default": 15,
        "Flag": 8,
        "Update Details": 25,
        "Review Comments": 30,
        "Initial_Date": 15,
        "Reviewer": 15
    },
    row_height=18
)


# ============================================================================
# RBQM 场景样式
# ============================================================================

RBQM_HEADER_STYLE = CellStyle(
    font=Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    fill=PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

RBQM_DATA_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(fill_type=None),
    alignment=ALIGN_LEFT,
    border=THIN_BORDER
)

RBQM_SYSTEM_COLUMN_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

RBQM_SPEC = SheetStyleSpec(
    header_style=RBQM_HEADER_STYLE,
    data_style=RBQM_DATA_STYLE,
    system_column_style=RBQM_SYSTEM_COLUMN_STYLE,
    freeze_panes="A2",
    column_widths={
        "default": 15,
        "Risk Level": 12,
        "Indicator": 20,
        "Threshold": 12,
        "Notes": 30
    },
    row_height=18
)


# ============================================================================
# Report 场景样式
# ============================================================================

REPORT_HEADER_STYLE = CellStyle(
    font=Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
    fill=PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

REPORT_DATA_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(fill_type=None),
    alignment=ALIGN_LEFT,
    border=THIN_BORDER
)

REPORT_SYSTEM_COLUMN_STYLE = CellStyle(
    font=Font(name='Calibri', size=10, color='000000'),
    fill=PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
    alignment=ALIGN_CENTER,
    border=THIN_BORDER
)

REPORT_SPEC = SheetStyleSpec(
    header_style=REPORT_HEADER_STYLE,
    data_style=REPORT_DATA_STYLE,
    system_column_style=REPORT_SYSTEM_COLUMN_STYLE,
    freeze_panes="A2",
    column_widths={
        "default": 15,
        "Category": 20,
        "Metric": 25,
        "Value": 15,
        "Notes": 30
    },
    row_height=18
)


# ============================================================================
# 场景样式映射
# ============================================================================

SCENARIO_STYLES: Dict[str, SheetStyleSpec] = {
    "manual": MANUAL_SPEC,
    "medical": MEDICAL_SPEC,
    "rbqm": RBQM_SPEC,
    "report": REPORT_SPEC
}


def get_style_spec(scenario: str) -> SheetStyleSpec:
    """获取场景样式规范"""
    return SCENARIO_STYLES.get(scenario.lower(), MANUAL_SPEC)


# ============================================================================
# 系统字段定义（按场景）
# ============================================================================

SYSTEM_FIELDS: Dict[str, List[str]] = {
    "manual": ["Flag", "Update Details", "Review Comments", "Initial_Date"],
    "medical": ["Flag", "Update Details", "Review Comments", "Initial_Date", "Reviewer"],
    "rbqm": ["Risk Level", "Indicator", "Threshold", "Notes"],
    "report": ["Category", "Metric", "Value", "Notes"]
}


def get_system_fields(scenario: str) -> List[str]:
    """获取场景的系统字段列表"""
    return SYSTEM_FIELDS.get(scenario.lower(), SYSTEM_FIELDS["manual"])


def is_system_field(column_name: str, scenario: str) -> bool:
    """判断是否为系统字段"""
    system_fields = get_system_fields(scenario)
    return column_name in system_fields
