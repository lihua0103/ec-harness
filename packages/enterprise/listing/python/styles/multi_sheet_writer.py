"""临床 Listing 单工作簿 Writer。

manual / medical 使用 RT01 Manual Listing 范例结构；report 使用 DM Status
Report 范例结构；rbqm 保留业务列自由度并复用 Manual Listing 视觉样式。
"""
from collections import Counter
from copy import copy
from datetime import datetime
import json
import os
from pathlib import Path
import re
import tempfile
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CONTENT_SHEET = "Content"
CONTENT_TITLE = "Comparison Summary"
CONTENT_COLUMNS = [
    "Listing Seq.", "Form Name", "New/Modified ?", "Total",
    "New", "Modified", "Old",
]
CONTENT_WIDTHS = [16.7109375, 50.7109375, 18.7109375, 9.7109375, 8.7109375, 12.7109375, 8.7109375]
COMPARISON_COLUMNS = [
    "Flag1", "__cmp_FLAG__", "__cmp_UpdateDetail__",
    "__cmp_RCcomment__", "__cmp_Idate__",
]
COMPARISON_LABELS = {
    "Flag1": "Flag1",
    "__cmp_FLAG__": "FLAG(New/Modified/Old)",
    "__cmp_UpdateDetail__": "Update Detail",
    "__cmp_RCcomment__": "Review Comments",
    "__cmp_Idate__": "Initial/Date",
}
STANDARD_SCENARIOS = {"manual", "medical"}
SUPPORTED_SCENARIOS = STANDARD_SCENARIOS | {"report", "rbqm"}
REPORT_COVER_SHEET = "Cover Page"
REPORT_TITLE = "数据管理状态报告\nDM Status Report"
REPORT_COVER_LABELS = [
    "申办方：\nSponsor:",
    "方案编号：\nProtocol No:",
    "康德弘翼项目编号：\nWuXi Project ID:",
    "最新报告生成日期：",
]
REPORT_METADATA_KEYS = ["sponsor", "protocol_no", "project_id", "report_date"]
REPORT_HEADER_HEIGHTS = {
    "Matrix by Study": 15.75,
    "Matrix by Site": 63.0,
    "Matrix by Subject": 63.0,
    "Missing Page": 31.5,
    "Missing Lab": 31.5,
    "UnSDV Page": 31.5,
    "Queries Not Resolved": 31.5,
    "All Queries Matrix by page": 15.75,
}
REPORT_COLUMN_WIDTHS = {
    "Matrix by Study": [43.7109375, 12.42578125, *([13.0] * 15)],
    "Matrix by Site": [12.7109375, *([13.0] * 40)],
    "Matrix by Subject": [12.7109375, *([13.0] * 43)],
    "Missing Page": [59.7109375, 13.7109375, 17.140625, 32.7109375, 7.7109375,
                     15.7109375, 49.7109375, 46.7109375, 12.42578125, 10.140625,
                     21.85546875, 18.42578125, 10.140625, 13.0, 16.0, 17.140625,
                     14.85546875, 12.42578125],
    "Missing Lab": [59.7109375, 21.7109375, 9.0, 32.7109375, 7.7109375,
                    15.7109375, 29.7109375, 24.7109375, 12.42578125, 11.7109375,
                    21.85546875, 18.42578125, 11.7109375, 10.140625, 16.0,
                    17.140625, 12.42578125, 64.7109375],
    "UnSDV Page": [59.7109375, 13.7109375, 17.140625, 32.7109375, 7.7109375,
                   15.7109375, 49.7109375, 72.7109375, 12.42578125, 14.85546875,
                   9.0, 18.42578125],
    "Queries Not Resolved": [59.7109375, 13.7109375, 17.140625, 32.7109375,
                             7.7109375, 15.7109375, 49.7109375, 12.42578125,
                             63.7109375, 18.42578125, 8.7109375, 10.140625,
                             17.7109375, 14.85546875, 100.7109375, 11.7109375,
                             14.7109375, 11.7109375, 13.7109375, 100.7109375,
                             10.140625, 13.7109375, 18.42578125, 17.140625,
                             16.0, 13.0, 17.140625, 18.42578125],
    "All Queries Matrix by page": [93.7109375, 16.0, 20.7109375, 14.85546875,
                                   19.5703125, 18.42578125, 6.5703125],
}
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

THIN_AUTO = Side(style="thin", color=Color(auto=True))
GRID_BORDER = Border(left=THIN_AUTO, right=THIN_AUTO, top=THIN_AUTO, bottom=THIN_AUTO)
PALE_BLUE = PatternFill(fill_type="solid", fgColor="FFEDF2F9")
WHITE = PatternFill(fill_type="solid", fgColor="FFFFFFFF")

CONTENT_TITLE_FONT = Font(name="Times New Roman", size=16, bold=True)
SHEET_TITLE_FONT = Font(name="Times New Roman", size=14, bold=True)
HEADER_FONT = Font(name="Times New Roman", size=13, bold=True)
DATA_FONT = Font(name="Times New Roman", size=13)
LINK_FONT = Font(name="Times New Roman", size=13, color="FF0000FF", underline="single")
BACK_LINK_FONT = Font(name="Times New Roman", size=13, color="FF0000FF")
REPORT_HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FF000000")
REPORT_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFC5D9F1")
REPORT_HEADER_BORDER = Border(top=THIN_AUTO, bottom=THIN_AUTO)
REPORT_COVER_LABEL_FONT = Font(name="宋体", size=14, bold=True, color="FF000000")
REPORT_COVER_TITLE_FONT = Font(name="宋体", size=16, bold=True, color="FF000000")
REPORT_COVER_VALUE_FONT = Font(name="微软雅黑", size=16, bold=True, color="FF000000")
REPORT_COVER_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
MEDIUM_AUTO = Side(style="medium", color=Color(auto=True))
REPORT_COVER_TITLE_BORDER = Border(right=MEDIUM_AUTO, bottom=MEDIUM_AUTO)
REPORT_COVER_ROW_BORDER = Border(right=THIN_AUTO, top=THIN_AUTO, bottom=THIN_AUTO)
CENTER = Alignment(horizontal="center", vertical="center")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def normalize_sheet_outputs(
    outputs: Dict[str, pd.DataFrame], reserved_names: Optional[List[str]] = None,
) -> Dict[str, pd.DataFrame]:
    """验证模型输出并规范 Excel 工作表名称；冲突时 fail-closed。"""
    if not isinstance(outputs, dict) or not outputs:
        raise ValueError("outputs 必须是非空 dict[str, pandas.DataFrame]")

    normalized: Dict[str, pd.DataFrame] = {}
    used = {name.casefold() for name in (reserved_names or [CONTENT_SHEET])}
    for raw_name, frame in outputs.items():
        if not isinstance(raw_name, str) or not raw_name.strip():
            raise ValueError("每个 outputs 键必须是非空工作表名称")
        if not isinstance(frame, pd.DataFrame):
            raise TypeError(f"outputs[{raw_name!r}] 必须是 pandas DataFrame")
        safe_name = _INVALID_SHEET_CHARS.sub("_", raw_name.strip()).strip("'")[:31].strip()
        if not safe_name:
            raise ValueError(f"工作表名称 {raw_name!r} 规范化后为空")
        key = safe_name.casefold()
        if key in used:
            raise ValueError(f"工作表名称冲突或使用保留名称: {raw_name!r} -> {safe_name!r}")
        used.add(key)
        normalized[safe_name] = frame.copy()
    return normalized


def _labels(frame: pd.DataFrame) -> Dict[str, str]:
    labels = frame.attrs.get("labels", {})
    if not isinstance(labels, dict):
        raise ValueError("DataFrame.attrs['labels'] 必须是 {变量名: 字段Label} 字典")
    return {str(column): str(labels.get(column, column)) for column in frame.columns}


def _prepare_outputs(outputs: Dict[str, pd.DataFrame], scenario: str) -> Dict[str, pd.DataFrame]:
    reserved = [REPORT_COVER_SHEET] if scenario == "report" else [CONTENT_SHEET]
    prepared = normalize_sheet_outputs(outputs, reserved)
    if scenario in STANDARD_SCENARIOS:
        for frame in prepared.values():
            labels = _labels(frame)
            for column in COMPARISON_COLUMNS:
                if column not in frame.columns:
                    frame[column] = ""
                labels[column] = COMPARISON_LABELS[column]
            frame.attrs["labels"] = labels
    return prepared


def _row_counter(frame: pd.DataFrame) -> Counter:
    values = frame.astype(object).where(pd.notna(frame), None)
    values = values.replace("", None)
    return Counter(tuple(row) for row in values.itertuples(index=False, name=None))


def load_previous_version(
    output_file: Path, index_sheet: str = CONTENT_SHEET, scenario: str = "manual",
) -> Optional[Dict[str, pd.DataFrame]]:
    """按场景读取上一版业务页；读取失败即不做变化比较。"""
    if not output_file.exists():
        return None
    try:
        previous = {}
        with pd.ExcelFile(output_file) as workbook:
            for sheet_name in workbook.sheet_names:
                if sheet_name == index_sheet:
                    continue
                if scenario == "report":
                    previous[sheet_name] = pd.read_excel(workbook, sheet_name=sheet_name, header=0)
                else:
                    # RT01：第 2 行是变量名，第 3 行仅为展示 Label，数据从第 4 行开始。
                    previous[sheet_name] = pd.read_excel(
                        workbook, sheet_name=sheet_name, header=1, skiprows=[2])
        return previous
    except Exception:
        return None


def calculate_changes(previous: Optional[Dict[str, pd.DataFrame]], current: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, int]]:
    """按完整行多重集计算新增/删除；无业务唯一键时不伪造 modified。"""
    changes: Dict[str, Dict[str, int]] = {}
    for sheet_name, frame in current.items():
        old = previous.get(sheet_name) if previous else None
        if old is None or list(old.columns) != list(frame.columns):
            changes[sheet_name] = {"new": len(frame), "modified": 0, "old": 0}
            continue
        old_rows = _row_counter(old)
        new_rows = _row_counter(frame)
        changes[sheet_name] = {
            "new": sum((new_rows - old_rows).values()),
            "modified": 0,
            "old": sum((old_rows - new_rows).values()),
        }
    return changes


def _style_cell(cell, *, font: Font, fill: PatternFill, alignment: Alignment) -> None:
    cell.font = copy(font)
    cell.fill = copy(fill)
    cell.border = copy(GRID_BORDER)
    cell.alignment = copy(alignment)


def _build_content(wb: Workbook, outputs: Dict[str, pd.DataFrame], changes: Dict[str, Dict[str, int]]) -> None:
    ws = wb.active
    ws.title = CONTENT_SHEET
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:G1")
    ws["A1"] = CONTENT_TITLE
    _style_cell(ws["A1"], font=CONTENT_TITLE_FONT, fill=PALE_BLUE, alignment=CENTER)
    for column in range(2, 8):
        _style_cell(ws.cell(1, column), font=CONTENT_TITLE_FONT, fill=PALE_BLUE, alignment=CENTER)

    for column, name in enumerate(CONTENT_COLUMNS, 1):
        cell = ws.cell(2, column, name)
        _style_cell(cell, font=HEADER_FONT, fill=PALE_BLUE, alignment=CENTER)
        ws.column_dimensions[get_column_letter(column)].width = CONTENT_WIDTHS[column - 1]

    for row, (sheet_name, frame) in enumerate(outputs.items(), 3):
        delta = changes[sheet_name]
        values = [
            row - 2,
            sheet_name,
            "Yes" if delta["new"] or delta["modified"] else "No",
            len(frame),
            delta["new"],
            delta["modified"],
            delta["old"],
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row, column, value)
            _style_cell(cell, font=DATA_FONT, fill=WHITE, alignment=CENTER)
        link = ws.cell(row, 2)
        link.value = f'=HYPERLINK("#\'{sheet_name}\'!A1","{sheet_name}")'
        link.font = copy(LINK_FONT)
        link.alignment = Alignment(horizontal="left", vertical="center")


def _build_listing(wb: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    labels = _labels(frame)
    columns = [str(column) for column in frame.columns]
    last_column = max(1, len(columns))

    ws["A1"] = '=HYPERLINK("#\'Content\'!A1","Go back")'
    ws["A1"].font = copy(BACK_LINK_FONT)
    ws["A1"].border = copy(GRID_BORDER)
    if last_column >= 2:
        merge_end = min(last_column, 6)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=merge_end)
        ws.cell(1, 2, sheet_name)
        for column in range(2, merge_end + 1):
            _style_cell(ws.cell(1, column), font=SHEET_TITLE_FONT, fill=PALE_BLUE, alignment=CENTER)
    for column in range(7, last_column + 1):
        _style_cell(ws.cell(1, column), font=DATA_FONT, fill=PALE_BLUE, alignment=CENTER)

    for column, variable in enumerate(columns, 1):
        _style_cell(ws.cell(2, column, variable), font=HEADER_FONT, fill=PALE_BLUE, alignment=CENTER)
        _style_cell(ws.cell(3, column, labels[variable]), font=HEADER_FONT, fill=PALE_BLUE, alignment=HEADER_ALIGNMENT)
        width = max(14.7109375, min(50.7109375, max(len(variable), len(labels[variable])) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[3].height = 60

    for row_index, values in enumerate(frame.itertuples(index=False, name=None), 4):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, None if pd.isna(value) else value)
            _style_cell(cell, font=DATA_FONT, fill=WHITE, alignment=Alignment(vertical="center"))
    ws.auto_filter.ref = f"A3:{get_column_letter(last_column)}{max(3, 3 + len(frame))}"


def _report_metadata(outputs: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
    """从首个业务表的 attrs 提取可选封面字段，不读取业务行。"""
    first = next(iter(outputs.values()))
    metadata = first.attrs.get("report_metadata", {})
    if not isinstance(metadata, dict):
        raise ValueError("DataFrame.attrs['report_metadata'] 必须是字典")
    return {key: metadata.get(key, "") for key in REPORT_METADATA_KEYS}


def _build_report_cover(wb: Workbook, metadata: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = REPORT_COVER_SHEET
    ws.merge_cells("A1:G1")
    ws["A1"] = REPORT_TITLE
    for column in range(1, 8):
        cell = ws.cell(1, column)
        cell.font = copy(REPORT_COVER_TITLE_FONT)
        cell.fill = copy(REPORT_COVER_FILL)
        cell.border = copy(REPORT_COVER_TITLE_BORDER)
        cell.alignment = copy(CENTER)
    ws.row_dimensions[1].height = 75
    ws.row_dimensions[2].height = 12.6
    ws.column_dimensions["A"].width = 38.140625
    ws.column_dimensions["B"].width = 38.42578125
    ws.column_dimensions["C"].width = 7.140625
    ws.column_dimensions["D"].width = 9.140625

    heights = [54, 47.25, 54, 39.75]
    for row, (label, key, height) in enumerate(zip(REPORT_COVER_LABELS, REPORT_METADATA_KEYS, heights), 3):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row, 1, label)
        label_cell = ws.cell(row, 1)
        label_cell.font = copy(REPORT_COVER_LABEL_FONT)
        label_cell.fill = copy(REPORT_COVER_FILL)
        label_cell.border = copy(REPORT_COVER_ROW_BORDER)
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row, 2, metadata[key])
        for column in range(2, 8):
            value_cell = ws.cell(row, column)
            value_cell.font = copy(REPORT_COVER_VALUE_FONT)
            value_cell.fill = copy(WHITE)
            value_cell.border = copy(REPORT_COVER_ROW_BORDER)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = height


def _report_column_width(sheet_name: str, column: int, header: str) -> float:
    template_widths = REPORT_COLUMN_WIDTHS.get(sheet_name, [])
    if column <= len(template_widths):
        return template_widths[column - 1]
    return min(100.7109375, max(13.0, len(header) + 2))


def _build_report_sheet(wb: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    """生成 report 单层表头业务页；数据始终从第 2 行开始。"""
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    columns = [str(column) for column in frame.columns]
    last_column = max(1, len(columns))
    for column, header in enumerate(columns, 1):
        cell = ws.cell(1, column, header)
        cell.font = copy(REPORT_HEADER_FONT)
        cell.fill = copy(REPORT_HEADER_FILL)
        cell.border = copy(REPORT_HEADER_BORDER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(column)].width = _report_column_width(sheet_name, column, header)
    ws.row_dimensions[1].height = REPORT_HEADER_HEIGHTS.get(sheet_name, 31.5)

    for row_index, values in enumerate(frame.itertuples(index=False, name=None), 2):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, None if pd.isna(value) else value)
            cell.font = Font(name="Calibri", size=11, color="FF000000")
            cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = f"A1:{get_column_letter(last_column)}{max(1, 1 + len(frame))}"


def _build_report_workbook(wb: Workbook, outputs: Dict[str, pd.DataFrame]) -> None:
    _build_report_cover(wb, _report_metadata(outputs))
    for sheet_name, frame in outputs.items():
        _build_report_sheet(wb, sheet_name, frame)


def create_multi_sheet_excel(
    outputs: Dict[str, pd.DataFrame],
    output_file: Path,
    scenario: str,
    unique_key_columns: Optional[Dict[str, List[str]]] = None,
    track_changes: bool = True,
) -> Dict[str, Any]:
    """原子生成唯一 Excel，并按场景选择固定工作簿结构。"""
    del unique_key_columns
    scenario = scenario.lower()
    if scenario not in SUPPORTED_SCENARIOS:
        raise ValueError(f"不支持的 Listing 场景: {scenario}")

    prepared = _prepare_outputs(outputs, scenario)
    index_sheet = REPORT_COVER_SHEET if scenario == "report" else CONTENT_SHEET
    previous = load_previous_version(output_file, index_sheet, scenario) if track_changes else None
    changes = calculate_changes(previous, prepared)

    wb = Workbook()
    if scenario == "report":
        _build_report_workbook(wb, prepared)
    else:
        _build_content(wb, prepared, changes)
        for sheet_name, frame in prepared.items():
            _build_listing(wb, sheet_name, frame)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(prefix=f".{output_file.stem}-", suffix=".xlsx", dir=output_file.parent)
    os.close(fd)
    temporary_file = Path(temporary_name)
    try:
        wb.save(temporary_file)
        wb.close()
        os.replace(temporary_file, output_file)
    finally:
        temporary_file.unlink(missing_ok=True)

    if track_changes:
        change_log = output_file.parent / f"{output_file.stem}_changes.json"
        change_log.write_text(json.dumps({
            "timestamp": datetime.now().isoformat(), "changes": changes,
        }, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "outputFile": str(output_file),
        "format": "single-workbook-multi-sheet-xlsx",
        "sheetNames": [index_sheet, *prepared.keys()],
        "listingSheetCount": len(prepared),
        "totalSheets": len(prepared) + 1,
        "totalRows": sum(len(frame) for frame in prepared.values()),
        "scenario": scenario,
        "standardStructureApplied": scenario in STANDARD_SCENARIOS,
        "reportStructureApplied": scenario == "report",
        "rbqmStructureFlexible": scenario == "rbqm",
    }


__all__ = [
    "CONTENT_COLUMNS", "CONTENT_SHEET", "COMPARISON_COLUMNS",
    "calculate_changes", "create_multi_sheet_excel", "normalize_sheet_outputs",
]



