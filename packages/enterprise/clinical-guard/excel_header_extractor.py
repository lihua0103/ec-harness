"""
Excel 表头提取器 - 提取元数据而不泄露数据值
"""
import sys
import json
import csv
import re
import openpyxl
from pathlib import Path

_SENSITIVE_HEADER = re.compile(
    r"^(?:[A-Z]{1,4}\d{4,}|\d{4}-\d{2}-\d{2}|\d{5,}|"
    r"screening|enrolled|已入组|受试者|患者)$",
    re.IGNORECASE,
)


def _safe_headers(values: list[str]) -> list[str]:
    return [
        f"COLUMN_{index + 1}" if _SENSITIVE_HEADER.match(value.strip()) else value
        for index, value in enumerate(values)
    ]

def extract_excel_headers(file_path: str) -> dict:
    """
    提取 Excel 文件的表头和元数据
    
    返回:
    {
        "sheets": [
            {
                "name": "Sheet1",
                "headers": ["COL1", "COL2", ...],
                "rowCount": 100,
                "columnCount": 10
            }
        ]
    }
    """
    path = Path(file_path)
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = csv.reader(handle)
            first_row = next(rows, [])
            row_count = 1 if first_row else 0
            for _ in rows:
                row_count += 1
        column_count = len(first_row)
        return {
            "sheets": [{
                "name": path.stem,
            "headers": [f"COLUMN_{index + 1}" for index in range(column_count)],
                "header_cells": [
                    {"row": 0, "col": index, "value": f"COLUMN_{index + 1}"}
                    for index in range(column_count)
                ],
                "rowCount": row_count,
                "columnCount": column_count,
            }]
        }

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result = {"sheets": []}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 获取表头（第一行）
        headers = []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            headers = [str(cell) if cell is not None else "" for cell in first_row]
        
        # 获取维度
        max_row = ws.max_row
        max_col = ws.max_column
        
        result["sheets"].append({
            "name": sheet_name,
            "headers": _safe_headers(headers),
            "header_cells": [
                {"row": 0, "col": index, "value": f"COLUMN_{index + 1}"}
                for index in range(max_col)
            ],
            "rowCount": max_row,
            "columnCount": max_col
        })
    
    wb.close()
    return result

if __name__ == "__main__":
    if len(sys.argv) not in (2, 4) or (len(sys.argv) == 4 and sys.argv[2] != "--max-scan-rows"):
        print(json.dumps({"error": "Usage: python excel_header_extractor.py <file_path> [--max-scan-rows <n>]"}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not Path(file_path).exists():
        print(json.dumps({"error": f"File not found: {file_path}"}))
        sys.exit(1)
    
    try:
        result = extract_excel_headers(file_path)
        print(json.dumps(result, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)
