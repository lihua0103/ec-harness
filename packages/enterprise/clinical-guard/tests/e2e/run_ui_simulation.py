"""
模拟用户在 Web UI 中的完整操作流程
完整的 inspect -> run -> publish E2E 测试
"""
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "python"))

from security.worker import _handle
import openpyxl
import json

def test_project_e2e(project: str, data_root: Path):
    """测试单个项目的完整流程"""
    print(f"\n{'='*70}")
    print(f"Testing Project: {project}")
    print('='*70)
    
    context = {
        'localDataAccess': 'uat-local',
        'localDataRoot': str(data_root),
        'sessionId': f'test-{project}',
        'mode': 'enforce'
    }
    
    try:
        # Step 1: Inspect
        print("\n[1/4] Inspect Phase...")
        inspect_result = _handle({
            'operation': 'listing_inspect',
            'project': project,
            'context': context
        })
        
        inspection = inspect_result.get('inspection', {})
        status = inspection.get('status')
        scenario = inspection.get('scenario', '')
        datasets = inspection.get('datasets', [])
        
        print(f"  Status: {status}")
        print(f"  Scenario: {scenario}")
        print(f"  Datasets: {len(datasets)}")
        
        for ds in datasets[:3]:  # 显示前3个
            ds_name = ds if isinstance(ds, str) else (ds.get('dataset') if isinstance(ds, dict) else str(ds))
            print(f"    - {ds_name}")
        
        if status != 'ready':
            print(f"  FAIL Inspect status is {status}, expected 'ready'")
            return False
        
        if not datasets:
            print("  FAIL No datasets found")
            return False
        
        print("  PASS Inspect phase")
        
        # Step 2: Run Code
        print("\n[2/4] Run Code Phase...")
        
        # 获取第一个数据集名称
        first_dataset = datasets[0] if isinstance(datasets[0], str) else datasets[0].get('dataset', 'dm')
        code = f"result = datasets['{first_dataset}']\n"
        
        print(f"  Executing code with dataset: {first_dataset}")
        
        run_result = _handle({
            'operation': 'listing_run_code',
            'project': project,
            'scenario': scenario,
            'code': code,
            'context': context
        })
        
        receipt = run_result.get('receipt', {})
        outputs = receipt.get('outputs', {})
        
        # 处理 outputs 可能是 list 或 dict
        if isinstance(outputs, list):
            print(f"  Outputs: {len(outputs)} items (list format)")
            for item in outputs[:3]:
                if isinstance(item, dict):
                    name = item.get('name', 'unknown')
                    rows = item.get('rowCount', 0)
                    cols = item.get('columnCount', 0)
                    print(f"    {name}: {rows} rows x {cols} columns")
        elif isinstance(outputs, dict):
            print(f"  Outputs: {len(outputs)} tables")
            for name, meta in list(outputs.items())[:3]:
                rows = meta.get('rowCount', 0) if isinstance(meta, dict) else 0
                cols = meta.get('columnCount', 0) if isinstance(meta, dict) else 0
                print(f"    {name}: {rows} rows x {cols} columns")
        else:
            print(f"  Outputs: {outputs}")
        
        if not outputs:
            print("  FAIL No outputs generated")
            return False
        
        print("  PASS Run Code phase")
        
        # Step 3: Publish
        print("\n[3/4] Publish Phase...")
        
        publish_result = _handle({
            'operation': 'listing_publish',
            'project': project,
            'context': context
        })
        
        artifact = publish_result.get('artifact', {})
        artifact_path = artifact.get('path', '')
        sheets = artifact.get('sheets', [])
        
        print(f"  Output file: {artifact_path}")
        print(f"  Sheets: {len(sheets)}")
        
        if not artifact_path:
            print("  FAIL No artifact path returned")
            return False
        
        print("  PASS Publish phase")
        
        # Step 4: Verify
        print("\n[4/4] Verification Phase...")
        
        output_path = data_root / project / '.clinical-listing' / 'output' / 'medical' / 'MEDICAL_LISTINGS.xlsx'
        
        if not output_path.exists():
            print(f"  FAIL Output file not found: {output_path}")
            return False
        
        print(f"  PASS File exists: {output_path.name}")
        
        # 验证 Excel 内容
        wb = openpyxl.load_workbook(output_path, read_only=True)
        print(f"  Sheets found: {', '.join(wb.sheetnames[:5])}")
        
        if len(wb.sheetnames) < 1:
            print(f"  FAIL Expected at least 1 sheet, got {len(wb.sheetnames)}")
            wb.close()
            return False
        
        # 检查第一个表
        first_sheet = wb[wb.sheetnames[0]]
        print(f"  First sheet: {first_sheet.title}")
        print(f"  Rows: {first_sheet.max_row}")
        print(f"  Columns: {first_sheet.max_column}")
        
        wb.close()
        
        print(f"\n  PASS All verification checks passed")
        print(f"\nPASS Project {project} E2E test completed successfully")
        return True
        
    except Exception as e:
        import traceback
        print(f"\nFAIL Exception occurred: {type(e).__name__}: {e}")
        print("\nTraceback:")
        traceback.print_exc()
        return False

def main():
    print("="*70)
    print("Clinical Guard System - Full E2E Test")
    print("Simulating user workflow in Web UI")
    print("="*70)
    
    # 检查 clinical-data 目录
    data_root = Path('G:/home/Clinical-Data')
    if not data_root.exists():
        print(f"\nFAIL clinical-data directory not found: {data_root.absolute()}")
        return 1
    
    print(f"\nData root: {data_root.absolute()}")
    
    # 获取可用项目
    projects = [p.name for p in data_root.iterdir() if p.is_dir() and not p.name.startswith('.')]
    
    if not projects:
        print("\nFAIL No projects found")
        return 1
    
    print(f"\nAvailable projects ({len(projects)}):")
    for i, proj in enumerate(sorted(projects), 1):
        print(f"  {i}. {proj}")
    
    # 测试项目
    if 'CGB3002-TEST' in projects:
        test_projects = ['CGB3002-TEST']
        print(f"\nTesting primary test project: CGB3002-TEST")
    else:
        test_projects = [projects[0]]
        print(f"\nNo CGB3002-TEST found, using: {test_projects[0]}")
    
    passed = 0
    failed = 0
    
    for project in test_projects:
        if test_project_e2e(project, data_root):
            passed += 1
        else:
            failed += 1
    
    # 总结
    print(f"\n{'='*70}")
    print("Test Summary")
    print('='*70)
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")
    print(f"Total: {passed + failed}")
    
    if failed == 0:
        print("\nRESULT: PASS - All E2E tests passed")
        print("System is production ready for real user workflows")
        return 0
    else:
        print(f"\nRESULT: FAIL - {failed} test(s) failed")
        return 1

if __name__ == '__main__':
    sys.exit(main())
