"""8 个真实临床项目的 inspect -> run code -> publish 端到端门禁。"""
from __future__ import annotations

import argparse
import json
import sys
import time
import uuid
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from security.listing_code_lane import reset_code_lane_state
from security.worker import _handle


PROJECTS = (
    "ADAV-008-CP4",
    "CGB3002-TEST",
    "DS5565-0002-NIS",
    "GQ1005-301",
    "H301",
    "RBQM",
    "RT01",
    "YL202-CN-301-01",
)


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _request(operation: str, project: str, context: dict[str, Any], **payload: Any) -> dict[str, Any]:
    response = _handle({
        "operation": operation,
        "project": project,
        "context": context,
        **payload,
    })
    _require(response.get("ok") is True, f"{operation} failed: {response.get('code')}")
    return response


def _verify_workbook(
    artifact_path: Path, artifact: dict[str, Any], expected_output: dict[str, Any],
) -> dict[str, int]:
    _require(artifact_path.is_file(), "published workbook is missing")
    workbook = openpyxl.load_workbook(artifact_path, read_only=True, data_only=False)
    try:
        _require(workbook.sheetnames and workbook.sheetnames[0].casefold() == "contents",
                 "published workbook has no Contents sheet")
        receipt_sheets = artifact.get("sheets") or []
        _require(len(receipt_sheets) == 1, "publish receipt has an unexpected sheet count")
        receipt_sheet = receipt_sheets[0]
        _require(receipt_sheet.get("name") in workbook.sheetnames,
                 "publish receipt sheet is absent from workbook")
        sheet = workbook[str(receipt_sheet["name"])]
        actual_rows = max(0, sheet.max_row - 2)
        actual_columns = sheet.max_column
        _require(actual_rows == int(receipt_sheet.get("rowCount") or 0),
                 "workbook row count differs from publish receipt")
        _require(actual_rows == int(expected_output.get("rowCount") or 0),
                 "publish replay row count differs from successful run")
        _require(actual_columns == int(receipt_sheet.get("columnCount") or 0),
                 "workbook column count differs from publish receipt")
        _require(actual_columns >= int(expected_output.get("columnCount") or 0),
                 "publish replay lost output columns")
        return {"rows": actual_rows, "columns": actual_columns, "sheets": len(workbook.sheetnames)}
    finally:
        workbook.close()


def run_project(data_root: Path, project: str, session_id: str) -> dict[str, Any]:
    context = {
        "localDataAccess": "uat-local",
        "localDataRoot": str(data_root),
        "sessionId": session_id,
        "mode": "enforce",
    }
    inspected = _request("listing_inspect", project, context)["inspection"]
    _require(inspected.get("clinicalGuard") == "CLINICAL_LISTING_INSPECTION",
             "inspect receipt marker is invalid")
    _require(inspected.get("status") == "ready", "inspect did not reach ready status")
    _require(not inspected.get("missing"), "inspect reports missing inputs")
    datasets = inspected.get("datasets") or []
    runnable = next((item for item in datasets if str(item.get("dataset") or "").strip()), None)
    _require(runnable is not None, "inspect returned no runnable dataset")
    dataset = str(runnable["dataset"])
    scenario = str(inspected.get("scenario") or "")
    _require(bool(scenario), "inspect did not resolve a scenario")

    code = f"result = datasets[{dataset!r}]\n"
    run_receipt = _request(
        "listing_run_code", project, context, scenario=scenario, code=code,
    )["receipt"]
    _require(run_receipt.get("clinicalGuard") == "CLINICAL_LISTING_CODE_RECEIPT",
             "run receipt marker is invalid")
    _require(run_receipt.get("status") == "ok", "run did not complete successfully")
    _require(run_receipt.get("dataClass") == "METADATA_ONLY", "run receipt is not metadata-only")
    outputs = run_receipt.get("outputs") or []
    _require(len(outputs) == 1, "run returned an unexpected output count")

    publish_receipt = _request(
        "listing_publish", project, context, scenario=scenario,
    )["receipt"]
    _require(publish_receipt.get("clinicalGuard") == "CLINICAL_LISTING_RECEIPT",
             "publish receipt marker is invalid")
    _require(publish_receipt.get("status") == "completed", "publish did not complete")
    artifacts = publish_receipt.get("artifacts") or []
    _require(len(artifacts) == 1, "publish returned an unexpected artifact count")
    artifact = artifacts[0]
    artifact_path = (
        data_root / project / ".clinical-listing" / "output" / scenario / str(artifact.get("name"))
    )
    workbook = _verify_workbook(artifact_path, artifact, outputs[0])
    return {
        "project": project,
        "scenario": scenario,
        "datasetCount": len(datasets),
        "rows": workbook["rows"],
        "columns": workbook["columns"],
        "sheets": workbook["sheets"],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-root", type=Path, default=Path(r"G:\home\Clinical-Data"))
    parser.add_argument("--rounds", type=int, default=3)
    args = parser.parse_args()
    _require(args.rounds > 0, "round count must be positive")
    data_root = args.data_root.resolve(strict=True)
    for project in PROJECTS:
        _require((data_root / project).is_dir(), f"real project is missing: {project}")

    reset_code_lane_state()
    started = time.monotonic()
    completed = 0
    try:
        for round_number in range(1, args.rounds + 1):
            round_started = time.monotonic()
            for project in PROJECTS:
                session_id = f"real-e2e-r{round_number}-{uuid.uuid4().hex}"
                result = run_project(data_root, project, session_id)
                completed += 1
                print(json.dumps({
                    "status": "PASS",
                    "round": round_number,
                    **result,
                }, ensure_ascii=True), flush=True)
            print(json.dumps({
                "status": "ROUND_PASS",
                "round": round_number,
                "projects": len(PROJECTS),
                "elapsedSeconds": round(time.monotonic() - round_started, 1),
            }), flush=True)
    except Exception as exc:
        print(json.dumps({
            "status": "FAIL",
            "completedProjectRuns": completed,
            "errorType": type(exc).__name__,
            "message": str(exc),
        }, ensure_ascii=True), flush=True)
        return 1
    finally:
        reset_code_lane_state()

    print(json.dumps({
        "status": "ALL_PASS",
        "rounds": args.rounds,
        "projectsPerRound": len(PROJECTS),
        "completedProjectRuns": completed,
        "elapsedSeconds": round(time.monotonic() - started, 1),
    }), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
