from __future__ import annotations

import io
import json
import os
import stat
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from unittest import mock


ROOT = Path(__file__).resolve().parents[2]
# 2026-08-25 架构迁移：Python 运行时（security/、assets/、JS 桥接 src/）
# 已移入 python/ 子目录；ROOT 仍指插件根（含 package.json、excel_header_extractor.py）。
PYTHON_ROOT = ROOT / "python"
sys.path.insert(0, str(PYTHON_ROOT))

from security.archive_passwords import extract_dataset_archive, password_candidates
from security.header_detect import process_csv, process_xls, process_xlsx
from security.local_data_inspector import LocalDataInspectionError, inspect_local_data
from security.listing_data_catalog import DatasetCatalog
from security.path_policy import PathPolicyError, read_credential, resolve_under_root, safe_extract_zip
from security import spec_parser
from security.spec_parser import find_spec_documents, parse_spec_document
from security.worker import _handle


def _xlsx(path: Path, mixed_record: bool = False) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "KRI Requirements"
    sheet.append(["说明"])
    sheet.append([])
    sheet.append(["KRI需求"])
    sheet.append([
        "KRI编号", "KRI名称", "维度", "计算方式", "KRI结果显示",
        "小数保留位数", "监控频率", "黄色预警阈值",
        "红色触发阈值", "关联风险", "数据来源",
    ])
    sheet.append([
        "KRI-001", "失访率", "质量", "失访人数/入组人数", "百分比",
        "2", "月度", "20%", "30%", "数据完整性", "DS/DM",
    ])
    if mixed_record:
        sheet.append(["101-001-0001 | 2026-08-19 | ALT 342", "不得出域"])
    workbook.save(path)


def _als_xlsx(path: Path, mappings: list[tuple[str, str, str]] | None = None) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ALS"
    sheet.append(["DatasetName", "ItemName", "PreText", "ItemOrder"])
    for order, (dataset, source, label) in enumerate(
        mappings or [("DM", "USUBJID", "Subject"), ("DM", "AGE", "Age")], start=1,
    ):
        sheet.append([dataset, source, label, order])
    workbook.save(path)


def _relational_als_xlsx(path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    forms = workbook.active
    forms.title = "Forms"
    forms.append(["FormOID", "FormName", "SASDatasetName"])
    forms.append(["FORM.DM", "人口统计学", "DM"])
    items = workbook.create_sheet("Items")
    items.append(["ItemOID", "SASFieldName", "ItemName", "SASLabel"])
    items.append(["ITEM.AGE", "AGE", "年龄", "年龄"])
    links = workbook.create_sheet("FormItem")
    links.append(["FormOID", "ItemOID"])
    links.append(["FORM.DM", "ITEM.AGE"])
    programming = workbook.create_sheet("AdvancedProgrammings")
    programming.append(["SourceCode"])
    programming.append(["const forbiddenContext = patientRows.map((row) => row.SUBJID);"])
    derivations = workbook.create_sheet("Derivations")
    derivations.append(["Formula"])
    derivations.append(["AGE = floor((VISDAT - BRTHDAT) / 365.25)"])
    workbook.save(path)


def _form_field_als_xlsx(path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    forms = workbook.active
    forms.title = "EDC Objects"
    forms.append(["OID", "Ordinal", "DraftFormName"])
    forms.append(["DM", 1, "人口统计学"])
    fields = workbook.create_sheet("EDC Attributes")
    fields.append([
        "FormOID", "FieldOID", "Ordinal", "DraftFieldName",
        "VariableOID", "PreText", "SASLabel",
    ])
    fields.append(["DM", "AGE_FIELD", 2, "年龄", "AGE", "受试者年龄", "年龄"])
    workbook.save(path)


def _layout_xlsx(path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Medical Listing"
    sheet.append(["Age from Spec", "Participant from Spec"])
    sheet.append(["AGE", "USUBJID"])
    workbook.save(path)


def _xpt(path: Path, subjects: list[str] | None = None) -> None:
    import pandas as pd
    import pyreadstat

    values = subjects or ["101-001-0001", "101-001-0002"]
    frame = pd.DataFrame({"USUBJID": values, "AGE": [42, 57][:len(values)]})
    pyreadstat.write_xport(
        frame,
        str(path),
        column_labels={"USUBJID": "Subject ID", "AGE": "Age"},
        table_name="DM",
    )


def _real_study(root: Path, project_name: str = "study") -> Path:
    project = root / project_name
    docs = project / "doc"
    data = project / "data"
    docs.mkdir(parents=True)
    data.mkdir()
    _als_xlsx(docs / "ALS.xlsx")
    _xpt(data / "DM.xpt")
    return project


def _zip(path: Path, members: dict[str, bytes]) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in members.items():
            archive.writestr(name, content)


def test_root_relative_path_policy() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory) / "root"
        project = root / "project"
        project.mkdir(parents=True)
        assert resolve_under_root(root, "project") == project.resolve()
        for rejected in (str(project.resolve()), "../outside", r"C:\clinical\project"):
            try:
                resolve_under_root(root, rejected)
            except PathPolicyError:
                pass
            else:
                raise AssertionError(f"unsafe path accepted: {rejected}")


def test_local_metadata_rejects_absolute_path_even_inside_root() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory) / "root"
        root.mkdir()
        source = root / "source.csv"
        source.write_text("Subject,Visit,Status\nA1234567,1,Open\n", encoding="utf-8")
        try:
            inspect_local_data(str(root), str(source.resolve()))
        except LocalDataInspectionError as exc:
            assert str(root) not in str(exc)
        else:
            raise AssertionError("local metadata accepted an absolute path")


def test_headerless_metadata_projects_values_and_counts_first_row() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        values = ["张三", "2026-08-19", "101-001-0001", 342]
        csv_path = root / "headerless.csv"
        csv_path.write_text(
            "张三,2026-08-19,101-001-0001,342\n李四,2026-08-20,101-001-0002,128\n",
            encoding="utf-8",
        )
        xlsx_path = root / "headerless.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(values)
        worksheet.append(["李四", "2026-08-20", "101-001-0002", 128])
        workbook.save(xlsx_path)

        for relative in ("headerless.csv", "headerless.xlsx"):
            result = inspect_local_data(str(root), relative)
            sheet = result["sheets"][0]
            blob = json.dumps(result, ensure_ascii=False)
            assert sheet["columns"] == ["COLUMN_1", "COLUMN_2", "COLUMN_3", "COLUMN_4"]
            assert sheet["rowCount"] == 2
            for value in values:
                assert str(value) not in blob

        names = root / "uppercase-names.csv"
        names.write_text("JOHN,SMITH\nJANE,DOE\n", encoding="utf-8")
        result = inspect_local_data(str(root), "uppercase-names.csv")
        assert result["sheets"][0]["columns"] == ["COLUMN_1", "COLUMN_2"]
        assert result["sheets"][0]["rowCount"] == 2
        assert "JOHN" not in json.dumps(result, ensure_ascii=False)


def test_metadata_preserves_proven_clinical_headers() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        source = root / "source.csv"
        source.write_text("Subject,Visit,Status\nA1234567,1,Open\n", encoding="utf-8")
        result = inspect_local_data(str(root), "source.csv")
        assert result["sheets"][0] == {
            "name": "data",
            "rowCount": 1,
            "columns": ["Subject", "Visit", "Status"],
        }
        cdisc = root / "cdisc.csv"
        cdisc.write_text("USUBJID,BRTHDTC,SEX\nA1234567,2024-03-05,M\n", encoding="utf-8")
        cdisc_result = inspect_local_data(str(root), "cdisc.csv")
        assert cdisc_result["sheets"][0]["columns"] == ["USUBJID", "BRTHDTC", "SEX"]


def test_header_extractor_has_no_full_read_mode() -> None:
    extractor = (ROOT / "excel_header_extractor.py").read_text(encoding="utf-8")
    detector = (PYTHON_ROOT / "security" / "header_detect.py").read_text(encoding="utf-8")
    node_guard = (PYTHON_ROOT / "src" / "tool-result-guard.js").read_text(encoding="utf-8")
    for obsolete in ("--mode", "--max-full-rows", "dump_xlsx_full", "dump_xls_full", "dump_csv_full"):
        assert obsolete not in extractor
        assert obsolete not in detector
        assert obsolete not in node_guard


def test_header_extractor_projects_unproven_text_cells() -> None:
    import subprocess

    with tempfile.TemporaryDirectory() as directory:
        source = Path(directory) / "names.csv"
        source.write_text("JOHN,SMITH\nJANE,DOE\n", encoding="utf-8")
        result = subprocess.run(
            [sys.executable, str(ROOT / "excel_header_extractor.py"), str(source)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=30,
        )
        assert result.returncode == 0, result.stderr
        payload = json.loads(result.stdout)
        blob = json.dumps(payload, ensure_ascii=False)
        assert "JOHN" not in blob and "SMITH" not in blob
        assert "JANE" not in blob and "DOE" not in blob
        cells = payload["sheets"][0]["header_cells"]
        assert cells
        assert all(cell["value"] == f"COLUMN_{cell['col'] + 1}" for cell in cells)


def test_horizontal_header_coordinates_match_across_supported_spreadsheets() -> None:
    """横向矩阵的观测标签必须固定在第 0 行，不能继承行循环的末尾索引。"""
    import openpyxl
    import xlwt

    rows = [
        ["Measure", "Observation A", "Observation B"],
        ["ALT", 12, 13],
        ["AST", 21, 22],
    ]
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        csv_path = root / "horizontal.csv"
        csv_path.write_text(
            "Measure,Observation A,Observation B\nALT,12,13\nAST,21,22\n",
            encoding="utf-8",
        )

        xlsx_path = root / "horizontal.xlsx"
        xlsx_book = openpyxl.Workbook()
        xlsx_sheet = xlsx_book.active
        for row in rows:
            xlsx_sheet.append(row)
        xlsx_book.save(xlsx_path)
        xlsx_book.close()

        xls_path = root / "horizontal.xls"
        xls_book = xlwt.Workbook()
        xls_sheet = xls_book.add_sheet("Sheet1")
        for row_index, row in enumerate(rows):
            for column_index, value in enumerate(row):
                xls_sheet.write(row_index, column_index, value)
        xls_book.save(str(xls_path))

        results = [
            process_csv(str(csv_path), 20)[0],
            process_xlsx(str(xlsx_path), None, 20)[0],
            process_xls(str(xls_path), None, 20)[0],
        ]
        for result in results:
            assert result["orientation"] == "HORIZONTAL"
            coordinates = {(cell["row"], cell["col"]) for cell in result["header_cells"]}
            assert {(0, 1), (0, 2)} <= coordinates
            assert (2, 1) not in coordinates and (2, 2) not in coordinates


class TestSkipped(Exception):
    """环境不具备前置条件；按发布门槛要求必须可见并说明原因，不得静默通过。"""


def test_reparse_point_escape_is_rejected() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory) / "root"
        outside = Path(directory) / "outside"
        root.mkdir()
        outside.mkdir()
        link = root / "linked"
        try:
            link.symlink_to(outside, target_is_directory=True)
        except OSError as symlink_error:
            if os.name != "nt":
                raise AssertionError("cannot create a symlink for the path escape test") from symlink_error
            # F-9: 无符号链接权限时回退 junction。symlink_to 失败可能已留下残留
            # 条目，导致 mklink 报 "Cannot create a file when that file already
            # exists"——那是测试清理缺陷，不是被测逻辑的缺陷。先防御性清理。
            for cleanup in (lambda: link.unlink(), lambda: link.rmdir()):
                try:
                    cleanup()
                    break
                except OSError:
                    continue
            junction = subprocess.run(
                ["cmd.exe", "/d", "/c", "mklink", "/J", str(link), str(outside)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            if junction.returncode or not link.exists():
                # 本机既无 symlink 权限也建不出 junction。路径越界保护本身由
                # test_zip_rejects_traversal_absolute_and_links_without_residue
                # 覆盖，这里显式 skip 而不是 FAIL——让门禁结果反映"环境缺前置
                # 条件"，而不是伪装成代码缺陷（发布门槛要求跳过项可见）。
                raise TestSkipped(
                    "no symlink privilege and junction creation failed: "
                    f"{junction.stderr.strip() or junction.stdout.strip()}"
                ) from symlink_error
        try:
            resolve_under_root(root, "linked")
        except PathPolicyError:
            return
        raise AssertionError("symlink escape was accepted")


def test_zip_rejects_traversal_absolute_and_links_without_residue() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        _zip(root / "traversal.zip", {"../escape.txt": b"x"})
        _zip(root / "absolute.zip", {r"C:\escape.txt": b"x"})
        link_zip = root / "link.zip"
        with zipfile.ZipFile(link_zip, "w") as archive:
            info = zipfile.ZipInfo("linked")
            info.create_system = 3
            info.external_attr = (stat.S_IFLNK | 0o777) << 16
            archive.writestr(info, "target")
        for archive in [root / "traversal.zip", root / "absolute.zip", link_zip]:
            destination = root / f"out-{archive.stem}"
            try:
                safe_extract_zip(archive, destination)
            except PathPolicyError:
                pass
            else:
                raise AssertionError(f"unsafe ZIP accepted: {archive.name}")
            assert not list(root.glob(".extract-*"))


def test_zip_extracts_atomically_and_rejects_nonempty_target() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        archive = root / "ok.zip"
        _zip(archive, {"data/source.sas7bdat": b"fixture"})
        target = root / "managed"
        extracted = safe_extract_zip(archive, target)
        assert extracted == [target / "data" / "source.sas7bdat"]
        assert extracted[0].read_bytes() == b"fixture"
        try:
            safe_extract_zip(archive, target)
        except PathPolicyError:
            pass
        else:
            raise AssertionError("non-empty managed target was overwritten")


def test_zip_accepts_realistic_high_compression_sas_with_size_caps() -> None:
    """稀疏 SAS 页可超过旧 200:1 阈值，不应被误判为 ZIP bomb。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        archive = root / "sparse-sas.zip"
        payload = bytes(range(256)) * 16 + b" " * (2 * 1024 * 1024 - 4096)
        with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as bundle:
            bundle.writestr("dm.sas7bdat", payload)
        with zipfile.ZipFile(archive) as bundle:
            info = bundle.getinfo("dm.sas7bdat")
            ratio = info.file_size / info.compress_size
            assert 200 < ratio < 1_000
        target = root / "managed"
        extracted = safe_extract_zip(archive, target)
        assert extracted == [target / "dm.sas7bdat"]
        assert extracted[0].stat().st_size == len(payload)


def test_nested_dataset_archive_is_extracted_under_managed_destination() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = root / "STUDY-001"
        project.mkdir()
        nested_bytes = io.BytesIO()
        with zipfile.ZipFile(nested_bytes, "w", zipfile.ZIP_DEFLATED) as nested:
            nested.writestr("data/dm.sas7bdat", b"fixture")
        archive = project / "delivery.zip"
        with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as outer:
            outer.writestr("nested.zip", nested_bytes.getvalue())

        destination = project / "_work" / "delivery"
        extracted = extract_dataset_archive(project, archive, destination)
        datasets = [path for path in extracted if path.suffix == ".sas7bdat"]
        assert len(datasets) == 1
        assert datasets[0].read_bytes() == b"fixture"
        assert datasets[0].is_relative_to(destination)


def test_catalog_collapses_only_byte_identical_dataset_copies() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "STUDY-001"
        (project / "one").mkdir(parents=True)
        (project / "two").mkdir()
        (project / "one" / "dm.csv").write_bytes(b"A\n1\n")
        (project / "two" / "dm.csv").write_bytes(b"A\n1\n")
        with DatasetCatalog(project, materialize_archives=False) as catalog:
            assert len(catalog.files()["dm"]) == 1

        (project / "two" / "dm.csv").write_bytes(b"A\n2\n")
        with DatasetCatalog(project, materialize_archives=False) as catalog:
            assert len(catalog.files()["dm"]) == 2


def test_spec_document_is_exempt_from_automatic_redaction() -> None:
    with tempfile.TemporaryDirectory() as directory:
        spec = Path(directory) / "requirements.xlsx"
        _xlsx(spec, mixed_record=True)
        parsed = parse_spec_document(str(spec), "kri")
        blob = json.dumps(parsed, ensure_ascii=False)
        assert "KRI-001" in blob
        assert "失访率" in blob
        assert "20%" in blob and "30%" in blob
        for raw in ("101-001-0001", "2026-08-19", "ALT 342"):
            assert raw in blob


def test_spec_requirement_text_reaches_harness_verbatim() -> None:
    """doc 内 spec 的判定规则必须原文到达 harness，不得 token 化。

    2026-08-23：`_safe_text` 对含日期/编号形态的单元格做 token 化，把
    "ALT: 3 倍正常上限" 打成 "[TEXT:..]: [NUM:..] [TEXT:..]"，AI 无法理解
    spec 需求。doc 内文档是规格来源不是 data 来源，按用户边界应完整放行。
    """
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        spec = Path(directory) / "listing_spec.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "MM listing要求"
        sheet.append(["序号", "MM Listing要求"])
        sheet.append([1, "ALT: 3 倍正常上限，AESTDTC 不早于 2024-01-01"])
        sheet.append([2, "访视日期格式 2024-03-05，受试者编号形如 A1234567"])
        sheet.append([3, "WBC = 4 时标记异常"])
        workbook.save(spec)
        workbook.close()

        parsed = parse_spec_document(str(spec), "spec")
        blob = json.dumps(parsed, ensure_ascii=False)
        for raw in (
            "ALT: 3 倍正常上限",
            "AESTDTC 不早于 2024-01-01",
            "访视日期格式 2024-03-05",
            "A1234567",
            "WBC = 4",
            "MM listing要求",
        ):
            assert raw in blob, f"spec 规格文本被脱敏，AI 无法理解: {raw}"
        for token_kind in ("[TEXT:", "[NUM:", "[DATE:", "[SUBJ:", "[VAL:"):
            assert token_kind not in blob, f"spec 解析产出了 token 化残留: {token_kind}"


def test_spec_parser_preserves_content_beyond_legacy_limits() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        spec = Path(directory) / "large.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Column", "Label"])
        long_label = "规格说明" * 300
        for index in range(10_001):
            sheet.append([f"FIELD_{index}", long_label if index == 10_000 else f"label-{index}"])
        for index in range(64):
            workbook.create_sheet(f"extra-{index}").append([f"sheet-{index}"])
        workbook.save(spec)
        workbook.close()

        parsed = parse_spec_document(str(spec), "spec")

        assert len(parsed["fields"]) == 10_001
        assert len(parsed["sheets"]) == 65
        assert parsed["fields"][-1]["label"] == long_label
        assert parsed["warnings"] == []


def test_spec_parser_preserves_columns_beyond_legacy_limits() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        spec = Path(directory) / "exact.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        columns = [f"COL_{index}" for index in range(300)]
        sheet.append(columns)
        workbook.save(spec)
        workbook.close()

        parsed = parse_spec_document(str(spec), "spec")

        assert len(parsed["sheets"][0]["columns"]) == 300
        assert len(parsed["requirements"][0]["cells"]) == 300
        assert parsed["requirements"][0]["cells"][-1] == "COL_299"




def test_data_example_rows_are_fully_readable_as_spec_requirements() -> None:
    # 2026-08-24 红线口径：doc/ 是受信规格域，内容必须全量供模型理解需求；
    # 数据域泄露由出域侧按来源域阻断，不再靠解析层丢弃"形似数据"的行。
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        spec = Path(directory) / "MM Listing要求.xlsx"
        workbook = openpyxl.Workbook()
        example = workbook.active
        example.title = "SV"
        example.append(["中心编号", "受试者编号", "表单名称", "访视日期"])
        example.append(["SITEID", "SUBJID", "FORMNM", "VISDAT"])
        example.append(["99", "101-001-0001", "访视日期", "2026-08-19"])
        requirements = workbook.create_sheet("MM listing要求")
        requirements.append(["序号", "MM Listing要求"])
        requirements.append([1, "删除 Code Value 列"])
        workbook.save(spec)
        workbook.close()

        parsed = parse_spec_document(str(spec), "spec")
        blob = json.dumps(parsed["requirements"], ensure_ascii=False)
        assert "删除 Code Value 列" in blob
        assert "101-001-0001" in blob
        sv = next(sheet for sheet in parsed["sheets"] if sheet["name"] == "SV")
        assert [column["name"] for column in sv["columns"]] == [
            "SITEID", "SUBJID", "FORMNM", "VISDAT",
        ]


def test_spec_parser_removes_invalid_auto_filter_ref_from_workbook_copy() -> None:
    with tempfile.TemporaryDirectory() as directory:
        spec = Path(directory) / "broken_filter.xlsx"
        _xlsx(spec)
        patched = Path(directory) / "patched.xlsx"
        with zipfile.ZipFile(spec) as source, zipfile.ZipFile(patched, "w") as target:
            for info in source.infolist():
                content = source.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    content = content.replace(b'ref="A1:INVALID"', b'ref="not-a-range"')
                    if b'ref="not-a-range"' not in content:
                        content = content.replace(b'<worksheet', b'<worksheet><autoFilter ref="not-a-range"', 1)
                        content = content.replace(b'</worksheet>', b'</autoFilter></worksheet>', 1)
                target.writestr(info, content)
        parsed = parse_spec_document(str(patched), "kri")
        assert parsed["kris"]


def test_relational_multi_edc_als_joins_forms_items_and_form_items() -> None:
    with tempfile.TemporaryDirectory() as directory:
        als = Path(directory) / "study_ALS.xlsx"
        _relational_als_xlsx(als)

        parsed = parse_spec_document(str(als), "als")
        assert parsed["datasets"] == ["DM"]
        assert parsed["forms"] == [{"formName": "人口统计学", "datasetName": "DM"}]
        assert parsed["mappings"] == [{
            "datasetName": "DM",
            "sourceColumn": "AGE",
            "displayLabel": "年龄",
            "formName": "人口统计学",
            "order": 0,
        }]
        assert parsed["requirements"] == []


def test_form_field_edc_als_is_detected_by_column_roles() -> None:
    with tempfile.TemporaryDirectory() as directory:
        als = Path(directory) / "study_ALS.xlsx"
        _form_field_als_xlsx(als)

        parsed = parse_spec_document(str(als), "als")
        assert parsed["datasets"] == ["DM"]
        assert parsed["forms"] == [{"formName": "人口统计学", "datasetName": "DM"}]
        assert parsed["mappings"] == [{
            "datasetName": "DM",
            "sourceColumn": "AGE",
            "displayLabel": "年龄",
            "formName": "人口统计学",
            "order": 2.0,
        }]
        assert parsed["requirements"] == []












def test_password_candidates_include_marker_stem_without_logging_values() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "GQ1005-301"
        project.mkdir()
        archive = project / "SAS_20250221.zip"
        archive.touch()
        marker = project / "A1234567.txt"
        marker.write_text("ignored-local-value", encoding="utf-8")

        candidates = password_candidates(project, archive)
        assert candidates[:3] == [
            b"GQ1005-301", b"GQ1005301", b"GQ1005",
        ]
        assert b"A1234567" in candidates
        assert b"SAS_20250221" in candidates
        assert candidates[-1] is None


def test_large_multiline_credential_stays_local_and_is_readable() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        secret = "first-line\n" + ("local-secret-material\n" * 2000)
        (root / "archive-password.txt").write_text(secret, encoding="utf-8")
        assert read_credential(root, "archive-password.txt") == secret.strip().encode("utf-8")


def test_archive_password_failure_is_actionable_without_candidate_values() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "STUDY-SECRET-42"
        project.mkdir()
        archive = project / "private-secret.zip"
        with zipfile.ZipFile(archive, "w") as bundle:
            bundle.writestr("dm.csv", "USUBJID\nA1234567\n")
        (project / "password.txt").write_text("never-return-this-password", encoding="utf-8")
        destination = project / "work"
        failed = PathPolicyError("archive extraction failed")
        failed.__cause__ = RuntimeError("bad password")
        with mock.patch("security.archive_passwords.safe_extract_zip", side_effect=failed):
            try:
                extract_dataset_archive(project, archive, destination)
            except PathPolicyError as exc:
                message = str(exc)
            else:
                raise AssertionError("密码候选耗尽后未返回结构化诊断")
        assert "credentialRef" in message
        for secret in ("STUDY-SECRET-42", "private-secret", "never-return-this-password"):
            assert secret not in message












def test_only_doc_directory_is_a_requirements_source() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory)
        for name in (
            "Medical Listing SPEC.xlsx",
            "listing_template.xlsx",
            "Data Validation Plan.xlsx",
            "patient_listing.xlsx",
            "DM Status Report.xlsx",
        ):
            (project / name).touch()
        doc = project / "doc"
        doc.mkdir()
        (doc / "GQ1005-301_MM Listing要求_20250211.xlsx").touch()
        (doc / "GQ1005-301_ALS_V1.0_20241219.xlsx").touch()
        (doc / "~$temporary.xlsx").touch()
        assert [path.name for path in find_spec_documents(project)] == [
            "GQ1005-301_ALS_V1.0_20241219.xlsx",
            "GQ1005-301_MM Listing要求_20250211.xlsx",
        ]




def test_worker_exposes_no_legacy_business_operations() -> None:
    # F-11: listing_workflow 是被新三阶段车道取代的旧操作。插件早已不注册它，
    # 但 worker 协议层此前仍可达，形成新旧双轨——误调会触发已被取代的失败模式。
    for operation in ("run_local_job", "parse_spec", "sas_metadata", "read_data_file",
                      "listing_workflow"):
        assert _handle({"operation": operation}) == {"ok": False, "code": "UNKNOWN_OPERATION"}


def test_legacy_listing_generator_module_is_gone() -> None:
    """F-11: 旧生成器已删除，不得被任何路径重新引入。"""
    import importlib

    assert not (PYTHON_ROOT / "security" / "emerald_listing_generator.py").exists()
    try:
        importlib.import_module("security.emerald_listing_generator")
    except ModuleNotFoundError:
        pass
    else:
        raise AssertionError("legacy listing generator is still importable")


def main() -> int:
    tests = [value for key, value in sorted(globals().items()) if key.startswith("test_")]
    failures = 0
    skipped = 0
    for test in tests:
        try:
            test()
            print(f"PASS {test.__name__}")
        except TestSkipped as reason:
            # 跳过项必须可见并说明环境原因（发布门槛第 1 条），但不计为失败。
            skipped += 1
            print(f"SKIP {test.__name__}: {reason}")
        except Exception as error:
            failures += 1
            print(f"FAIL {test.__name__}: {error}")
    summary = f"RESULT {len(tests) - failures - skipped}/{len(tests)}"
    print(f"{summary} (SKIPPED {skipped})" if skipped else summary)
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
