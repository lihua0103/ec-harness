"""代码车道（run → publish）工作流合同测试。

红线验证重点：
- run 收据是 METADATA_ONLY：受试者标记（含被走私进列名的值）绝不出现
- publish 前必须有成功 run；产物由固定 Writer 产出（CONTENTS + 复核列）
- run/publish 双预算持续审计但不阻断；IR 车道操作已从 worker 退役
"""
from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from security import listing_budget
from security.listing_code_lane import (
    publish_listing_code,
    reset_code_lane_state,
    run_listing_code,
)
from security.worker import _handle

SENTINEL = "101-001-0001"

# standalone（run_all.py 直跑）没有 conftest 的清理容忍补丁：openpyxl 引用环
# 与杀毒对新建 xlsx 的毫秒级扫描锁会让 TemporaryDirectory.cleanup 撞
# WinError 32。与 conftest._gc_tolerant_cleanup 同款：先打断引用环再短重试。
_original_cleanup = tempfile.TemporaryDirectory.cleanup


def _gc_tolerant_cleanup(self) -> None:
    import gc
    import time

    gc.collect()
    for attempt in range(5):
        try:
            _original_cleanup(self)
            return
        except PermissionError:
            if attempt == 4:
                raise
            gc.collect()
            time.sleep(0.3)


tempfile.TemporaryDirectory.cleanup = _gc_tolerant_cleanup


def _project(root: Path) -> str:
    project = root / "GQ9001"
    (project / "doc").mkdir(parents=True)
    (project / "dm.csv").write_text(
        f"USUBJID,AGE,SEX\n{SENTINEL},44,M\n101-001-0002,51,F\n", encoding="utf-8")
    (project / "vs.csv").write_text(
        f"USUBJID,VSTESTCD,VSSTRESN\n{SENTINEL},TEMP,36.8\n101-001-0002,TEMP,37.1\n"
        "101-001-0002,WEIGHT,62.5\n",
        encoding="utf-8")
    return "GQ9001"


HAPPY_CODE = (
    "dm = datasets['dm']\n"
    "vs = datasets['vs']\n"
    "merged = dm.merge(vs, on='USUBJID', how='left')\n"
    "outputs = {'vital_signs': merged}\n"
)


class CodeLaneWorkflowTest(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory(prefix="code-lane-test-")
        self.root = Path(self._tmp.name)
        self.project = _project(self.root)
        reset_code_lane_state()
        listing_budget.reset_budget()
        listing_budget.reset_run_budget()

    def tearDown(self):
        self._tmp.cleanup()
        reset_code_lane_state()
        listing_budget.reset_budget()
        listing_budget.reset_run_budget()

    def _run(self, code, scenario="medical", session="s1"):
        return run_listing_code(
            local_data_root=str(self.root), project=self.project,
            scenario=scenario, code=code, session_id=session)

    def _publish(self, scenario="medical", session="s1"):
        return publish_listing_code(
            local_data_root=str(self.root), project=self.project,
            scenario=scenario, session_id=session)

    def test_run_receipt_is_metadata_only(self):
        receipt = self._run(HAPPY_CODE)
        self.assertEqual(receipt["status"], "ok")
        self.assertEqual(receipt["dataClass"], "METADATA_ONLY")
        self.assertEqual(receipt["scenario"], "medical")
        outputs = {item["name"]: item for item in receipt["outputs"]}
        self.assertIn("vital_signs", outputs)
        self.assertEqual(outputs["vital_signs"]["rowCount"], 3)
        serialized = json.dumps(receipt, ensure_ascii=False)
        self.assertNotIn(SENTINEL, serialized)

    def test_column_name_smuggling_is_scrubbed(self):
        # 把单元格值走私进列名是代码车道独有的字符串通道，必须被 scrub 拦截。
        receipt = self._run(
            "dm = datasets['dm']\n"
            f"dm = dm.rename(columns={{'AGE': str(dm.iloc[0, 0])}})\n"
            "result = dm\n")
        self.assertEqual(receipt["status"], "ok")
        serialized = json.dumps(receipt, ensure_ascii=False)
        self.assertNotIn(SENTINEL, serialized)

    def test_rejected_code_returns_structured_receipt(self):
        receipt = self._run("import os\nresult = datasets['dm']")
        self.assertEqual(receipt["status"], "rejected")
        self.assertEqual(receipt["code"], "SANDBOX_CODE_REJECTED")

    def test_publish_requires_successful_run(self):
        receipt = self._publish()
        self.assertEqual(receipt["status"], "rejected")
        self.assertEqual(receipt["code"], "NO_SUCCESSFUL_RUN")

    def test_publish_writes_fixed_writer_workbook(self):
        run = self._run(HAPPY_CODE)
        self.assertEqual(run["status"], "ok")
        receipt = self._publish()
        self.assertEqual(receipt["status"], "completed")
        self.assertEqual(receipt["dataClass"], "REAL")
        artifact_dir = self.root / self.project / ".clinical-listing" / "output" / "medical"
        files = list(artifact_dir.glob("MEDICAL_LISTINGS.xlsx"))
        self.assertEqual(len(files), 1)
        import openpyxl

        workbook = openpyxl.load_workbook(files[0], read_only=True)
        try:
            self.assertIn("Contents", workbook.sheetnames)
            sheet = workbook[workbook.sheetnames[1]]
            header = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
            self.assertIn("USUBJID", header)
            # medical 场景固定追加复核列。
            self.assertIn("Flag", header)
            self.assertIn("Review Comments", header)
        finally:
            workbook.close()

    def test_run_budget_warns_without_blocking(self):
        with mock.patch.dict(os.environ, {"EMERALD_LISTING_MAX_CODE_RUNS": "1"}):
            first = self._run(HAPPY_CODE, session="budget")
            second = self._run(HAPPY_CODE, session="budget")
        self.assertEqual(first["status"], "ok")
        self.assertEqual(second["status"], "ok")

    def test_publish_budget_uses_execute_channel(self):
        with mock.patch.dict(os.environ, {"EMERALD_LISTING_MAX_EXECUTIONS": "1"}):
            self._run(HAPPY_CODE, session="pb")
            first = self._publish(session="pb")
            second = self._publish(session="pb")
        self.assertEqual(first["status"], "completed")
        self.assertEqual(second["status"], "completed")


class WorkerDispatchTest(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory(prefix="code-lane-worker-")
        self.root = Path(self._tmp.name)
        self.project = _project(self.root)
        reset_code_lane_state()
        listing_budget.reset_budget()
        listing_budget.reset_run_budget()

    def tearDown(self):
        self._tmp.cleanup()
        reset_code_lane_state()
        listing_budget.reset_budget()
        listing_budget.reset_run_budget()

    def _context(self):
        return {"localDataAccess": "uat-local", "localDataRoot": str(self.root),
                "sessionId": "w1", "mode": "enforce"}

    def test_run_code_dispatch(self):
        response = _handle({
            "operation": "listing_run_code", "project": self.project,
            "scenario": "medical", "code": HAPPY_CODE, "context": self._context(),
        })
        self.assertTrue(response["ok"])
        self.assertEqual(response["action"], "listing-run-code")
        self.assertEqual(response["receipt"]["status"], "ok")

    def test_run_code_requires_uat_local(self):
        response = _handle({
            "operation": "listing_run_code", "project": self.project,
            "scenario": "medical", "code": HAPPY_CODE,
            "context": {"localDataAccess": "cloud", "localDataRoot": str(self.root)},
        })
        self.assertFalse(response["ok"])
        self.assertEqual(response["code"], "LOCAL_DATA_ACCESS_REQUIRED")

    def test_publish_dispatch_after_run(self):
        _handle({
            "operation": "listing_run_code", "project": self.project,
            "scenario": "medical", "code": HAPPY_CODE, "context": self._context(),
        })
        response = _handle({
            "operation": "listing_publish", "project": self.project,
            "scenario": "medical", "context": self._context(),
        })
        self.assertTrue(response["ok"])
        self.assertEqual(response["action"], "listing-publish")
        self.assertEqual(response["receipt"]["status"], "completed")

    def test_ir_lane_operations_are_retired(self):
        for operation in ("listing_validate_plan", "listing_execute"):
            response = _handle({
                "operation": operation, "project": self.project,
                "scenario": "medical", "plan": {}, "context": self._context(),
            })
            self.assertFalse(response["ok"])
            self.assertEqual(response["code"], "UNKNOWN_OPERATION")


if __name__ == "__main__":
    unittest.main()
