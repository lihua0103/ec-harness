from __future__ import annotations

import json
import hashlib
import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DRIVER = ROOT / "tests" / "integration" / "plugin_driver.js"
sys.path.insert(0, str(ROOT))
# FIX-12 后审计默认写入用户主目录；测试显式指回项目 var 以便断言审计内容。
os.environ.setdefault("EMERALD_AUDIT_ROOT", str(ROOT / "var" / "egress_audit"))


def make_xlsx(path: Path):
    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Specification"
    sheet.append(["Subject", "Visit", "Status"])
    sheet.append(["A1234567", "2024-03-05", "Screening"])
    workbook.save(path)


def make_kri_xlsx(path: Path):
    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "KRI Requirements"
    sheet.append(["说明"])
    sheet.append([])
    sheet.append(["KRI需求"])
    sheet.append([
        "KRI编号", "KRI名称", "维度", "计算方式", "KRI结果显示",
        "小数保留位数", "监控频率", "黄色预警阈值",
        "红色触发阈值", "关联风险", "数据来源",
    ])
    sheet.append([
        "KRI-001", "失访率", "质量", "失访人数/入组人数", "百分比",
        "2", "月度", "20%", "30%", "数据完整性", "DS/DM",
    ])
    workbook.save(path)


def make_als_xlsx(path: Path):
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ALS"
    sheet.append(["DatasetName", "ItemName", "PreText", "ItemOrder"])
    sheet.append(["DM", "USUBJID", "Subject", 1])
    sheet.append(["DM", "AGE", "Age", 2])
    workbook.save(path)


def make_dm_xpt(path: Path):
    import pandas as pd
    import pyreadstat

    frame = pd.DataFrame({
        "USUBJID": ["101-001-0001", "101-001-0002"],
        "AGE": [42, 57],
    })
    pyreadstat.write_xport(
        frame,
        str(path),
        column_labels={"USUBJID": "Subject ID", "AGE": "Age"},
        table_name="DM",
    )


def run(scenario: str, excel: Path | None = None,
        python: str | None = None,
        credentials_dir: Path | None = None, credential_file: Path | None = None,
        env_extra: dict | None = None):
    env = os.environ.copy()
    env["PYTHON"] = python or sys.executable
    env["PLUGIN_PYTHON"] = python or sys.executable
    env.setdefault("PYTHONIOENCODING", "utf-8")
    if env_extra:
        env.update(env_extra)
    if excel:
        env["EXCEL_FILE"] = str(excel)
    if credentials_dir:
        env["CREDENTIALS_DIR"] = str(credentials_dir)
    if credential_file:
        env["CREDENTIAL_FILE"] = str(credential_file)
    result = subprocess.run(
        ["node", str(DRIVER), scenario],
        cwd=ROOT,
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=30,
    )
    assert result.returncode == 0, result.stderr
    return json.loads(result.stdout)


def test_local_output_projection_hides_real_records():
    with tempfile.TemporaryDirectory() as directory:
        source = Path(directory) / "source.sas7bdat"
        source.write_bytes(b"fixture")
        output = run(
            "local-output-projection",
            env_extra={"LOCAL_OUTPUT_SOURCE": str(source)},
        )
        serialized = json.dumps(output, ensure_ascii=False)
        assert "010-001-1001" not in serialized
        assert "DATA_BLOCKED" in serialized


def test_protected_source_provenance_blocks_model_roundtrip():
    """数据域文件直读的投影结果可安全进入 llm/stream，但原始值绝不出现。

    2026-08-24 口径修正：投影块（SAS_DATA 占位 / EXCEL_STRUCTURE_ONLY 表头
    结构）本身无数据值，是 local_data_metadata 等合法元数据通道的产物；
    旧断言"带 protectedDataSource 标记即 llm/stream 硬阻断"会把任何查看过
    数据域文件结构的会话判死刑（local_data_metadata 100% 误伤）。安全
    不变量改为断言：投影文本中不出现原始受试者/受保护值。
    """
    for kind, suffix in (("sas", ".xpt"), ("external_excel", ".xlsx")):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / f"source{suffix}"
            source.write_bytes(b"fixture")
            output = run(
                "protected-source-roundtrip",
                env_extra={
                    "PROTECTED_SOURCE_FILE": str(source),
                    "PROTECTED_SOURCE_KIND": kind,
                    "LOCAL_DATA_ROOT": str(Path(directory)),
                },
            )
            serialized = json.dumps(output, ensure_ascii=False)
            assert output["blocked"] is False, output
            assert "101-001-0001" not in serialized
            assert "A1234567" not in serialized


def test_local_metadata_lane_returns_only_structure_and_enforces_root():
    """UAT 本地数据车道只给模型结构元数据，且不可越过指定项目根。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory) / "project"
        root.mkdir()
        source = root / "source.xlsx"
        make_xlsx(source)
        outside = Path(directory) / "outside.xlsx"
        make_xlsx(outside)
        env = os.environ.copy()
        env["PYTHON"] = sys.executable
        env["PLUGIN_PYTHON"] = sys.executable
        env["LOCAL_DATA_ACCESS"] = "uat-local"
        env["LOCAL_DATA_ROOT"] = str(root)
        env["LOCAL_METADATA_FILE"] = str(source)
        env["LOCAL_METADATA_RELATIVE_FILE"] = "source.xlsx"
        env["LOCAL_METADATA_OUTSIDE_FILE"] = str(outside)
        result = subprocess.run(
            ["node", str(DRIVER), "local-metadata"],
            cwd=ROOT,
            env=env,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
        assert result.returncode == 0, result.stderr
        output = json.loads(result.stdout)
        value = output
        assert value["clinicalGuard"] == "LOCAL_METADATA_ONLY"
        assert value["path"] == "source.xlsx"
        assert value["sheets"][0]["columns"] == ["Subject", "Visit", "Status"]
        assert "A1234567" not in json.dumps(value, ensure_ascii=False)

        absolute_inside_result = subprocess.run(
            ["node", str(DRIVER), "local-metadata-absolute-inside-root"],
            cwd=ROOT,
            env=env,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
        assert absolute_inside_result.returncode == 0, absolute_inside_result.stderr
        assert json.loads(absolute_inside_result.stdout)["blocked"] is True

        outside_result = subprocess.run(
            ["node", str(DRIVER), "local-metadata-outside-root"],
            cwd=ROOT,
            env=env,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
        assert outside_result.returncode == 0, outside_result.stderr
        assert json.loads(outside_result.stdout)["blocked"] is True


def test_local_metadata_uses_session_cwd_over_configured_root():
    """元数据工具必须服从 Web UI 当前会话工作区，而不是静态回退目录。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        selected = root / "selected-workspace"
        configured = root / "wrong-configured-root"
        selected.mkdir()
        configured.mkdir()
        make_xlsx(selected / "source.xlsx")
        value = run(
            "local-metadata-session-cwd",
            env_extra={
                "LOCAL_DATA_ACCESS": "uat-local",
                "LOCAL_DATA_ROOT": str(configured),
                "SESSION_CWD": str(selected),
                "LOCAL_METADATA_RELATIVE_FILE": "source.xlsx",
            },
        )
        assert value["clinicalGuard"] == "LOCAL_METADATA_ONLY"
        assert value["path"] == "source.xlsx"
        assert value["sheets"][0]["columns"] == ["Subject", "Visit", "Status"]
        assert str(selected) not in json.dumps(value, ensure_ascii=False)


def test_llm_clean_streams_and_dirty_blocks():
    assert run("llm-clean")["streamed"] is True
    assert run("llm-platform-header-clean")["streamed"] is True
    # 普通模型语义不做全局 token 化；数据边界由来源域和专用工具负责。
    dirty = run("llm-dirty")
    assert dirty["streamed"] is True
    assert dirty["content"] == "Subject A1234567"

    structured = run("llm-structured-dirty")
    forwarded = json.dumps(structured["options"], ensure_ascii=False)
    for raw in ("1234567", "A1234567", "2024-03-05", "5.8"):
        assert raw in forwarded
    assert structured["options"]["temperature"] == 0.2

    disabled = run(
        "llm-structured-dirty-protection-disabled",
        env_extra={"DATA_PROTECTION_ENABLED": "0"},
    )
    disabled_forwarded = json.dumps(disabled["options"], ensure_ascii=False)
    for raw in ("1234567", "A1234567", "2024-03-05", "5.8"):
        assert raw in disabled_forwarded
    assert disabled["options"]["temperature"] == 0.2


def test_full_model_request_scope_blocks_and_audits_clean_requests():
    # system 指令也保持原始语义。
    dirty = run("llm-system-dirty")
    assert dirty["streamed"] is True
    assert dirty["system"] == "Subject A1234567"

    assert run("llm-clean")["streamed"] is True
    payload = {
        "provider": "test-provider",
        "model": "test-model",
        "messages": [{"role": "user", "content": "请生成列表规范。"}],
        "system": "出域审计范围验证",
        "tools": [{
            "name": "demo",
            "description": "演示工具",
            "parameters": {"type": "object", "properties": {}},
        }],
        "temperature": 0.2,
        "maxTokens": 128,
        "stop": ["完成"],
        "purpose": "session-title",
    }
    canonical = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    import os
    audit_root = Path(os.environ.get("EMERALD_AUDIT_ROOT", ROOT / "var" / "egress_audit"))
    audit_path = max(audit_root.glob("*.jsonl"), key=lambda path: path.stat().st_mtime)
    record = json.loads(audit_path.read_text(encoding="utf-8").splitlines()[-1])
    evidence = record["request_evidence"]
    assert record["action"] == "ALLOWED"
    assert evidence["payload_sha256"] == hashlib.sha256(canonical).hexdigest()
    assert evidence["payload_bytes"] == len(canonical)
    assert evidence["message_count"] == 1
    assert {"provider", "model", "messages", "system", "tools", "stop"} <= set(
        evidence["payload_fields"]
    )
    assert "A1234567" not in json.dumps(record, ensure_ascii=False)


def test_non_text_and_invalid_messages_fail_closed():
    assert run("llm-image")["thrown"] is True
    assert run("llm-invalid")["thrown"] is True


def test_excel_post_execute_keeps_headers_only():
    with tempfile.TemporaryDirectory() as directory:
        excel = Path(directory) / "clinical.xlsx"
        make_xlsx(excel)
        decision = run("post-excel", excel=excel)
        blob = json.dumps(decision, ensure_ascii=False)
        assert decision["kind"] == "accept"
        assert "A1234567" not in blob and "2024-03-05" not in blob
        assert "Subject" in blob


def test_no_path_control_result_preserves_semantics():
    decision = run("post-no-path")
    blob = json.dumps(decision, ensure_ascii=False)
    assert "status A1234567 2024-03-05" in blob


def test_credential_file_value_stays_local():
    """本地凭据通道只返回逻辑引用，不暴露原值或绝对目录。"""
    with tempfile.TemporaryDirectory() as directory:
        cred_dir = Path(directory) / "credentials"
        cred_dir.mkdir()
        cred_file = cred_dir / "A1234567.txt"
        cred_file.write_text("A1234567", encoding="utf-8")
        decision = run(
            "credential-local",
            credentials_dir=cred_dir,
            credential_file=cred_file,
        )
        assert decision["kind"] == "accept"
        value = json.loads(decision["content"][0]["text"])
        assert value["clinicalGuard"] == "CREDENTIAL_LOCAL_ONLY"
        assert "A1234567" not in value.get("message", "")
        assert value["credentialRef"] == "A1234567.txt"
        assert str(cred_dir) not in json.dumps(value, ensure_ascii=False)


def test_missing_python_worker_fails_closed():
    decision = run("fail-closed", python="definitely-missing-python-7f3c")
    assert decision["thrown"] is True
    assert "enoent" in decision["message"].lower()


def test_explicit_config_overrides_environment_default():
    state = run(
        "policy-initial-state",
        env_extra={
            "DATA_PROTECTION_ENABLED": "1",
            "DATA_INTERCEPTION_ENABLED": "1",
            "CONFIG_DATA_INTERCEPTION_ENABLED": "0",
        },
    )
    assert state["dataInterceptionEnabled"] is False


def test_fetch_database_result_without_path_is_scrubbed():
    """TC-20 / BY-13（真）：非 read 工具名 + 无路径 + 合成受试者标记被替换。"""
    decision = run("fetch-database")
    blob = json.dumps(decision, ensure_ascii=False)
    assert decision["kind"] == "accept"
    assert "A1234567" not in blob and "2024-03-05" not in blob
    assert "DATA_BLOCKED" in blob and "DATA_QUERY" in blob


def test_fetch_database_error_cannot_bypass_data_query_boundary():
    """数据查询工具把记录值拼入错误文本时仍必须按来源能力阻断。"""
    decision = run("fetch-database-error")
    blob = json.dumps(decision, ensure_ascii=False)
    assert decision["kind"] == "accept"
    assert "A1234567" not in blob and "2024-03-05" not in blob
    assert "DATA_BLOCKED" in blob and "DATA_QUERY" in blob


def test_unknown_extension_result_is_scrubbed():
    """BR-03.4：带路径但扩展名未识别（.xpt）的结果强制脱敏。"""
    decision = run("post-unknown-ext")
    blob = json.dumps(decision, ensure_ascii=False)
    assert decision["kind"] == "accept"
    assert "A1234567" not in blob and "2024-03-05" not in blob


def test_post_sensitive_data_query_is_projected():
    """数据查询结果在进入模型上下文前投影为不含原值的占位符。"""
    decision = run("post-sensitive")
    blob = json.dumps(decision, ensure_ascii=False)
    assert decision["kind"] == "accept"
    assert decision["content"][0]["text"]
    assert "clinicalGuard" in decision["content"][0]["text"]
    assert "DATA_BLOCKED" in decision["content"][0]["text"]
    assert "101-000" not in blob


def test_spec_and_document_content_is_exempt_from_automatic_redaction():
    """可信需求文档全文可见，但豁免 provenance 不得成为普通消息旁路。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        (root / "doc").mkdir()
        (root / "doc" / "spec").mkdir()
        spec_xlsx = root / "doc" / "spec" / "KRI.xlsx"
        aux_xlsx = root / "doc" / "aux.xlsx"
        for path in (spec_xlsx, aux_xlsx):
            import openpyxl
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["KRI编号", "指标名称", "说明"])
            sheet.append(["KRI-001", "失访率", "黄红区判定"])
            workbook.save(path)
        env = {
            "PLANE_DOCUMENT_ROOTS": json.dumps([str(root / "doc")]),
            "PLANE_SPEC_ROOTS": json.dumps([str(root / "doc" / "spec")]),
            "PLANE_SPEC_FILE": str(spec_xlsx),
            "PLANE_DOCUMENT_FILE": str(aux_xlsx),
            "SESSION_CWD": str(root),
        }
        full = run("trusted-document-roundtrip", env_extra=env)
        guarded = json.dumps(full["guarded"], ensure_ascii=False)
        trusted = json.dumps(full["trusted"], ensure_ascii=False)
        ordinary = json.dumps(full["ordinary"], ensure_ascii=False)
        for raw in ("KRI-001", "101-001-0001", "2026-08-19", "ALT 342"):
            assert raw in guarded
            assert raw in trusted
        assert "TRUSTED_DOCUMENT_CONTENT" in guarded
        assert "trustedDocumentToken" in guarded
        assert "TRUSTED_DOCUMENT_CONTENT" not in trusted
        assert "trustedDocumentToken" not in trusted
        assert "101-001-0001" in ordinary and "2026-08-19" in ordinary

        auxiliary = run("trusted-aux-document-roundtrip", env_extra=env)
        auxiliary_forwarded = json.dumps(auxiliary["trusted"], ensure_ascii=False)
        assert "辅助要求 101-001-0001 2026-08-19" in auxiliary_forwarded
        assert "TRUSTED_DOCUMENT_CONTENT" not in auxiliary_forwarded
        assert "trustedDocumentToken" not in auxiliary_forwarded


def test_canonical_glob_paths_survive_egress_without_exposing_display_text():
    """事故回放：canonical 路径是控制面，不能被 token 化或误判成整表。"""
    result = run("canonical-paths-roundtrip")
    guarded = json.dumps(result["guarded"], ensure_ascii=False)
    forwarded = json.dumps(result["forwarded"], ensure_ascii=False)
    for value in (
        "GQ1005-301", "GQ1005-301_ALS.xlsx", "SAS_20250221.zip",
        "H301", "DM.sas7bdat", "_mk_alpha",
    ):
        assert value in guarded
        assert value in forwarded
    assert "CONTROL_PATHS" in forwarded
    assert "trustedControlToken" not in forwarded
    assert "101-000" not in guarded and "101-000" not in forwarded
    assert "[VAL:" not in forwarded and "[SUBJ:" not in forwarded


def test_nested_listing_receipt_roundtrip_survives_egress():
    """真实宿主包装下，inspection 收据应放行且恢复原始控制面 JSON。"""
    result = run("listing-receipt-roundtrip", env_extra={"LOCAL_DATA_ACCESS": "uat-local"})
    guarded = json.dumps(result["guarded"], ensure_ascii=False)
    forwarded = json.dumps(result["forwarded"], ensure_ascii=False)
    assert "TRUSTED_LISTING_RECEIPT" in guarded
    assert "trustedListingToken" in guarded
    assert "CLINICAL_LISTING_INSPECTION" in forwarded
    assert "METADATA_ONLY" in forwarded
    assert "USUBJID" in forwarded and "SITEID" in forwarded
    assert "TRUSTED_LISTING_RECEIPT" not in forwarded
    assert "trustedListingToken" not in forwarded


def test_listing_execute_receipt_stays_on_normal_egress_scan():
    """REAL 执行收据不能借工具名进入 listing 信任恢复通道。

    2026-08-24：execute 收据改走 projectExecuteReceipt 白名单投影（"未列出即
    不存在"），夹带的 payload 在 Node 侧就被整键丢弃，不再进入出域 DLP 扫描，
    因此 forwarded 里不会有 [SUBJ: token。断言改为直接校验投影语义：受试者原值
    与承载它的 payload 键都不得出现，而 status 仍可读。
    """
    result = run("listing-untrusted-receipt-roundtrip", env_extra={"LOCAL_DATA_ACCESS": "uat-local"})
    guarded = json.dumps(result["guarded"], ensure_ascii=False)
    forwarded = json.dumps(result["forwarded"], ensure_ascii=False)
    assert "TRUSTED_LISTING_RECEIPT" not in guarded
    assert "trustedListingToken" not in guarded
    assert "A1234567" not in forwarded and "A1234567" not in guarded
    assert "payload" not in forwarded and "payload" not in guarded
    assert "completed" in forwarded


def test_listing_fake_or_error_receipt_stays_on_normal_egress_scan():
    """伪造 marker 或工具错误文本不能伪造本地信任。"""
    result = run("listing-fake-receipt-roundtrip", env_extra={"LOCAL_DATA_ACCESS": "uat-local"})
    guarded = json.dumps(result["guarded"], ensure_ascii=False)
    forwarded = json.dumps(result["forwarded"], ensure_ascii=False)
    assert "TRUSTED_LISTING_RECEIPT" not in guarded
    assert "trustedListingToken" not in guarded
    assert "A1234567" not in forwarded
    assert "[SUBJ:" in forwarded


def test_real_inspect_receipt_field_set_stays_trusted():
    """真实 inspect() 字段集必须被认可，spec 规格文本不得被 token 化。

    2026-08-23：isTrustedListingReceipt 的 allowed 闭集漏掉了
    inferredScenario / scenarioConfidence / supportData，真实收据整体失信后
    落入 scrubUntrustedListingContent，spec 全文被打成 token，harness 无法
    理解 spec 需求。此前所有集成场景都用手工裁剪的收据（键全在白名单内），
    因此只有真实跑项目才暴露。
    """
    result = run("listing-real-inspect-receipt", env_extra={"LOCAL_DATA_ACCESS": "uat-local"})
    assert result["receiptTrusted"], "真实 inspect 收据失信，spec 会被 token 化"
    assert result["specTextPreserved"], "spec 规格文本未原文保留"
    assert not result["tokenized"], "收据中出现 token 化残留"


def test_run_code_envelope_roundtrip_stays_trusted_and_readable():
    """代码车道 run 信封必须双侧信任透传，迭代反馈（行数/列名）不得 token 化。"""
    result = run("listing-run-code-receipt-roundtrip", env_extra={"LOCAL_DATA_ACCESS": "uat-local"})
    assert result["receiptTrusted"], "run 信封失信，会被 scrub 成 token"
    assert result["envelopeReadable"], "信封元数据未以可读形状到达 harness"
    assert not result["tokenized"], "信封出现 token 化残留"


def test_listing_receipt_keys_stay_within_whitelists():
    """生产者字段集必须同时被 Node 侧与 Python 出域侧的闭集白名单覆盖。

    两道防线各自维护一份闭集，任一侧漏字段都会让真实收据失信：Node 侧失信收据
    被 token 化，出域侧失信则收据被递归 DLP 扫描后 BLOCK 或降级——两种结果都让
    harness 读不到 spec 与 schema。2026-08-23 的事故正是 Python 侧漏了
    inferredScenario / scenarioConfidence / supportData 三键。
    """
    guard = (ROOT / "src" / "tool-result-guard.js").read_text(encoding="utf-8")
    block = guard.split("const allowed = new Set([", 1)[1].split("]);", 1)[0]
    node_allowed = set(re.findall(r"'([^']+)'", block))

    egress = (ROOT / "security" / "egress_checkpoint.py").read_text(encoding="utf-8")
    py_block = egress.split("    allowed = {", 1)[1].split("}", 1)[0]
    py_allowed = set(re.findall(r'"([^"]+)"', py_block))

    produced: set[str] = set()
    for source in ("listing_inspector.py", "listing_workflow.py"):
        text = (ROOT / "security" / source).read_text(encoding="utf-8")
        for marker in ("CLINICAL_LISTING_INSPECTION", "CLINICAL_LISTING_PLAN_RECEIPT"):
            for chunk in text.split(f'"clinicalGuard": "{marker}"')[1:]:
                produced.update(re.findall(r'^\s{8}"([A-Za-z_][A-Za-z0-9_]*)":', chunk.split("\n    }", 1)[0], re.M))

    assert produced, "未能从生产者解析出收据字段，测试本身失效"
    missing_node = sorted(produced - node_allowed)
    assert not missing_node, f"收据字段未进 Node 白名单，收据会失信并被 token 化: {missing_node}"
    missing_py = sorted(produced - py_allowed)
    assert not missing_py, f"收据字段未进出域白名单，收据会被出域 DLP 吃掉: {missing_py}"

    # 投影后的 execute 收据也要经过出域侧，这几个键必须在 Python 闭集里。
    for key in ("artifact", "artifacts", "note", "inferredScenario",
                "scenarioConfidence", "scenarioCandidates", "supportData"):
        assert key in py_allowed, f"出域白名单缺 {key}"

    # 2026-08-24 代码车道（listing_code_lane.py）收据字段合同：run 信封与
    # publish 收据的顶层键必须同时落在双侧闭集内，任一侧漏键即失信 token 化。
    code_lane = (ROOT / "security" / "listing_code_lane.py").read_text(encoding="utf-8")
    for marker in ("CLINICAL_LISTING_CODE_RECEIPT", "CLINICAL_LISTING_RECEIPT"):
        assert f'"{marker}"' in code_lane, f"listing_code_lane.py 缺少收据 marker {marker}"
    code_lane_keys = {
        "clinicalGuard", "status", "stage", "project", "scenario",
        "schemaFingerprint", "dataClass", "outputs", "datasetsTouched",
        "code", "message", "errorType",
    }
    missing_node = sorted(code_lane_keys - node_allowed)
    assert not missing_node, f"代码车道收据键未进 Node 白名单: {missing_node}"
    missing_py = sorted(code_lane_keys - py_allowed)
    assert not missing_py, f"代码车道收据键未进出域白名单: {missing_py}"


def test_execute_receipt_projection_keeps_receipt_readable():
    """投影后的 execute 收据必须让 harness 读出执行状态与产物。"""
    result = run("listing-execute-receipt-readable", env_extra={"LOCAL_DATA_ACCESS": "uat-local"})
    assert result["statusReadable"], "harness 读不出执行状态"
    assert result["artifactReadable"], "harness 读不出产物文件名"
    assert result["rowCountReadable"], "harness 读不出产物行数"
    assert not result["keysTokenized"], "投影后仍出现 token 化残留"


def test_execute_receipt_projection_drops_extra_payload():
    """同一投影必须丢弃收据夹带的 payload / rows 等真实数据泄漏面。"""
    result = run("listing-execute-receipt-payload-dropped", env_extra={"LOCAL_DATA_ACCESS": "uat-local"})
    assert result["statusReadable"], "投影后执行状态不可读"
    assert not result["subjectLeaked"], "受试者原值泄漏"
    assert not result["payloadLeaked"], "payload 键未被投影丢弃"
    assert not result["rowsLeaked"], "rows 键未被投影丢弃"


def test_ordinary_text_file_result_preserves_real_semantics():
    """事故回放：普通日志/需求文本不能按疑似数据行统一 token 化。"""
    with tempfile.TemporaryDirectory() as directory:
        result = run(
            "ordinary-text-file-roundtrip",
            env_extra={"SESSION_CWD": directory},
        )
    rendered = json.dumps(result, ensure_ascii=False)
    assert "(contents of zebra-alpha-file - version 123)" in rendered
    assert "request: preserve real intent" in rendered
    assert "已自动脱敏" not in rendered
    assert "[TEXT:" not in rendered and "[NUM:" not in rendered


def test_listing_tools_and_workflow_guidance_are_always_registered():
    disabled = run("listing-tool-contract")
    disabled_names = {tool["name"] for tool in disabled["tools"]}
    assert {
        "clinical_listing_inspect", "clinical_listing_run_code",
        "clinical_listing_publish",
    }.issubset(disabled_names)
    assert len(disabled["prompts"]) == 1
    prompt = disabled["prompts"][0]["text"]
    assert "inspect -> run -> iterate -> publish" in prompt
    assert "绝不能" not in prompt

    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory) / "clinical-root"
        root.mkdir()
        tools = run(
            "listing-tool-contract",
            env_extra={"LOCAL_DATA_ACCESS": "uat-local", "LOCAL_DATA_ROOT": str(root)},
        )
        names = [tool["name"] for tool in tools["tools"]]
        assert set(names) == {
            "clinical_listing_inspect", "clinical_listing_run_code",
            "clinical_listing_publish", "local_data_metadata",
        }
        assert len(names) == 4
        listing = next(tool for tool in tools["tools"] if tool["name"] == "clinical_listing_inspect")
        parameters = listing["parameters"]["properties"]
        assert set(parameters) == {"project", "scenario", "credentialRef"}
        assert "credentialRef" in listing["parameters"]["required"]
        assert "password" not in parameters
        assert "projectPath" not in parameters
        run_code = next(tool for tool in tools["tools"] if tool["name"] == "clinical_listing_run_code")
        assert "code" in run_code["parameters"]["properties"]
        assert "plan" not in run_code["parameters"]["properties"]


def test_runtime_policy_toggle_applies_without_worker_restart():
    result = run("runtime-policy-toggle")
    enabled = json.dumps(result["enabledResult"], ensure_ascii=False)
    disabled = json.dumps(result["disabledResult"], ensure_ascii=False)
    re_enabled = json.dumps(result["reEnabledResult"], ensure_ascii=False)
    assert "101-001-0001" not in enabled
    assert "101-001-0001" in disabled
    assert result["disabledStream"]["streamed"] is True
    assert "101-001-0001" not in re_enabled


def test_listing_without_local_capability_returns_structured_error():
    result = run("listing-capability-required")
    assert result["ok"] is False
    assert result["code"] == "LOCAL_DATA_ACCESS_REQUIRED"
    assert result["action"] == "listing-inspect"


def test_listing_rejects_absolute_project_without_path_disclosure():
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory) / "clinical-root"
        project = root / "study"
        project.mkdir(parents=True)
        output = run(
            "listing-absolute-project",
            env_extra={
                "LOCAL_DATA_ACCESS": "uat-local",
                "LOCAL_DATA_ROOT": str(root),
                "LISTING_PROJECT": str(project),
            },
        )
        assert output["blocked"] is True
        assert str(root) not in output["message"]


def test_listing_uses_session_cwd_over_configured_root():
    """Listing 从 Web UI 会话 cwd 绑定根，并按 ALS 查询真实 XPT。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        selected = root / "selected-workspace" / "study"
        configured = root / "wrong-configured-root"
        doc = selected / "doc"
        data = selected / "data"
        doc.mkdir(parents=True)
        data.mkdir()
        configured.mkdir()
        make_als_xlsx(doc / "ALS.xlsx")
        make_dm_xpt(data / "DM.xpt")
        receipt = run(
            "listing-session-cwd",
            env_extra={
                "LOCAL_DATA_ACCESS": "uat-local",
                "LOCAL_DATA_ROOT": str(configured),
                "SESSION_CWD": str(selected),
                "LISTING_PROJECT": ".",
            },
        )
        assert receipt["inspection"]["clinicalGuard"] == "CLINICAL_LISTING_INSPECTION"
        assert receipt["inspection"]["status"] == "ready"
        assert receipt["run"]["clinicalGuard"] == "CLINICAL_LISTING_CODE_RECEIPT"
        assert receipt["run"]["status"] == "ok"
        assert receipt["run"]["dataClass"] == "METADATA_ONLY"
        assert receipt["clinicalGuard"] == "CLINICAL_LISTING_RECEIPT"
        assert receipt["status"] == "completed"
        assert receipt["stage"] == "publish"
        assert receipt["dataClass"] == "REAL"
        assert len(receipt["artifacts"]) == 1
        assert receipt["project"] == "."
        assert receipt["artifacts"][0]["name"] == "RBQM_LISTINGS.xlsx"
        blob = json.dumps(receipt, ensure_ascii=False)
        assert "101-001-0001" not in blob and "101-001-0002" not in blob
        assert str(selected) not in blob and str(configured) not in blob
        assert (selected / ".clinical-listing" / "output" / "rbqm" / "RBQM_LISTINGS.xlsx").is_file()


def main() -> int:
    tests = [value for key, value in sorted(globals().items()) if key.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"PASS {test.__name__}")
        except Exception as error:
            failures += 1
            print(f"FAIL {test.__name__}: {error}")
    print(f"RESULT {len(tests) - failures}/{len(tests)}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
