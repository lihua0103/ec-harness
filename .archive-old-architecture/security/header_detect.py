"""表头识别单一来源（header_detect）— 评分式算法 + 三车道适配器。

2026-08-21 架构决定（用户定调）：表头识别合并进本插件，作为独立模块而非
独立 dsh 插件——它的输出直接决定放行范围，跨插件传递等于把边界判定移出
信任域。原先三处实现（excel_header_extractor.py 评分式 /
历史上分散在多个模块的关键词式与首行式表头识别现统一收敛到本模块：

  score_row / detect_orientation / find_header_end_row  评分式核心
  dlp_scan_cell                                          表头输出白名单兜底
  detect_header_row                                      逐行分类适配器
  header_names                                           元数据车道列名提取
  process_xlsx/xls/csv                                   表头提取（辅助档）

口径保证：同一份文件在任何车道判出同样的表头。
"""

import io
import os
import re
import sys
from pathlib import Path
from typing import Any

from security.patterns import (
    DATE_PATTERNS,
    NUMERIC_SUBJECT_ID_RE,
    SUBJECT_ID_PATTERNS,
)

MAX_SCAN_ROWS_DEFAULT = 20

HEADER_DLP_PATTERNS = [*SUBJECT_ID_PATTERNS, *DATE_PATTERNS]

HEADER_SAFE_PATTERNS = [
    re.compile(r'\bDay\s*\d+\b', re.I),
    re.compile(r'\bWeek\s*\d+\b', re.I),
    re.compile(r'\bCycle\s*\d+\b', re.I),
    re.compile(r'\bVisit\s*\d+\b', re.I),
    re.compile(r'\bMonth\s*\d+\b', re.I),
    re.compile(r'\bHour\s*\d+\b', re.I),
    re.compile(r'\bBaseline\b', re.I),
    re.compile(r'\bScreening\b', re.I),
    re.compile(r'\bEOT\b', re.I),
    re.compile(r'\bEOS\b', re.I),
]


def _is_numeric(val: Any) -> bool:
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        try:
            float(val.replace(',', ''))
            return True
        except ValueError:
            return False
    return False


def _is_string_like(val: Any) -> bool:
    if val is None or val == '':
        return False
    if isinstance(val, str) and val.strip():
        return not _is_numeric(val)
    return False


def _score_row(row: list[Any], total_cols: int, merged_cols: set[int]) -> float:
    if not row or all(v is None or v == '' for v in row):
        return -10.0

    non_empty = [v for v in row if v is not None and v != '']
    if not non_empty:
        return -10.0

    score = 0.0

    str_count = sum(1 for v in non_empty if _is_string_like(v))
    num_count = sum(1 for v in non_empty if _is_numeric(v))

    if str_count == len(non_empty):
        score += 3.0
    elif num_count / len(non_empty) >= 0.5:
        # 真实故障修复（crViewer.xls 数据泄露）：数值占比 ==0.5 的行同样是数据行，
        # 边界必须扣分，否则无表头数据表的行全被判为"表头"整表泄露。
        score -= 4.0

    if len(non_empty) / max(total_cols, 1) >= 0.5:
        score += 2.0

    if merged_cols:
        score += 2.0

    row_text = ' '.join(str(v) for v in non_empty)
    for pattern, _ in HEADER_DLP_PATTERNS:
        match = pattern.search(row_text)
        if match:
            val = match.group()
            if not any(sp.search(val) for sp in HEADER_SAFE_PATTERNS):
                if re.search(r'\b\d{3,4}-\d{4,6}-\d{3,6}\b', val):
                    score -= 5.0
                else:
                    score -= 3.0

    # 辅助档收紧（2026-08-21）：表头列名极少是"数字为主"的短码
    # （_looks_like_data_value 口径：纯数值/数字占比>=0.4 的 <=12 字符串）。
    # 纯文本规格表的数据行（KRI-001/101/A12 编号形态）借此被挡在
    # 表头区域之外；Day 3 / Visit 1 等安全词形态不受影响。
    for v in non_empty:
        if isinstance(v, str) and _looks_like_data_value(v.strip()) \
                and not any(sp.search(v) for sp in HEADER_SAFE_PATTERNS):
            score -= 6.0
            break

    return score


def _detect_orientation(scan_rows: list[list[Any]]) -> str:
    if not scan_rows or not scan_rows[0]:
        return 'VERTICAL'

    total_rows = len(scan_rows)
    col0_str = sum(
        1 for r in scan_rows
        if r and _is_string_like(r[0])
    )
    col0_str_ratio = col0_str / total_rows

    body_cells = 0
    body_numeric = 0
    for r in scan_rows:
        for v in r[1:]:
            body_cells += 1
            if _is_numeric(v):
                body_numeric += 1

    numeric_ratio = body_numeric / max(body_cells, 1)

    col0_text = ' '.join(str(r[0]) for r in scan_rows if r and r[0] is not None)
    for pattern, _ in HEADER_DLP_PATTERNS:
        if pattern.search(col0_text):
            return 'VERTICAL'

    if col0_str_ratio > 0.8 and numeric_ratio > 0.5:
        return 'HORIZONTAL'
    return 'VERTICAL'


def _find_header_end_row(scan_rows: list[list[Any]], merged_by_row: dict[int, set[int]]) -> int:
    total_cols = max((len(r) for r in scan_rows), default=1)
    for i, row in enumerate(scan_rows):
        score = _score_row(row, total_cols, merged_by_row.get(i, set()))
        if score < 0:
            return i
    return len(scan_rows)


def select_header_row(scan_rows: list[list[Any]], max_rows: int = 20) -> int | None:
    """从可信文档的前若干行选择最可能的字段表头，返回零基行号。"""
    rows = [list(row) for row in scan_rows[:max_rows]]
    if not rows:
        return None
    total_cols = max((len(row) for row in rows), default=1)
    candidates: list[tuple[float, int]] = []
    for index, row in enumerate(rows):
        score = _score_row(row, total_cols, set())
        if score >= 0:
            non_empty = [str(value).strip() for value in row if value not in (None, '')]
            machine_names = sum(
                1 for value in non_empty
                if re.fullmatch(r"[A-Z][A-Z0-9_]{1,31}", value)
            )
            if non_empty:
                score += machine_names / len(non_empty)
            candidates.append((score, -index))
    if not candidates:
        return None
    _, negative_index = max(candidates)
    return -negative_index


def _looks_like_data_value(text: str) -> bool:
    """ST-P1-9: 表头输出白名单兜底——判断一个单元格值"像数据而非列名"。

    启发式打分可能把无表头数据表首行误判为表头，而其中不触发 DLP 模式的值
    （纯数值编号、测量值、含数字的短码）会被原值输出，形成泄露通道。列名极少
    是纯数字或"数字为主"的串，故这类值即便不命中 DLP 也一律 REDACTED。
    """
    # 纯数值（整数/小数/千分位）——列名几乎不会是纯数字，数据值常是。
    stripped = text.replace(',', '').replace(' ', '')
    try:
        float(stripped)
        return True
    except ValueError:
        pass
    # 数字占比过半的短码（如 A12、3-5mg）——数据编号形态，非列名词。
    digits = sum(c.isdigit() for c in text)
    if len(text) <= 12 and digits >= 1 and digits / max(len(text), 1) >= 0.4:
        return True
    return False


def _dlp_scan_cell(val: Any) -> tuple[Any, str | None]:
    if val is None:
        return val, None
    text = str(val).strip()
    if not text:
        return val, None
    if NUMERIC_SUBJECT_ID_RE.fullmatch(text):
        return '[REDACTED]', 'SUBJECT_ID'
    for pattern, label in HEADER_DLP_PATTERNS:
        m = pattern.search(text)
        if m:
            matched = m.group()
            if any(sp.search(matched) for sp in HEADER_SAFE_PATTERNS):
                continue
            return '[REDACTED]', label
    # ST-P1-9: DLP 未命中也要挡住"像数据值"的单元格，防止误判表头时原值泄露。
    if _looks_like_data_value(text) and not any(
        sp.search(text) for sp in HEADER_SAFE_PATTERNS
    ):
        return '[REDACTED]', 'DATA_VALUE'
    return val, None


def _extract_merged_info(ws) -> dict[int, set[int]]:
    merged_by_row: dict[int, set[int]] = {}
    try:
        for rng in ws.merged_cells.ranges:
            for row_idx in range(rng.min_row - 1, rng.max_row):
                merged_by_row.setdefault(row_idx, set())
                for col_idx in range(rng.min_col - 1, rng.max_col):
                    merged_by_row[row_idx].add(col_idx)
    except Exception:
        pass
    return merged_by_row


def process_xlsx(file_path: str, sheet_name: str | None, max_scan: int) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print('missing openpyxl: pip install openpyxl', file=sys.stderr)
        sys.exit(2)

    workbook_bytes = Path(file_path).read_bytes()
    wb_ro = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet_names = [sheet_name] if sheet_name else wb_ro.sheetnames

    # P3: 一次性加载merged cells信息（循环外），避免每sheet重复开销
    merged_by_file: dict[str, dict[int, set[int]]] = {}
    try:
        wb_merge = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=False, data_only=True)
        for sn in wb_merge.sheetnames:
            merged_by_file[sn] = _extract_merged_info(wb_merge[sn])
        wb_merge.close()
    except Exception:
        pass

    results = []

    for sname in sheet_names:
        if sname not in wb_ro.sheetnames:
            continue
        ws_ro = wb_ro[sname]
        total_rows = ws_ro.max_row or 0
        total_cols = ws_ro.max_column or 0

        scan_rows: list[list[Any]] = []
        # P2: 不依赖 total_rows（read_only模式常返回None/0），直接用max_scan上限
        for row in ws_ro.iter_rows(min_row=1, max_row=max_scan, values_only=True):
            scan_rows.append(list(row))
            if len(scan_rows) >= max_scan:
                break

        # P0修复: merged_by_row从预加载的字典取，不在循环内重新加载workbook
        merged_by_row = merged_by_file.get(sname, {})

        orientation = _detect_orientation(scan_rows)

        header_cells = []
        redacted = []
        warnings_list = []

        if orientation == 'HORIZONTAL':
            for ri, row in enumerate(scan_rows):
                if not row:
                    continue
                raw = row[0]
                cleaned, label = _dlp_scan_cell(raw)
                if label:
                    redacted.append({'row': ri, 'col': 0, 'type': label})
                    warnings_list.append(f'row {ri} col 0: DLP hit {label}')
                if cleaned is not None and str(cleaned).strip():
                    header_cells.append({'row': ri, 'col': 0, 'value': _safe_header_value(cleaned, 0)})
            if scan_rows:
                for ci, val in enumerate(scan_rows[0][1:], start=1):
                    cleaned, label = _dlp_scan_cell(val)
                    if label:
                        redacted.append({'row': 0, 'col': ci, 'type': label})
                    if cleaned is not None and str(cleaned).strip():
                        header_cells.append({'row': 0, 'col': ci, 'value': _safe_header_value(cleaned, ci)})
            header_rows = list(range(len(scan_rows)))
            data_start_row = 1
        else:
            header_end = _find_header_end_row(scan_rows, merged_by_row)
            header_rows = list(range(header_end))
            data_start_row = header_end

            for ri in range(header_end):
                row = scan_rows[ri] if ri < len(scan_rows) else []
                for ci, val in enumerate(row):
                    cleaned, label = _dlp_scan_cell(val)
                    if label:
                        redacted.append({'row': ri, 'col': ci, 'type': label})
                        warnings_list.append(f'row {ri} col {ci}: DLP hit {label}')
                    if cleaned is not None and str(cleaned).strip():
                        header_cells.append({'row': ri, 'col': ci, 'value': _safe_header_value(cleaned, ci)})

        results.append({
            'sheet': sname,
            'orientation': orientation,
            'header_rows': header_rows,
            'data_start_row': data_start_row,
            'total_rows': total_rows,
            'total_cols': total_cols,
            'header_cells': header_cells,
            'redacted_in_header': redacted,
            'warnings': warnings_list,
        })

    # P0修复: wb_ro.close() 移到循环外，不在每次迭代关闭
    wb_ro.close()
    return results


def process_csv(file_path: str, max_scan: int) -> list[dict]:
    import csv

    with open(file_path, newline='', encoding='utf-8-sig', errors='replace') as f:
        reader = csv.reader(f)
        scan_rows = []
        total_rows = 0
        total_cols = 0
        for row in reader:
            total_rows += 1
            col_count = len(row)
            if col_count > total_cols:
                total_cols = col_count
            if len(scan_rows) < max_scan:
                scan_rows.append(row)
    orientation = _detect_orientation(scan_rows)

    header_cells = []
    redacted = []
    warnings_list = []

    if orientation == 'HORIZONTAL':
        header_rows = list(range(len(scan_rows)))
        data_start_row = 1
        for ri, row in enumerate(scan_rows):
            if not row:
                continue
            cleaned, label = _dlp_scan_cell(row[0])
            if label:
                redacted.append({'row': ri, 'col': 0, 'type': label})
                warnings_list.append(f'row {ri} col 0: DLP hit {label}')
            if cleaned is not None and str(cleaned).strip():
                header_cells.append({'row': ri, 'col': 0, 'value': _safe_header_value(cleaned, 0)})
        for ci, val in enumerate(scan_rows[0][1:] if scan_rows else [], start=1):
            cleaned, label = _dlp_scan_cell(val)
            if label:
                redacted.append({'row': 0, 'col': ci, 'type': label})
            if cleaned is not None and str(cleaned).strip():
                header_cells.append({'row': 0, 'col': ci, 'value': _safe_header_value(cleaned, ci)})
    else:
        header_end = _find_header_end_row(scan_rows, {})
        header_rows = list(range(min(header_end, len(scan_rows))))
        data_start_row = header_end
        for ri in range(min(header_end, len(scan_rows))):
            for ci, val in enumerate(scan_rows[ri]):
                cleaned, label = _dlp_scan_cell(val)
                if label:
                    redacted.append({'row': ri, 'col': ci, 'type': label})
                    warnings_list.append(f'row {ri} col {ci}: DLP hit {label}')
                if cleaned is not None and str(cleaned).strip():
                    header_cells.append({'row': ri, 'col': ci, 'value': _safe_header_value(cleaned, ci)})

    return [{
        'sheet': os.path.basename(file_path),
        'orientation': orientation,
        'header_rows': header_rows,
        'data_start_row': data_start_row,
        'total_rows': total_rows,
        'total_cols': total_cols,
        'header_cells': header_cells,
        'redacted_in_header': redacted,
        'warnings': warnings_list,
    }]


def process_xls(file_path: str, sheet_name: str | None, max_scan: int) -> list[dict]:
    """FIX-8 (FR-06-03 / TC-15): .xls 经 xlrd 只读解析，交付与 .xlsx 相同的表头结构。

    xlrd 缺失时 fail-closed（退出码 2），不静默回退到数据读取。
    """
    try:
        import xlrd
    except ImportError:
        print('missing xlrd: pip install xlrd', file=sys.stderr)
        sys.exit(2)

    workbook = xlrd.open_workbook(file_path, on_demand=True, formatting_info=False)
    try:
        names = [sheet_name] if sheet_name else workbook.sheet_names()
        results = []
        for sname in names:
            if sname not in workbook.sheet_names():
                continue
            sheet = workbook.sheet_by_name(sname)
            total_rows = sheet.nrows
            total_cols = sheet.ncols
            scan_rows = [
                [sheet.cell_value(r, c) if c < total_cols else None
                 for c in range(total_cols)]
                for r in range(min(total_rows, max_scan))
            ]

            orientation = _detect_orientation(scan_rows)
            header_cells = []
            redacted = []
            warnings_list = []

            if orientation == 'HORIZONTAL':
                for ri, row in enumerate(scan_rows):
                    if not row:
                        continue
                    raw = row[0]
                    cleaned, label = _dlp_scan_cell(raw)
                    if label:
                        redacted.append({'row': ri, 'col': 0, 'type': label})
                        warnings_list.append(f'row {ri} col 0: DLP hit {label}')
                    if cleaned is not None and str(cleaned).strip():
                        header_cells.append({'row': ri, 'col': 0, 'value': _safe_header_value(cleaned, 0)})
                if scan_rows:
                    for ci, val in enumerate(scan_rows[0][1:], start=1):
                        cleaned, label = _dlp_scan_cell(val)
                        if label:
                            redacted.append({'row': 0, 'col': ci, 'type': label})
                        if cleaned is not None and str(cleaned).strip():
                            header_cells.append({'row': 0, 'col': ci, 'value': _safe_header_value(cleaned, ci)})
                header_rows = list(range(len(scan_rows)))
                data_start_row = 1
            else:
                header_end = _find_header_end_row(scan_rows, {})
                header_rows = list(range(header_end))
                data_start_row = header_end
                for ri in range(header_end):
                    row = scan_rows[ri] if ri < len(scan_rows) else []
                    for ci, val in enumerate(row):
                        cleaned, label = _dlp_scan_cell(val)
                        if label:
                            redacted.append({'row': ri, 'col': ci, 'type': label})
                            warnings_list.append(f'row {ri} col {ci}: DLP hit {label}')
                        if cleaned is not None and str(cleaned).strip():
                            header_cells.append({'row': ri, 'col': ci, 'value': _safe_header_value(cleaned, ci)})

            results.append({
                'sheet': sname,
                'orientation': orientation,
                'header_rows': header_rows,
                'data_start_row': data_start_row,
                'total_rows': total_rows,
                'total_cols': total_cols,
                'header_cells': header_cells,
                'redacted_in_header': redacted,
                'warnings': warnings_list,
            })
        return results
    finally:
        workbook.release_resources()


# ---------------------------------------------------------------------------
# 临床表头关键词（标注用词汇表，不参与 is_header 判定）
# ---------------------------------------------------------------------------
CLINICAL_HEADER_KEYWORDS = {
    "subject", "subj", "usubjid", "subjid", "patient", "ptid",
    "participant", "screening", "screen",
    "visit", "visitnum", "visitdy", "day", "date", "dt", "dtc",
    "time", "timestamp", "rfstdtc", "rfendtc",
    "status", "state", "aestdat", "aeendat", "cm", "ae", "lb",
    "value", "result", "lborres", "lbstresc", "measure",
    "site", "siteid", "center", "country",
    "term", "category", "severity", "grade", "arm", "treatment",
    "dose", "unit", "parameter", "test", "code", "label", "description",
}

CLINICAL_HEADER_PHRASES = {
    "adverse event", "subject id", "patient id", "visit name", "visit number",
    "study day", "start date", "end date", "result value", "reference range",
}

CLINICAL_HEADER_CODES = {
    "STUDYID", "DOMAIN", "USUBJID", "SUBJID", "SITEID", "INVID",
    "AGE", "AGEU", "SEX", "RACE", "ETHNIC", "COUNTRY",
    "ARMCD", "ARM", "ACTARMCD", "ACTARM", "RFSTDTC", "RFENDTC",
    "VISIT", "VISITNUM", "VISITDY", "EPOCH", "DTC", "DY", "SEQ",
    "TERM", "DECOD", "BODSYS", "CAT", "SCAT", "TESTCD", "TEST",
    "ORRES", "ORRESU", "STRESC", "STRESN", "STRESU", "STAT", "REASND",
    "BRTHDTC", "DMDTC", "DMDY", "AESTDTC", "AEENDTC", "AESTDY", "AEENDY",
    "AEDECOD", "AETERM", "AESEV", "AESER", "LBTESTCD", "LBTEST",
    "LBORRES", "LBORRESU", "LBSTRESC", "LBSTRESN", "LBSTRESU",
    "PRETEXT", "ITEMORDER", "DATASETNAME", "SASLABEL", "FORMOID", "ITEMOID",
}

# 常见 EDC 导出系统字段统一映射。这里识别字段角色，不读取记录值；调用方可用
# canonical role 跨 Rave、InForm、Veeva 与 CDISC 导出匹配同一业务字段。
EDC_FIELD_ROLES = {
    "study": {"STUDYID", "STUDYNAME", "PROTOCOLNAME", "PROTOCOL"},
    "site": {"SITEID", "SITENUMBER", "SITENAME", "CENTER", "CENTERID"},
    "subject": {
        "USUBJID", "SUBJID", "SUBJECTID", "SUBJECTNAME", "PATIENTID",
        "PATIENTNUMBER", "SCREENINGNUMBER",
    },
    "visit": {
        "VISIT", "VISITNAME", "VISITNUM", "FOLDERNAME", "FOLDEROID",
        "EVENTNAME", "EVENTOID",
    },
    "form": {"FORM", "FORMNAME", "FORMOID", "PAGENAME", "CRFNAME"},
    "field": {
        "FIELDNAME", "ITEMNAME", "ITEMOID", "VARIABLENAME", "SASNAME",
        "COLUMNNAME", "PRETEXT", "ITEMORDER", "DATASETNAME", "SASLABEL",
    },
    "repeat": {
        "RECORDPOSITION", "REPEATKEY", "FORMREPEATKEY", "ITEMGROUPREPEATKEY",
        "INSTANCEID",
    },
    "status": {
        "STATUS", "RECORDSTATUS", "FORMSTATUS", "SUBJECTSTATUS", "LOCKSTATUS",
        "SDVSTATUS", "FREEZESTATUS",
    },
    "date": {
        "CREATEDDATE", "MODIFIEDDATE", "LASTUPDATED", "ENTRYDATE", "DATAPAGENAME",
    },
}


def canonical_edc_field(value: Any) -> str | None:
    """返回跨 EDC 系统的统一字段角色；未知字段返回 None。"""
    if not isinstance(value, str):
        return None
    normalized = re.sub(r"[^A-Z0-9]", "", value.upper())
    for role, aliases in EDC_FIELD_ROLES.items():
        if normalized in aliases:
            return role
    return None

# SDTM 变量命名规范：绝大多数字段名是「域前缀 + 标准后缀」的合成词。枚举合成
# 规则而不是枚举成品，才能覆盖 EXDOSE、VSORRES、QSORRES 这类词表永远列不全的
# 长尾。这是结构规范识别，不读取任何记录值。
SDTM_DOMAIN_PREFIXES = {
    "DM", "AE", "CM", "EX", "LB", "VS", "QS", "SC", "SE", "SV", "DS", "DV",
    "MH", "PE", "PR", "RS", "SU", "TU", "TR", "EG", "IE", "FA", "MB", "MS",
    "OE", "CE", "BS", "CP", "GF", "ML", "NV", "OI", "PC", "PP", "RE", "RP",
    "SR", "SS", "TA", "TE", "TI", "TS", "TV",
}

SDTM_VARIABLE_SUFFIXES = {
    "SEQ", "STDTC", "ENDTC", "DTC", "DY", "STDY", "ENDY", "TERM", "DECOD",
    "MODIFY", "BODSYS", "SEV", "SER", "ACN", "REL", "OUT", "TESTCD", "TEST",
    "CAT", "SCAT", "ORRES", "ORRESU", "ORNRLO", "ORNRHI", "STRESC", "STRESN",
    "STRESU", "STNRLO", "STNRHI", "NRIND", "STAT", "REASND", "BLFL", "DRVFL",
    "TOXGR", "SPEC", "METHOD", "POS", "LOC", "LAT", "DIR", "DOSE", "DOSU",
    "DOSFRM", "DOSFRQ", "ROUTE", "TRT", "ADJ", "INDC", "CLAS", "PRESP",
    "OCCUR", "EVAL", "REFID", "SPID", "GRPID", "LNKID", "LNKGRP", "STRF",
    "ENRF", "STRTPT", "ENRTPT", "STTPT", "ENTPT", "DUR", "TPT", "TPTNUM",
    "ELTM", "TPTREF", "RFTDTC", "EPOCH", "PRESPEC",
}

# SUPPQUAL 与关联型数据集的固定变量名，不遵循域前缀规则。
SUPPQUAL_VARIABLES = {
    "RDOMAIN", "IDVAR", "IDVARVAL", "QNAM", "QLABEL", "QVAL", "QORIG",
    "QEVAL", "RELID", "USUBJID", "STUDYID",
}

# 常见实验室分析物代码。这些是检查项目名（结构元数据），不是检查结果值。
# 只收 3 字符以上的代码：NA、K、CL 这类超短代码与数据行取值难以区分。
LAB_TEST_CODES = {
    "ALT", "AST", "ALP", "ALB", "TBIL", "DBIL", "IBIL", "GGT", "LDH",
    "CREAT", "BUN", "UREA", "GLUC", "CHOL", "TRIG", "HDL", "LDL", "URATE",
    "SODIUM", "POTASSIUM", "CHLORIDE", "CALCIUM", "MAGNESIUM", "PHOS",
    "WBC", "RBC", "HGB", "HCT", "PLAT", "NEUT", "LYM", "MONO", "EOS",
    "BASO", "RETI", "INR", "APTT", "FIBRINOGEN", "TSH", "HBA1C", "CRP",
    "EGFR", "AMYLASE", "LIPASE", "PROT", "BILI", "TEMP", "PULSE",
    "SYSBP", "DIABP", "RESP", "HEIGHT", "WEIGHT", "BMI",
}


def _matches_sdtm_naming(text: str) -> bool:
    """按 SDTM/SUPPQUAL 命名规范证明这是字段名，而非记录值。

    枚举合成规则（域前缀 + 标准后缀）而非成品清单，覆盖词表列不全的长尾。
    人名、地名、分类取值不满足任何合成规则，因此 JOHN、SMITH 不会被证明。
    """
    upper = re.sub(r"[^A-Z0-9]", "", text.upper())
    if not upper or len(upper) > 8 or not upper.isalpha():
        return False
    if upper in LAB_TEST_CODES or upper in SUPPQUAL_VARIABLES:
        return True
    # SUPP<域> 是补充限定符数据集名，如 SUPPDM、SUPPAE。
    if upper.startswith("SUPP") and upper[4:] in SDTM_DOMAIN_PREFIXES:
        return True
    prefix, suffix = upper[:2], upper[2:]
    return prefix in SDTM_DOMAIN_PREFIXES and suffix in SDTM_VARIABLE_SUFFIXES


CHINESE_HEADER_TERMS = {
    "受试者", "患者", "研究", "中心", "访视", "日期", "时间", "状态", "结果",
    "数值", "单位", "参数", "检查", "事件", "术语", "严重程度", "分级", "治疗",
    "剂量", "组别", "编号", "代码", "名称", "描述", "类别", "指标", "项目",
    "需求", "阈值", "变量", "标签", "国家", "部位", "基线",
}


def detect_header_row(row_cells, total_cols=None, merged_cols=frozenset()):
    """逐行表头分类（scan_xlsx_sheet_safe 车道适配器）。

    判据（评分式，取代旧关键词命中式）：
      score_row > 0 且行内无 DLP 命中 → 表头。
    安全方向：误判为表头的行在调用方被整体跳过（不出现在输出里），
    不会泄露单元格值；含 DLP 形态或数值占比高的行永远走数据行脱敏。

    Returns:
        (is_header, confidence, matched_keywords)
        keywords 仅作标注展示，不含单元格原值。
    """
    cells = [c for c in (row_cells or []) if c is not None and str(c).strip()]
    if len(cells) < 2:
        return False, 0.0, []
    score = _score_row(list(row_cells or []), total_cols or len(row_cells or []), set(merged_cols))
    row_text = ' '.join(str(c) for c in cells)
    dlp_hit = False
    for p, _ in HEADER_DLP_PATTERNS:
        m = p.search(row_text)
        if m and not any(sp.search(m.group()) for sp in HEADER_SAFE_PATTERNS):
            dlp_hit = True
            break
    is_header = score > 0 and not dlp_hit
    confidence = min(0.95, max(score, 0.0) / 5.0)
    lowered = ' '.join(str(c).lower() for c in cells)
    keywords = sorted(kw for kw in CLINICAL_HEADER_KEYWORDS if kw in lowered)
    return is_header, confidence, keywords


def _is_proven_header_name(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text or len(text) > 256:
        return False
    # 冒号和等号通常表示“字段: 值”或“字段=值”，不能仅因其中含 site、date
    # 等临床词就把整段自由文本证明为字段名。
    if ":" in text or "=" in text:
        return False
    cleaned, label = _dlp_scan_cell(text)
    if label or cleaned != text:
        return False
    normalized = re.sub(r"[_/()-]+", " ", text).strip().lower()
    words = set(re.findall(r"[a-z]+", normalized))
    if normalized in CLINICAL_HEADER_PHRASES or words & CLINICAL_HEADER_KEYWORDS:
        return True
    if text.upper() in CLINICAL_HEADER_CODES:
        return True
    if canonical_edc_field(text):
        return True
    # 词表永远列不全字段名，改用命名规范作证据：EXDOSE、VSORRES、QNAM、SUPPDM
    # 这类合法字段名不在任何词表里，但满足 SDTM 合成规则。
    if _matches_sdtm_naming(text):
        return True
    return any(term in text for term in CHINESE_HEADER_TERMS)


def _looks_like_data_value_or_dlp(text: str) -> bool:
    """元数据车道的列名降级判据（黑名单）：这一格更像记录值而不是列名。

    与 _is_proven_header_name 的白名单证明制相反：这里只拦真实风险形态，
    合法但不在词表内的字段名（ALT、QNAM、SUPPDM 等）照常原样输出。
    """
    # 冒号/等号是"字段: 值"或"字段=值"的自由文本形态，右半边就是真实取值
    # （真实泄露案例 crViewer.xls 的 "Site: UAT_006 (Site Number:23)"）。
    # 这是形态判据而非词表判据：列名本身从不含冒号或等号，故一律降级。
    if ":" in text or "=" in text:
        return True
    _, label = _dlp_scan_cell(text)
    return bool(label)


def _safe_header_value(value: Any, column_index: int) -> str:
    """投影 process_xlsx/xls/csv 车道的表头单元格值。

    2026-08-24：这条车道逐格调用，没有整行上下文，无法做 header_names 那样的
    行级证明继承，因此保留单格证明制——扫的是未知表格的前若干行，"是否表头"
    尚未确定，放行未经证明的短文本会直接漏出数据行取值（JOHN、SMITH）。

    证明制本身已随 _matches_sdtm_naming 增强：ALT、EXDOSE、QNAM、SUPPDM 这类
    此前失守的合法字段名现在能被命名规范证明，不再退化成 COLUMN_n。
    """
    text = str(value).strip() if value is not None else ""
    if not text:
        return f"COLUMN_{column_index + 1}"
    if _is_proven_header_name(text):
        return text[:256]
    return f"COLUMN_{column_index + 1}"


def header_names(values, *, with_verdict=False):
    """投影元数据车道的列名：只降级"像数据值"的单元格，其余原样保留。

    2026-08-24 修复：这里原先用 _is_proven_header_name 作为白名单证明制——列名
    必须命中临床词表才输出原值，否则一律替换成 COLUMN_n。实测 32 个标准 SDTM
    列名丢 13 个（ALT、AST、TBIL、EXDOSE、EXTRT、CMTRT、VSORRES、VSTESTCD、
    QSORRES、SUPPDM、IDVAR、QNAM、QVAL），harness 拿到这种 schema 无法按 ALS
    字段定位数据列，inspect 与 execute 的列名口径也会分叉。

    列名是结构元数据，不是记录值，不能套用面向数据值的保护启发式——这与
    local_data_inspector._sas_metadata / _xpt_metadata 对 SAS/XPT descriptor
    的豁免理由完全相同，xlsx/csv 此前漏掉了同一豁免。

    修法不是废弃证明制——ALT 与 JOHN 在词法上不可区分（都是纯大写短字母串、
    不命中 DLP），纯黑名单制无法把人名数据行和 SDTM 列名分开。改为两步加强：

    1. 把"证明"从零散词表升级为命名规范识别（_matches_sdtm_naming）：枚举
       「域前缀 + 标准后缀」的合成规则，覆盖词表列不全的长尾。
    2. 行级证明继承：过半非空格被证明为字段名且无一格呈数据形态时，整行判定
       为表头，剩余未被单独证明的格（自定义列名、缩写）继承该证明。

    数据行不满足行级条件——记录值不会成片命中字段名证据，因此 ST-P1-9 的原始
    安全目标（无表头数据表首行不泄露真实值）仍然成立。

    verdict 沿用同一个 recognized 计数：它决定"该行能否被证明为表头"，进而决定
    调用方 rowCount 是否减 1，与投影共享判据可保证两者口径一致。
    """
    source = list(values or [])
    texts = [str(value).strip() if value is not None else "" for value in source]
    non_empty = sum(1 for text in texts if text)
    proven = [_is_proven_header_name(value) for value in source]
    recognized = sum(1 for flag in proven if flag)
    risky = [bool(text) and _looks_like_data_value_or_dlp(text) for text in texts]

    # 行级证明：过半非空格被证明为字段名，且无一格呈数据形态，这一行整体就是
    # 表头。此时未被单独证明的格（自定义列名、缩写、派生变量）继承行级证明。
    # 数据行不满足这个条件——记录值不会成片命中字段名证据。
    required = max(2, (non_empty + 1) // 2)
    row_is_data = any(risky)
    row_proven = non_empty > 0 and recognized >= required and not row_is_data

    projected = []
    for index, text in enumerate(texts, start=1):
        # 行级否决优先于单格证明：只要行内有格呈数据形态（受试者 ID、日期），
        # 整行就是数据行，此时"受试者"这类既像列名又是常见取值的格不能放行。
        allow = text and not row_is_data and (proven[index - 1] or row_proven)
        projected.append(text[:256] if allow else f"COLUMN_{index}")

    row_is_header, _, _ = detect_header_row(source)
    verdict = row_is_header and recognized >= required
    if with_verdict:
        return projected, verdict
    return projected
