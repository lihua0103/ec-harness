"""经 ListingPlan validator 放行后的本地数据执行器。"""
from __future__ import annotations

import math
import os
import re
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import pyreadstat
from security.listing_data_catalog import DatasetCatalog
from security.header_detect import header_names
from security.listing_plan import REVIEW_COLUMNS, _is_code_value_column
from security.project_profile import ProjectProfile, load_project_profile


class ListingExecutionError(RuntimeError):
    pass


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
_CONTENTS_HEADERS = [
    "Listing Seq.", "Listing Name(Please Click Down)", "Data Set Label",
    "Report Description", "New/Modified ?", "Total Row Count", "New Count",
    "Modified Count",
]


def _sheet_name(name: str, used: set[str], contents_sheet_name: str = "Contents") -> str:
    base = _INVALID_SHEET_CHARS.sub("_", str(name).strip()) or "Listing"
    base = base[:31]
    candidate = base
    index = 2
    while candidate.casefold() in used or candidate.casefold() == str(contents_sheet_name).casefold():
        suffix = f"_{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate.casefold())
    return candidate


def _protect_cell(cell: openpyxl.cell.cell.Cell) -> None:
    if isinstance(cell.value, str) and cell.value.lstrip()[:1] in {"=", "+", "-", "@"}:
        cell.data_type = "s"


def _style_header(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, count: int) -> None:
    style = openpyxl.styles.NamedStyle(name=f"listing_header_{id(sheet)}")
    style.font = openpyxl.styles.Font(bold=True)
    style.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    style.fill = openpyxl.styles.PatternFill("solid", fgColor="D9E1F2")
    for cell in sheet[row][:count]:
        cell.font = style.font
        cell.alignment = style.alignment
        cell.fill = style.fill


def _finish_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, header_row: int, column_count: int, freeze_rows: int, freeze_columns: int) -> None:
    if column_count <= 0:
        raise ListingExecutionError("output header shape is invalid")
    last_column = openpyxl.utils.get_column_letter(column_count)
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{sheet.max_row}"
    freeze_row = max(header_row + 1, header_row + int(freeze_rows))
    freeze_column = max(1, int(freeze_columns) + 1)
    sheet.freeze_panes = f"{openpyxl.utils.get_column_letter(freeze_column)}{freeze_row}"
    for column in range(1, column_count + 1):
        values = [sheet.cell(row, column).value for row in range(1, min(sheet.max_row, 200) + 1)]
        width = max((len(str(value or "")) for value in values), default=0) + 3
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = max(10, min(60, width))
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row, min_col=1, max_col=column_count):
        for cell in row:
            _protect_cell(cell)


def _atomic_save(workbook: openpyxl.Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".tmp.{os.getpid()}.{path.name}")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _read(path: Path) -> pd.DataFrame:
    try:
        source_columns = None
        if path.suffix.lower() == ".xpt":
            frame, metadata = pyreadstat.read_xport(str(path))
            source_columns = list(metadata.column_names or frame.columns)
        elif path.suffix.lower() == ".sas7bdat":
            frame, metadata = pyreadstat.read_sas7bdat(str(path))
            source_columns = list(metadata.column_names or frame.columns)
        elif path.suffix.lower() == ".csv":
            frame = pd.read_csv(path)
        else:
            raise ListingExecutionError("unsupported local data source")
        if path.suffix.lower() in {".sas7bdat", ".xpt"}:
            # SAS/XPT descriptor 的列名是结构元数据，不是未知表格的首行数据。
            # 必须与 local_data_inspector 使用同一 canonical 名称，否则 inspect
            # 和 validate 虽然通过，execute 会在取列时得到 COLUMN_n 并失败。
            projected = [str(column).strip()[:256] for column in (source_columns or frame.columns)]
        else:
            # csv/xlsx 必须与 local_data_inspector._header_projection 调同一个
            # header_names，两侧口径分叉会让 validate 通过而 execute 取列失败。
            # header_names 已改为黑名单降级，合法 SDTM 字段名原样保留。
            projected = header_names([str(column) for column in (source_columns or frame.columns)])
        used: set[str] = set()
        names: list[str] = []
        for index, name in enumerate(projected, start=1):
            candidate = name
            if candidate.casefold() in used:
                candidate = f"{name}_{index}"
            used.add(candidate.casefold())
            names.append(candidate)
        frame.columns = names
        return frame
    except ListingExecutionError:
        raise
    except Exception as exc:
        raise ListingExecutionError("a local dataset could not be read") from exc


def _column(frame: pd.DataFrame, ref: str) -> str:
    parts = ref.split(".")
    name = parts[-1]
    matches = {str(column).casefold(): str(column) for column in frame.columns}
    if len(parts) == 2:
        qualified = f"{parts[0]}__{name}"
        mapped = frame.attrs.get("qualified_columns", {}).get(qualified.casefold())
        if mapped in frame.columns:
            return mapped
        if qualified in frame.columns:
            return qualified
        if qualified.casefold() in matches:
            return matches[qualified.casefold()]
    if name.casefold() not in matches:
        raise ListingExecutionError("validated field is unavailable")
    return matches[name.casefold()]


def _literal(value: dict[str, Any]) -> Any:
    return value.get("value")


def _apply_filters(frame: pd.DataFrame, filters: list[dict[str, Any]]) -> pd.DataFrame:
    result = frame
    for item in filters:
        column = _column(result, item["column"])
        operator = item["operator"]
        series = result[column]
        if operator == "is_null":
            mask = series.isna()
        elif operator == "not_null":
            mask = series.notna()
        else:
            # F-2: valueRef 是字段引用，必须比较该列的**值**。此前只解析出列名
            # 字符串就参与比较，使 eq 静默返回空表、ne 静默返回全表——validator
            # 放行的合法计划产出错误临床交付物且无任何报错信号。
            right = result[_column(result, item["valueRef"])] if item.get("valueRef") else _literal(item["literal"])
            if operator == "eq": mask = series == right
            elif operator == "ne": mask = series != right
            elif operator == "gt": mask = series > right
            elif operator == "gte": mask = series >= right
            elif operator == "lt": mask = series < right
            else: mask = series <= right
        result = result.loc[mask.fillna(False) if hasattr(mask, "fillna") else mask]
    return result


def _derive(frame: pd.DataFrame, items: list[dict[str, Any]]) -> pd.DataFrame:
    result = frame.copy()
    for item in items:
        refs = [_column(result, ref) for ref in item["refs"]]
        op = item["operation"]
        if op == "copy": result[item["name"]] = result[refs[0]]
        elif op == "concat": result[item["name"]] = result[refs].fillna("").astype(str).agg(item.get("separator", "").join, axis=1)
        elif op == "coalesce": result[item["name"]] = result[refs].bfill(axis=1).iloc[:, 0]
        elif op == "date_diff_days": result[item["name"]] = (pd.to_datetime(result[refs[0]]) - pd.to_datetime(result[refs[1]])).dt.days
        else:
            value = result[refs[0]]
            for other in refs[1:]:
                if op == "add": value = value + result[other]
                elif op == "subtract": value = value - result[other]
                elif op == "multiply": value = value * result[other]
                elif op == "divide": value = value / result[other].replace(0, math.nan)
            result[item["name"]] = value
    return result


NUMERIC_AGGREGATIONS = {"sum", "mean"}


def _require_numeric(series: pd.Series, operation: str) -> None:
    """sum/mean 只对数值字段有意义。

    pandas 对字符串 series 的 sum 是**拼接**（'a'+'b' → 'ab'），mean 则抛
    TypeError。拼接结果会静默变成临床交付物里的一个"合计"单元格——这是比报错
    严重得多的正确性缺陷，必须显式拒绝而不是交给 pandas 的隐式语义。
    """
    if operation in NUMERIC_AGGREGATIONS and not pd.api.types.is_numeric_dtype(series):
        raise ListingExecutionError("a numeric aggregation requires a numeric field")


def _execute_output(output: dict[str, Any], files: dict[str, list[Path]], profile: ProjectProfile) -> pd.DataFrame:
    matches = files.get(output["source"].casefold(), [])
    if len(matches) != 1:
        raise ListingExecutionError("a required dataset is missing or ambiguous")
    frame = _read(matches[0])
    for join in output["joins"]:
        right_matches = files.get(join["dataset"].casefold(), [])
        if len(right_matches) != 1:
            raise ListingExecutionError("a joined dataset is missing or ambiguous")
        right = _read(right_matches[0])
        left_on = [_column(frame, key) for key in join["leftKeys"]]
        right_on = [_column(right, key) for key in join["rightKeys"]]
        dataset_prefix = join["dataset"]
        # 右表列始终保留唯一内部名，并登记限定引用映射。不能把内部列再改回
        # DATASET__COLUMN，否则左表原生同名列会让 pandas 返回 DataFrame 而非 Series。
        right_renamed = {
            column: f"__JOIN_{len(frame.columns)}_{dataset_prefix}__{column}"
            for column in right.columns
            if column not in right_on
        }
        right = right.rename(columns=right_renamed)
        qualified_columns = dict(frame.attrs.get("qualified_columns", {}))
        frame = frame.merge(right, how=join["type"], left_on=left_on,
                            right_on=right_on, suffixes=("", "_JOIN"))
        qualified_columns.update({
            f"{dataset_prefix}__{column}".casefold(): internal
            for column, internal in right_renamed.items()
        })
        frame.attrs["qualified_columns"] = qualified_columns
    # F-5: 顺序必须与 validator 一致（listing_plan 先注册 derivations 再校验
    # filters，因此合法计划允许 filter 引用派生列）。此前执行期先 filter 后
    # derive，会让这类已 validated 的计划必然抛 "validated field is unavailable"，
    # 破坏"validated 即可执行"契约。
    frame = _derive(frame, output["derivations"])
    frame = _apply_filters(frame, output["filters"])
    layout = output.get("layout", {})
    status_filter = layout.get("statusFilter", "")
    if status_filter:
        status_column = next((column for column in frame.columns if str(column).casefold() == profile.status_column_name.casefold()), None)
        if status_column is None:
            raise ListingExecutionError("status filter requires a Status field")
        frame = frame.loc[frame[status_column].astype(str).eq(status_filter)]
    if output["aggregations"]:
        group = [_column(frame, key) for key in output["groupBy"]]
        grouped = frame.groupby(group, dropna=False, sort=False) if group else None
        values: dict[str, Any] = {}
        for item in output["aggregations"]:
            source = _column(frame, item["column"])
            _require_numeric(frame[source], item["operation"])
            if grouped is None:
                series = frame[source]
                # N-9: count 语义两条分支必须一致。pandas 的 grouped.count() 不计
                # NaN，而 len(series) 计入，使"有无 groupBy"改变同一计划的 count
                # 定义。统一为 pandas 口径（非空计数）。
                # 运算必须按需求值：此前构造包含全部六种聚合的 dict 再取一个，
                # 使"只求 count"的计划也会在文本列上执行 sum/mean 并抛 TypeError。
                operations = {
                    "count": lambda: int(series.count()),
                    "count_distinct": lambda: series.nunique(),
                    "sum": lambda: series.sum(),
                    "mean": lambda: series.mean(),
                    "min": lambda: series.min(),
                    "max": lambda: series.max(),
                }
                try:
                    values[item["name"]] = operations[item["operation"]]()
                except (TypeError, ValueError) as exc:
                    raise ListingExecutionError(
                        "an aggregation is not applicable to its field type"
                    ) from exc
            else:
                method = "nunique" if item["operation"] == "count_distinct" else item["operation"]
                try:
                    values[item["name"]] = getattr(grouped[source], method)()
                except (TypeError, ValueError) as exc:
                    raise ListingExecutionError(
                        "an aggregation is not applicable to its field type"
                    ) from exc
        if grouped is not None:
            frame = grouped.first().reset_index()[group]
            for name, series in values.items(): frame[name] = series.to_numpy()
        else:
            frame = pd.DataFrame([values])
    selected = {}
    for item in output["columns"]:
        if layout.get("dropCodeValue") and _is_code_value_column(item):
            continue
        selected[item["name"]] = frame[_column(frame, item["source"])]
    if not selected:
        raise ListingExecutionError("layout removed every output field")
    result = pd.DataFrame(selected)
    if layout.get("appendReviewColumns"):
        # 2026-08-25：frame 追加列必须与 names/labels 同源（profile.review_columns）。
        # 此前误用全局 REVIEW_COLUMNS，自定义复核列项目上 frame 5 列、表头 2 列，
        # 数据行宽于表头（test_execute_uses_custom_contents_and_review_columns 抓获）。
        for review in profile.review_columns:
            result[review] = ""
    for item in output["sort"]:
        sort_name = item["column"].split(".")[-1]
        if sort_name not in result.columns:
            raise ListingExecutionError("sort field must be present in output columns")
        result = result.sort_values(by=sort_name, ascending=item["direction"] == "asc")
    return result.reset_index(drop=True)


def execute_listing_plan(
    project_dir: str, output_dir: str, plan: dict[str, Any],
    credential: bytes | str | None = None,
) -> dict[str, Any]:
    project = Path(project_dir).resolve(strict=True)
    output = Path(output_dir).resolve()
    output.mkdir(parents=True, exist_ok=True)
    profile = load_project_profile(project)
    artifacts = []
    required_datasets = {
        str(definition["source"])
        for definition in plan["outputs"]
    }
    required_datasets.update(
        str(join["dataset"])
        for definition in plan["outputs"]
        for join in definition["joins"]
    )
    with DatasetCatalog(project, credential, required_datasets=required_datasets) as catalog:
        files = catalog.files()
        rendered = []
        for definition in plan["outputs"]:
            frame = _execute_output(definition, files, profile)
            visible_columns = [item for item in definition["columns"] if not (
                definition["layout"].get("dropCodeValue") and _is_code_value_column(item)
            )]
            names = [item["name"] for item in visible_columns]
            labels = [item["label"] or item["name"] for item in visible_columns]
            if definition["layout"].get("appendReviewColumns"):
                names.extend(profile.review_columns)
                labels.extend(profile.review_columns.values())
            rendered.append((definition, frame, names, labels))

        include_contents = any(item[0]["layout"].get("includeContents") for item in rendered)
        if not include_contents:
            for index, (definition, frame, names, labels) in enumerate(rendered, start=1):
                path = output / f"{plan['scenario'].upper()}_{index:03d}_{definition['name']}.xlsx"
                workbook = openpyxl.Workbook()
                workbook.iso_dates = True
                try:
                    sheet = workbook.active
                    sheet.title = _sheet_name(definition["name"], set(), profile.contents_sheet_name)
                    sheet.append(labels)
                    _style_header(sheet, 1, len(labels))
                    for row in frame.itertuples(index=False, name=None):
                        sheet.append(list(row) + ([""] * (len(names) - len(frame.columns))))
                    _finish_sheet(sheet, 1, len(labels), definition["layout"].get("freezeRows", 1), definition["layout"].get("freezeColumns", 0))
                    _atomic_save(workbook, path)
                finally:
                    workbook.close()
                artifacts.append({"path": str(path), "name": path.name, "kind": "xlsx", "rowCount": len(frame), "columnCount": len(names)})
            return {"status": "completed", "dataClass": "REAL", "artifacts": artifacts, "artifact": artifacts[0]}

        path = output / f"{plan['scenario'].upper()}_LISTINGS.xlsx"
        workbook = openpyxl.Workbook()
        workbook.iso_dates = True
        try:
            if include_contents:
                contents = workbook.active
                contents.title = profile.contents_sheet_name
                contents.append(list(profile.contents_headers))
                _style_header(contents, 1, len(profile.contents_headers))
                used_sheet_names = {profile.contents_sheet_name.casefold()}
                rendered_names = []
                for index, (definition, frame, _, _) in enumerate(rendered, start=1):
                    name = _sheet_name(definition["name"], used_sheet_names, profile.contents_sheet_name)
                    rendered_names.append(name)
                    contents.append([index, name, definition["source"], "", "", len(frame), 0, 0])
                    cell = contents.cell(contents.max_row, 2)
                    cell.hyperlink = f"#'{name}'!A1"
                    cell.style = "Hyperlink"
                _finish_sheet(contents, 1, len(profile.contents_headers), 1, 1)
            else:
                contents = None
                workbook.remove(workbook.active)

            if include_contents:
                used_sheet_names = {profile.contents_sheet_name.casefold()}
            else:
                used_sheet_names = set()
            for output_index, (definition, frame, names, labels) in enumerate(rendered):
                sheet_name = rendered_names[output_index] if include_contents else _sheet_name(definition["name"], used_sheet_names, profile.contents_sheet_name)
                sheet = workbook.create_sheet(sheet_name)
                header_row = 2 if include_contents else 1
                if include_contents:
                    cell = sheet.cell(1, 1, "Go back")
                    cell.hyperlink = f"#'{profile.contents_sheet_name}'!A1"
                    cell.style = "Hyperlink"
                sheet.append(labels)
                if len(names) != len(labels):
                    raise ListingExecutionError("output header shape is invalid")
                _style_header(sheet, header_row, len(labels))
                for row in frame.itertuples(index=False, name=None):
                    values = list(row) + ([""] * (len(names) - len(frame.columns)))
                    sheet.append(values)
                _finish_sheet(
                    sheet, header_row, len(labels),
                    definition["layout"].get("freezeRows", 1),
                    definition["layout"].get("freezeColumns", 0),
                )

            if not workbook.sheetnames:
                raise ListingExecutionError("workbook has no sheets")
            _atomic_save(workbook, path)
        finally:
            workbook.close()
        artifact = {
            "path": str(path),
            "name": path.name,
            "kind": "xlsx",
            "rowCount": sum(len(frame) for _, frame, _, _ in rendered),
            "columnCount": max((len(names) for _, _, names, _ in rendered), default=0),
        }
        artifacts.append(artifact)
    return {"status": "completed", "dataClass": "REAL", "artifacts": artifacts, "artifact": artifacts[0]}
