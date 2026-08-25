"""结构化解析可信规格文档，不返回任意整表内容。"""
from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from security.project_profile import DEFAULT_SPEC_DIRECTORY, ProjectProfile


class SpecParseError(ValueError):
    """Spec 解析失败的安全原因（不含数据值）。"""


SPEC_DIRECTORY = DEFAULT_SPEC_DIRECTORY
SUPPORTED_DOCUMENT_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".txt", ".pdf"}

_IDENTIFIER = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]{0,127}$")


def _safe_text(value: Any) -> str:
    # 2026-08-23：不再对 spec 单元格做 token 化。这里处理的是 doc 目录内的
    # 规格文档（ALS / listing 要求 / KRI 定义），属规格来源而不是 data 来源。
    # 原实现对含日期或编号形态的单元格调 smart_scrub_text，把
    # "ALT: 3 倍正常上限" 打成 "[TEXT:..]: [NUM:..] [TEXT:..]"，
    # 判定阈值、变量名、访视日期格式全部丢失，harness 无法理解 spec 需求。
    # 真实 data 数据由 protectedDataSource 按来源域在出域侧阻断，不靠此处。
    # smart_scrub_text 按"不启用，先不删"保留，仅断开调用。
    if value is None:
        return ""
    return str(value).strip()


def _headers(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        re.sub(r"[\s_\-]+", "", _safe_text(value).lower()): index
        for index, value in enumerate(row)
        if value is not None and str(value).strip()
    }


def _column(headers: dict[str, int], *names: str, default: int | None = None) -> int | None:
    for name in names:
        key = re.sub(r"[\s_\-]+", "", name.lower())
        if key in headers:
            return headers[key]
    return default


def _value(row: tuple[Any, ...], index: int | None) -> str:
    return _safe_text(row[index]) if index is not None and index < len(row) else ""


def _raw_value(row: tuple[Any, ...], index: int | None) -> Any:
    return row[index] if index is not None and index < len(row) else None


def _number(value: Any, default: int | float = 0) -> int | float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return default
    return int(number) if number.is_integer() else number


def _rows(sheet) -> list[tuple[Any, ...]]:
    return list(sheet.iter_rows(values_only=True))


def _header_index(rows: list[tuple[Any, ...]], required: tuple[tuple[str, ...], ...]) -> int | None:
    for index, row in enumerate(rows):
        headers = _headers(row)
        if all(_column(headers, *roles) is not None for roles in required):
            return index
    return None


def _parse_flat_als_sheet(sheet) -> dict[str, Any]:
    rows = _rows(sheet)
    header_index = _header_index(rows, (("datasetname", "sasdatasetname", "dataset"),))
    if header_index is None:
        return {"forms": [], "datasets": [], "mappings": [], "warnings": []}
    headers = _headers(rows[header_index])
    dataset_index = _column(headers, "datasetname", "sasdatasetname", "dataset")
    source_index = _column(headers, "sasfieldname", "itemname", "variableoid", "variable", "field", "column")
    label_index = _column(headers, "saslabel", "pretext", "itemlabel", "label", "draftfieldname")
    form_index = _column(headers, "formname", "draftformname")
    order_index = _column(headers, "itemorder", "ordinal", "order")
    forms: list[dict[str, Any]] = []
    datasets: list[str] = []
    mappings: list[dict[str, Any]] = []
    warnings: list[str] = []
    seen_forms: set[tuple[str, str]] = set()
    for row in rows[header_index + 1:]:
        dataset = _value(row, dataset_index)
        if not dataset:
            continue
        if dataset not in datasets:
            datasets.append(dataset)
        form = _value(row, form_index)
        form_key = (form, dataset)
        if form and form_key not in seen_forms:
            forms.append({"formName": form, "datasetName": dataset})
            seen_forms.add(form_key)
        source = _value(row, source_index)
        if source:
            mapping = {
                "datasetName": dataset,
                "sourceColumn": source,
                "displayLabel": _value(row, label_index) or source,
                "formName": form,
                "order": _number(_raw_value(row, order_index), len(mappings)),
            }
            mappings.append(mapping)
    return {"forms": forms, "datasets": datasets, "mappings": mappings, "warnings": warnings}


def _parse_relational_als(workbook) -> dict[str, Any]:
    sheet_rows = {sheet.title.casefold(): _rows(sheet) for sheet in workbook.worksheets}
    forms_by_oid: dict[str, dict[str, str]] = {}
    items_by_oid: dict[str, dict[str, str]] = {}
    links: list[tuple[str, str, int | float]] = []
    direct_fields: list[dict[str, Any]] = []

    for rows in sheet_rows.values():
        forms_header = _header_index(rows, (("formoid", "oid"), ("formname", "draftformname")))
        if forms_header is not None:
            headers = _headers(rows[forms_header])
            oid_index = _column(headers, "formoid", "oid")
            name_index = _column(headers, "formname", "draftformname")
            dataset_index = _column(headers, "sasdatasetname", "datasetname", "dataset", "oid")
            for row in rows[forms_header + 1:]:
                oid = _value(row, oid_index)
                name = _value(row, name_index)
                dataset = _value(row, dataset_index)
                if oid and name and dataset:
                    forms_by_oid[oid] = {"formName": name, "datasetName": dataset}

        items_header = _header_index(rows, (("itemoid",), ("sasfieldname", "variableoid")))
        if items_header is not None:
            headers = _headers(rows[items_header])
            oid_index = _column(headers, "itemoid")
            source_index = _column(headers, "sasfieldname", "variableoid")
            label_index = _column(headers, "saslabel", "itemname", "draftfieldname", "pretext")
            for row in rows[items_header + 1:]:
                oid = _value(row, oid_index)
                source = _value(row, source_index)
                if oid and source:
                    items_by_oid[oid] = {"sourceColumn": source, "displayLabel": _value(row, label_index) or source}

        link_header = _header_index(rows, (("formoid",), ("itemoid",)))
        if link_header is not None:
            headers = _headers(rows[link_header])
            form_index = _column(headers, "formoid")
            item_index = _column(headers, "itemoid")
            order_index = _column(headers, "ordinal", "itemorder", "order")
            for order, row in enumerate(rows[link_header + 1:]):
                form_oid = _value(row, form_index)
                item_oid = _value(row, item_index)
                if form_oid and item_oid:
                    links.append((form_oid, item_oid, _number(_raw_value(row, order_index), order)))

        direct_header = _header_index(rows, (("formoid",), ("variableoid", "sasfieldname"), ("draftfieldname", "saslabel", "pretext")))
        if direct_header is not None:
            headers = _headers(rows[direct_header])
            form_index = _column(headers, "formoid")
            source_index = _column(headers, "variableoid", "sasfieldname")
            label_index = _column(headers, "saslabel", "draftfieldname", "pretext")
            order_index = _column(headers, "ordinal", "itemorder", "order")
            for order, row in enumerate(rows[direct_header + 1:]):
                form_oid = _value(row, form_index)
                source = _value(row, source_index)
                if form_oid and source:
                    direct_fields.append({
                        "formOID": form_oid,
                        "sourceColumn": source,
                        "displayLabel": _value(row, label_index) or source,
                        "order": _number(_raw_value(row, order_index), order),
                    })

    if not forms_by_oid or (not links and not direct_fields):
        return {"forms": [], "datasets": [], "mappings": [], "warnings": []}

    warnings: list[str] = []
    forms: list[dict[str, Any]] = []
    datasets: list[str] = []
    mappings: list[dict[str, Any]] = []
    for form in forms_by_oid.values():
        dataset = form["datasetName"]
        if dataset not in datasets:
            datasets.append(dataset)
        forms.append(dict(form))

    for form_oid, item_oid, order in links:
        form = forms_by_oid.get(form_oid)
        item = items_by_oid.get(item_oid)
        if not form or not item:
            continue
        mappings.append({
            "datasetName": form["datasetName"],
            "sourceColumn": item["sourceColumn"],
            "displayLabel": item["displayLabel"],
            "formName": form["formName"],
            "order": order,
        })

    for field in direct_fields:
        form = forms_by_oid.get(field["formOID"])
        if not form:
            continue
        mappings.append({
            "datasetName": form["datasetName"],
            "sourceColumn": field["sourceColumn"],
            "displayLabel": field["displayLabel"],
            "formName": form["formName"],
            "order": field["order"],
        })

    return {"forms": forms, "datasets": datasets, "mappings": mappings, "warnings": warnings}


def _parse_spec_sheet(sheet) -> dict[str, Any]:
    rows = _rows(sheet)
    header_index = _header_index(rows, (("column", "variable", "field", "name"),))
    if header_index is None:
        return {"fields": [], "warnings": []}
    headers = _headers(rows[header_index])
    name_index = _column(headers, "column", "variable", "field", "name")
    label_index = _column(headers, "label", "description", "requirement")
    type_index = _column(headers, "type", "datatype", "format")
    rule_index = _column(headers, "rule", "derivation", "logic")
    fields: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row in rows[header_index + 1:]:
        name = _value(row, name_index)
        if not name:
            continue
        fields.append({
            "column": name,
            "label": _value(row, label_index),
            "dataType": _value(row, type_index),
            "rule": _value(row, rule_index),
        })
    return {"fields": fields, "warnings": warnings}


def _parse_kri_sheet(sheet) -> dict[str, Any]:
    rows = _rows(sheet)
    header_index = next((
        index
        for index, row in enumerate(rows)
        if sum("kri" in str(value).strip().casefold() for value in row if value is not None) >= 2
    ), None)
    if header_index is None:
        return {"kris": [], "warnings": []}
    headers = [str(value).strip() if value is not None else "" for value in rows[header_index]]
    kris: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row in rows[header_index + 1:]:
        cells = [str(value).strip() if value is not None else "" for value in row]
        if not any(cells):
            continue
        definition = {header or f"column{index + 1}": cells[index] if index < len(cells) else "" for index, header in enumerate(headers) if header}
        definition["name"] = cells[0]
        definition["threshold"] = " | ".join(value for value in cells[1:] if value)
        kris.append(definition)
    return {"kris": kris, "warnings": warnings}


def _parse_layout_sheet(sheet) -> dict[str, Any]:
    rows = _rows(sheet)
    for index, row in enumerate(rows[:32]):
        values = [_safe_text(value) for value in row]
        populated = [value for value in values if value]
        if not populated:
            continue
        identifier_count = sum(bool(_IDENTIFIER.fullmatch(value)) for value in populated)
        if identifier_count < max(1, len(populated) // 2):
            continue
        previous = [_safe_text(value) for value in rows[index - 1]] if index else []
        columns = []
        for column_index, name in enumerate(values):
            if not name or not _IDENTIFIER.fullmatch(name):
                continue
            label = previous[column_index] if column_index < len(previous) else ""
            columns.append({"name": name, "label": label if label != name else ""})
        if columns:
            return {"headerRow": index + 1, "columns": columns}
    return {"headerRow": 0, "columns": []}


def _parse_requirement_rows(sheet) -> dict[str, Any]:
    requirements: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [_safe_text(value) for value in row]
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        requirements.append({"sheet": _safe_text(sheet.title), "row": row_number, "cells": cells})
    return {"requirements": requirements, "warnings": warnings}


class _XlrdRequirementSheetAdapter:
    """把 xlrd sheet 适配为 _parse_requirement_rows 需要的 openpyxl 风格接口。"""

    def __init__(self, sheet) -> None:
        self.title = sheet.name
        self._sheet = sheet

    def iter_rows(self, values_only=True):
        for index in range(self._sheet.nrows):
            yield tuple(self._sheet.row_values(index))


def parse_xls_spec_document(path: Path) -> dict[str, Any]:
    """doc/ 规格域内 .xls 文件的受信全量解析（2026-08-24 红线口径）。

    doc/ 内文件全量供模型理解需求；.xls 此前只报 "could not be parsed"，
    是"AI 读不到需求"的真实来源之一。复用 requirement 行解析器的全部边界
    （sheet 数、单元格数、行数上限），产出与 xlsx spec 同构的 requirements。
    .txt 不走此通道：doc/ 内可能有密码 sidecar（如 A1234567.txt），
    全量读文本等于把凭据送进模型。
    """
    import xlrd

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    parsed: dict[str, Any] = {"requirements": [], "warnings": []}
    try:
        for sheet in workbook.sheets():
            result = _parse_requirement_rows(_XlrdRequirementSheetAdapter(sheet))
            parsed["requirements"].extend(result["requirements"])
            parsed["warnings"].extend(result["warnings"])
    finally:
        workbook.release_resources()
    return parsed


def find_spec_documents(project: Path, profile: ProjectProfile | None = None) -> list[Path]:
    root = project / (profile or ProjectProfile()).spec_directory
    if not root.is_dir():
        return []
    return sorted(
        (
            path for path in root.rglob("*")
            if path.is_file()
            and path.suffix.casefold() in SUPPORTED_DOCUMENT_EXTENSIONS
            and not path.name.startswith("~$")
        ),
        key=lambda path: path.as_posix().casefold(),
    )


def classify_spec_document(path: Path, profile: ProjectProfile | None = None) -> str:
    profile = profile or ProjectProfile()
    stem = path.stem.casefold()
    parts = {part.casefold() for part in path.parts[:-1]}
    if profile.spec_directory.casefold() in parts:
        return "als" if "als" in stem else "specification"
    if any(term in stem for term in profile.spec_keywords):
        return "specification"
    if any(term in stem for term in profile.report_support_keywords):
        return "report_support_data"
    return "requirement_note"


_FILTER_REF = re.compile(
    r"^[$]?[A-Za-z]{1,3}[$]?\d+(:[$]?[A-Za-z]{1,3}[$]?\d+?)?$|^[A-Za-z]{1,3}:[A-Za-z]{1,3}$"
)


def _is_valid_filter_ref(value: str) -> bool:
    return bool(_FILTER_REF.fullmatch(value))


def _workbook_bytes_for_parser(path: Path) -> bytes:
    raw = path.read_bytes()
    if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        return raw
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            members = {name: archive.read(name) for name in archive.namelist()}
    except (OSError, zipfile.BadZipFile) as exc:
        raise SpecParseError("无法打开规格文档") from exc

    changed = False
    filter_tag = re.compile(rb"(<autoFilter\b)([^>]*)(>)", re.IGNORECASE)
    ref_attribute = re.compile(rb"\s+ref=(['\"])(.*?)\1", re.IGNORECASE)
    for name in tuple(members):
        if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name, re.IGNORECASE):
            continue

        def clean_filter(match: re.Match[bytes]) -> bytes:
            nonlocal changed
            attributes = match.group(2)
            ref_match = ref_attribute.search(attributes)
            if ref_match is None:
                return match.group(0)
            ref = ref_match.group(2).decode("utf-8", "ignore")
            if _is_valid_filter_ref(ref):
                return match.group(0)
            changed = True
            attributes = attributes[:ref_match.start()] + attributes[ref_match.end():]
            return match.group(1) + attributes + match.group(3)

        members[name] = filter_tag.sub(clean_filter, members[name])

    if not changed:
        return raw
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in members.items():
            archive.writestr(name, content)
    return output.getvalue()


def parse_spec_document(spec_path: str, doc_type: Optional[str] = None) -> dict[str, Any]:
    if not HAS_OPENPYXL:
        raise SpecParseError("openpyxl 未安装，无法解析 Excel 文档")
    path = Path(spec_path)
    if not path.exists():
        raise SpecParseError("规格文档不存在")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise SpecParseError("当前结构化解析器仅支持 XLSX/XLSM 规格文档")
    try:
        workbook_bytes = _workbook_bytes_for_parser(path)
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    except SpecParseError:
        raise
    except Exception as exc:
        raise SpecParseError("无法打开规格文档") from exc

    result: dict[str, Any] = {
        "forms": [],
        "datasets": [],
        "fields": [],
        "kris": [],
        "requirements": [],
        "mappings": [],
        "warnings": [],
        "sheets": [],
    }
    try:
        if doc_type == "als":
            relational = _parse_relational_als(workbook)
            if relational["forms"] or relational["mappings"]:
                for key in ("forms", "datasets", "mappings"):
                    result[key] = relational[key]
                result["warnings"].extend(relational["warnings"])
            else:
                for sheet in workbook.worksheets:
                    parsed = _parse_flat_als_sheet(sheet)
                    for key in ("forms", "datasets", "mappings"):
                        result[key].extend(parsed[key])
                    result["warnings"].extend(parsed["warnings"])
        for sheet in workbook.worksheets:
            layout = _parse_layout_sheet(sheet)
            sheet_result = {
                "name": _safe_text(sheet.title),
                "rowCount": sheet.max_row or 0,
                "headerRow": layout["headerRow"],
                "columns": layout["columns"],
                "parsed": bool(layout["columns"]),
            }
            result["sheets"].append(sheet_result)
            if doc_type == "als":
                continue
            if doc_type == "kri" or "kri" in sheet.title.casefold():
                parsed = _parse_kri_sheet(sheet)
                result["kris"].extend(parsed["kris"])
                result["warnings"].extend(parsed["warnings"])
                continue
            parsed = _parse_spec_sheet(sheet)
            result["fields"].extend(parsed["fields"])
            result["warnings"].extend(parsed["warnings"])
            # 2026-08-24：doc/ 是受信规格域，必须全量供模型理解需求。此前
            # _is_data_example_layout 命中（sheet 含日期/编号形态）就整表跳过
            # requirements，真实项目里上万行的 doc/ 明细表因此只剩表头一行——
            # "doc 文件没有全量读取" 的直接根因。数据域泄露由 protectedDataSource
            # 在出域侧阻断，不靠牺牲规格域完整性。
            requirements = _parse_requirement_rows(sheet)
            result["requirements"].extend(requirements["requirements"])
            result["warnings"].extend(requirements["warnings"])

        result["datasets"] = sorted(set(result["datasets"]), key=str.casefold)
        result["warnings"] = sorted(set(result["warnings"]))
        return result
    finally:
        workbook.close()
