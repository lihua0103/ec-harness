"""临床 Listing inspect 阶段。

这里仅输出需求语义和数据结构元数据。数据文件只以 metadata-only 方式读取，
任何记录、单元格、样本值和绝对路径都不会进入返回对象。
"""
from __future__ import annotations

import hashlib
import io
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None

from security.local_data_inspector import LocalDataInspectionError, inspect_local_data
from security.listing_data_catalog import DatasetCatalog
from security.listing_plan import SCENARIOS
from security.path_policy import PathPolicyError, relative_display_path, resolve_under_root
from security.project_profile import ProjectProfile, load_project_profile
from security.spec_parser import (
    HAS_OPENPYXL,
    SpecParseError,
    classify_spec_document,
    find_spec_documents,
    parse_spec_document,
    parse_xls_spec_document,
    _safe_text,
)


def _schema_fingerprint(schema: dict[str, set[str]]) -> str:
    material = "|".join(
        f"{name}:{','.join(sorted(columns, key=str.casefold))}"
        for name, columns in sorted(schema.items())
    )
    return hashlib.sha256(material.encode("utf-8")).hexdigest()[:16]


def _safe_spec(parsed: dict[str, Any], kind: str) -> dict[str, Any]:
    # 需求正文和结构来自受信 spec 域；排除 layout 中被识别为真实数据示例的页面。
    result: dict[str, Any] = {
        "forms": parsed.get("forms", []),
        "datasets": parsed.get("datasets", []),
        "fields": parsed.get("fields", []),
        "mappings": parsed.get("mappings", []),
        "kris": parsed.get("kris", []),
        "requirements": parsed.get("requirements", []) if kind != "als" else [],
        "sheets": [],
    }
    for sheet in parsed.get("sheets", []):
        result["sheets"].append({
            "name": str(sheet.get("name") or ""),
            "headerRow": int(sheet.get("headerRow") or 0),
            "columns": [
                {"name": str(item.get("name") or ""), "label": str(item.get("label") or "")}
                for item in sheet.get("columns", [])
            ],
        })
    return result


def _infer_scenario(
    paths: list[Path], profile: ProjectProfile | None = None,
) -> tuple[str, float, list[str]]:
    profile = profile or ProjectProfile()
    scores = {scenario: 0 for scenario in SCENARIOS}
    for path in paths:
        name = path.name.casefold()
        for scenario, terms in profile.scenario_keywords.items():
            if scenario in scores and any(term in name for term in terms):
                scores[scenario] += 2
    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    if not ranked[0][1]:
        return profile.default_scenario, 0.0, [profile.default_scenario]
    if scores["manual"] and scores["manual"] == scores["report"] and not scores["rbqm"] and not scores["medical"]:
        return "manual", min(1.0, scores["manual"] / 2), ["manual", "report"]
    if len(ranked) > 1 and ranked[0][1] == ranked[1][1]:
        candidates = [name for name, score in ranked if score == ranked[0][1]]
        return profile.default_scenario, 0.0, candidates
    return ranked[0][0], min(1.0, ranked[0][1] / 2), [ranked[0][0]]


def _support_metadata(path: Path) -> dict[str, Any]:
    result = {
        "name": path.name, "role": "report_support_data",
        "fileType": path.suffix.casefold().lstrip("."), "sheets": [],
    }
    if path.suffix.casefold() == ".xlsx" and HAS_OPENPYXL:
        workbook = openpyxl.load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                total = sheet.max_row or 0
                rows: list[list[str]] = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [_safe_text(value) for value in row]
                    # 去掉行尾空单元格，减少收据噪音。
                    while cells and not cells[-1]:
                        cells.pop()
                    rows.append(cells)
                result["sheets"].append({
                    "name": sheet.title,
                    "rowCount": total,
                    "truncated": False,
                    "rows": rows,
                })
        finally:
            workbook.close()
    elif path.suffix.casefold() == ".xls":
        # .xls 辅助文件同样属于 doc/ 规格域，走 xlrd 全量有界读取
        # （openpyxl 不支持 .xls；SpreadsheetML 伪装文件由 xlrd 抛错跳过）。
        try:
            import xlrd

            workbook = xlrd.open_workbook(str(path), on_demand=True)
            try:
                for sheet in workbook.sheets():
                    total = sheet.nrows
                    rows = []
                    for row_index in range(total):
                        cells = [
                            _safe_text(value)
                            for value in sheet.row_values(row_index)
                        ]
                        while cells and not cells[-1]:
                            cells.pop()
                        rows.append(cells)
                    result["sheets"].append({
                        "name": sheet.name,
                        "rowCount": total,
                        "truncated": False,
                        "rows": rows,
                    })
            finally:
                workbook.release_resources()
        except Exception:
            # 打不开的辅助文件降级为仅名称清单，不让单个文件杀死 inspect。
            result["sheets"] = []
    return result


def inspect_listing_context(
    local_data_root: str, project: str, scenario: str | None = None,
    credential: bytes | str | None = None,
) -> dict[str, Any]:
    if scenario is not None and scenario not in SCENARIOS:
        raise ValueError("unsupported listing scenario")
    root = Path(local_data_root).resolve(strict=True)
    try:
        project_path = resolve_under_root(str(root), project, allow_root=True)
    except PathPolicyError as exc:
        raise ValueError("project must be a relative path under the local root") from exc
    if not project_path.is_dir():
        raise ValueError("project is not a directory")

    documents: list[dict[str, Any]] = []
    support_data: list[dict[str, Any]] = []
    schema: dict[str, set[str]] = {}
    warnings: list[str] = []
    profile = load_project_profile(project_path)
    warnings.extend(profile.warnings)
    spec_paths = find_spec_documents(project_path, profile)
    specification_paths = [path for path in spec_paths if classify_spec_document(path, profile) == "specification"]
    inferred_scenario, scenario_confidence, scenario_candidates = _infer_scenario(
        specification_paths, profile)
    scenario = scenario or inferred_scenario
    spec_path_set = {path.resolve() for path in spec_paths}
    for path in spec_paths:
        kind = classify_spec_document(path, profile)
        if kind == "report_support_data":
            support_data.append(_support_metadata(path))
            continue
        if path.suffix.casefold() == ".xls":
            # 2026-08-24 红线口径：doc/ 内 .xls 也是规格域，全量解析供模型
            # 理解需求（此前 warning-only，AI 根本读不到内容）。
            try:
                parsed = parse_xls_spec_document(path)
            except Exception:
                warnings.append("a specification document could not be parsed")
                continue
        elif path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            # .txt/.pdf 保持不解析：doc/ 内可能有密码 sidecar，全量读文本
            # 等于把凭据送进模型。
            warnings.append("a specification document could not be parsed")
            continue
        else:
            try:
                parsed = parse_spec_document(str(path), "als" if kind == "als" else "spec")
            except SpecParseError:
                warnings.append("a specification document could not be parsed")
                continue
        safe = _safe_spec(parsed, kind)
        warnings.extend(str(warning) for warning in parsed.get("warnings", []))
        documents.append({
            "name": relative_display_path(project_path, path),
            "kind": kind,
            "content": safe,
        })
        for dataset in parsed.get("datasets", []):
            schema.setdefault(str(dataset).casefold(), set())
        # inspect 不解压归档。ALS/spec 是受信字段目录，ZIP central directory
        # 只证明数据集成员存在；两者交叉后形成计划校验所需的 schema。
        for mapping in parsed.get("mappings", []):
            dataset = str(mapping.get("datasetName") or "").strip()
            column = str(mapping.get("sourceColumn") or "").strip()
            if dataset and column:
                schema.setdefault(dataset.casefold(), set()).add(column)

    data_files: list[dict[str, Any]] = []
    missing_archives: list[str] = []
    # inspect 阶段只读取归档的 descriptor 元数据，不读取或返回记录值。
    # 只有 central directory 无法提供真实字段名；ALS 映射不能证明归档内
    # 实际列一定存在，否则会出现 ready/validated 但 execute 失败。
    with DatasetCatalog(project_path, credential, materialize_archives=True) as catalog:
        # E-3: 打不开的归档降级为结构化 missing（credential:<项目相对名>），
        # 其余明文数据集照常进入 schema——不再让单个加密归档杀死整个 inspect。
        missing_archives = list(catalog.missing_archives)
        for dataset_name, paths in sorted(catalog.files().items()):
            # 原始目录可能同时有同名副本；必须明确报告歧义，不能猜选一个。
            if len(paths) != 1:
                warnings.append("a local dataset name is ambiguous")
                continue
            path = paths[0]
            try:
                # N-7: 归档数据集被解包到 .clinical-listing 下的随机临时目录
                # （.listing-catalog-<random>），直接用相对路径当显示名会让
                # AI 每次看到不同的名字，无法在多轮对话里稳定引用同一数据集。
                # 解包产物统一显示为 archive/<file>；此前那个 is_relative_to
                # 分支永远为真（临时目录就在项目内），archive/ 分支是死代码。
                catalog_root = project_path / ".clinical-listing"
                if catalog_root in path.parents or project_path not in path.parents:
                    display = f"archive/{path.name}"
                    metadata_root, requested = path.parent, path.name
                else:
                    display = relative_display_path(project_path, path)
                    metadata_root, requested = project_path, display
                metadata = inspect_local_data(str(metadata_root), requested)
            except (LocalDataInspectionError, OSError):
                warnings.append("a local data source could not be inspected")
                continue
            for sheet in metadata.get("sheets", []):
                dataset = path.stem
                columns = {str(column) for column in sheet.get("columns", []) if str(column).strip()}
                schema.setdefault(dataset.casefold(), set()).update(columns)
                data_files.append({
                    "name": display,
                    "dataset": dataset,
                    "fileType": metadata.get("fileType", ""),
                    "rowCount": max(0, int(sheet.get("rowCount") or 0)),
                    "columns": sorted(columns, key=str.casefold),
                })
        # central directory 只提供成员名称；不读取、解压或返回任何记录。
        indexed_files = catalog.files()
        for dataset_name, archives in sorted(catalog.archive_datasets().items()):
            # 已通过 metadata-only descriptor 建立真实 schema 的归档成员不再
            # 用 ALS 映射重复添加；仅对无法解包/缺凭据的成员保留结构化提示。
            if indexed_files.get(dataset_name):
                continue
            if len(archives) != 1:
                warnings.append("a local dataset name is ambiguous")
                continue
            columns = schema.get(dataset_name, set())
            if not columns:
                warnings.append("a local dataset has no trusted field mapping")
                continue
            data_files.append({
                "name": f"archive/{Path(archives[0]).name}/{dataset_name}",
                "dataset": dataset_name,
                "fileType": "archive-member",
                "rowCount": 0,
                "columns": sorted(columns, key=str.casefold),
            })
    if missing_archives:
        warnings.append("a local data archive could not be opened")

    return {
        "clinicalGuard": "CLINICAL_LISTING_INSPECTION",
        "status": "ready" if documents and schema else "needs_input",
        "stage": "inspect",
        "project": project,
        "scenario": scenario,
        "inferredScenario": inferred_scenario,
        "scenarioConfidence": scenario_confidence,
        "scenarioCandidates": scenario_candidates,
        "documents": documents,
        "supportData": support_data,
        "datasets": data_files,
        "schema": {name: sorted(columns, key=str.casefold) for name, columns in sorted(schema.items())},
        "schemaFingerprint": _schema_fingerprint(schema),
        "missing": ([] if documents else ["specification"])
        + [f"credential:{name}" for name in sorted(set(missing_archives))],
        "warnings": sorted(set(warnings)),
        "dataClass": "METADATA_ONLY",
    }
