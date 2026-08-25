"""沙箱子进程入口：`python -m security.sandbox_runner`。

协议：stdin 一行 job JSON → stdout 一行 result JSON。本进程是**受信固定代码**：
- 执行前重新跑白名单 AST（纵深防御，父层已查过一遍）
- 数据集只能经注入的 ``datasets`` 注册表按需读取（复用 IR 执行器的 canonical
  列名投影，避免 inspect/执行口径分叉）
- run 模式只产出聚合元数据信封；publish 模式由本文件的固定 Writer 落盘 Excel，
  模型代码永远不接触文件写出
- 任何异常收敛为 {type, message}，消息经 sanitize_error 脱敏后仍由父层二次 scrub

stdout 除结果行外不得写任何内容；pandas 警告走 stderr，不回传。
"""
from __future__ import annotations

import json
import os
import sys
from collections.abc import Mapping
from pathlib import Path
from typing import Any

class SandboxDataError(RuntimeError):
    """数据集不可用等模型可行动的错误（文案为元数据，不含路径）。"""


class SandboxSecurityError(RuntimeError):
    """沙盒安全违规：路径遍历或未授权访问。"""


class DatasetRegistry(Mapping):
    """P0-2 Fix: 名称 → DataFrame 的懒加载只读注册表。

    安全加固：
    1. 只允许从白名单目录读取文件
    2. 规范化路径后验证无路径遍历
    3. 不导入外部模块，在沙盒内部实现数据读取
    4. 缓存结果防止重复读取
    """

    _ALLOWED_DATA_DIRS: list[Path] = []

    @classmethod
    def set_allowed_dirs(cls, dirs: list[str]) -> None:
        # P0-FIX (RBQM run_code): 过滤空字符串与纯空白项。Path("").resolve() 会
        # 静默解析为当前工作目录（沙箱子进程的 cwd 是包根），把整个包目录变成
        # 可读白名单——既是越权，也让真实的"未配置"故障伪装成"配置成功"。
        cleaned = [str(d).strip() for d in dirs if str(d).strip()]
        if not cleaned:
            raise SandboxSecurityError("allowed data directories are empty or invalid")
        cls._ALLOWED_DATA_DIRS = [Path(d).resolve() for d in cleaned]

    def __init__(self, files: dict[str, str]) -> None:
        self._files: dict[str, Path] = {}
        for name, path in files.items():
            resolved = self._validate_path(path)
            if resolved:
                self._files[str(name).casefold()] = resolved
        self._cache: dict[str, Any] = {}

    @classmethod
    def _validate_path(cls, path: str) -> Path | None:
        """验证路径安全性：只允许白名单目录内的文件。

        防止路径遍历攻击：攻击者可能通过 ../../../etc/passwd 等方式
        尝试读取项目目录外的敏感文件。
        """
        try:
            resolved = Path(path).resolve()
        except (OSError, ValueError):
            raise SandboxSecurityError(f"invalid path: {path[:50]}")

        if not cls._ALLOWED_DATA_DIRS:
            raise SandboxSecurityError("no allowed data directories configured")

        for allowed_dir in cls._ALLOWED_DATA_DIRS:
            try:
                resolved.relative_to(allowed_dir)
                return resolved
            except ValueError:
                continue

        raise SandboxSecurityError(f"path outside allowed directories: {resolved.name}")

    def __getitem__(self, key: str) -> Any:
        name = str(key).casefold()
        if name not in self._cache:
            path = self._files.get(name)
            if path is None:
                available = sorted(self._files)
                raise SandboxDataError(
                    f"dataset '{name}' is not available; available datasets: "
                    + (", ".join(available) if available else "(none)"))
            self._cache[name] = _read_local(path)
        return self._cache[name]

    def __iter__(self):
        return iter(self._files)

    def __len__(self) -> int:
        return len(self._files)


def _read_local(path: Path) -> Any:
    """沙盒内部数据读取（不依赖外部模块）。

    P0-2 Fix: 移除对 listing_executor._read 的依赖，
    在沙盒内部实现安全的数据读取。
    """
    import pandas as pd

    try:
        suffix = path.suffix.lower()
        if suffix == ".xpt":
            import pyreadstat
            frame, metadata = pyreadstat.read_xport(str(path))
            return frame
        elif suffix == ".sas7bdat":
            import pyreadstat
            frame, _metadata = pyreadstat.read_sas7bdat(str(path))
            return frame
        elif suffix == ".csv":
            return pd.read_csv(path)
        elif suffix in {".xlsx", ".xls", ".xlsm"}:
            return pd.read_excel(path, engine="openpyxl" if suffix != ".xls" else None)
        raise SandboxDataError(f"unsupported dataset format: {suffix or '(none)'}")
    except SandboxDataError:
        raise
    except Exception as exc:
        raise SandboxDataError(f"failed to read dataset: {type(exc).__name__}") from exc


def _import_pandas():
    import pandas as pd

    return pd


def _collect_outputs(namespace: dict[str, Any]) -> dict[str, Any]:
    pd = _import_pandas()
    if "outputs" in namespace:
        raw = namespace["outputs"]
        if not isinstance(raw, dict) or not raw:
            raise SandboxDataError(
                "'outputs' must be a non-empty dict of {listing name: DataFrame}")
        frames = {}
        for name, frame in raw.items():
            if not isinstance(frame, pd.DataFrame):
                raise SandboxDataError(
                    f"output '{str(name)}' is not a DataFrame")
            frames[str(name)] = frame
        return frames
    if "result" in namespace:
        frame = namespace["result"]
        if not isinstance(frame, pd.DataFrame):
            raise SandboxDataError("'result' must be a DataFrame")
        return {"result": frame}
    raise SandboxDataError(
        "code must assign 'result' (a DataFrame) or 'outputs' (a dict of DataFrames)")


def _envelope(frames: dict[str, Any], datasets_touched: list[str]) -> dict[str, Any]:
    outputs = []
    for name, frame in frames.items():
        columns = []
        for column in list(frame.columns):
            series = frame[column]
            columns.append({
                "name": str(column),
                "dtype": str(series.dtype),
                "nullCount": int(series.isna().sum()),
            })
        outputs.append({
            "name": str(name),
            "rowCount": int(len(frame)),
            "columnCount": int(len(frame.columns)),
            "columns": columns,
        })
    return {
        "status": "ok",
        "mode": "run",
        "outputs": outputs,
        "datasetsTouched": datasets_touched,
    }


def _clean_cell(value: Any) -> Any:
    pd = _import_pandas()
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _write_workbook(
    staging: Path, scenario: str, frames: dict[str, Any],
    review_columns: list[str], contents_sheet_name: str,
) -> dict[str, Any]:
    """固定 Writer：CONTENTS 目录页 + 每 listing 一 sheet + 复核列。

    复用 IR 执行器的格式红线（表头样式、冻结/筛选、公式注入防护、原子落盘），
    保证两条车道产出物格式一致；模型代码不参与写文件。
    """
    import openpyxl

    from security.listing_executor import (
        _CONTENTS_HEADERS,
        _atomic_save,
        _finish_sheet,
        _sheet_name,
        _style_header,
    )

    staging.mkdir(parents=True, exist_ok=True)
    path = staging / f"{scenario.upper()}_LISTINGS.xlsx"
    workbook = openpyxl.Workbook()
    workbook.iso_dates = True
    try:
        contents = workbook.active
        contents.title = "Contents"
        contents.append(_CONTENTS_HEADERS)
        _style_header(contents, 1, len(_CONTENTS_HEADERS))
        used = {str(contents_sheet_name).casefold()}
        for index, (name, frame) in enumerate(frames.items(), start=1):
            sheet_name = _sheet_name(name, used)
            contents.append([index, sheet_name, "", "", "", int(len(frame)), 0, 0])
            cell = contents.cell(contents.max_row, 2)
            cell.hyperlink = f"#'{sheet_name}'!A1"
            cell.style = "Hyperlink"
        _finish_sheet(contents, 1, len(_CONTENTS_HEADERS), 1, 1)

        used = {str(contents_sheet_name).casefold()}
        for name, frame in frames.items():
            sheet_name = _sheet_name(name, used, contents_sheet_name)
            sheet = workbook.create_sheet(sheet_name)
            back = sheet.cell(1, 1, "Go back")
            back.hyperlink = f"#'{contents_sheet_name}'!A1"
            back.style = "Hyperlink"
            labels = [str(column) for column in frame.columns] + list(review_columns)
            sheet.append(labels)
            _style_header(sheet, 2, len(labels))
            padding = [""] * len(review_columns)
            for row in frame.itertuples(index=False, name=None):
                sheet.append([_clean_cell(value) for value in row] + padding)
            _finish_sheet(sheet, 2, len(labels), 1, 0)
        _atomic_save(workbook, path)
    finally:
        workbook.close()
    return {
        "status": "ok",
        "mode": "publish",
        "artifacts": [{
            "file": path.name,
            "sheets": [
                {"name": str(name), "rowCount": int(len(frame)),
                 "columnCount": int(len(frame.columns)) + len(review_columns)}
                for name, frame in frames.items()
            ],
        }],
    }


def _error_envelope(exc: BaseException) -> dict[str, Any]:
    from security.patterns import sanitize_error

    try:
        message = sanitize_error(exc) or type(exc).__name__
    except Exception:
        message = type(exc).__name__
    return {
        "status": "error",
        "error": {"type": type(exc).__name__, "message": str(message)},
    }


def execute_job(job: dict[str, Any]) -> dict[str, Any]:
    from security.code_sandbox import SandboxViolation, build_namespace, check_code

    mode = str(job.get("mode") or "run")
    code = job.get("code") or ""
    try:
        check_code(code)

        # P0-FIX (RBQM run_code): 区分三种情形，任何一种缺失都必须 fail-closed。
        # 过去 `job.get(...) or []` 把"键不存在"与"显式空列表"折叠成同一分支，
        # 再由 `elif not _ALLOWED_DATA_DIRS` 兜底，导致真实故障（调用方算出空
        # 路径）只在后续 _validate_path 时才以另一句文案暴露，掩盖了根因。
        if "allowedDataDirs" not in job:
            raise SandboxSecurityError("no allowed data directories provided")
        DatasetRegistry.set_allowed_dirs(job["allowedDataDirs"] or [])

        datasets = DatasetRegistry(job.get("datasets") or {})
        namespace = build_namespace(datasets)
        exec(compile(code, "<listing-code>", "exec"), namespace)  # noqa: S102
        frames = _collect_outputs(namespace)
        if mode == "publish":
            staging = job.get("staging")
            if not staging:
                raise SandboxDataError("publish mode requires a staging directory")
            result = _write_workbook(
                Path(staging), str(job.get("scenario") or "listing"), frames,
                [str(item) for item in (job.get("reviewColumns") or [])],
                str(job.get("contentsSheetName") or "Contents"),
            )
        else:
            result = _envelope(frames, sorted(datasets._cache))
        return result
    except SandboxSecurityError as exc:
        return {
            "status": "rejected",
            "error": {"type": "SandboxSecurityError", "message": str(exc)},
        }
    except SandboxViolation as exc:
        return {
            "status": "rejected",
            "error": {"type": "SandboxViolation", "message": str(exc)},
        }
    except BaseException as exc:  # noqa: BLE001 - 收敛一切异常为脱敏信封
        return _error_envelope(exc)


def main() -> int:
    if hasattr(sys.stdin, "reconfigure"):
        sys.stdin.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    line = sys.stdin.readline()
    try:
        job = json.loads(line)
        if not isinstance(job, dict):
            raise ValueError("job must be an object")
        result = execute_job(job)
    except Exception as exc:  # noqa: BLE001
        result = _error_envelope(exc)
    sys.stdout.write(json.dumps(result, ensure_ascii=False, separators=(",", ":")) + "\n")
    sys.stdout.flush()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
