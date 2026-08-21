"""DSH 临床数据守卫的常驻安全检查进程。

协议：每行一个 JSON 请求，每行一个 JSON 响应。进程只输出脱敏摘要，
任何导入或检查异常都转换为 fail-closed 响应，绝不放行未知故障。

FIX-3 (R-6 / AR-2.9): 所有异常回执经 sanitize_error 统一脱敏（路径→[PATH]、
受试者/日期→占位），绝不外泄 str(exc) 原文与本地路径。
FIX-6 (AR-2.6): 单行解析失败返回 SECURITY_UNAVAILABLE 并继续服务后续请求。
FIX-8 (FR-06-03): inspect_file 遍历全部 sheet 并使用配置 max_scan_rows。
FIX-9 (BR-06.5): 上下文键名统一（camelCase/snake_case 均可消费）。
FIX-4 (FR-13): L3_ALLOW_AUDITED 授权检查侧一次性消费。
"""
from __future__ import annotations

import json
import sys
from typing import Any

from security.ai_operations_monitor import (
    AIOperationMonitor,
    DangerousOperationBlocked,
    check_tool_call,
)
from security.data_egress_guard import (
    DataRiskLevel,
    ClinicalDataDetector,
    StreamingScrubber,
)
from security.egress_authz import authorize_category, consume_category, is_authorizable
from security.egress_checkpoint import EgressViolation, check_egress
from security.local_data_inspector import LocalDataInspectionError, inspect_local_data
from security.data_egress_guard import scan_xlsx_sheet_safe
from security.patterns import clean_surrogates, sanitize_error
from security.smart_guard import (
    is_mass_data_dump,
    smart_scrub_structure,
    smart_scrub_text,
)


def _result(ok: bool, **fields: Any) -> dict[str, Any]:
    return {"ok": ok, **fields}


def _emit(response: dict[str, Any]) -> None:
    """安全写出响应行（真实故障修复：孤立代理字符）。

    DSH harness 的 JSON 可携带 \\udXXX 孤立代理（读 GBK 文件名等场景），
    直接 ensure_ascii=False 写 stdout 会抛 UnicodeEncodeError 杀死 worker，
    导致后续全部请求挂起。任何失败都不允许中断服务循环。
    """
    try:
        line = clean_surrogates(json.dumps(response, ensure_ascii=False, separators=(",", ":")))
        sys.stdout.write(line + "\n")
    except UnicodeEncodeError:
        sys.stdout.write(json.dumps(
            {"ok": False, "code": "SECURITY_UNAVAILABLE",
             "reason": "response encoding failure", "requestId": response.get("requestId")},
            ensure_ascii=True,
        ) + "\n")
    sys.stdout.flush()


class _XlrdSheetAdapter:
    """把 xlrd sheet 适配为 scan_xlsx_sheet_safe 需要的 openpyxl 风格接口。"""

    def __init__(self, sheet):
        self._sheet = sheet
        self.max_row = sheet.nrows

    def iter_rows(self, min_row=1, values_only=True):
        for index in range(min_row - 1, self._sheet.nrows):
            yield tuple(self._sheet.row_values(index))


class _XlrdWorkbook:
    """把 xlrd workbook 适配为 openpyxl 风格的 sheetnames/下标访问。"""

    def __init__(self, workbook):
        self._workbook = workbook

    @property
    def sheetnames(self):
        return self._workbook.sheet_names()

    def __getitem__(self, name):
        return _XlrdSheetAdapter(self._workbook.sheet_by_name(name))


def _scan_workbook(workbook, max_rows):
    """FIX-8 (FR-06-03): 遍历全部 sheet，使用配置 max_scan_rows，返回首个用户决策提示。"""
    prompt = None
    options = None
    for sheet_name in workbook.sheetnames:
        report = scan_xlsx_sheet_safe(workbook[sheet_name], sheet_name, max_rows=max_rows)
        found = report.get("user_prompt") or {}
        if found:
            prompt = found.get("message")
            options = found.get("options")
            break
    return prompt, options


def _normalize_context(context: Any) -> dict[str, Any]:
    """FIX-9: Node 侧 camelCase 上下文键名统一为 Python 消费侧 snake_case。"""
    if not isinstance(context, dict):
        return {}
    normalized = dict(context)
    for camel, snake in (("sessionId", "session_id"), ("userId", "user_id")):
        if snake not in normalized or normalized.get(snake) in (None, ""):
            normalized[snake] = normalized.get(camel)
    return normalized


def _handle(request: dict[str, Any]) -> dict[str, Any]:
    operation = request.get("operation")
    context = _normalize_context(request.get("context"))
    mode = str(context.get("mode", "enforce")).lower()

    if operation == "check_tool":
        try:
            tool_name = str(request.get("tool", ""))
            args = request.get("args", {})
            # The dedicated local metadata tool is itself the capability boundary:
            # it is allowed only in explicit UAT-local mode and only its worker
            # implementation can open local source files.
            if tool_name == "local_data_metadata":
                if context.get("localDataAccess") != "uat-local":
                    return _result(False, code="LOCAL_DATA_ACCESS_REQUIRED",
                                   reason="local metadata inspection is disabled")
                return _result(True, action="allow")
            # First apply the UAT capability policy, then the generic dangerous-
            # operation monitor. Generic shell/read/write access to source files
            # is never a substitute for local_data_metadata.
            monitor = AIOperationMonitor()
            try:
                monitor.check_local_data_policy(tool_name, args, context)
            except DangerousOperationBlocked as exc:
                return _result(False, code="LOCAL_DATA_ACCESS_REQUIRED",
                               audit_id=exc.audit_id, reason=exc.threat.reason)
            check_tool_call(tool_name, args, context)
            return _result(True, action="allow")
        except DangerousOperationBlocked as exc:
            if mode == "shadow":
                return _result(True, action="observed", audit_id=exc.audit_id,
                               reason=exc.threat.reason)
            return _result(False, code="DANGEROUS_OPERATION", audit_id=exc.audit_id,
                           reason=exc.threat.reason)

    if operation == "check_llm":
        payload = request.get("payload", {})
        try:
            evidence = check_egress(payload, context)
            return _result(True, action="allow", **evidence)
        except EgressViolation as exc:
            if mode == "shadow":
                return _result(True, action="observed", audit_id=exc.audit_id,
                               threats=len(exc.threats))
            # 白名单自愈车道（smart_guard 接线）：出域命中不再一刀切拦死——
            # 把载荷统一 token 化后放行。模型拿到 [SUBJ:xx]/[DATE:xx] 仍可
            # join/推理，真实值不出域；token 形态幂等，历史误报不再钉死会话。
            # 唯一硬红线：整表转储体量（is_mass_data_dump）仍拒绝。
            try:
                scrubbed, stats = smart_scrub_structure(payload)
            except Exception:
                # 脱敏自身失败 → fail-closed，维持原拦截判定。
                return _result(False, code="EGRESS_VIOLATION", audit_id=exc.audit_id,
                               threats=len(exc.threats))
            if is_mass_data_dump(stats):
                return _result(False, code="EGRESS_VIOLATION", audit_id=exc.audit_id,
                               threats=len(exc.threats),
                               reason="mass data dump: 数据行体量超阈值，拒绝出域")
            # token 化后的载荷以 shadow 口径复扫留痕：审计链呈现
            # BLOCKED(原文) → OBSERVED/ALLOWED(token 化后)，可用于调优，
            # 不再产生第二次阻断。
            try:
                evidence = check_egress(
                    scrubbed,
                    {**context, "mode": "shadow", "egress_stage": "post_scrub"},
                )
            except Exception:
                evidence = {}
            return _result(True, action="scrubbed", payload=scrubbed,
                           blocked_audit_id=exc.audit_id,
                           tokens_hashed=stats.tokens_hashed,
                           data_lines=stats.data_lines, **evidence)

    if operation == "inspect_local_data":
        if context.get("localDataAccess") != "uat-local":
            return _result(
                False,
                code="LOCAL_DATA_ACCESS_REQUIRED",
                reason="this UAT task requires the explicit local metadata inspection lane",
            )
        try:
            metadata = inspect_local_data(
                str(context.get("localDataRoot") or ""),
                str(request.get("path") or ""),
            )
        except LocalDataInspectionError as exc:
            return _result(False, code="LOCAL_DATA_INSPECTION_DENIED", reason=str(exc))
        except Exception as exc:
            return _result(False, code="SECURITY_UNAVAILABLE", reason=sanitize_error(exc))
        return _result(True, action="local-metadata", metadata=metadata)

    if operation == "scrub_row":
        detector = ClinicalDataDetector()
        scrubber = StreamingScrubber(detector)
        row = request.get("row", [])
        scrubbed, detection = scrubber.scrub_row(
            row if isinstance(row, list) else [row],
            int(request.get("row_index", 0)),
            bool(request.get("after_header", False)),
        )
        return _result(
            True,
            row=scrubbed,
            risk_level=detection.risk_level.name,
            patterns=detection.patterns_matched,
            evidence=detection.evidence,
            recommendation=detection.recommendation,
            needs_user=detection.risk_level == DataRiskLevel.SENSITIVE,
        )

    if operation == "scrub_text":
        # smart_guard 接线：白名单式统一 token 化取代旧 StreamingScrubber 逐行
        # 黑名单。散文/表头/路径原样，其余含数字值一律 token；与 check_llm
        # 出域自愈车道同一口径，写入/出域不再不对称。
        text = str(request.get("text", ""))
        scrubbed_text, stats = smart_scrub_text(text)
        needs_user = is_mass_data_dump(stats)
        return _result(
            True,
            text=scrubbed_text,
            scrubbed_rows=stats.lines_changed,
            data_lines=stats.data_lines,
            tokens_hashed=stats.tokens_hashed,
            needs_user=needs_user,
            user_prompt=(
                "数据安全检查：检测到整表级数据转储。选项：跳过（默认）/ 脱敏后继续 / 允许（需授权）"
                if needs_user else None
            ),
        )

    if operation == "inspect_file":
        path = str(request.get("path", ""))
        if not path.lower().endswith((".xlsx", ".xls")):
            return _result(True, needs_user=False)
        max_rows = max(1, int(request.get("max_scan_rows", 200)))
        try:
            if path.lower().endswith(".xls"):
                # 真实故障修复：openpyxl 不支持 .xls（真实 .xls 预检一律失败导致
                # agent 被拒）。.xls 走 xlrd 只读解析；SpreadsheetML XML 伪装的
                # .xls（如 *PROD.xls）由 xlrd 抛错进入 fail-closed。
                import xlrd
                workbook = xlrd.open_workbook(path, on_demand=True)
                try:
                    prompt, options = _scan_workbook(_XlrdWorkbook(workbook), max_rows)
                finally:
                    workbook.release_resources()
            else:
                import openpyxl
                workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
                try:
                    prompt, options = _scan_workbook(workbook, max_rows)
                finally:
                    workbook.close()
        except Exception as exc:
            # FIX-3: 异常原文不得回传（可能含本地路径/受试者标记/孤立代理）。
            return _result(False, code="SECURITY_UNAVAILABLE",
                           reason=sanitize_error(exc))
        return _result(
            True,
            needs_user=bool(prompt),
            user_prompt=prompt,
            options=options,
        )

    if operation == "authorize":
        category = str(request.get("category", ""))
        if not is_authorizable(category):
            return _result(False, code="INVALID_CATEGORY")
        record = authorize_category(
            request.get("root"),
            request.get("user"),
            request.get("session"),
            category,
            request.get("operator"),
        )
        return _result(bool(record.get("ok")), categories=record.get("categories", []))

    if operation == "ping":
        # FIX-11: 心跳探活——只需返回响应即证明 worker 存活。
        return _result(True, action="pong")

    if operation == "consume_authorization":
        # FIX-4 (FR-13): L3_ALLOW_AUDITED 仅当次有效——消费即移除。
        consumed = consume_category(
            request.get("root"),
            request.get("user"),
            request.get("session"),
            str(request.get("category", "L3_ALLOW_AUDITED")),
        )
        return _result(consumed)

    return _result(False, code="UNKNOWN_OPERATION")


def main() -> int:
    # 真实故障修复（P0）：zh-CN Windows 下 sys.stdin/stdout 默认 cp936，
    # Node 侧按 UTF-8 写入的中文请求会被解码为乱码并产生孤立代理字符，
    # 既使出域指纹 .encode('utf-8') 崩溃（全线 fail-closed 停摆），
    # 也会让中文检测在乱码上运行而静默失效。协议层强制 UTF-8，
    # 不依赖 PYTHONIOENCODING/PYTHONUTF8 外部环境变量。
    if hasattr(sys.stdin, "reconfigure"):
        sys.stdin.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    for line in sys.stdin:
        if not line.strip():
            continue
        request_id = None
        try:
            request = json.loads(line)
            if not isinstance(request, dict):
                raise ValueError("request must be an object")
            request_id = request.get("requestId")
            response = _handle(request)
        except Exception as exc:
            # FIX-3 / FIX-6 (AR-2.6): 畸形行不再让进程崩溃（UnboundLocalError 已修），
            # 返回 SECURITY_UNAVAILABLE 并继续服务后续请求。
            response = _result(False, code="SECURITY_UNAVAILABLE",
                               reason=sanitize_error(exc))
        response["requestId"] = request_id
        _emit(response)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
