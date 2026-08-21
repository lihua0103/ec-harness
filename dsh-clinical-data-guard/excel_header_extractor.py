#!/usr/bin/env python3
"""
excel_header_extractor.py

从 Excel/CSV 文件中识别表头区域，输出结构化 JSON，不输出数据区任何值。
被 Node 端 guard 插件通过子进程调用：
    python3 excel_header_extractor.py <file_path> [<sheet_name>] [--max-scan-rows N]

stdout: JSON  (见下方 schema)
stderr: 错误信息
exit 0: 成功
exit 1: 文件读取失败
exit 2: 依赖缺失
"""

import sys
import json
import re
import os
import argparse
from typing import Any

from security.patterns import (
    DATE_PATTERNS,
    NUMERIC_SUBJECT_ID_RE,
    SUBJECT_ID_PATTERNS,
    clean_surrogates,
    sanitize_error,
)

MAX_SCAN_ROWS_DEFAULT = 20

HEADER_DLP_PATTERNS = [*SUBJECT_ID_PATTERNS, *DATE_PATTERNS]

HEADER_SAFE_PATTERNS = [
    re.compile(r'\bDay\s*\d+\b', re.I),
    re.compile(r'\bWeek\s*\d+\b', re.I),
    re.compile(r'\bCycle\s*\d+\b', re.I),
    re.compile(r'\bVisit\s*\d+\b', re.I),
    re.compile(r'\bMonth\s*\d+\b', re.I),
    re.compile(r'\bHour\s*\d+\b', re.I),
    re.compile(r'\bBaseline\b', re.I),
    re.compile(r'\bScreening\b', re.I),
    re.compile(r'\bEOT\b', re.I),
    re.compile(r'\bEOS\b', re.I),
]


def _is_numeric(val: Any) -> bool:
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        try:
            float(val.replace(',', ''))
            return True
        except ValueError:
            return False
    return False


def _is_string_like(val: Any) -> bool:
    if val is None or val == '':
        return False
    if isinstance(val, str) and val.strip():
        return not _is_numeric(val)
    return False


def _score_row(row: list[Any], total_cols: int, merged_cols: set[int]) -> float:
    if not row or all(v is None or v == '' for v in row):
        return -10.0

    non_empty = [v for v in row if v is not None and v != '']
    if not non_empty:
        return -10.0

    score = 0.0

    str_count = sum(1 for v in non_empty if _is_string_like(v))
    num_count = sum(1 for v in non_empty if _is_numeric(v))

    if str_count == len(non_empty):
        score += 3.0
    elif num_count / len(non_empty) >= 0.5:
        # 真实故障修复（crViewer.xls 数据泄露）：数值占比 ==0.5 的行同样是数据行，
        # 边界必须扣分，否则无表头数据表的行全被判为"表头"整表泄露。
        score -= 4.0

    if len(non_empty) / max(total_cols, 1) >= 0.5:
        score += 2.0

    if merged_cols:
        score += 2.0

    row_text = ' '.join(str(v) for v in non_empty)
    for pattern, _ in HEADER_DLP_PATTERNS:
        match = pattern.search(row_text)
        if match:
            val = match.group()
            if not any(sp.search(val) for sp in HEADER_SAFE_PATTERNS):
                if re.search(r'\b\d{3,4}-\d{4,6}-\d{3,6}\b', val):
                    score -= 5.0
                else:
                    score -= 3.0

    return score


def _detect_orientation(scan_rows: list[list[Any]]) -> str:
    if not scan_rows or not scan_rows[0]:
        return 'VERTICAL'

    total_rows = len(scan_rows)
    col0_str = sum(
        1 for r in scan_rows
        if r and _is_string_like(r[0])
    )
    col0_str_ratio = col0_str / total_rows

    body_cells = 0
    body_numeric = 0
    for r in scan_rows:
        for v in r[1:]:
            body_cells += 1
            if _is_numeric(v):
                body_numeric += 1

    numeric_ratio = body_numeric / max(body_cells, 1)

    col0_text = ' '.join(str(r[0]) for r in scan_rows if r and r[0] is not None)
    for pattern, _ in HEADER_DLP_PATTERNS:
        if pattern.search(col0_text):
            return 'VERTICAL'

    if col0_str_ratio > 0.8 and numeric_ratio > 0.5:
        return 'HORIZONTAL'
    return 'VERTICAL'


def _find_header_end_row(scan_rows: list[list[Any]], merged_by_row: dict[int, set[int]]) -> int:
    total_cols = max((len(r) for r in scan_rows), default=1)
    for i, row in enumerate(scan_rows):
        score = _score_row(row, total_cols, merged_by_row.get(i, set()))
        if score < 0:
            return i
    return len(scan_rows)


def _looks_like_data_value(text: str) -> bool:
    """ST-P1-9: 表头输出白名单兜底——判断一个单元格值"像数据而非列名"。

    启发式打分可能把无表头数据表首行误判为表头，而其中不触发 DLP 模式的值
    （纯数值编号、测量值、含数字的短码）会被原值输出，形成泄露通道。列名极少
    是纯数字或"数字为主"的串，故这类值即便不命中 DLP 也一律 REDACTED。
    """
    # 纯数值（整数/小数/千分位）——列名几乎不会是纯数字，数据值常是。
    stripped = text.replace(',', '').replace(' ', '')
    try:
        float(stripped)
        return True
    except ValueError:
        pass
    # 数字占比过半的短码（如 A12、3-5mg）——数据编号形态，非列名词。
    digits = sum(c.isdigit() for c in text)
    if len(text) <= 12 and digits >= 1 and digits / max(len(text), 1) >= 0.4:
        return True
    return False


def _dlp_scan_cell(val: Any) -> tuple[Any, str | None]:
    if val is None:
        return val, None
    text = str(val).strip()
    if not text:
        return val, None
    if NUMERIC_SUBJECT_ID_RE.fullmatch(text):
        return '[REDACTED:SUBJECT_ID]', 'SUBJECT_ID'
    for pattern, label in HEADER_DLP_PATTERNS:
        m = pattern.search(text)
        if m:
            matched = m.group()
            if any(sp.search(matched) for sp in HEADER_SAFE_PATTERNS):
                continue
            return f'[REDACTED:{label}]', label
    # ST-P1-9: DLP 未命中也要挡住"像数据值"的单元格，防止误判表头时原值泄露。
    if _looks_like_data_value(text) and not any(
        sp.search(text) for sp in HEADER_SAFE_PATTERNS
    ):
        return '[REDACTED:DATA_VALUE]', 'DATA_VALUE'
    return val, None


def _extract_merged_info(ws) -> dict[int, set[int]]:
    merged_by_row: dict[int, set[int]] = {}
    try:
        for rng in ws.merged_cells.ranges:
            for row_idx in range(rng.min_row - 1, rng.max_row):
                merged_by_row.setdefault(row_idx, set())
                for col_idx in range(rng.min_col - 1, rng.max_col):
                    merged_by_row[row_idx].add(col_idx)
    except Exception:
        pass
    return merged_by_row


def process_xlsx(file_path: str, sheet_name: str | None, max_scan: int) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print('missing openpyxl: pip install openpyxl', file=sys.stderr)
        sys.exit(2)

    wb_ro = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = [sheet_name] if sheet_name else wb_ro.sheetnames

    # P3: 一次性加载merged cells信息（循环外），避免每sheet重复开销
    merged_by_file: dict[str, dict[int, set[int]]] = {}
    try:
        wb_merge = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        for sn in wb_merge.sheetnames:
            merged_by_file[sn] = _extract_merged_info(wb_merge[sn])
        wb_merge.close()
    except Exception:
        pass

    results = []

    for sname in sheet_names:
        if sname not in wb_ro.sheetnames:
            continue
        ws_ro = wb_ro[sname]
        total_rows = ws_ro.max_row or 0
        total_cols = ws_ro.max_column or 0

        scan_rows: list[list[Any]] = []
        # P2: 不依赖 total_rows（read_only模式常返回None/0），直接用max_scan上限
        for row in ws_ro.iter_rows(min_row=1, max_row=max_scan, values_only=True):
            scan_rows.append(list(row))
            if len(scan_rows) >= max_scan:
                break

        # P0修复: merged_by_row从预加载的字典取，不在循环内重新加载workbook
        merged_by_row = merged_by_file.get(sname, {})

        orientation = _detect_orientation(scan_rows)

        header_cells = []
        redacted = []
        warnings_list = []

        if orientation == 'HORIZONTAL':
            for ri, row in enumerate(scan_rows):
                if not row:
                    continue
                raw = row[0]
                cleaned, label = _dlp_scan_cell(raw)
                if label:
                    redacted.append({'row': ri, 'col': 0, 'type': label})
                    warnings_list.append(f'row {ri} col 0: DLP hit {label}')
                if cleaned is not None and str(cleaned).strip():
                    header_cells.append({'row': ri, 'col': 0, 'value': str(cleaned)})
            if scan_rows:
                for ci, val in enumerate(scan_rows[0][1:], start=1):
                    cleaned, label = _dlp_scan_cell(val)
                    if label:
                        redacted.append({'row': 0, 'col': ci, 'type': label})
                    if cleaned is not None and str(cleaned).strip():
                        header_cells.append({'row': 0, 'col': ci, 'value': str(cleaned)})
            header_rows = list(range(len(scan_rows)))
            data_start_row = 1
        else:
            header_end = _find_header_end_row(scan_rows, merged_by_row)
            header_rows = list(range(header_end))
            data_start_row = header_end

            for ri in range(header_end):
                row = scan_rows[ri] if ri < len(scan_rows) else []
                for ci, val in enumerate(row):
                    cleaned, label = _dlp_scan_cell(val)
                    if label:
                        redacted.append({'row': ri, 'col': ci, 'type': label})
                        warnings_list.append(f'row {ri} col {ci}: DLP hit {label}')
                    if cleaned is not None and str(cleaned).strip():
                        header_cells.append({'row': ri, 'col': ci, 'value': str(cleaned)})

        results.append({
            'sheet': sname,
            'orientation': orientation,
            'header_rows': header_rows,
            'data_start_row': data_start_row,
            'total_rows': total_rows,
            'total_cols': total_cols,
            'header_cells': header_cells,
            'redacted_in_header': redacted,
            'warnings': warnings_list,
        })

    # P0修复: wb_ro.close() 移到循环外，不在每次迭代关闭
    wb_ro.close()
    return results


def process_csv(file_path: str, max_scan: int) -> list[dict]:
    import csv

    with open(file_path, newline='', encoding='utf-8-sig', errors='replace') as f:
        reader = csv.reader(f)
        scan_rows = []
        total_rows = 0
        total_cols = 0
        for row in reader:
            total_rows += 1
            col_count = len(row)
            if col_count > total_cols:
                total_cols = col_count
            if len(scan_rows) < max_scan:
                scan_rows.append(row)
    orientation = _detect_orientation(scan_rows)

    header_cells = []
    redacted = []
    warnings_list = []

    if orientation == 'HORIZONTAL':
        header_rows = list(range(len(scan_rows)))
        data_start_row = 1
        for ri, row in enumerate(scan_rows):
            if not row:
                continue
            cleaned, label = _dlp_scan_cell(row[0])
            if label:
                redacted.append({'row': ri, 'col': 0, 'type': label})
                warnings_list.append(f'row {ri} col 0: DLP hit {label}')
            if cleaned is not None and str(cleaned).strip():
                header_cells.append({'row': ri, 'col': 0, 'value': str(cleaned)})
        for ci, val in enumerate(scan_rows[0][1:] if scan_rows else [], start=1):
            cleaned, label = _dlp_scan_cell(val)
            if label:
                redacted.append({'row': 0, 'col': ci, 'type': label})
            if cleaned is not None and str(cleaned).strip():
                header_cells.append({'row': ri, 'col': ci, 'value': str(cleaned)})
    else:
        header_end = _find_header_end_row(scan_rows, {})
        header_rows = list(range(min(header_end, len(scan_rows))))
        data_start_row = header_end
        for ri in range(min(header_end, len(scan_rows))):
            for ci, val in enumerate(scan_rows[ri]):
                cleaned, label = _dlp_scan_cell(val)
                if label:
                    redacted.append({'row': ri, 'col': ci, 'type': label})
                    warnings_list.append(f'row {ri} col {ci}: DLP hit {label}')
                if cleaned is not None and str(cleaned).strip():
                    header_cells.append({'row': ri, 'col': ci, 'value': str(cleaned)})

    return [{
        'sheet': os.path.basename(file_path),
        'orientation': orientation,
        'header_rows': header_rows,
        'data_start_row': data_start_row,
        'total_rows': total_rows,
        'total_cols': total_cols,
        'header_cells': header_cells,
        'redacted_in_header': redacted,
        'warnings': warnings_list,
    }]


def process_xls(file_path: str, sheet_name: str | None, max_scan: int) -> list[dict]:
    """FIX-8 (FR-06-03 / TC-15): .xls 经 xlrd 只读解析，交付与 .xlsx 相同的表头结构。

    xlrd 缺失时 fail-closed（退出码 2），不静默回退到数据读取。
    """
    try:
        import xlrd
    except ImportError:
        print('missing xlrd: pip install xlrd', file=sys.stderr)
        sys.exit(2)

    workbook = xlrd.open_workbook(file_path, on_demand=True, formatting_info=False)
    try:
        names = [sheet_name] if sheet_name else workbook.sheet_names()
        results = []
        for sname in names:
            if sname not in workbook.sheet_names():
                continue
            sheet = workbook.sheet_by_name(sname)
            total_rows = sheet.nrows
            total_cols = sheet.ncols
            scan_rows = [
                [sheet.cell_value(r, c) if c < total_cols else None
                 for c in range(total_cols)]
                for r in range(min(total_rows, max_scan))
            ]

            orientation = _detect_orientation(scan_rows)
            header_cells = []
            redacted = []
            warnings_list = []

            if orientation == 'HORIZONTAL':
                for ri, row in enumerate(scan_rows):
                    if not row:
                        continue
                    raw = row[0]
                    cleaned, label = _dlp_scan_cell(raw)
                    if label:
                        redacted.append({'row': ri, 'col': 0, 'type': label})
                        warnings_list.append(f'row {ri} col 0: DLP hit {label}')
                    if cleaned is not None and str(cleaned).strip():
                        header_cells.append({'row': ri, 'col': 0, 'value': str(cleaned)})
                if scan_rows:
                    for ci, val in enumerate(scan_rows[0][1:], start=1):
                        cleaned, label = _dlp_scan_cell(val)
                        if label:
                            redacted.append({'row': 0, 'col': ci, 'type': label})
                        if cleaned is not None and str(cleaned).strip():
                            header_cells.append({'row': 0, 'col': ci, 'value': str(cleaned)})
                header_rows = list(range(len(scan_rows)))
                data_start_row = 1
            else:
                header_end = _find_header_end_row(scan_rows, {})
                header_rows = list(range(header_end))
                data_start_row = header_end
                for ri in range(header_end):
                    row = scan_rows[ri] if ri < len(scan_rows) else []
                    for ci, val in enumerate(row):
                        cleaned, label = _dlp_scan_cell(val)
                        if label:
                            redacted.append({'row': ri, 'col': ci, 'type': label})
                            warnings_list.append(f'row {ri} col {ci}: DLP hit {label}')
                        if cleaned is not None and str(cleaned).strip():
                            header_cells.append({'row': ri, 'col': ci, 'value': str(cleaned)})

            results.append({
                'sheet': sname,
                'orientation': orientation,
                'header_rows': header_rows,
                'data_start_row': data_start_row,
                'total_rows': total_rows,
                'total_cols': total_cols,
                'header_cells': header_cells,
                'redacted_in_header': redacted,
                'warnings': warnings_list,
            })
        return results
    finally:
        workbook.release_resources()


def main():
    # 真实故障修复（P0 同源）：zh-CN Windows 默认 cp936 会让中文表头输出
    # 乱码并被 Node 侧按 UTF-8 误读，协议层强制 UTF-8。
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description='Extract header structure from Excel/CSV for DSH guard')
    parser.add_argument('file_path', help='Path to .xlsx / .xls / .csv file')
    parser.add_argument('sheet', nargs='?', default=None, help='Sheet name (xlsx only)')
    parser.add_argument('--max-scan-rows', type=int, default=MAX_SCAN_ROWS_DEFAULT)
    args = parser.parse_args()

    fp = args.file_path
    if not os.path.isfile(fp):
        # FIX-3 (AR-2.9): stderr 不回显原始路径（可能含受试者标记）。
        print(f'file not found: [PATH]', file=sys.stderr)
        sys.exit(1)

    ext = os.path.splitext(fp)[1].lower()
    try:
        if ext == '.xlsx':
            sheets = process_xlsx(fp, args.sheet, args.max_scan_rows)
        elif ext == '.xls':
            sheets = process_xls(fp, args.sheet, args.max_scan_rows)
        elif ext == '.csv':
            sheets = process_csv(fp, args.max_scan_rows)
        else:
            print(f'unsupported extension: {ext}', file=sys.stderr)
            sys.exit(1)
    except SystemExit:
        raise
    except Exception as e:
        # FIX-3 (AR-2.9): 异常原文统一脱敏后再写 stderr。
        print(sanitize_error(e), file=sys.stderr)
        sys.exit(1)

    # FIX-3: 输出 file 字段对文件名做脱敏（文件名可能含受试者标记/孤立代理）。
    output = json.dumps(
        {'file': sanitize_error(os.path.basename(fp)), 'sheets': sheets},
        ensure_ascii=False,
    )
    # 真实故障修复：cell/sheet 名可携带孤立代理（\udXXX），UTF-8 stdout 直接崩溃。
    print(clean_surrogates(output))


if __name__ == '__main__':
    main()
