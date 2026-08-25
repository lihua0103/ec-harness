"""2026-08-22 E2E 第二轮审计 E-1~E-4 修复的回归测试。

对应 docs/EMERALD_LISTING_E2E_AUDIT_20260822.md：
- E-1 worker 导入失败不再以 UnboundLocalError 掩盖真因；
- E-2 worker 依赖预检 fail-fast + Node 侧横幅捕获；
- E-3 归档打不开降级为结构化 missing（credential:<名>）+ 拒绝码透传；
- E-4 execute 前清扫历史残留临时目录。
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path
from unittest import mock

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from security import worker
from security import spec_parser
from security import listing_inspector
from security.listing_inspector import inspect_listing_context
from security.listing_workflow import _sweep_stale_transient, inspect_listing


def _als_xlsx(path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ALS"
    sheet.append(["DatasetName", "ItemName", "PreText", "ItemOrder"])
    sheet.append(["DM", "USUBJID", "Subject", 1])
    sheet.append(["DM", "AGE", "Age", 2])
    workbook.save(path)


def _xpt(path: Path) -> None:
    import pandas as pd
    import pyreadstat

    frame = pd.DataFrame({"USUBJID": ["101-001-0001", "101-001-0002"], "AGE": [42, 57]})
    pyreadstat.write_xport(
        frame, str(path),
        column_labels={"USUBJID": "Subject ID", "AGE": "Age"},
        table_name="DM",
    )


def _project_with_locked_archive(root: Path) -> Path:
    """带 1 个打不开归档 + 明文数据集 + ALS 的最小项目。"""
    project = root / "study"
    (project / "doc").mkdir(parents=True)
    (project / "data").mkdir()
    _als_xlsx(project / "doc" / "ALS.xlsx")
    _xpt(project / "data" / "DM.xpt")
    # 损坏 ZIP：extract_dataset_archive 抛 PathPolicyError("archive extraction failed")
    (project / "locked.zip").write_bytes(b"PK\x03\x04 not a real archive payload")
    return project


# --- E-1 ---------------------------------------------------------------------

def test_worker_import_failure_returns_structured_error() -> None:
    """E-1: listing 栈导入失败必须回 LISTING_STACK_UNAVAILABLE，而非 UnboundLocalError。"""
    saved = sys.modules.get("security.listing_workflow")
    sys.modules["security.listing_workflow"] = None  # from-import 触发 ImportError
    try:
        response = worker._handle({
            "operation": "listing_inspect",
            "project": "study",
            "scenario": "rbqm",
            "context": {"localDataAccess": "uat-local", "localDataRoot": "unused"},
        })
    finally:
        if saved is None:
            sys.modules.pop("security.listing_workflow", None)
        else:
            sys.modules["security.listing_workflow"] = saved
    assert response["ok"] is False
    assert response["code"] == "LISTING_STACK_UNAVAILABLE"
    assert "listing_workflow" in response["reason"], response


# --- E-2 ---------------------------------------------------------------------

def test_missing_worker_dependencies_reports_missing_names() -> None:
    import importlib

    real = importlib.import_module

    def fake_importer(name: str):
        if name == "pyreadstat":
            raise ImportError("No module named 'pyreadstat'", name="pyreadstat")
        return real(name)

    assert worker.missing_worker_dependencies(fake_importer) == ["pyreadstat"]
    # 运行环境（requirements.txt 齐备的 venv）必须通过预检。
    assert worker.missing_worker_dependencies() == []


def test_worker_protocol_smoke_after_preflight() -> None:
    """依赖预检通过后协议层不受影响：ping → pong。"""
    proc = subprocess.Popen(
        [sys.executable, "-m", "security.worker"],
        cwd=ROOT, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL, text=True, encoding="utf-8",
    )
    try:
        proc.stdin.write(json.dumps({"requestId": "r1", "operation": "ping"}) + "\n")
        proc.stdin.flush()
        response = json.loads(proc.stdout.readline())
        assert response["ok"] is True and response["action"] == "pong"
    finally:
        proc.stdin.close()
        proc.wait(timeout=30)


def _run_node(script: str, extra_env: dict[str, str] | None = None) -> dict:
    env = {**os.environ, "PYTHON": sys.executable, "PLUGIN_PYTHON": sys.executable}
    for key, value in (extra_env or {}).items():
        env[key] = value
    result = subprocess.run(
        ["node", "--input-type=module", "-e", script],
        cwd=ROOT, env=env, capture_output=True, text=True,
        encoding="utf-8", errors="replace", timeout=90,
    )
    assert result.returncode == 0, result.stderr
    return json.loads(result.stdout)


def test_worker_dependency_banner_fails_fast_with_actionable_reason() -> None:
    """E-2: 缺依赖 worker 启动即横幅退出，Node 捕获并把真因并入错误信息。"""
    with tempfile.TemporaryDirectory() as directory:
        fake = Path(directory) / "site"
        (fake / "pandas").mkdir(parents=True)
        (fake / "pandas" / "__init__.py").write_text(
            "raise ImportError(\"No module named 'pandas' (stub)\", name='pandas')",
            encoding="utf-8",
        )
        previous = os.environ.get("PYTHONPATH")
        os.environ["PYTHONPATH"] = str(fake)
        try:
            output = _run_node(
                "import { SecurityRuntime } from './src/index.js';\n"
                "const rt = new SecurityRuntime({});\n"
                "await new Promise((resolve) => setTimeout(resolve, 5000));\n"
                "const notice = rt.startupNotice;\n"
                "let failure = null;\n"
                "try {\n"
                "  await rt.request({ operation: 'ping' }, { timeoutMs: 15000 });\n"
                "} catch (error) {\n"
                "  failure = error.message;\n"
                "}\n"
                "process.stdout.write(JSON.stringify({\n"
                "  failure,\n"
                "  noticeCode: notice ? notice.code : null,\n"
                "  noticeReason: notice ? notice.reason : null,\n"
                "  broken: rt.broken === true,\n"
                "}));\n"
                "rt.dispose();\n"
            )
        finally:
            if previous is None:
                os.environ.pop("PYTHONPATH", None)
            else:
                os.environ["PYTHONPATH"] = previous
        assert output["noticeCode"] == "WORKER_DEPENDENCY_MISSING", output
        assert "pandas" in output["noticeReason"], output
        assert output["failure"], "依赖缺失时请求必须被拒绝（fail-closed）"
        assert "WORKER_DEPENDENCY_MISSING" in output["failure"], output
        assert output["broken"] is True


# --- E-3 ---------------------------------------------------------------------

def test_inspect_degrades_unreadable_archive_to_structured_missing() -> None:
    """E-3: 单个打不开的归档不得杀死 inspect；明文数据集照常入 schema。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = _project_with_locked_archive(root)
        result = inspect_listing(
            local_data_root=str(root), project=project.name, scenario="rbqm")
        assert result["clinicalGuard"] == "CLINICAL_LISTING_INSPECTION"
        assert "dm" in result["schema"], "明文数据集应照常进入 schema"
        assert "credential:locked.zip" in result["missing"], result["missing"]
        assert "a local data archive could not be opened" in result["warnings"]


def test_inspection_receipt_preserves_all_spec_definitions() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = root / "study"
        (project / "doc").mkdir(parents=True)
        _als_xlsx(project / "doc" / "ALS.xlsx")

        result = inspect_listing_context(str(root), "study", "rbqm")

        als = next(document for document in result["documents"] if document["kind"] == "als")
        assert len(als["content"]["mappings"]) > 1
        assert "a specification definition limit was reached" not in result["warnings"]
        assert result["dataClass"] == "METADATA_ONLY"


def test_worker_listing_inspect_surfaces_missing_archive() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = _project_with_locked_archive(root)
        response = worker._handle({
            "operation": "listing_inspect",
            "project": project.name,
            "scenario": "rbqm",
            "context": {"localDataAccess": "uat-local", "localDataRoot": str(root)},
        })
        assert response["ok"] is True, response
        assert "credential:locked.zip" in response["inspection"]["missing"]


def test_worker_preserves_structured_workflow_rejection_code() -> None:
    """E-3: 具体拒绝（凭据目录未配置）必须带结构化 code 与可行动文案。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        _project_with_locked_archive(root)
        response = worker._handle({
            "operation": "listing_inspect",
            "project": "study",
            "scenario": "rbqm",
            "credentialRef": "proj-pass.txt",
            "context": {"localDataAccess": "uat-local", "localDataRoot": str(root)},
        })
        assert response["ok"] is False
        assert response["code"] == "CREDENTIALS_DIR_NOT_CONFIGURED", response
        assert response["reason"] == "credentials directory is not configured"


# --- Real specification/scenario defects ------------------------------------

def _spec_xlsx(path: Path, title: str = "Report Listing Specification") -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(["Requirement ID", "Requirement"])
    sheet.append(["R-1", "Report listing of DM status"])
    workbook.save(path)
    workbook.close()


def test_report_specification_beats_als_header_heuristic() -> None:
    with tempfile.TemporaryDirectory() as directory:
        path = Path(directory) / "RT01_DM Status Report Specification_11Aug2026.xlsx"
        _spec_xlsx(path)
        assert spec_parser.classify_spec_document(path) == "specification"


def test_doc_files_are_requirement_documents_even_when_names_resemble_support_data() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = root / "CGB3002-TEST"
        (project / "doc").mkdir(parents=True)
        names = (
            "RT01_DM Status Report Specification_11Aug2026.xlsx",
            "Page_Details.xlsx",
            "Query_Details.xlsx",
            "Coding Results.xlsx",
            "Listing_Export.xlsx",
            "Report_Details.xlsx",
        )
        for name in names:
            _spec_xlsx(project / "doc" / name)
        kinds = {path.name: spec_parser.classify_spec_document(path) for path in spec_parser.find_spec_documents(project)}
        assert kinds["RT01_DM Status Report Specification_11Aug2026.xlsx"] == "specification"
        assert all(kinds[name] in {"specification", "als", "template", "requirement_note"} for name in names)
        result = inspect_listing_context(str(root), project.name, None)
        assert {item["name"] for item in result["documents"]} == {f"doc/{name}" for name in names}
        assert result["supportData"] == []
        assert all("R-1" in str(item) for item in result["documents"])


def test_support_data_is_discovered_outside_doc() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = root / "CGB3002-TEST"
        (project / "doc").mkdir(parents=True)
        (project / "report_data").mkdir()
        _spec_xlsx(project / "doc" / "Status Specification.xlsx")
        _spec_xlsx(project / "report_data" / "Page_Details.xlsx")
        assert spec_parser.classify_spec_document(project / "report_data" / "Page_Details.xlsx") == "report_support_data"


def test_rbqm_and_manual_scenarios_are_inferred_without_input() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        for name, expected in (("RBQM KRI Specification.xlsx", "rbqm"), ("Manual Review Specification.xlsx", "manual")):
            project = root / expected
            (project / "doc").mkdir(parents=True)
            _spec_xlsx(project / "doc" / name)
            result = inspect_listing_context(str(root), project.name, None)
            assert result["scenario"] == expected
            assert result["inferredScenario"] == expected


def test_xls_report_support_files_are_discovered() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "CGB3002-TEST"
        (project / "doc").mkdir(parents=True)
        for name in ("RT01_V1.0_29JUN2026_PROD.xls", "crViewer.xls"):
            (project / "doc" / name).write_bytes(b"placeholder")
        assert {path.name for path in spec_parser.find_spec_documents(project)} == {
            "RT01_V1.0_29JUN2026_PROD.xls", "crViewer.xls"
        }


def test_doc_xls_specification_is_fully_parsed() -> None:
    """2026-08-24 红线口径：doc/ 内 .xls 全量解析，不再 warning-only。

    此前 inspect 只解析 .xlsx，doc 内 .xls 规格文件返回空内容 + 一条
    warning——AI 拿不到任何需求文本。现在走 xlrd 全量 requirement 行解析。
    """
    import xlwt

    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = root / "CGB3002-TEST"
        (project / "doc").mkdir(parents=True)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("Requirements")
        sheet.write(0, 0, "KRI编号")
        sheet.write(0, 1, "计算方式")
        sheet.write(1, 0, "KRI-001")
        sheet.write(1, 1, "失访人数/入组人数")
        workbook.save(str(project / "doc" / "RT01_Requirements_PROD.xls"))

        result = inspect_listing_context(str(root), project.name, None)
        documents = [item for item in result["documents"] if item["name"].endswith(".xls")]
        assert len(documents) == 1, ".xls 规格文件未进入 documents"
        requirements = documents[0]["content"]["requirements"]
        flat = [cell for item in requirements for cell in item["cells"]]
        assert "KRI-001" in flat, ".xls 规格正文未被全量解析"
        assert "失访人数/入组人数" in flat, ".xls 规格正文未被全量解析"
        assert "a specification document could not be parsed" not in result["warnings"]


def test_doc_txt_sidecar_is_never_parsed_into_documents() -> None:
    """doc/ 内 .txt 可能是密码 sidecar（如 A1234567.txt），绝不进 documents。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = root / "GQ9001"
        (project / "doc").mkdir(parents=True)
        (project / "doc" / "A1234567.txt").write_text("SECRET-PASSWORD-123", encoding="utf-8")
        result = inspect_listing_context(str(root), project.name, None)
        assert result["documents"] == [], ".txt sidecar 不应被解析"
        serialized = json.dumps(result, ensure_ascii=False)
        assert "SECRET-PASSWORD-123" not in serialized, "凭据 sidecar 泄漏进 inspect 收据"


def test_real_project_filenames_map_to_expected_roles_and_scenarios() -> None:
    cases = (
        ("test_Final.xlsx", "specification", "rbqm"),
        ("Manual Review.xlsx", "specification", "manual"),
        ("Data Validation Plan.xlsx", "specification", "manual"),
        ("数据核查计划.xlsx", "specification", "manual"),
        ("H301中文报表明细.xlsx", "report_support_data", "report"),
        ("odm.xlsx", "report_support_data", "report"),
    )
    for name, expected_kind, expected_scenario in cases:
        path = Path(name)
        assert spec_parser.classify_spec_document(path) == expected_kind
        assert spec_parser.classify_spec_document(path) != "als" or expected_scenario != "report"
        if expected_kind == "specification":
            assert listing_inspector._infer_scenario([path])[0] == expected_scenario


def test_manual_validation_plan_wins_report_keyword() -> None:
    paths = [
        Path("CGB3002-RT01_Data Validation Plan_Manual Review_V1.1_20260715.xlsx"),
        Path("YL202-CN-301-01_数据核查计划_V1.0_20260226_for programming.xlsx"),
    ]
    assert all(listing_inspector._infer_scenario([path])[0] == "manual" for path in paths)


def test_scenario_is_inferred_and_conflicts_are_structured() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = root / "CGB3002-TEST"
        (project / "doc").mkdir(parents=True)
        _spec_xlsx(project / "doc" / "RT01_DM Status Report Specification.xlsx")
        report = inspect_listing_context(str(root), project.name, None)
        assert report["inferredScenario"] == "report"
        assert report["scenarioConfidence"] > 0
        _spec_xlsx(project / "doc" / "Medical Listing Specification.xlsx")
        conflict = inspect_listing_context(str(root), project.name, None)
        assert set(conflict["scenarioCandidates"]) == {"medical", "report"}
        assert conflict["scenarioConfidence"] == 0.0


# --- E-4 ---------------------------------------------------------------------

def test_sweep_stale_transient_removes_leftover_tmp_and_staging() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "study"
        output = project / ".clinical-listing" / "output"
        staging = project / ".clinical-listing" / "staging"
        (output / ".rbqm-tmp-26041d78").mkdir(parents=True)
        (output / ".rbqm-tmp-26041d78" / "MEDICAL_001.xlsx").write_bytes(b"x")
        (output / ".rbqm-backup-deadbeef").mkdir()
        (staging / "deadbeef").mkdir(parents=True)
        keep = output / "rbqm"
        keep.mkdir()
        (keep / "RBQM_001.xlsx").write_bytes(b"x")
        _sweep_stale_transient(project, ".clinical-listing/output")
        assert not (output / ".rbqm-tmp-26041d78").exists()
        assert not (output / ".rbqm-backup-deadbeef").exists()
        assert not staging.exists() or not any(staging.iterdir())
        assert (keep / "RBQM_001.xlsx").exists(), "正式发布产物不得被清扫"


def main() -> int:
    tests = [value for key, value in sorted(globals().items()) if key.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"PASS {test.__name__}")
        except Exception as error:
            failures += 1
            print(f"FAIL {test.__name__}: {error}")
    print(f"RESULT {len(tests) - failures}/{len(tests)}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
