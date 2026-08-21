from __future__ import annotations

import json
import hashlib
import re
import os
import subprocess
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)
# FIX-12 后审计/授权默认写入用户主目录；测试显式指回项目 var 以便断言审计内容。
os.environ.setdefault("EMERALD_AUDIT_ROOT", str(ROOT / "var" / "egress_audit"))

from security.ai_operations_monitor import DangerousOperationBlocked, check_bash, check_python_code, check_tool_call  # noqa: E402
from security.audit_log import write_audit_record  # noqa: E402
from security.data_egress_guard import ClinicalDataDetector, DataRiskLevel, StreamingScrubber  # noqa: E402
from security.egress_authz import authorize_category, consume_category  # noqa: E402
from security.egress_checkpoint import EgressCheckpoint, EgressViolation, check_egress  # noqa: E402
from security.patterns import sanitize_error  # noqa: E402

CASES = []


def case(func):
    CASES.append(func)
    return func


@case
def subject_id_blocks():
    try:
        check_egress([{"role": "user", "content": "Subject A1234567"}])
    except EgressViolation as exc:
        assert exc.audit_id
        return
    raise AssertionError("字母前缀受试者编号未拦截")


@case
def lowercase_subject_and_sas_date_block():
    """ST-P1-1 红队：小写受试者号/SAS 日期不得绕过检测（检测与脱敏口径统一）。"""
    for content in ["subject a1234567", "visit 01jan2024", "pt: 10001234"]:
        try:
            check_egress([{"role": "user", "content": content}])
        except EgressViolation:
            continue
        raise AssertionError(f"小写绕过未拦截: {content}")


@case
def invisible_and_nfkc_bypass_block():
    """ST-P1-2 红队：全角数字/软连字符/BOM/bidi 插入不得切断受试者号检测。"""
    for content in [
        "Subject Ａ１２３４５６７",  # 全角 A1234567
        "A123­4567",   # 软连字符
        "A﻿1234567",   # BOM
        "A​1234567",   # 零宽空格
    ]:
        try:
            check_egress([{"role": "user", "content": content}])
        except EgressViolation:
            continue
        raise AssertionError(f"不可见字符/NFKC 绕过未拦截: {content!r}")


@case
def platform_hyphenated_identifiers_are_not_subject_ids():
    """真实 DSH 请求头回归：平台 kebab-case 标识不得被 USUBJID 模式误杀。"""
    payload = {
        "messages": [{"role": "user", "content": "读取项目需求文档。"}],
        "sessionId": "session-6ad2d0e2-c715-4a38-bdc2-2def4bdfe804",
        "system": "Current fs-observation-policy uses modification-time-ordered files.",
        "tools": [{
            "name": "pwsh",
            "description": "Run with danger-full-access only after approval.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sandbox_permissions": {
                        "type": "string",
                        "enum": ["workspace-write", "danger-full-access"],
                    }
                },
            },
        }],
    }
    check_egress(payload)

    # 仅 GenerateOptions 顶层 sessionId 是协议元数据；嵌套同名业务字段不能借此绕过。
    try:
        check_egress({"messages": [{"role": "user", "content": "正常请求"}],
                      "metadata": {"sessionId": "A1234567"}})
    except EgressViolation:
        pass
    else:
        raise AssertionError("嵌套 sessionId 绕过临床数据扫描")

    # 收紧误报不能削弱真实复合 USUBJID 检测。
    for sample in [
        "STUDY001-SITE01-SUBJ001",
        "STUDY-SITE-001",
        "ABC-DEF-001",
    ]:
        try:
            check_egress([{"role": "user", "content": sample}])
        except EgressViolation:
            continue
        raise AssertionError(f"真实 USUBJID 未拦截: {sample}")


@case
def site_subject_and_date_composite_blocks():
    try:
        check_egress([{"role": "user", "content": "101-001234 on 2024-03-05"}])
    except EgressViolation:
        return
    raise AssertionError("复合临床信号未拦截")


@case
def base64_payload_blocks():
    import base64
    encoded = base64.b64encode("USUBJID: A1234567".encode()).decode()
    try:
        check_egress([{"role": "user", "content": encoded}])
    except EgressViolation:
        return
    raise AssertionError("base64 临床载荷未拦截")


@case
def full_generate_options_fields_are_scanned():
    clean_messages = [{"role": "user", "content": "请生成列表规范。"}]
    auxiliary_fields = {
        "system": "Subject A1234567",
        "tools": [{"name": "demo", "description": "Subject A1234567", "parameters": {}}],
        "stop": ["Subject A1234567"],
    }
    for field, value in auxiliary_fields.items():
        try:
            check_egress({"messages": clean_messages, field: value})
        except EgressViolation:
            continue
        raise AssertionError(f"完整模型请求字段未拦截: {field}")


@case
def clean_model_request_audit_keeps_only_fingerprint():
    payload = {
        "provider": "test-provider",
        "model": "test-model",
        "messages": [{"role": "user", "content": "clean request evidence"}],
        "system": "clean system evidence",
        "tools": [{"name": "demo", "description": "clean tool", "parameters": {}}],
        "stop": ["finish"],
    }
    canonical = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    with tempfile.TemporaryDirectory() as directory:
        checkpoint = EgressCheckpoint(directory)
        evidence = checkpoint.check(payload)
        assert evidence["payload_sha256"] == hashlib.sha256(canonical).hexdigest()
        assert evidence["payload_bytes"] == len(canonical)
        assert evidence["message_count"] == 1
        audit_path = next(Path(directory).glob("*.jsonl"))
        blob = audit_path.read_text(encoding="utf-8")
        assert "clean request evidence" not in blob
        assert "clean system evidence" not in blob


@case
def filename_date_is_allowed():
    check_egress([{"role": "user", "content": "report_v2024-08-18.xlsx"}])


@case
def dangerous_tools_and_bash_block():
    for tool, args in [
        ("read_expected_output", {}),
        ("bash", {"command": "cat data.sas7bdat"}),
    ]:
        try:
            check_tool_call(tool, args)
        except DangerousOperationBlocked:
            continue
        raise AssertionError(f"危险操作未拦截: {tool}")


@case
def dynamic_exec_and_quote_split_bypass_block():
    """ST-P1-8 红队：__import__/eval/exec/getattr/marshal 与 shell 引号拼接不得绕过。"""
    for code in [
        "__import__('pickle').load(open('x','rb'))",
        "eval('1+1')",
        "exec('x=1')",
        "getattr(o,'load')()",
        "import marshal\nmarshal.load(f)",
    ]:
        try:
            check_python_code(code)
        except DangerousOperationBlocked:
            continue
        raise AssertionError(f"危险代码未拦截: {code!r}")
    # bash 引号拼接绕过
    try:
        check_bash("c''at data.sas7bdat")
    except DangerousOperationBlocked:
        pass
    else:
        raise AssertionError("引号拼接绕过未拦截")
    # 正常 pandas 代码不误杀
    check_python_code("import pandas as pd\ndf = pd.DataFrame()")


@case
def ai_operation_audit_has_no_raw_filename_or_identity():
    try:
        check_tool_call(
            "read_file",
            {"path": "patient-A1234567.sas7bdat"},
            {"session_id": "session-raw-9f", "user_id": "user-raw-9f"},
        )
    except DangerousOperationBlocked:
        pass
    else:
        raise AssertionError("敏感路径未拦截")
    latest_records = [
        line
        for path in (ROOT / "var" / "ai_ops_audit").glob("*.jsonl")
        for line in path.read_text(encoding="utf-8").splitlines()
    ]
    audit_blob = latest_records[-1]
    assert "A1234567" not in audit_blob
    assert "session-raw-9f" not in audit_blob
    assert "user-raw-9f" not in audit_blob


@case
def graded_scrub_removes_values():
    detector = ClinicalDataDetector()
    scrubber = StreamingScrubber(detector)
    scrubbed, result = scrubber.scrub_row(
        ["A1234567", "2024-03-05", "Screening", "已入组"], 1, True
    )
    joined = " ".join(scrubbed)
    assert "A1234567" not in joined and "2024-03-05" not in joined
    assert result.risk_level == DataRiskLevel.SENSITIVE


@case
def authorization_stores_no_raw_values():
    with tempfile.TemporaryDirectory() as root:
        record = authorize_category(root, "user@example.com", "session-1", "L3_ALLOW_AUDITED", "human-operator-9f")
        assert record["ok"]
        files = list(Path(root).rglob("*"))
        blob = "\n".join(path.read_text(encoding="utf-8") for path in files if path.is_file())
        assert "user@example.com" not in blob, "授权文件包含原始用户"
        assert "session-1" not in blob, "授权文件包含原始会话"
        assert "human-operator-9f" not in blob, "授权文件包含原始操作人"


@case
def normal_request_is_fast():
    """NFR-1: 正常请求检查路径 <10ms。

    使用独立临时审计目录（隔离宿主盘 IO 抖动）+ 5 次采样最小值，
    度量检查逻辑本身的成本。
    """
    with tempfile.TemporaryDirectory() as directory:
        checkpoint = EgressCheckpoint(directory)
        best = float("inf")
        for _ in range(5):
            start = time.perf_counter()
            checkpoint.check([{"role": "user", "content": "请帮我生成列表规范说明。"}])
            best = min(best, time.perf_counter() - start)
        assert best * 1000 < 10


@case
def normal_requests_have_low_false_positive_rate():
    samples = [
        "请生成列表规范说明。",
        "更新 README 文档。",
        "运行 npm test。",
        "检查 CI 状态。",
        "准备开发计划。",
    ] * 20
    blocked = [text for text in samples if not _allows(text)]
    assert not blocked, f"正常请求误拦: {blocked}"


def _allows(text: str) -> bool:
    try:
        check_egress([{"role": "user", "content": text}])
        return True
    except EgressViolation:
        return False


@case
def every_request_is_audited_without_raw_values():
    audit_dir = ROOT / "var" / "egress_audit"
    before = sum(
        len(path.read_text(encoding="utf-8").splitlines())
        for path in audit_dir.glob("*.jsonl")
    )
    _allows("请生成列表规范说明。")
    assert not _allows("Subject A1234567")
    records = []
    for path in audit_dir.glob("*.jsonl"):
        records.extend(line for line in path.read_text(encoding="utf-8").splitlines() if line)
    assert len(records) - before == 2
    blob = "\n".join(records[-2:])
    assert '"action":"ALLOWED"' in blob or '"action": "ALLOWED"' in blob
    assert '"action":"BLOCKED"' in blob or '"action": "BLOCKED"' in blob
    assert "A1234567" not in blob


@case
def audit_rotation_has_disk_cap():
    with tempfile.TemporaryDirectory() as root:
        # ST-D-4: 文件名跟随当前月份，不能写死 egress_202608（否则跨月必红）。
        stem = f"egress_{datetime.now().strftime('%Y%m')}.jsonl"
        for index in range(6):
            archive = Path(root) / f"{stem}.{index:02d}-00000000.rotated"
            archive.write_text("{}", encoding="utf-8")
        write_audit_record(
            root, "egress", {"audit_id": "rotation-test"},
            max_bytes=1, max_archives=2,
        )
        files = list(Path(root).iterdir())
        archives = [path for path in files if path.name.endswith(".rotated")]
        current = Path(root) / stem
        assert len(archives) == 2
        assert current.read_text(encoding="utf-8").strip()


# ============================================================================
# FIX-1~FIX-12 新增验收用例
# ============================================================================


def _run_worker(lines, timeout=20):
    """直接以子进程驱动 security.worker，逐行喂请求并收集响应。"""
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONPATH"] = str(ROOT)
    proc = subprocess.Popen(
        [sys.executable, "-m", "security.worker"],
        cwd=ROOT, env=env,
        stdin=subprocess.PIPE, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, text=True, encoding="utf-8",
    )
    try:
        out, err = proc.communicate("\n".join(lines) + "\n", timeout=timeout)
        responses = [json.loads(line) for line in out.splitlines() if line.strip()]
        return responses, err
    finally:
        if proc.poll() is None:
            proc.kill()


@case
def worker_malformed_line_returns_unavailable_and_survives():
    """AR-2.6 / FIX-6: 非法 JSON 行返回 SECURITY_UNAVAILABLE 且继续服务。"""
    good = json.dumps({
        "requestId": "r1", "operation": "check_tool",
        "tool": "read_file", "args": {"path": "safe.txt"}, "context": {},
    })
    responses, _ = _run_worker(["{{{not-json", good])
    assert responses[0]["ok"] is False
    assert responses[0]["code"] == "SECURITY_UNAVAILABLE"
    assert responses[0]["requestId"] is None
    assert responses[1]["ok"] is True
    assert responses[1]["requestId"] == "r1"


@case
def worker_error_receipt_is_sanitized():
    """R-6 / AR-2.9 / FIX-3: 异常回执不含路径原文与受试者标记。"""
    request = json.dumps({
        "requestId": "r2", "operation": "inspect_file",
        "path": r"C:\secret\patient-A1234567.xlsx", "context": {},
    })
    responses, _ = _run_worker([request])
    blob = json.dumps(responses[0], ensure_ascii=False)
    assert responses[0]["ok"] is False
    assert "A1234567" not in blob and "C:\\secret" not in blob


@case
def sanitize_error_strips_paths_subjects_and_dates():
    """FIX-3: sanitize_error 覆盖 Windows 路径、受试者编号与日期。"""
    raw = r"open failed: C:\data\patient-A1234567.xlsx at 2024-03-05 visit 101-0012"
    cleaned = sanitize_error(raw)
    assert "A1234567" not in cleaned
    assert "2024-03-05" not in cleaned
    assert "101-0012" not in cleaned
    assert "C:" not in cleaned


@case
def sensitive_key_names_block_and_audit_never_leak_raw_key():
    """R-7 / FR-09-03 / FR-12-05 / FIX-2: 顶层/嵌套敏感键名阻断；审计只留哈希。"""
    with tempfile.TemporaryDirectory() as directory:
        checkpoint = EgressCheckpoint(directory)
        for payload in (
            {"A1234567": "value", "messages": [{"role": "user", "content": "hi"}]},
            {"messages": [{"role": "user", "content": "hi", "meta": {"SUBJ-01-001234": 1}}]},
        ):
            try:
                checkpoint.check(payload)
            except EgressViolation:
                continue
            raise AssertionError(f"敏感键名未拦截: {payload}")
        blob = "\n".join(
            path.read_text(encoding="utf-8")
            for path in Path(directory).glob("*.jsonl")
        )
        assert "A1234567" not in blob
        assert "SUBJ-01-001234" not in blob


@case
def audit_payload_fields_use_known_whitelist_and_hash_unknown():
    """FR-12-05 / FIX-2: 已知字段呈现名称，未知字段只留截断哈希。"""
    with tempfile.TemporaryDirectory() as directory:
        checkpoint = EgressCheckpoint(directory)
        checkpoint.check({
            "model": "m", "messages": [{"role": "user", "content": "clean"}],
            "secretCustomField": "anything",
        })
        record = json.loads(
            next(Path(directory).glob("*.jsonl")).read_text(encoding="utf-8").splitlines()[-1]
        )
        fields = record["request_evidence"]["payload_fields"]
        assert "model" in fields and "messages" in fields
        assert "secretCustomField" not in fields
        assert any(str(f).startswith("sha256:") for f in fields)


@case
def write_file_with_malicious_python_is_blocked():
    """FR-03-13 / TC-25 / FIX-5: write_file 写入恶意 Python 代码被 AST 检查阻断。"""
    for content in [
        "import pickle\npickle.load(open('data.pkl','rb'))",
        "import pickle as p\np.load(open('data.pkl','rb'))",
        "from pickle import load\nload(open('data.pkl','rb'))",
        "import pandas as pd\npd.read_excel('expected/output.xlsx')",
    ]:
        try:
            check_tool_call("write_file", {"path": "gen.py", "content": content})
        except DangerousOperationBlocked:
            continue
        raise AssertionError(f"恶意代码写入未拦截: {content[:40]}")


@case
def benign_write_file_is_allowed():
    """FIX-5: 正常代码写入不受影响。"""
    check_tool_call("write_file", {
        "path": "report.py",
        "content": "def build():\n    return [i for i in range(10)]\n",
    })


@case
def strings_reading_data_files_is_blocked():
    """FR-03-10 / FIX-12: strings/xxd/od 读取数据文件被阻断。"""
    for command in [
        "strings data.xlsx",
        "xxd patients.csv",
        "od -c trial.sas7bdat",
    ]:
        try:
            check_bash(command)
        except DangerousOperationBlocked:
            continue
        raise AssertionError(f"数据转储命令未拦截: {command}")


@case
def light_scrub_covers_code_dates_and_usubjid():
    """FR-11: 轻度脱敏把编码/日期/USUBJID 复合格式 token 化（不可逆，同值同 token）。"""
    scrubber = StreamingScrubber(ClinicalDataDetector())
    scrubbed = scrubber._light_scrub(
        ["PT: 10001234", "03/05/2024", "2024年3月5日", "STUDY001-SITE01-SUBJ001"]
    )
    joined = " ".join(scrubbed)
    assert "10001234" not in joined
    assert "03/05/2024" not in joined
    assert "2024年3月5日" not in joined
    assert "STUDY001-SITE01-SUBJ001" not in joined
    # 原值被 HMAC token 替换（形如 [CODE:xxxxxxxx]），前缀标注语义类型。
    assert "[CODE:" in joined and "[DATE:" in joined and "[SUBJ:" in joined


@case
def tokenizer_is_deterministic_within_session_and_irreversible():
    """用户方案核验：同值同 token（LLM 可关联），不同值不同 token，无原值。"""
    from security.tokenizer import token_for
    a = token_for("101-001", "SUBJ")
    b = token_for("101-001", "SUBJ")
    c = token_for("101-002", "SUBJ")
    assert a == b            # 同值同 token → LLM 可 join/去重/计数
    assert a != c            # 不同值不同 token → 不塌缩信息
    assert "101-001" not in a  # 不含原值 → 不可逆
    assert a.startswith("[SUBJ:") and a.endswith("]")


@case
def authorization_requires_identity_and_consumes_once():
    """ST-P1-4 / ST-P1-3：缺 user/session 不授权(不共享 anonymous 桶)；
    有身份时授权一次、消费一次后失效。"""
    with tempfile.TemporaryDirectory() as root:
        # 缺身份 → fail-closed
        assert authorize_category(root, "u", None, "L3_ALLOW_AUDITED", "op")["ok"] is False
        assert authorize_category(root, None, "s", "L3_ALLOW_AUDITED", "op")["ok"] is False
        assert consume_category(root, "u", None, "L3_ALLOW_AUDITED") is False
        # 有身份 → 授权后一次性消费
        assert authorize_category(root, "u", "s", "L3_ALLOW_AUDITED", "op")["ok"] is True
        assert consume_category(root, "u", "s", "L3_ALLOW_AUDITED") is True
        assert consume_category(root, "u", "s", "L3_ALLOW_AUDITED") is False
        # 不同 session 不共享授权
        assert authorize_category(root, "u", "s2", "L3_ALLOW_AUDITED", "op")["ok"] is True
        assert consume_category(root, "u", "s-other", "L3_ALLOW_AUDITED") is False


@case
def metadata_evidence_has_no_raw_cell_text():
    """FIX-9 / R-6: METADATA 行 evidence 不含原始单元格文本。"""
    detector = ClinicalDataDetector()
    result = detector.detect_data_row(["请生成列表规范说明文档"], False)
    assert result.risk_level == DataRiskLevel.METADATA
    assert "请生成列表规范说明文档" not in result.evidence


@case
def l3_allow_audited_consumed_once():
    """FR-13 / TC-28 / FIX-4: L3_ALLOW_AUDITED 消费一次后失效。"""
    with tempfile.TemporaryDirectory() as root:
        authorize_category(root, "u", "s", "L3_ALLOW_AUDITED", "op")
        assert consume_category(root, "u", "s", "L3_ALLOW_AUDITED") is True
        assert consume_category(root, "u", "s", "L3_ALLOW_AUDITED") is False


def _concurrent_writer(root, worker_id):
    for i in range(40):
        write_audit_record(
            root, "egress",
            {"worker": worker_id, "seq": i},
            # max_bytes=400 使并发期间触发多次轮转 + 归档清理，覆盖 TC-34 全路径。
            max_bytes=400, max_archives=5,
        )


@case
def concurrent_audit_writes_keep_every_record():
    """BR-06.10 / TC-34 / NFR-13 / FIX-7: 双进程并发追加 + 轮转不丢记录。"""
    import multiprocessing

    with tempfile.TemporaryDirectory() as root:
        procs = [
            multiprocessing.Process(target=_concurrent_writer, args=(root, w))
            for w in range(2)
        ]
        for p in procs:
            p.start()
        for p in procs:
            p.join(timeout=60)
            assert p.exitcode == 0

        records = []
        for path in Path(root).glob("egress_*.jsonl*"):
            for line in path.read_text(encoding="utf-8").splitlines():
                if line.strip():
                    records.append(json.loads(line))
        assert len(records) == 80, f"并发丢记录: {len(records)}/80"
        seen = {(r["worker"], r["seq"]) for r in records}
        assert len(seen) == 80
        archives = list(Path(root).glob("*.rotated"))
        assert len(archives) <= 5


@case
def base64_short_candidates_are_not_flagged():
    """NFR-2 / FIX-11: 短 base64 候选（<24 总长）不触发拦截，误报不回归。"""
    for token in ["aGVsbG8gd29ybGQh", "dGVzdDEyMw", "QUJDREVGRw"]:
        try:
            check_egress([{"role": "user", "content": f"ref {token} done"}])
        except EgressViolation:
            raise AssertionError(f"短 base64 候选误报: {token}")


@case
def identity_context_hash_is_non_empty():
    """BR-06.5 / FIX-9: camelCase 上下文身份真实进入审计（哈希非空）。"""
    with tempfile.TemporaryDirectory() as directory:
        checkpoint = EgressCheckpoint(directory)
        checkpoint.check(
            [{"role": "user", "content": "clean"}],
            {"mode": "enforce", "sessionId": "sess-1", "userId": "user-1"},
        )
        record = json.loads(
            next(Path(directory).glob("*.jsonl")).read_text(encoding="utf-8").splitlines()[-1]
        )
        assert record["context"]["session_id"]
        assert record["context"]["session_id"].startswith("sha256:")
        assert record["context"]["user_id"].startswith("sha256:")


@case
def node_patterns_json_matches_python_source_of_truth():
    """Claude P2-3 / FIX-11: node_patterns.json 与 patterns.py 保持一致。"""
    from security.patterns import NODE_DLP_PATTERNS
    synced = json.loads(
        (ROOT / "security" / "node_patterns.json").read_text(encoding="utf-8")
    )
    expected = [
        {"source": p["re"], "flags": p.get("flags", ""), "label": p["label"]}
        for p in NODE_DLP_PATTERNS
    ]
    assert synced == expected, "运行 scripts/sync_patterns.py 同步 Node 模式副本"


@case
def xls_header_extraction_delivers_structure_without_values():
    """FR-06-03 / TC-15 / FIX-8: .xls 经 xlrd 解析表头结构，数据值不泄露。"""
    import xlwt
    with tempfile.TemporaryDirectory() as directory:
        path = Path(directory) / "legacy.xls"
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("DM")
        for col, name in enumerate(["USUBJID", "BRTHDTC", "SEX"]):
            sheet.write(0, col, name)
        sheet.write(1, 0, "A1234567")
        sheet.write(1, 1, "2024-03-05")
        workbook.save(str(path))

        result = subprocess.run(
            [sys.executable, str(ROOT / "excel_header_extractor.py"), str(path)],
            capture_output=True, text=True, encoding="utf-8", timeout=30,
        )
        assert result.returncode == 0, result.stderr
        payload = json.loads(result.stdout)
        blob = json.dumps(payload, ensure_ascii=False)
        assert payload["sheets"][0]["sheet"] == "DM"
        assert any(c["value"] == "USUBJID" for c in payload["sheets"][0]["header_cells"])
        assert "A1234567" not in blob and "2024-03-05" not in blob


@case
def authorization_and_audit_share_hash_context():
    """FIX-9.4 / R-6: 授权与审计使用同一哈希上下文，可通过哈希关联。"""
    from security.patterns import stable_hash

    with tempfile.TemporaryDirectory() as root:
        authorize_category(root, "link-user", "link-session", "L3_SKIP", "op")
        record = next(Path(root).rglob("egress_authz.json"))
        # 身份哈希同时体现在授权记录的目录层级中，与审计 stable_hash 同源。
        record_path = str(record).replace(str(root), "")
        assert stable_hash("link-user") in record_path
        assert stable_hash("link-session") in record_path
        assert "link-user" not in record_path and "link-session" not in record_path


# ============================================================================
# 真实项目（CGB3002-TEST）暴露缺陷的回归用例
# ============================================================================


@case
def worker_surrogate_input_survives():
    """真实故障（CGB3002 运行一炮报错）：JSON 携带 \\udcae 孤立代理。

    worker 响应写 stdout 曾直接 UnicodeEncodeError 崩溃、后续请求全部挂起；
    现在必须返回响应（代理→U+FFFD）且继续服务。
    """
    dirty = json.dumps({
        "requestId": "s1", "operation": "scrub_text",
        "text": "subject 01001 on 08 Jun 2026 \udcae trailing", "context": {},
    })
    ping = json.dumps({"requestId": "s2", "operation": "ping"})
    responses, _ = _run_worker([dirty, ping])
    assert len(responses) == 2, "worker 遇孤立代理崩溃，未继续服务"
    assert responses[0]["ok"] is True
    assert "\udcae" not in responses[0].get("text", "")
    assert responses[1]["ok"] is True and responses[1]["requestId"] == "s2"


@case
def sanitize_error_and_clean_surrogates_handle_lone_surrogates():
    """sanitize_error / clean_surrogates 清除孤立代理且保留可读文本。"""
    from security.patterns import clean_surrogates
    raw = "read failed: \udcae\udcff at C:\\data\\A1234567.xlsx"
    cleaned = sanitize_error(raw)
    assert "\udcae" not in cleaned and "\udcff" not in cleaned
    assert "A1234567" not in cleaned and "C:" not in cleaned
    assert "\ud800" not in clean_surrogates("ok \ud800\udfff done")


@case
def dd_mmm_yyyy_clinical_date_is_detected_and_scrubbed():
    """真实缺陷（crViewer.xls）：DD-MMM-YYYY 临床日期未识别。

    08 Jun 2026 / 08 Jun 2026 05:19:50 是 Rave/EDC 导出标准格式，
    检测（出域阻断）与轻度脱敏双侧必须覆盖。
    """
    for text in ["Visit 08 Jun 2026 done", "Entry 08 Jun 2026 05:19:50 done"]:
        try:
            check_egress([{"role": "user", "content": text}])
        except EgressViolation:
            continue
        raise AssertionError(f"临床报告日期未拦截: {text}")
    scrubbed = StreamingScrubber(ClinicalDataDetector())._light_scrub(
        ["08 Jun 2026", "08 Jun 2026 05:19:50"]
    )
    joined = " ".join(scrubbed)
    assert "Jun" not in joined and "[DATE:" in joined


@case
def headerless_data_table_does_not_leak_rows():
    """真实红线泄露（crViewer.xls）：无表头数据表整表被当作表头输出。

    结构复刻：2 行标题 + 数值型数据行（编号/数值/日期/状态）。
    修复后 header_cells 只含标题行，data_start_row 停在首个数据行。
    """
    import openpyxl
    from excel_header_extractor import process_xlsx
    with tempfile.TemporaryDirectory() as directory:
        path = str(Path(directory) / "crviewer_like.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["SiteGroup: World"])
        sheet.append(["Site: UAT_006 (Site Number:23)"])
        for subject in ["01001", "01002", "S005"]:
            sheet.append([subject, "346.0", "08 Jun 2026 05:19:50", "Screening"])
        workbook.save(path)
        result = subprocess.run(
            [sys.executable, str(ROOT / "excel_header_extractor.py"), path],
            capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30,
        )
        assert result.returncode == 0, result.stderr
        blob = result.stdout
        for marker in ["01001", "01002", "S005", "346.0", "08 Jun 2026"]:
            assert marker not in blob, f"无表头数据表泄露数据行: {marker}"
        payload = json.loads(blob)
        entry = payload["sheets"][0]
        assert entry["data_start_row"] == 2
        assert entry["header_cells"] == [
            {"row": 0, "col": 0, "value": "SiteGroup: World"},
            {"row": 1, "col": 0, "value": "Site: UAT_006 (Site Number:23)"},
        ]


@case
def leading_zero_subject_cell_is_redacted():
    """真实缺陷（crViewer.xls）：前导零 5 位受试者号（01001）未脱敏。

    NUMERIC_SUBJECT_ID_RE 覆盖 EDC 前导零形态。
    ST-P1-9 输出白名单收紧后：普通短数值（346.0/002/2024）在表头输出侧
    也一律 REDACTED（DATA_VALUE 兜底）——列名极少是纯数字，数据值常是，
    误判表头时这类值原值输出即泄露通道。
    """
    from excel_header_extractor import _dlp_scan_cell
    cleaned, label = _dlp_scan_cell("01001")
    assert label == "SUBJECT_ID" and "01001" not in str(cleaned)
    for data_like in ["346.0", "002", "2024"]:
        value, hit = _dlp_scan_cell(data_like)
        assert hit == "DATA_VALUE" and data_like not in str(value), \
            f"数据形态值原值输出: {data_like}"
    # 合法列名（含安全词/中文列名）不误伤
    for header in ["Subject", "Visit Date", "受试者编号", "Visit 1", "Day 8"]:
        value, hit = _dlp_scan_cell(header)
        assert hit is None and value == header, f"列名误报: {header}"


@case
def worker_inspects_real_xls_via_xlrd():
    """真实缺陷：worker inspect_file 对 .xls 用 openpyxl 必失败（agent 被拒）。

    .xls 现在走 xlrd；含 SENSITIVE 组合的 .xls 触发用户决策提示。
    """
    import xlwt
    with tempfile.TemporaryDirectory() as directory:
        path = str(Path(directory) / "legacy.xls")
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("DM")
        sheet.write(0, 0, "Subject")
        sheet.write(1, 0, "A1234567")
        sheet.write(1, 1, "2024-03-05")
        sheet.write(1, 2, "Screening")
        workbook.save(path)
        request = json.dumps({
            "requestId": "x1", "operation": "inspect_file",
            "path": path, "max_scan_rows": 20, "context": {"mode": "enforce"},
        })
        responses, _ = _run_worker([request])
        assert responses[0]["ok"] is True, responses[0]
        assert responses[0]["needs_user"] is True
        assert responses[0]["user_prompt"]


@case
def windows_path_separator_does_not_bypass_blacklist():
    """ST-P2-6: Windows 路径分隔符 \\ 不能绕过危险路径黑名单。

    ai_operations_monitor._assess_file_threat 归一化后匹配，
    C:\\project\\docment\\dm\\data\\sas_data_cache.pkl 必须被拦截。
    """
    from security.ai_operations_monitor import AIOperationMonitor, RiskLevel
    monitor = AIOperationMonitor()
    threat = monitor._assess_file_threat(
        r"C:\project\docment\dm\data\sas_data_cache.pkl", "read"
    )
    assert threat.risk_level >= RiskLevel.HIGH, (
        f"Windows 路径绕过黑名单: risk={threat.risk_level}"
    )


@case
def message_id_uuid_does_not_trigger_egress_block():
    """E2E-4 回归: 消息元数据 UUID 字段（payload.messages[n].id）不引发误拦截。

    UUID 形如 abc12345-0001-0002-0003-def456789012，与受试者编号模式部分重叠；
    _METADATA_KEY_FIELDS 白名单确保此类标量值跳过内容 DLP。
    """
    payload = {
        "model": "gpt-4o",
        "messages": [
            {"role": "user", "content": "请生成列表", "id": "abc12345-0001-0002-0003-def456789012"},
            {"role": "assistant", "content": "好的", "id": "bcd23456-0001-0002-0003-ef0123456789"},
        ],
    }
    try:
        check_egress(payload)
    except EgressViolation as exc:
        raise AssertionError(f"消息 id UUID 误报触发拦截: {exc}") from exc


@case
def authz_replay_is_rejected():
    """ST-D-3: 授权重放——同一授权消费两次，第二次必须返回 False。"""
    with tempfile.TemporaryDirectory() as root:
        authorize_category(root, "u1", "s1", "L3_ALLOW_AUDITED", "op")
        first = consume_category(root, "u1", "s1", "L3_ALLOW_AUDITED")
        second = consume_category(root, "u1", "s1", "L3_ALLOW_AUDITED")
        assert first is True, "首次消费应成功"
        assert second is False, "重放消费未拒绝"


@case
def concurrent_double_consume_is_rejected():
    """ST-D-3: 并发双消费——两个进程竞争消费同一授权，恰好一个成功。"""
    import multiprocessing
    with tempfile.TemporaryDirectory() as root:
        authorize_category(root, "u2", "s2", "L3_ALLOW_AUDITED", "op")

        results = multiprocessing.Manager().list()

        def _consume(r):
            results.append(consume_category(root, "u2", "s2", "L3_ALLOW_AUDITED"))

        procs = [multiprocessing.Process(target=_consume, args=(results,)) for _ in range(2)]
        for p in procs: p.start()
        for p in procs: p.join(timeout=10)
        successes = [x for x in results if x is True]
        assert len(successes) == 1, f"并发双消费应仅一次成功，实际 {len(successes)} 次"


@case
def singleton_get_egress_checkpoint_is_thread_safe():
    """ST-P3-x 回归: 并发 get_egress_checkpoint() 返回同一对象。"""
    import threading
    from security.egress_checkpoint import get_egress_checkpoint
    instances = []
    def _get(): instances.append(id(get_egress_checkpoint()))
    threads = [threading.Thread(target=_get) for _ in range(8)]
    for t in threads: t.start()
    for t in threads: t.join()
    assert len(set(instances)) == 1, f"单例不唯一，得到 {len(set(instances))} 个实例"


@case
def document_version_number_is_not_subject_id():
    """真实缺陷回归：'标识+YYYYMMDD' 文档版本号不得被误判为受试者编号。

    历史故障：`\\b[A-Z]{1,4}\\d{6,8}\\b` 把 DVP20260610 判为受试者编号。因
    llm/stream 每轮重扫完整 messages 历史，一条误报即把整个会话永久钉死
    （审计 20260820_135432-c4f47e1a89：109 个阻断威胁全是该模式）。
    """
    for text in ("DVP20260610 修订说明", "SPEC20260610", "ALS20251231 定稿",
                 "参见 DVP 20260610 第 3 节"):
        assert _allows(text), f"文档版本号被误拦: {text!r}"


@case
def document_version_exemption_does_not_weaken_subject_detection():
    """反向回归：文档版本号豁免绝不能放宽真实受试者编号的检出。

    豁免是纯格式判据（尾部 8 位须为合法 YYYYMMDD），因此 7 位数字的
    A1234567、月份非法的 S0001234、非日期的 EC12345678 必须仍被拦截。
    """
    for text in ("Subject A1234567", "subj S0001234", "Edit Check EC12345678",
                 "site 101-001"):
        assert not _allows(text), f"真实受试者编号漏放: {text!r}"


@case
def spec_and_header_terms_are_allowed():
    """spec/ALS/template 文本与 SAS 表头字段名允许出域（用户规则：表头可读）。"""
    for text in ("spec ALS DVP template 表头 USUBJID AESTDTC VISITNUM",
                 "读取 CGB3002 项目的 ALS 文档与 DVP 模板",
                 "SAS 数据集表头: USUBJID, SITEID, VISIT, AEDECOD"):
        assert _allows(text), f"spec/表头文本被误拦: {text!r}"


@case
def tokenizer_is_idempotent_and_shared_by_both_lanes():
    """脱敏保底必须真实生效且幂等——这是会话可自愈的前提。

    历史故障：token_for/token_sub 只在 post-execute 车道被调用，出域车道从不
    token 化，"读到 data 先 hash 再送 AI"的保底在出域层是死代码。现在两个车道
    共用 tokenizer.tokenize_clinical_text 单一来源。
    """
    from security.tokenizer import tokenize_clinical_text
    from security.data_egress_guard import ClinicalDataDetector, StreamingScrubber

    raw = "Subject A1234567 visit 2024-01-15"
    once = tokenize_clinical_text(raw)
    assert "A1234567" not in once and "2024-01-15" not in once, f"原值未脱敏: {once!r}"
    assert "[SUBJ:" in once and "[DATE:" in once, f"token 前缀缺失: {once!r}"
    # 幂等：token 化结果重扫不再变化，否则历史重扫无法自愈
    assert tokenize_clinical_text(once) == once, "token 化非幂等"
    # 同值同 token（LLM 仍可 join/去重/计数）
    assert tokenize_clinical_text("A1234567") == tokenize_clinical_text("a1234567")
    # 两车道同一口径：post-execute 的 _light_scrub 走同一函数
    scrubber = StreamingScrubber(ClinicalDataDetector())
    assert scrubber._light_scrub([raw]) == [once], "两车道脱敏口径不一致"


@case
def tokenized_text_passes_egress():
    """脱敏后的 token 文本必须能通过出域检查，否则保底手段等于无效。"""
    from security.tokenizer import tokenize_clinical_text
    assert _allows(tokenize_clinical_text("Subject A1234567 visit 2024-01-15")), \
        "token 化后仍被拦截，脱敏保底无效"


@case
def pdf_spec_visit_window_terms_are_allowed():
    """真实缺陷回归（审计 20260820_150945-85697ec72e）：PDF 规格文档的
    访视窗术语 "Visit Date(D1)"、"Visit Date- Screening (D-56~D-3)" 被
    CDISC字段:visit 判为数据行，13 次 BLOCK + 复合威胁拦死整个请求。
    用户规则：spec/方案文档可读。纯格式判据（术语/括号天数标签），无关键词豁免。
    """
    for text in ("Visit Date- Screening (D-56~D-3)",
                 "Visit Date- Baseline (D-2)",
                 "D3 | Visit Date(D1) +2 days",
                 "ADA Follow-up | Visit Date(D1) +89 days",
                 "If the Visit Date of D1 is entered, Expected date follows"):
        assert _allows(text), f"PDF spec 访视术语被误拦: {text!r}"


@case
def code_identifier_after_cdisc_field_is_allowed():
    """真实缺陷回归（审计 20260820_155650-bbbef95ac1）：读代码文件结果里的
    注释 "cohort per subject (DSCOHORT@5218; STD@5278)" 被 CDISC字段:subject
    误判（值 (DSCOHORT@5218 短且含数字）。含 @ 的值是代码变量注解，
    CDISC 数据值形态从不包含 @。真实数据值不受影响。
    """
    for text in ("    # ---- DS1: cohort per subject (DSCOHORT@5218; STD@5278) ----",
                 'per subject (AESTDTC@1234) annotation'):
        assert _allows(text), f"代码标识被误拦: {text!r}"


@case
def file_paths_and_filenames_are_operational():
    """真实缺陷回归（2026-08-20 工作台实测）：路径中的模式形态被 token 化，
    模型拿 "G:\...\CGB3002-TEST\[SUBJ:d1b1c9f9].txt" 假路径读文件直接
    not found，工作流断裂。用户规则：路径/文件名是辅助读取的操作性数据，
    写入/出域/工具参数三车道一律原样放行；数据值照旧脱敏/拦截。
    """
    from security.data_egress_guard import ClinicalDataDetector, StreamingScrubber
    scrubber = StreamingScrubber(ClinicalDataDetector())
    cases = [
        r'G:\home\Clinical-Data\CGB3002-TEST\SUBJ123456_report.txt',
        'G:\\\\home\\\\Clinical-Data\\\\CGB3002-TEST\\\\AE1234567.txt',
        r'\\server\\share\\010-001-1001.csv',
        '/data/exports/S0001234/2024-01-15.txt',
        'bare AE123456.txt reference',
    ]
    for op in cases:
        scrubbed, _ = scrubber.scrub_row([f'{{"path": "{op}"}}'], 1, False)
        assert op in scrubbed[0], f"路径被改写: {op!r} -> {scrubbed[0]!r}"
        assert _allows(f'Error: cannot read "{op}": not found'), f"路径被拦: {op!r}"
    # 数据值（无操作性形态）照旧：写入侧脱敏消除原值、出域侧拦截
    scrubbed, _ = scrubber.scrub_row(['subject 101-001 visit 2024-01-15'], 1, False)
    assert '101-001' not in scrubbed[0] and '2024-01-15' not in scrubbed[0], scrubbed[0]
    assert re.search(r'\[(SUBJ|DATE|TEXT|NUM):[0-9a-f]{8}\]', scrubbed[0]), scrubbed[0]
    assert not _allows('subject 101-001 visit 2024-01-15')


@case
def numeric_visit_values_still_block():
    """反向回归：结构字段后跟纯数字值（真实数据行形态）必须仍拦。"""
    for text in ("VISIT: 3", "VISITNUM 4.0", "visit 12 结果"):
        assert not _allows(text), f"真实数据形态漏放: {text!r}"


@case
def uuid_and_docid_follow_unified_tokenization_architecture():
    """统一 token 化架构回归（用户规则）：
    - UUID（技术标识，E2E-4 口径）在出域检测入口剥离——原文与 normalized
      拼接形态（'id c25e...' 融合后按位置豁免会越界失效）都不得误拦。
    - 文档编号（DS5565-0002-NIS-MA 等）出域侧按原值拦（fail-closed），
      由写入侧统一 token 化消化：token 化后的文本必须放行。真实 USUBJID
      （含 STUDY-SITE-001 纯字母段形态）任何时候都拦。
    """
    from security.tokenizer import tokenize_clinical_text
    # UUID：两种形态都放行
    for text in ('id c25e2638-0ced-4330-86ae-728287fcdeaa isError false',
                 'call_6ad2d0e2-c715-4a38-bdc2-2def4bdfe804'):
        assert _allows(text), f"UUID 被误拦: {text!r}"
    # 编号结构判据：字母末段=文档/项目编号（含 meta.lines 投影场景）放行
    for text in ("DS5565-0002-NIS-MA DM Status Report Specification",
                 "CGB3002-TEST 项目目录", "读取 CGB3002-TEST spec"):
        assert _allows(text), f"文档/项目编号被误拦: {text!r}"
    # 数字末段=真实 USUBJID，任何时候都拦
    for text in ("USUBJID 010-001-1001", "STUDY001-SITE01-SUBJ001",
                 "STUDY-SITE-001", "010-001-1001-S1 随机"):
        assert not _allows(text), f"真实 USUBJID 漏放: {text!r}"
    # 写入侧统一 token 化后的文本必须放行（保底闭环）
    for text in ("DS5565-0002-NIS-MA DM Status Report Specification",
                 "USUBJID 010-001-1001 随机 2024-01-15"):
        tokenized = tokenize_clinical_text(text)
        assert tokenized != text, f"token 化未生效: {text!r}"
        assert _allows(tokenized), f"token 化后仍被拦: {tokenized!r}"


@case
def uuid_is_preserved_by_tokenizer():
    """token 化不得吃掉 UUID（LLM 需要消息 id 引用），USUBJID 正常 token 化。"""
    from security.tokenizer import tokenize_clinical_text
    uuid_text = "id c25e2638-0ced-4330-86ae-728287fcdeaa ok"
    assert tokenize_clinical_text(uuid_text) == uuid_text
    assert "[SUBJ:" in tokenize_clinical_text("subj 010-001-1001")


def main() -> int:
    failures = 0
    for test in CASES:
        try:
            test()
            print(f"PASS {test.__name__}")
        except Exception as exc:
            failures += 1
            print(f"FAIL {test.__name__}: {exc}")
    print(f"RESULT {len(CASES) - failures}/{len(CASES)}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
