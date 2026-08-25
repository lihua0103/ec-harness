"""ListingPlan validator 与本地执行器的对抗性测试网（审计发现 F-3）。

## 为什么需要独立的测试文件

新架构的安全论证是："模型只能提交结构化 IR，validator 是死命令，因此 SAS 数据值
不可能出域"。这个论证完全依赖 validator 的完备性——它是唯一的结构性保证。
审计发现此前 `tests/` 中对 `validate_listing_plan` / `execute_listing_plan`
零命中，F-2（valueRef 静默产出错误临床交付物）正是因此漏网。

## 三族用例

1. **合法族**：正常计划必须通过并产出正确数据（防止收紧校验时误伤真实业务）。
2. **越界族**：未知属性/数据集/字段、非法 join、超限资源、公式注入必须被拒。
3. **混淆族**：类型伪装（bool 冒充 number）、大小写与限定名混淆、
   "validated 却执行不了"的双侧语义不一致。

第 3 族是最关键的：validator 与 executor 对同一份计划的理解若有分歧，
"validated 即可执行"契约就是空的，而契约是模型自我纠正的唯一依据。
"""
from __future__ import annotations

import os
import io
import json
import shutil
import sys
import tempfile
from pathlib import Path
from unittest import mock

import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from security.listing_budget import charge_execution, reset_budget
from security.listing_executor import ListingExecutionError, execute_listing_plan
from security.listing_plan import (
    MAX_ITEMS_PER_OUTPUT,
    MAX_OUTPUTS,
    ListingPlanError,
    validate_listing_plan,
)
from security.listing_workflow import ListingWorkflowError, execute_listing_plan_workflow

SCHEMA = {
    "ae": {"USUBJID", "AETERM", "AESEV", "AESTDAT", "AEENDAT", "Status"},
    "dm": {"USUBJID", "AGE", "SITEID"},
}


def _plan(**overrides):
    output = {
        "name": "REPLAY",
        "source": "AE",
        "columns": [{"source": "AETERM", "name": "AETERM", "label": "不良事件"}],
    }
    output.update(overrides.pop("output", {}))
    plan = {"version": 1, "scenario": "rbqm", "outputs": [output]}
    plan.update(overrides)
    return plan


def _validate(plan, scenario: str = "rbqm", **kwargs):
    return validate_listing_plan(plan, SCHEMA, scenario, **kwargs)


def _rejects(plan, code: str, scenario: str = "rbqm", **kwargs) -> None:
    try:
        _validate(plan, scenario, **kwargs)
    except ListingPlanError as error:
        assert error.code == code, f"期望 {code}，实际 {error.code}（{error}）"
        # 拒绝原因不得回传数据值：path 是结构位置，message 是规则说明。
        assert error.path.startswith("plan"), error.path
        return
    raise AssertionError(f"计划本应以 {code} 被拒绝")


# ---------------------------------------------------------------------------
# 1. 合法族
# ---------------------------------------------------------------------------

def test_minimal_plan_normalizes_with_stable_defaults() -> None:
    normalized = _validate(_plan())
    assert normalized["version"] == 1
    assert normalized["scenario"] == "rbqm"
    output = normalized["outputs"][0]
    assert output["source"] == "AE"
    assert output["columns"] == [{"source": "AETERM", "name": "AETERM", "label": "不良事件"}]
    # 未声明的布局项必须有确定默认值，避免执行器读到 None 再各自兜底。
    assert output["layout"]["freezeRows"] == 1
    assert output["layout"]["freezeColumns"] == 0
    assert output["layout"]["appendReviewColumns"] is False
    assert output["layout"]["statusFilter"] == ""


def test_qualified_and_case_insensitive_refs_resolve_to_real_columns() -> None:
    normalized = _validate(_plan(output={"columns": [
        {"source": "ae.aeterm", "name": "TERM", "label": ""},
        {"source": "AESEV", "name": "SEV", "label": ""},
    ]}))
    sources = [item["source"] for item in normalized["outputs"][0]["columns"]]
    # 限定名与大小写混淆都必须归一到 schema 中的真实字段名。
    assert sources == ["AE.AETERM", "AESEV"]


def test_join_derive_filter_aggregate_plan_is_accepted() -> None:
    normalized = _validate(_plan(output={
        "joins": [{"dataset": "DM", "type": "left", "leftKeys": ["USUBJID"], "rightKeys": ["USUBJID"]}],
        "derivations": [{"name": "DURATION", "operation": "date_diff_days", "refs": ["AEENDAT", "AESTDAT"]}],
        "filters": [{"column": "DURATION", "operator": "gte", "literal": {"type": "number", "value": 1}}],
        "groupBy": ["AESEV"],
        "aggregations": [{"name": "EVENTS", "operation": "count", "column": "AETERM"}],
        "columns": [
            {"source": "AESEV", "name": "AESEV", "label": "严重程度"},
            {"source": "EVENTS", "name": "EVENTS", "label": "事件数"},
        ],
    }))
    output = normalized["outputs"][0]
    assert output["derivations"][0]["refs"] == ["AEENDAT", "AESTDAT"]
    # F-6 反向：聚合别名必须可以作为输出列被引用。
    assert [item["name"] for item in output["columns"]] == ["AESEV", "EVENTS"]


# ---------------------------------------------------------------------------
# 2. 越界族
# ---------------------------------------------------------------------------

def test_unknown_properties_are_rejected_at_every_level() -> None:
    _rejects({**_plan(), "script": "rm -rf /"}, "UNKNOWN_PROPERTY")
    _rejects(_plan(output={"path": "C:/data/ae.sas7bdat"}), "UNKNOWN_PROPERTY")
    _rejects(_plan(output={"filters": [
        {"column": "AETERM", "operator": "eq", "literal": {"type": "string", "value": "x"}, "raw": "1"},
    ]}), "UNKNOWN_PROPERTY")
    _rejects(_plan(output={"layout": {"exec": True}}), "UNKNOWN_PROPERTY")


def test_unknown_dataset_and_column_are_rejected() -> None:
    _rejects(_plan(output={"source": "SECRET"}), "UNKNOWN_DATASET")
    _rejects(_plan(output={"columns": [{"source": "SSN", "name": "SSN", "label": ""}]}), "UNKNOWN_COLUMN")
    # 未 join 的数据集字段不得可见——否则计划能引用任意数据集的列。
    _rejects(_plan(output={"columns": [{"source": "DM.AGE", "name": "AGE", "label": ""}]}), "UNKNOWN_COLUMN")


def test_version_and_scenario_are_strictly_bound() -> None:
    _rejects({**_plan(), "version": 2}, "UNSUPPORTED_VERSION")
    _rejects({**_plan(), "version": "1"}, "UNSUPPORTED_VERSION")
    # 计划声明的场景必须与请求一致，防止用 rbqm 计划走 medical 质量门禁。
    _rejects(_plan(), "SCENARIO_MISMATCH", scenario="medical")


def test_invalid_joins_are_rejected() -> None:
    base = {"dataset": "DM", "leftKeys": ["USUBJID"], "rightKeys": ["USUBJID"]}
    _rejects(_plan(output={"joins": [{**base, "type": "cross"}]}), "INVALID_JOIN")
    _rejects(_plan(output={"joins": [{**base, "rightKeys": ["USUBJID", "AGE"]}]}), "INVALID_JOIN")
    _rejects(_plan(output={"joins": [{**base, "dataset": "GHOST"}]}), "UNKNOWN_DATASET")
    _rejects(_plan(output={"joins": [{**base, "leftKeys": ["NOPE"]}]}), "UNKNOWN_COLUMN")
    _rejects(_plan(output={"joins": [{**base, "rightKeys": ["NOPE"]}]}), "UNKNOWN_COLUMN")


def test_identifier_shape_blocks_paths_and_expressions() -> None:
    for hostile in ("../../etc/passwd", "AE; DROP TABLE", "AE AETERM", "1AE", "", "A" * 200):
        _rejects(_plan(output={"source": hostile}), "INVALID_IDENTIFIER"
                 if hostile else "INVALID_IDENTIFIER")


def test_plan_resource_limits_are_enforced() -> None:
    # N-1: 无上限计划是本地 DoS/磁盘填满面。
    many_outputs = {**_plan(), "outputs": [
        {"name": f"OUT{index}", "source": "AE",
         "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}]}
        for index in range(MAX_OUTPUTS + 1)
    ]}
    _rejects(many_outputs, "PLAN_TOO_LARGE")
    _rejects(_plan(output={"columns": [
        {"source": "AETERM", "name": f"C{index}", "label": ""}
        for index in range(MAX_ITEMS_PER_OUTPUT + 1)
    ]}), "PLAN_TOO_LARGE")


def test_formula_prefixes_are_rejected_in_free_text() -> None:
    # N-10: openpyxl 会把 = 开头的字符串存为公式，污染交付物。
    for hostile in ("=1+1", "+HYPERLINK(1)", "-2", "@SUM(A1)"):
        _rejects(_plan(output={"columns": [
            {"source": "AETERM", "name": "AETERM", "label": hostile},
        ]}), "INVALID_LABEL")
        _rejects(_plan(output={"layout": {"statusFilter": hostile}}),
                 "INVALID_LAYOUT_REQUIREMENT")


def test_layout_numbers_reject_garbage_with_structured_errors() -> None:
    # N-2: 裸 ValueError 会被降级为 WORKFLOW_UNAVAILABLE，模型看不到哪个字段非法。
    for hostile in ("abc", {}, [], True):
        _rejects(_plan(output={"layout": {"freezeColumns": hostile}}),
                 "INVALID_LAYOUT_REQUIREMENT")


def test_empty_plan_and_empty_output_are_rejected() -> None:
    _rejects({**_plan(), "outputs": []}, "EMPTY_PLAN")
    _rejects(_plan(output={"columns": []}), "EMPTY_OUTPUT")


# ---------------------------------------------------------------------------
# 3. 混淆族：字面量通道
# ---------------------------------------------------------------------------

def test_literal_types_cannot_be_spoofed() -> None:
    def literal_filter(literal):
        return _plan(output={"filters": [
            {"column": "AESEV", "operator": "eq", "literal": literal},
        ]})

    # bool 是 int 的子类；若不显式排除，True 能以 number 身份混入。
    _rejects(literal_filter({"type": "number", "value": True}), "INVALID_LITERAL")
    _rejects(literal_filter({"type": "string", "value": "x" * 257}), "INVALID_LITERAL")
    _rejects(literal_filter({"type": "number", "value": "1"}), "INVALID_LITERAL")
    _rejects(literal_filter({"type": "object", "value": {}}), "INVALID_LITERAL")
    # 多余/缺失键都必须拒绝，避免夹带额外载荷。
    _rejects(literal_filter({"type": "number", "value": 1, "raw": "x"}), "INVALID_LITERAL")
    _rejects(literal_filter({"type": "number"}), "INVALID_LITERAL")


def test_filter_must_carry_exactly_one_comparison_source() -> None:
    _rejects(_plan(output={"filters": [
        {"column": "AESEV", "operator": "eq", "valueRef": "AETERM",
         "literal": {"type": "string", "value": "x"}},
    ]}), "INVALID_FILTER")
    _rejects(_plan(output={"filters": [{"column": "AESEV", "operator": "eq"}]}),
             "MISSING_REFERENCE")
    _rejects(_plan(output={"filters": [{"column": "AESEV", "operator": "matches",
                                        "literal": {"type": "string", "value": "x"}}]}),
             "INVALID_FILTER")


def test_null_checks_reject_smuggled_comparison_values() -> None:
    # N-4: 此前 is_null 携带的 literal 被静默丢弃，模型的显式意图被无声忽略。
    _rejects(_plan(output={"filters": [
        {"column": "AESEV", "operator": "is_null", "literal": {"type": "string", "value": "x"}},
    ]}), "INVALID_FILTER")
    _rejects(_plan(output={"filters": [
        {"column": "AESEV", "operator": "not_null", "valueRef": "AETERM"},
    ]}), "INVALID_FILTER")
    # 不携带比较值的空值判断是合法的。
    assert _validate(_plan(output={"filters": [
        {"column": "AESEV", "operator": "is_null"},
    ]}))["outputs"][0]["filters"][0]["literal"] is None


def test_derivation_arity_is_validated() -> None:
    # N-5: date_diff_days 单 ref 此前在执行期 IndexError，诊断信息全丢。
    for operation in ("date_diff_days", "subtract", "divide"):
        _rejects(_plan(output={"derivations": [
            {"name": "D", "operation": operation, "refs": ["AESTDAT"]},
        ]}), "INVALID_DERIVATION")
    _rejects(_plan(output={"derivations": [
        {"name": "D", "operation": "shell", "refs": ["AESTDAT"]},
    ]}), "INVALID_DERIVATION")
    _rejects(_plan(output={"derivations": [
        {"name": "D", "operation": "copy", "refs": []},
    ]}), "MISSING_REFERENCE")


# ---------------------------------------------------------------------------
# 3. 混淆族：validator ↔ executor 契约一致性
# ---------------------------------------------------------------------------

def _study(root: Path) -> Path:
    """最小真实项目：一个 CSV 数据集 + 一个可 join 的数据集。"""
    project = root / "study"
    (project / "data").mkdir(parents=True)
    pd.DataFrame({
        "USUBJID": ["101-001", "101-002", "101-003"],
        "AETERM": ["headache", "nausea", "headache"],
        "AESEV": ["MILD", "SEVERE", "MILD"],
        "AESTDAT": ["2024-01-01", "2024-01-05", "2024-02-01"],
        "AEENDAT": ["2024-01-03", "2024-01-06", "2024-02-10"],
        "Status": ["已编码", "未编码", "已编码"],
    }).to_csv(project / "data" / "AE.csv", index=False)
    pd.DataFrame({
        "USUBJID": ["101-001", "101-002", "101-003"],
        "AGE": [42, 57, 33],
        "SITEID": ["101", "101", "101"],
    }).to_csv(project / "data" / "DM.csv", index=False)
    return project


def _execute(project: Path, plan: dict) -> pd.DataFrame:
    """执行计划并读回产物第一个 sheet 的数据区（跳过标签/字段名两行）。"""
    import openpyxl

    output = project / "out"
    result = execute_listing_plan(str(project), str(output), plan)
    assert result["dataClass"] == "REAL"
    path = Path(result["artifacts"][0]["path"])
    # Python 3.13 + openpyxl 关闭后可能仍持有源文件描述符；使用内存副本，
    # 确保 Windows 能清理 TemporaryDirectory。
    workbook = openpyxl.load_workbook(io.BytesIO(path.read_bytes()))
    rows = list(workbook.active.values)
    workbook.close()
    return pd.DataFrame(list(rows[1:]), columns=list(rows[0]))


def test_valueref_filter_compares_values_not_column_names() -> None:
    """F-2: 这是本次审计的核心正确性缺陷。

    `_column()` 返回列名字符串，此前 filter 直接拿它与数据 series 比较：
    `eq` 恒不成立（静默空表）、`ne` 恒成立（静默全表）。计划完全合法、
    无任何报错，交付物却是错的。
    """
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        plan = _validate(_plan(output={
            "filters": [{"column": "AESTDAT", "operator": "ne", "valueRef": "AEENDAT"}],
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
        }))
        frame = _execute(project, plan)
        # 三行的起止日期都不同 → ne 全部命中；错误实现在这里也返回 3 行，
        # 因此必须同时断言 eq 分支（错误实现返回 0 行）。
        assert len(frame) == 3

        plan = _validate(_plan(output={
            "filters": [{"column": "AESTDAT", "operator": "eq", "valueRef": "AESTDAT"}],
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
        }))
        frame = _execute(project, plan)
        # 列与自身相等 → 必须是全部 3 行。错误实现比较 "AESTDAT" 字面量，得 0 行。
        assert len(frame) == 3, "valueRef 过滤把数据值与列名比较了（F-2 回归）"


def test_filter_can_reference_a_derived_column() -> None:
    """F-5: validator 先注册 derivations 再校验 filters，执行器必须同序。"""
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        plan = _validate(_plan(output={
            "derivations": [{"name": "DURATION", "operation": "date_diff_days",
                             "refs": ["AEENDAT", "AESTDAT"]}],
            "filters": [{"column": "DURATION", "operator": "gt",
                         "literal": {"type": "number", "value": 5}}],
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
        }))
        # 执行期若仍先 filter 后 derive，这里会抛 "validated field is unavailable"。
        frame = _execute(project, plan)
        assert list(frame["AETERM"]) == ["headache"]


def test_validated_sort_is_always_executable() -> None:
    """F-6: validator 允许的排序列必须真的存在于产物中。"""
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        # 源字段存在但不在输出列中 → validator 必须拒绝（旧实现放行、执行期报错）。
        _rejects(_plan(output={
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
            "sort": [{"column": "AESEV", "direction": "asc"}],
        }), "UNKNOWN_COLUMN")

        # 输出列排序必须可执行。
        plan = _validate(_plan(output={
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
            "sort": [{"column": "AETERM", "direction": "desc"}],
        }))
        frame = _execute(project, plan)
        assert list(frame["AETERM"]) == ["nausea", "headache", "headache"]


def test_review_columns_are_sortable_and_written_once() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        plan = _validate(_plan(output={
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
            "layout": {"appendReviewColumns": True},
            "sort": [{"column": "Flag", "direction": "asc"}],
        }))
        frame = _execute(project, plan)
        assert list(frame.columns) == [
            "AETERM", "Flag", "Update Details", "Review Comments", "Initial/Date",
        ]


def test_dropped_code_value_column_cannot_be_sorted_on() -> None:
    # dropCodeValue 会把列从产物移除；validator 的排序域必须同口径。
    _rejects(_plan(output={
        "columns": [
            {"source": "AETERM", "name": "AETERM", "label": ""},
            {"source": "AESEV", "name": "SEV_CODE_VALUE", "label": "Code Value"},
        ],
        "layout": {"dropCodeValue": True},
        "sort": [{"column": "SEV_CODE_VALUE", "direction": "asc"}],
    }), "UNKNOWN_COLUMN")


def test_status_filter_selects_rows_without_exposing_values() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        plan = _validate(_plan(output={
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
            "layout": {"statusFilter": "已编码"},
        }))
        frame = _execute(project, plan)
        assert len(frame) == 2


def test_aggregation_count_semantics_match_across_grouping() -> None:
    """N-9: 有无 groupBy 不得改变 count 的定义（含/不含空值）。"""
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = _study(root)
        frame = pd.read_csv(project / "data" / "AE.csv")
        frame.loc[0, "AETERM"] = None
        frame.to_csv(project / "data" / "AE.csv", index=False)

        grouped = _execute(project, _validate(_plan(output={
            "groupBy": ["AESEV"],
            "aggregations": [{"name": "N", "operation": "count", "column": "AETERM"}],
            "columns": [
                {"source": "AESEV", "name": "AESEV", "label": ""},
                {"source": "N", "name": "N", "label": ""},
            ],
        })))
        ungrouped = _execute(project, _validate(_plan(output={
            "name": "TOTAL",
            "aggregations": [{"name": "N", "operation": "count", "column": "AETERM"}],
            "columns": [{"source": "N", "name": "N", "label": ""}],
        })))
        # 两条分支都应按"非空计数"报告：3 行里有 1 个空 → 合计 2。
        assert sum(int(value) for value in grouped["N"]) == int(ungrouped["N"].iloc[0]) == 2


def test_aggregation_rejects_type_incompatible_operations() -> None:
    """文本列求和必须报错，绝不能静默产出 pandas 的字符串拼接结果。

    pandas 对字符串 series 的 sum 是拼接（'a'+'b' → 'ab'）。若不拦，交付物里
    会出现一个看起来像"合计"、实际是所有不良事件名首尾相接的单元格——静默
    错误，比抛异常危险得多。
    """
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        for operation in ("sum", "mean"):
            plan = _validate(_plan(output={
                "name": "TOTAL",
                "aggregations": [{"name": "S", "operation": operation, "column": "AETERM"}],
                "columns": [{"source": "S", "name": "S", "label": ""}],
            }))
            try:
                frame = _execute(project, plan)
            except ListingExecutionError:
                continue
            raise AssertionError(
                f"{operation} 在文本列上未被拒绝，产出了 {frame['S'].iloc[0]!r}"
            )


def test_numeric_aggregation_still_works_on_numeric_fields() -> None:
    # 收紧类型校验不得误伤真实数值聚合。
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        schema = {"dm": {"USUBJID", "AGE", "SITEID"}}
        plan = validate_listing_plan({
            "version": 1, "scenario": "rbqm", "outputs": [{
                "name": "AGES", "source": "DM",
                "aggregations": [{"name": "TOTAL", "operation": "sum", "column": "AGE"}],
                "columns": [{"source": "TOTAL", "name": "TOTAL", "label": ""}],
            }],
        }, schema, "rbqm")
        frame = _execute(project, plan)
        assert int(frame["TOTAL"].iloc[0]) == 132


def test_join_qualified_reference_never_silently_picks_a_native_column() -> None:
    """N-8: 左表原生 `DM__AGE` 列不得让 `DM.AGE` 静默错列。"""
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        frame = pd.read_csv(project / "data" / "AE.csv")
        frame["DM__AGE"] = [-1, -2, -3]
        frame.to_csv(project / "data" / "AE.csv", index=False)

        schema = {
            "ae": set(frame.columns),
            "dm": {"USUBJID", "AGE", "SITEID"},
        }
        plan = validate_listing_plan(_plan(output={
            "joins": [{"dataset": "DM", "type": "left",
                       "leftKeys": ["USUBJID"], "rightKeys": ["USUBJID"]}],
            "columns": [{"source": "DM.AGE", "name": "AGE", "label": ""}],
        }), schema, "rbqm")
        result = _execute(project, plan)
        # 必须取到 join 进来的真实年龄，而不是左表那个同名干扰列。
        assert sorted(int(value) for value in result["AGE"]) == [33, 42, 57]


def test_executor_rejects_fields_absent_from_local_data() -> None:
    """schema 与真实数据不一致时必须报错，不能静默产出缺列产物。"""
    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        plan = _validate(_plan(output={
            "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
        }))
        plan["outputs"][0]["columns"][0]["source"] = "NOT_A_COLUMN"
        try:
            _execute(project, plan)
        except ListingExecutionError:
            return
        raise AssertionError("执行器接受了本地数据中不存在的字段")


# ---------------------------------------------------------------------------
# medical 来源确认（F-11 迁移后的等价物）
# ---------------------------------------------------------------------------

CHANGE_BASELINE_SPEC = "New/Modified的信息请标识"
CODING_STATUS_SPEC = "编码页面只呈现Status为已编码的信息"


def test_medical_provenance_requires_review_columns_and_status_filter() -> None:
    """F-11: 旧生成器用 ListingNeedsInput 拦路；现在是结构化计划校验规则。

    旧行为把整轮打回 needs_input，模型看不到"缺什么"；新行为指明缺失的计划字段。
    """
    medical = {**_plan(), "scenario": "medical"}
    _rejects(medical, "MEDICAL_PROVENANCE_REQUIRED", scenario="medical",
             requirement_text=CHANGE_BASELINE_SPEC)
    _rejects(medical, "MEDICAL_PROVENANCE_REQUIRED", scenario="medical",
             requirement_text=CODING_STATUS_SPEC)

    satisfied = {**_plan(output={"layout": {
        "appendReviewColumns": True, "statusFilter": "已编码",
    }}), "scenario": "medical"}
    normalized = _validate(satisfied, "medical",
                           requirement_text=f"{CHANGE_BASELINE_SPEC}\n{CODING_STATUS_SPEC}")
    assert normalized["outputs"][0]["layout"]["appendReviewColumns"] is True
    assert normalized["outputs"][0]["layout"]["statusFilter"] == "已编码"


def test_medical_provenance_rule_does_not_fire_without_the_requirement() -> None:
    # spec 没提这些规则时不得凭空要求复核列——否则每个 medical 计划都被误拒。
    assert _validate({**_plan(), "scenario": "medical"}, "medical",
                     requirement_text="仅列出所有不良事件")["scenario"] == "medical"


# ---------------------------------------------------------------------------
# F-4: 存在性预言机预算
# ---------------------------------------------------------------------------

def test_execute_budget_counts_and_warns_without_blocking(monkeypatch=None) -> None:
    import os

    reset_budget()
    os.environ["EMERALD_LISTING_MAX_EXECUTIONS"] = "3"
    with tempfile.TemporaryDirectory() as directory:
        os.environ["EMERALD_AUDIT_ROOT"] = directory
        try:
            plan = _validate(_plan())
            for expected in (1, 2, 3):
                assert charge_execution(session_id="s1", project="study",
                                        scenario="rbqm", plan=plan) == expected
            assert charge_execution(session_id="s1", project="study",
                                    scenario="rbqm", plan=plan) == 4
            # 另一个会话/项目有独立预算，不被前者拖累。
            assert charge_execution(session_id="s2", project="study",
                                    scenario="rbqm", plan=plan) == 1
            audit = list(Path(directory).glob("listing_ops_*.jsonl"))
            assert audit, "execute 必须留下审计记录"
            body = audit[0].read_text(encoding="utf-8")
            assert '"event":"listing_execute"' in body
            assert '"thresholdExceeded":true' in body
            assert '"allowed":true' in body
        finally:
            os.environ.pop("EMERALD_LISTING_MAX_EXECUTIONS", None)
            os.environ.pop("EMERALD_AUDIT_ROOT", None)
            reset_budget()


def test_audit_record_never_contains_literal_values() -> None:
    """审计里抄下 literal 等于把推断出的数据值写进文件，反而扩大出域面。"""
    import json
    import os

    reset_budget()
    with tempfile.TemporaryDirectory() as directory:
        os.environ["EMERALD_AUDIT_ROOT"] = directory
        try:
            plan = _validate(_plan(output={"filters": [
                {"column": "AETERM", "operator": "eq",
                 "literal": {"type": "string", "value": "SECRET-PATIENT-VALUE"}},
            ]}))
            charge_execution(session_id="s1", project="study", scenario="rbqm", plan=plan)
            body = next(Path(directory).glob("listing_ops_*.jsonl")).read_text(encoding="utf-8")
            assert "SECRET-PATIENT-VALUE" not in body
            record = json.loads(body.splitlines()[0])
            assert record["filterCount"] == 1
            assert len(record["planFingerprint"]) == 16
        finally:
            os.environ.pop("EMERALD_AUDIT_ROOT", None)
            reset_budget()


# ---------------------------------------------------------------------------
# 发布路径（F-7）与工作流收据
# ---------------------------------------------------------------------------

def _publishable_study(root: Path) -> Path:
    """带 spec 文档的项目，可走完整 inspect → validate → execute 链路。"""
    import openpyxl

    project = _study(root)
    (project / "doc").mkdir()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Listing要求"
    sheet.append(["序号", "要求"])
    sheet.append([1, "列出所有不良事件"])
    workbook.save(project / "doc" / "Listing要求.xlsx")
    workbook.close()
    return project


WORKFLOW_PLAN = {
    "version": 1, "scenario": "rbqm", "outputs": [{
        "name": "AE", "source": "AE",
        "columns": [{"source": "AETERM", "name": "AETERM", "label": "术语"}],
    }],
}


def _publish(root: Path, **kwargs):
    return execute_listing_plan_workflow(
        local_data_root=str(root), project="study", scenario="rbqm",
        plan=WORKFLOW_PLAN, session_id="publish-session", **kwargs,
    )


def test_workflow_publishes_relative_artifacts_without_absolute_paths() -> None:
    reset_budget()
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = _publishable_study(root)
        receipt = _publish(root)
        assert receipt["status"] == "completed"
        assert receipt["dataClass"] == "REAL"
        assert receipt["artifacts"][0]["id"] == ".clinical-listing/output/rbqm/RBQM_001_AE.xlsx"
        # 收据不得含绝对路径或数据值。
        assert str(root) not in repr(receipt)
        assert (project / ".clinical-listing" / "output" / "rbqm").is_dir()
        # staging 与 catalog 临时目录必须清理干净。
        assert not list((project / ".clinical-listing").glob("staging/*"))
        assert not list((project / ".clinical-listing").glob(".listing-catalog-*"))
    reset_budget()


def test_republish_replaces_scenario_output_without_stale_artifacts() -> None:
    reset_budget()
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = _publishable_study(root)
        _publish(root)
        output_dir = project / ".clinical-listing" / "output" / "rbqm"
        stale = output_dir / "RBQM_999_STALE.xlsx"
        stale.write_bytes(b"stale")
        second = _publish(root)
        assert second["status"] == "completed"
        # 重新发布是整目录替换：上一轮的孤儿产物不得残留。
        assert not stale.exists()
        assert [path.name for path in sorted(output_dir.iterdir())] == ["RBQM_001_AE.xlsx"]
        # 备份目录也不得留下。
        assert not list((project / ".clinical-listing" / "output").glob(".rbqm-backup-*"))
    reset_budget()


def test_publish_failure_restores_the_previous_listing() -> None:
    """F-7: output→backup 后若 staging→output 失败，旧产物必须被放回原位。

    此前新发布路径丢了回滚分支：异常上抛、旧产物留在孤儿 backup 目录，
    用户既拿不到新产物也丢了上一版。
    """
    reset_budget()
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        project = _publishable_study(root)
        _publish(root)
        output_dir = project / ".clinical-listing" / "output" / "rbqm"
        marker = output_dir / "RBQM_001_AE.xlsx"
        previous = marker.read_bytes()

        real_rename = Path.rename
        failed_once = []

        def failing_rename(self, target):
            # 只让第一次"staging → output"失败；回滚那次同名改名必须放过，
            # 否则测的就不是回滚而是"回滚也坏了"。
            if Path(target).name == "rbqm" and not failed_once:
                failed_once.append(True)
                raise OSError("publish interrupted")
            return real_rename(self, target)

        with mock.patch.object(Path, "rename", failing_rename):
            try:
                _publish(root)
            except ListingWorkflowError:
                pass
            else:
                raise AssertionError("发布失败必须上抛工作流错误")
        assert failed_once, "注入的发布失败没有触发"

        assert marker.exists(), "回滚未恢复上一版产物"
        assert marker.read_bytes() == previous
        assert not list((project / ".clinical-listing" / "output").glob(".rbqm-backup-*"))
    reset_budget()


def test_published_listing_stays_completed_when_backup_cleanup_fails() -> None:
    """备份清理失败不影响本次发布结果，只降级为收据 warning。"""
    reset_budget()
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        _publishable_study(root)
        _publish(root)
        real_rmtree = shutil.rmtree

        def failing_rmtree(path, *args, **kwargs):
            # 只让备份清理失败。catalog 与 staging 的清理必须照常进行，
            # 否则测出来的是"临时目录清不掉"而不是"备份清不掉"。
            if Path(path).name.startswith(".rbqm-backup-"):
                raise OSError("locked")
            return real_rmtree(path, *args, **kwargs)

        with mock.patch("security.listing_workflow.shutil.rmtree", failing_rmtree):
            receipt = _publish(root)
        assert receipt["status"] == "completed"
        assert receipt["warnings"] == ["a previous listing backup could not be removed"]
    reset_budget()


def test_workflow_returns_structured_rejection_for_invalid_plans() -> None:
    """非法计划必须是 invalid 收据（带 code/path），不是异常。"""
    reset_budget()
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        _publishable_study(root)
        receipt = execute_listing_plan_workflow(
            local_data_root=str(root), project="study", scenario="rbqm",
            plan={"version": 1, "scenario": "rbqm", "outputs": [{
                "name": "AE", "source": "GHOST",
                "columns": [{"source": "AETERM", "name": "AETERM", "label": ""}],
            }]},
            session_id="publish-session",
        )
        assert receipt["status"] == "invalid"
        assert receipt["code"] == "UNKNOWN_DATASET"
        assert receipt["dataClass"] == "METADATA_ONLY"
        # 被拒计划不得留下任何产物目录。
        assert not (root / "study" / ".clinical-listing" / "output").exists()
    reset_budget()


def test_exhausted_budget_is_audited_without_blocking_execution() -> None:
    """预算阈值用于持续审计，不能阻断本地交付。"""
    reset_budget()
    os.environ["EMERALD_LISTING_MAX_EXECUTIONS"] = "1"
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        os.environ["EMERALD_AUDIT_ROOT"] = str(root / "audit")
        try:
            _publishable_study(root)
            assert _publish(root)["status"] == "completed"
            receipt = _publish(root)
            assert receipt["status"] == "completed"
            records = []
            for path in (root / "audit").glob("listing_ops*.jsonl"):
                records.extend(json.loads(line) for line in path.read_text(encoding="utf-8").splitlines())
            assert records[-1]["sequence"] == 2
            assert records[-1]["thresholdExceeded"] is True
            assert records[-1]["allowed"] is True
        finally:
            os.environ.pop("EMERALD_LISTING_MAX_EXECUTIONS", None)
            os.environ.pop("EMERALD_AUDIT_ROOT", None)
            reset_budget()


def test_report_rejects_duplicate_output_names() -> None:
    first = _plan(output={"name": "AE_REPORT"})["outputs"][0]
    second = {**first, "name": "ae_report"}
    _rejects({"version": 1, "scenario": "report", "outputs": [first, second]}, "DUPLICATE_OUTPUT_NAME", "report")


def test_report_rejects_names_colliding_after_excel_truncation() -> None:
    first = _plan(output={"name": "A" * 31 + "_ONE"})["outputs"][0]
    second = {**first, "name": "A" * 31 + "_TWO"}
    _rejects({"version": 1, "scenario": "report", "outputs": [first, second]}, "DUPLICATE_SHEET_NAME", "report")


def test_report_rejects_reserved_contents_sheet_name() -> None:
    _rejects({**_plan(output={"name": "Contents"}), "scenario": "report"}, "RESERVED_SHEET_NAME", "report")


def test_report_rejects_duplicate_output_column_names() -> None:
    _rejects({**_plan(output={"columns": [
        {"source": "AETERM", "name": "TERM", "label": ""},
        {"source": "AESEV", "name": "term", "label": ""},
    ]}), "scenario": "report"}, "DUPLICATE_COLUMN_NAME", "report")


def test_report_rejects_review_column_and_aggregation_alias_collisions() -> None:
    _rejects({**_plan(output={
        "aggregations": [{"name": "Flag", "operation": "count", "column": "AETERM"}],
        "columns": [{"source": "Flag", "name": "Flag", "label": ""}],
        "layout": {"appendReviewColumns": True},
    }), "scenario": "report"}, "DUPLICATE_REVIEW_COLUMN", "report")


def test_report_defaults_to_contents_navigation() -> None:
    normalized = _validate({**_plan(), "scenario": "report"}, "report")
    assert normalized["outputs"][0]["layout"]["includeContents"] is True


def main() -> int:
    tests = [value for key, value in sorted(globals().items()) if key.startswith("test_")]
    failures = 0
    for test in tests:
        try:
            test()
            print(f"PASS {test.__name__}")
        except Exception as error:
            failures += 1
            print(f"FAIL {test.__name__}: {error}")
    print(f"RESULT {len(tests) - failures}/{len(tests)}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())


def test_report_layout_writes_contents_navigation_and_filters() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        project = _study(Path(directory))
        plan = validate_listing_plan({
            "version": 1,
            "scenario": "report",
            "outputs": [{
                "name": "AE_REPORT",
                "source": "AE",
                "columns": [
                    {"source": "USUBJID", "name": "Subject", "label": "Subject"},
                    {"source": "AETERM", "name": "Term", "label": "Event Term"},
                ],
                "layout": {"includeContents": True, "freezeColumns": 1},
            }],
        }, SCHEMA, "report")
        result = execute_listing_plan(str(project), str(project / "out"), plan)
        path = Path(result["artifacts"][0]["path"])
        workbook = openpyxl.load_workbook(io.BytesIO(path.read_bytes()))
        try:
            assert workbook.sheetnames == ["Contents", "AE_REPORT"]
            contents = workbook["Contents"]
            report = workbook["AE_REPORT"]
            assert [cell.value for cell in contents[1]][:8] == [
                "Listing Seq.", "Listing Name(Please Click Down)", "Data Set Label",
                "Report Description", "New/Modified ?", "Total Row Count", "New Count", "Modified Count",
            ]
            assert report["A1"].value == "Go back"
            assert report["A1"].hyperlink.target == "#'Contents'!A1"
            assert [cell.value for cell in report[2]] == ["Subject", "Event Term"]
            assert report.auto_filter.ref == "A2:B5"
            assert report.freeze_panes == "B3"
            assert report.column_dimensions["A"].width >= 10
        finally:
            workbook.close()
