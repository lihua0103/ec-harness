"""per-project 配置词表（security/project_profile.py）的单元测试。

覆盖：默认值与原硬编码一致、.clinical-listing/listing-profile.json 覆写、
EMERALD_LISTING_PROFILE 环境变量覆写、非法配置逐字段回退并给出 warning、
自定义词表在 classify/场景推断/validate/execute 全链路生效。
"""
from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from security import listing_inspector, spec_parser
from security.listing_executor import execute_listing_plan
from security.listing_plan import REVIEW_COLUMNS, ListingPlanError, validate_listing_plan
from security.project_profile import (
    PROFILE_INVALID_WARNING,
    ProjectProfile,
    load_project_profile,
)


def _write_profile(project: Path, payload: dict) -> None:
    directory = project / ".clinical-listing"
    directory.mkdir(parents=True, exist_ok=True)
    (directory / "listing-profile.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )


def _normalized_plan(**layout) -> dict:
    base_layout = {
        "freezeColumns": 0, "freezeRows": 1, "includeContents": True, "toc": False,
        "dropCodeValue": False, "titleLanguage": "", "appendReviewColumns": True,
        "statusFilter": "", "unsupportedRequirements": [],
    }
    base_layout.update(layout)
    return {
        "version": 1, "scenario": "report", "toc": False, "assumptions": [],
        "outputs": [{
            "name": "DMList", "source": "dm", "joins": [],
            "columns": [{"source": "USUBJID", "name": "USUBJID", "label": ""}],
            "filters": [], "derivations": [], "groupBy": [], "aggregations": [],
            "sort": [], "layout": base_layout,
        }],
    }


def test_default_profile_matches_legacy_hardcoding() -> None:
    profile = ProjectProfile()
    assert profile.spec_directory == "doc"
    assert profile.default_scenario == "report"
    assert profile.scenario_keywords["report"] == ("report", "status", "rt01")
    assert profile.scenario_keywords["rbqm"] == ("rbqm", "quality", "kri", "test_final")
    assert "crviewer" in profile.report_support_keywords
    assert "test_final" in profile.spec_keywords
    assert profile.review_columns == REVIEW_COLUMNS
    assert profile.contents_sheet_name == "Contents"
    assert profile.status_column_name == "status"
    assert profile.warnings == ()


def test_missing_profile_file_yields_defaults() -> None:
    with tempfile.TemporaryDirectory() as directory:
        assert load_project_profile(Path(directory)) == ProjectProfile()


def test_invalid_profile_json_falls_back_with_warning() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory)
        _write_profile(project, {})
        (project / ".clinical-listing" / "listing-profile.json").write_text("not json", encoding="utf-8")
        profile = load_project_profile(project)
        assert profile.review_columns == REVIEW_COLUMNS
        assert profile.warnings == (PROFILE_INVALID_WARNING,)


def test_invalid_field_falls_back_per_field() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory)
        _write_profile(project, {
            "reviewColumns": {"=evil": "=cmd"},
            "contentsSheetName": "TOC",
        })
        profile = load_project_profile(project)
        assert profile.review_columns == REVIEW_COLUMNS
        assert profile.contents_sheet_name == "TOC"
        assert profile.warnings == (PROFILE_INVALID_WARNING,)


def test_env_var_points_to_explicit_profile(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "project"
        project.mkdir()
        profile_file = Path(directory) / "central.json"
        profile_file.write_text(json.dumps({"specDirectory": "specs"}), encoding="utf-8")
        monkeypatch.setenv("EMERALD_LISTING_PROFILE", str(profile_file))
        assert load_project_profile(project).spec_directory == "specs"


def test_custom_keywords_drive_classification_and_scenario() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory)
        _write_profile(project, {
            "scenarioKeywords": {"rbqm": ["qrm"]},
            "specKeywords": ["dvp"],
            "reportSupportKeywords": ["extract"],
            "defaultScenario": "manual",
        })
        profile = load_project_profile(project)
        assert spec_parser.classify_spec_document(Path("study_dvp_v1.xlsx"), profile) == "specification"
        assert spec_parser.classify_spec_document(Path("site_extract.xlsx"), profile) == "report_support_data"
        assert spec_parser.classify_spec_document(Path("crviewer.xls"), profile) == "requirement_note"
        assert listing_inspector._infer_scenario([Path("QRM_plan.xlsx")], profile) == ("rbqm", 1.0, ["rbqm"])
        # SCENARIO_AMBIGUOUS 拆除后：无任何关键词命中时回传默认场景作为唯一
        # 候选（E2E GQ1005-301 实测 candidates=["report"]），不再是空候选硬拒。
        assert listing_inspector._infer_scenario([Path("notes.xlsx")], profile) == ("manual", 0.0, ["manual"])


def test_custom_spec_directory_is_discovered() -> None:
    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory)
        (project / "specs").mkdir()
        (project / "specs" / "anything.xlsx").write_bytes(b"placeholder")
        _write_profile(project, {"specDirectory": "specs"})
        profile = load_project_profile(project)
        found = spec_parser.find_spec_documents(project, profile)
        assert [path.name for path in found] == ["anything.xlsx"]
        assert spec_parser.classify_spec_document(found[0], profile) == "specification"
        assert spec_parser.find_spec_documents(project) == []


def test_validate_plan_uses_custom_review_columns() -> None:
    schema = {"dm": {"USUBJID"}}
    plan = {
        "version": 1, "scenario": "report",
        "outputs": [{
            "name": "DMList", "source": "dm",
            "columns": [{"source": "USUBJID", "name": "QC", "label": ""}],
            "layout": {"appendReviewColumns": True},
        }],
    }
    validate_listing_plan(plan, schema, "report", review_columns={"Flag": "Flag"})
    try:
        validate_listing_plan(plan, schema, "report", review_columns={"QC": "QC"})
    except ListingPlanError as exc:
        assert exc.code == "DUPLICATE_REVIEW_COLUMN"
    else:
        raise AssertionError("custom review column collision must be rejected")


def test_execute_uses_custom_contents_and_review_columns() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "project"
        output = Path(directory) / "out"
        project.mkdir()
        (project / "dm.csv").write_text("USUBJID\nS1\n", encoding="utf-8")
        _write_profile(project, {
            "reviewColumns": {"QC": "QC"},
            "contentsHeaders": ["Seq", "Name"],
            "contentsSheetName": "TOC",
        })
        result = execute_listing_plan(str(project), str(output), _normalized_plan())
        workbook = openpyxl.load_workbook(result["artifact"]["path"])
        try:
            assert "TOC" in workbook.sheetnames
            contents = workbook["TOC"]
            assert [contents.cell(1, 1).value, contents.cell(1, 2).value] == ["Seq", "Name"]
            listing = workbook["DMList"]
            headers = [cell.value for cell in listing[2]]
            assert headers == ["USUBJID", "QC"]
        finally:
            workbook.close()


def test_execute_defaults_preserve_legacy_contents_template() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as directory:
        project = Path(directory) / "project"
        output = Path(directory) / "out"
        project.mkdir()
        (project / "dm.csv").write_text("USUBJID\nS1\n", encoding="utf-8")
        result = execute_listing_plan(str(project), str(output), _normalized_plan())
        workbook = openpyxl.load_workbook(result["artifact"]["path"])
        try:
            assert "Contents" in workbook.sheetnames
            contents = workbook["Contents"]
            assert contents.cell(1, 2).value == "Listing Name(Please Click Down)"
            listing = workbook["DMList"]
            headers = [cell.value for cell in listing[2]]
            # 表头呈现的是 REVIEW_COLUMNS 的 value（表头标签），真实交付样例的
            # 末列即 "Initial/Date"；key（Initial_Date）只是产物内部列名。
            assert headers == ["USUBJID", *REVIEW_COLUMNS.values()]
        finally:
            workbook.close()
